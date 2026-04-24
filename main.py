"""dashboard4dx — local-only integrated development dashboard.

All processing runs in-process; no data leaves the machine.
Single-file by design: every loader, KPI, and UI helper lives here.
"""
from __future__ import annotations

import base64
import hashlib
import io
import json
import logging
import math
import re
import time
import traceback
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Callable, Optional

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# =============================================================================
# Persistent input store
# =============================================================================
SCRIPT_DIR = Path(__file__).resolve().parent
INPUT_DIR = SCRIPT_DIR / "input"
LOG_DIR = SCRIPT_DIR / "log"

# Cap the on-screen error block at this many characters so it fits a typical
# laptop screen (≈50 lines × 60 cols ≈ 3000 chars) without scrolling. The
# log file always retains the full untruncated entry.
_MAX_ERROR_DETAIL_CHARS = 3000


def _get_log_file_path() -> Path:
    """Return the per-Streamlit-session log path. Filename is fixed at first
    use (`log_YYYYMMDDhhmmss.log`) and reused across reruns within the same
    session so a single session's events stay in one file."""
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    if "_log_file_path" not in st.session_state:
        st.session_state._log_file_path = LOG_DIR / (
            f"log_{datetime.now().strftime('%Y%m%d%H%M%S')}.log"
        )
    return st.session_state._log_file_path


def _get_logger() -> logging.Logger:
    """Return a logger whose FileHandler points at the current session's log
    file. Idempotent across Streamlit reruns — old handlers are replaced if
    the log path changes."""
    log_path = _get_log_file_path()
    logger = logging.getLogger("dashboard4dx")
    has_correct_handler = any(
        isinstance(h, logging.FileHandler)
        and Path(getattr(h, "baseFilename", "")) == log_path
        for h in logger.handlers
    )
    if not has_correct_handler:
        for h in list(logger.handlers):
            if isinstance(h, logging.FileHandler):
                logger.removeHandler(h)
                try:
                    h.close()
                except Exception:
                    pass
        h = logging.FileHandler(log_path, encoding="utf-8")
        h.setFormatter(logging.Formatter(
            "%(asctime)s [%(levelname)s] %(message)s"
        ))
        logger.addHandler(h)
        logger.setLevel(logging.INFO)
        logger.info("=" * 60)
        logger.info("Session opened — dashboard4dx")
        logger.info("=" * 60)
    return logger


def log_error(category: str, summary: str, *,
              exc: Optional[BaseException] = None,
              context: Optional[dict] = None) -> str:
    """Append a structured error entry to the session log file and return a
    compact (≤_MAX_ERROR_DETAIL_CHARS) string suitable for inline display.

    The log file gets the **full** entry (no truncation); the returned string
    is the shortest readable form that still includes the reason, context,
    and either the full or truncated traceback so a user can paste it into
    a bug report without scrolling.
    """
    logger = _get_logger()
    sep = "─" * 60

    head = [sep, f"[{category}]  {summary}", sep]
    ctx_lines: list[str] = []
    if context:
        ctx_lines.append("Context:")
        for k, v in context.items():
            v_str = str(v)
            if len(v_str) > 250:
                v_str = v_str[:250] + "…"
            ctx_lines.append(f"  • {k}: {v_str}")
        ctx_lines.append("")

    tb_lines: list[str] = []
    if exc is not None:
        tb_lines.append("Stack trace:")
        tb_text = "".join(traceback.format_exception(
            type(exc), exc, exc.__traceback__))
        for tb_line in tb_text.rstrip("\n").split("\n"):
            tb_lines.append(f"  {tb_line}")

    full_body = "\n".join(head + ctx_lines + tb_lines)
    logger.error(full_body)

    # Trim for screen if needed: always keep the head and context, then
    # collapse the traceback's middle frames as needed to fit.
    fixed_part = "\n".join(head + ctx_lines)
    budget = _MAX_ERROR_DETAIL_CHARS - len(fixed_part) - 80
    if not tb_lines or budget <= 0:
        return full_body[:_MAX_ERROR_DETAIL_CHARS]
    if sum(len(line) + 1 for line in tb_lines) <= budget:
        return full_body
    # Keep first 6 and last 6 lines of traceback, drop middle.
    keep_n = 6
    if len(tb_lines) > 2 * keep_n + 2:
        kept = (tb_lines[: keep_n + 1]
                + [f"  …({len(tb_lines) - 2 * keep_n - 1} frames omitted; "
                   "see full log file)"]
                + tb_lines[-keep_n:])
    else:
        kept = tb_lines
    return "\n".join(head + ctx_lines + kept)

# Matches `..._YYYYMMDDhhmmss.<ext>` in a filename.
_FILENAME_TS_RE = re.compile(r"_(\d{8})(\d{6})\.[A-Za-z0-9]+$")


def _ensure_input_dir() -> Path:
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    return INPUT_DIR


def _snapshot_date_from_filename(filename: str) -> Optional[date]:
    """Extract YYYY-MM-DD from a *_YYYYMMDDhhmmss.<ext>-style filename."""
    m = _FILENAME_TS_RE.search(filename)
    if not m:
        return None
    try:
        return datetime.strptime(m.group(1), "%Y%m%d").date()
    except ValueError:
        return None


def save_uploaded_bytes(slot: str, filename: str, data: bytes) -> Path:
    """Persist `data` to `input/<YYYY-MM-DD>/<slot>/<filename>`.

    Folder date is derived from the filename's `_YYYYMMDDhhmmss` suffix
    (the snapshot moment); if absent, today's date is used. Same filename
    overwrites — that is by design so re-importing the same snapshot is
    idempotent, while different timestamps in the filename produce distinct
    history entries used downstream for trend analysis.
    """
    snap = _snapshot_date_from_filename(filename) or date.today()
    folder = _ensure_input_dir() / snap.isoformat() / slot
    folder.mkdir(parents=True, exist_ok=True)
    target = folder / filename
    target.write_bytes(data)
    return target


def find_latest_for_slot(slot: str) -> Optional[Path]:
    """Return the newest file (by mtime) under `input/*/<slot>/`, or None."""
    if not INPUT_DIR.exists():
        return None
    candidates: list[Path] = []
    for date_dir in INPUT_DIR.iterdir():
        if not date_dir.is_dir():
            continue
        slot_dir = date_dir / slot
        if not slot_dir.is_dir():
            continue
        for f in slot_dir.iterdir():
            if f.is_file():
                candidates.append(f)
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


# ----- Manual design-pages persistence ---------------------------------------
# Stored at the root of input/ (next to the daily snapshot folders) since it's
# a single piece of slowly-changing user state, not a dated snapshot.
DESIGN_PAGES_FILE = INPUT_DIR / "design_pages.json"
USER_SETTINGS_FILE = INPUT_DIR / "user_settings.json"

# Keys in session_state whose value is user-facing and should survive
# an app restart. Everything else (dfs cache, signatures, error state,
# …) is deliberately volatile.
_PERSISTENT_SETTING_KEYS: tuple[str, ...] = (
    "test_density_threshold",
    "incident_rate_threshold_pct",
    "wbs_attach_after_dup",
    "lang",
)


def load_user_settings() -> dict:
    """Best-effort read of the persisted user-settings JSON. Returns ``{}``
    on missing/invalid file so a fresh install still boots with defaults."""
    if not USER_SETTINGS_FILE.exists():
        return {}
    try:
        with USER_SETTINGS_FILE.open(encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            return data
    except Exception:
        pass
    return {}


def save_user_settings() -> None:
    """Persist the currently-set user-facing settings to disk. Called by
    widget ``on_change`` callbacks so each adjustment is durable; silent
    on IO errors since settings persistence is best-effort."""
    try:
        payload = {
            k: st.session_state[k]
            for k in _PERSISTENT_SETTING_KEYS
            if k in st.session_state
        }
        _ensure_input_dir()
        with USER_SETTINGS_FILE.open("w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def load_design_pages() -> dict[str, int]:
    """Load saved design page counts. Returns {} on missing/invalid file.

    Keys are routed through the same `_normalize_fid` helper that every
    other source uses so a pre-change JSON with lowercase / mixed-case
    entries merges cleanly against the (now uppercased) master FIDs.
    When two keys collapse to the same canonical FID the later value
    wins — mirrors how a user with duplicates would expect merge-last
    semantics from an edit flow.
    """
    if not DESIGN_PAGES_FILE.exists():
        return {}
    try:
        with DESIGN_PAGES_FILE.open(encoding="utf-8") as f:
            data = json.load(f)
    except (OSError, json.JSONDecodeError):
        return {}
    out: dict[str, int] = {}
    if isinstance(data, dict):
        for k, v in data.items():
            try:
                if v is None or v == "":
                    continue
                canon = _normalize_fid(k) or str(k)
                out[canon] = int(v)
            except (TypeError, ValueError):
                continue
    return out


def save_design_pages(
    visible_ids: set[str], edited: dict[str, Optional[int]]
) -> Path:
    """Persist design pages to disk.

    For Function IDs currently in the master (`visible_ids`), the editor is
    authoritative: a numeric value sets/updates it, a blank deletes it. For
    IDs NOT in the current master (orphans from earlier masters), existing
    values are preserved on disk so they reappear if the ID returns later.
    """
    _ensure_input_dir()
    existing = load_design_pages()
    for fid in visible_ids:
        v = edited.get(fid)
        if v is None:
            existing.pop(fid, None)
        else:
            try:
                existing[fid] = int(v)
            except (TypeError, ValueError):
                existing.pop(fid, None)
    with DESIGN_PAGES_FILE.open("w", encoding="utf-8") as f:
        json.dump(existing, f, indent=2, ensure_ascii=False, sort_keys=True)
    return DESIGN_PAGES_FILE


def list_history_for_slot(slot: str) -> list[Path]:
    """Return all stored files for `slot`, newest first by mtime."""
    if not INPUT_DIR.exists():
        return []
    out: list[Path] = []
    for date_dir in INPUT_DIR.iterdir():
        if not date_dir.is_dir():
            continue
        slot_dir = date_dir / slot
        if not slot_dir.is_dir():
            continue
        out.extend(p for p in slot_dir.iterdir() if p.is_file())
    out.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return out


def load_all_snapshots_for_slot(
    slot: str, loader: Callable[[bytes], pd.DataFrame]
) -> list[tuple[date, Path, pd.DataFrame]]:
    """Load every saved file for `slot`, return [(snapshot_date, path, df), ...]
    sorted ascending by snapshot date.

    The snapshot date prefers the `_YYYYMMDDhhmmss` stamp in the filename; if
    that's absent (e.g. master/wbs/defects exports), the parent folder name
    is used. Files that fail to parse are silently skipped.
    """
    if not INPUT_DIR.exists():
        return []
    out: list[tuple[date, Path, pd.DataFrame]] = []
    for date_dir in INPUT_DIR.iterdir():
        if not date_dir.is_dir():
            continue
        try:
            folder_date = date.fromisoformat(date_dir.name)
        except ValueError:
            continue
        slot_dir = date_dir / slot
        if not slot_dir.is_dir():
            continue
        for f in slot_dir.iterdir():
            if not f.is_file():
                continue
            snap = _snapshot_date_from_filename(f.name) or folder_date
            try:
                df = loader(f.read_bytes())
            except Exception:
                continue
            out.append((snap, f, df))
    out.sort(key=lambda x: x[0])
    return out


def delete_snapshot_file(path: Path) -> bool:
    """Delete a single saved snapshot and prune now-empty parent dirs.

    Removing a Code/Test snapshot also removes that point from the trend
    charts on the next render, since the trend builders rebuild from
    `load_all_snapshots_for_slot()` each time. Returns True on success.
    The path is bounded to INPUT_DIR for safety.
    """
    if not path.exists() or not path.is_file():
        return False
    try:
        path.resolve().relative_to(INPUT_DIR.resolve())
    except ValueError:
        return False
    path.unlink()
    # Prune empty <slot> dir, then empty <date> dir. Stop at INPUT_DIR.
    for parent in (path.parent, path.parent.parent):
        try:
            if (parent != INPUT_DIR and parent.is_dir()
                    and not any(parent.iterdir())):
                parent.rmdir()
        except OSError:
            break
    return True


def storage_summary_for_slot(slot: str) -> dict:
    """Aggregate counts/sizes/last-modified for a slot's stored files."""
    files = list_history_for_slot(slot)
    total = sum(f.stat().st_size for f in files)
    last = max((f.stat().st_mtime for f in files), default=0.0)
    return {
        "count": len(files),
        "size": total,
        "last": datetime.fromtimestamp(last) if last else None,
        "files": files,
    }


def _human_size(n: float) -> str:
    if n < 1024:
        return f"{int(n)} B"
    for unit in ("KB", "MB", "GB"):
        n /= 1024
        if n < 1024:
            return f"{n:.1f} {unit}"
    return f"{n / 1024:.1f} TB"

# =============================================================================
# Constants
# =============================================================================
WBS_SHEET = "メイン"
WBS_DATA_START_ROW = 16
WBS_FUNC_ID_COLS = ("E", "F", "G", "H", "I")  # scan range for 機能ID
# Phase date anchors. Real WBS files write row dates as 月/日 only (no year)
# and put the absolute phase start/end in these merged row-6 cells (as
# 年/月/日) so the year of each per-task 月/日 can be resolved.
WBS_PHASE_START_CELL = ("J", 6)  # merged J6:L6
WBS_PHASE_END_CELL = ("N", 6)    # merged N6:O6
# Sub-task marker column. A data row (row 16+) without its own 機能ID is
# treated as a sub-task of the most recent valid parent only when this
# column contains this mark character. Any other row without 機能ID is
# skipped entirely.
WBS_SUBTASK_MARK_COL = "L"
WBS_SUBTASK_MARK = "●"

MASTER_SHEET = "機能一覧"
MASTER_FID_COL = "F"
MASTER_NAME_COL = "G"

CODE_SHEET = "機能ID別サマリ"

DEFECT_TRACKER_FILTER = "不具合管理"

# Function ID format (per real data): 1–10 ASCII letters followed by 1–10
# ASCII digits, e.g. SYM1010 / AD44020 / F001 / AUTH001.
_FID_BARE_RE = re.compile(r"^[A-Za-z]{1,10}\d{1,10}$")
# Labeled form: "機能ID：XXXX" / "機能ID:XXXX". Capture liberally and validate
# against _FID_BARE_RE afterwards.
_FID_LABELED_RE = re.compile(r"機能ID\s*[：:]\s*(\S+)")
# "FID<sep>name" — a bare FID followed by a separator then trailing title.
# Accepted separators (after NFKC normalises full-width): colon (:), hyphen,
# or any whitespace (half-width ASCII space, tab, or full-width U+3000 —
# Python's \s is Unicode-aware so U+3000 matches without special-casing).
# Captures just the FID portion.
_FID_PREFIX_RE = re.compile(r"^([A-Za-z]{1,10}\d{1,10})[\s:\-]")


# =============================================================================
# Helpers
# =============================================================================
def _col_to_idx(letter: str) -> int:
    """Excel column letter -> 1-based index. Supports A..ZZ."""
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _normalize_fid(value) -> Optional[str]:
    """Extract a Function ID from a cell value.

    Accepts (after NFKC-normalising full-width → half-width):
      • '機能ID：XXXX' / '機能ID:XXXX'               (labeled)
      • 'XXXX<sep>何かの機能名' where <sep> is a colon, hyphen, or
        any whitespace (half-width, tab, or full-width space)
      • bare 'XXXX'
    where XXXX is 1–10 ASCII letters followed by 1–10 ASCII digits
    (e.g. SYM1010 / AD44020 / F001 / AUTH001 / ADM01010).

    Returns the extracted ID **uppercased** so cross-source joins don't
    split AUTH001 / auth001 / Auth001 into three separate features.
    Returns None for empty or non-ID-shaped strings.
    """
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    # NFKC turns ＳＹＭ１０１０ → SYM1010, and full-width colon ： → :,
    # so downstream regexes only need to match the half-width form.
    s = unicodedata.normalize("NFKC", s)

    m = _FID_LABELED_RE.search(s)
    if m:
        cand = m.group(1).strip()
        # Trim trailing punctuation that may follow the ID in free-text cells.
        cand = cand.rstrip("、。,.;:")
        if _FID_BARE_RE.match(cand):
            return cand.upper()
        # Fall through: cand may be "XXXX:title" (labeled + titled).
        m2 = _FID_PREFIX_RE.match(cand)
        return m2.group(1).upper() if m2 else None

    # "XXXX<sep>title" (no '機能ID' label) — separator is colon / hyphen /
    # whitespace per _FID_PREFIX_RE.
    m_prefix = _FID_PREFIX_RE.match(s)
    if m_prefix:
        return m_prefix.group(1).upper()

    return s.upper() if _FID_BARE_RE.match(s) else None


def _to_date(value) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return pd.to_datetime(value, errors="coerce").date()
    except Exception:
        return None


def _parse_us_date(value) -> Optional[date]:
    """Parse MM/DD/YYYY explicitly (the format used by the defect tracker
    export). Falls back to the generic parser when the format doesn't match."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    s = s.split()[0]  # tolerate trailing time
    try:
        return datetime.strptime(s, "%m/%d/%Y").date()
    except ValueError:
        return _to_date(s)


def _to_pydate(v) -> Optional[date]:
    """Coerce a date-ish value (date / datetime / pd.Timestamp / NaT / None /
    string) to a plain `datetime.date`, or None if it can't be parsed."""
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, pd.Timestamp):
        return v.date()
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        ts = pd.to_datetime(v, errors="coerce")
        return None if pd.isna(ts) else ts.date()
    except Exception:
        return None


# =============================================================================
# Cute B&W dinosaur icons — pixel-art SVGs generated from ASCII grids
# =============================================================================
# Each value is a multi-line ASCII grid. 'X' is a filled pixel, anything else
# (including '.') is empty. The grids are intentionally small (10–18 cells
# wide) so each icon stays under ~1 KB and renders crisply at any size via
# `shape-rendering="crispEdges"`.
DINO_GRIDS: dict[str, str] = {
    # T-Rex: chunky upright pose, big head, small arms, thick tail.
    # Reused for the favicon AND for the page title.
    "trex": """
.........XXXXX
.........X.XXX
.........XXXXX
.........XXXX.
.........X....
XX......XXXXX.
XXX....XXXXXX.
.XX...XXXXXXX.
..XXXXXXXXXX..
..X.XXXXXX....
..X..XX.XX....
.....X..X.....
""",
    # Brontosaurus: long neck arching up, round body, long tail.
    "bronto": """
.........XX....
.........XXX...
.........XX....
.........XX....
.........XX....
.........XX....
.....XXXXXX....
..XXXXXXXXXX...
.XXXXXXXXXXXXX.
XXXXXXXXXXXXXXX
.X..X.....X..X.
.X..X.....X..X.
""",
    # Triceratops: 3 horns plus bony frill, stocky body.
    "trike": """
............X..
.....X......X..
....XX....XXX..
...XXXXXXXXXX..
..XXXXXXXXXXX..
.XXXXXXXXXXXX..
XXXXXXXXXXXX...
.X..X.X..X.....
.X..X.X..X.....
""",
    # Stegosaurus: row of plates along back.
    "stego": """
.....X..X..X..X
.....X..X..X..X
....XX..X..X..X
.XXXXXXXXXXXXXX
XXXXXXXXXXXXXXX
XXXXXXXXXXXXXXX
.X..X..X.X..X..
.X..X..X.X..X..
""",
    # Pterodactyl: wings spread.
    "ptero": """
.X.............X
XXX...........XX
XXXXX.......XXXX
.XXXXXX...XXXXX.
..XXXXXXXXXXX...
....XXXXXXX.....
......XXX.......
......X.XXXXXXXX
......X.........
""",
    # Velociraptor: sleek leaning-forward pose.
    "raptor": """
............XXX.
............XXXX
............X.XX
...........XXXXX
..........XXXX..
.XXXXXXXXXXXX...
XXXXXXXXXXX.....
.X.X.XXX........
.X.X..X.........
""",
    # Plesiosaurus: long neck on swimming body, flippers below.
    "plesio": """
............XXX
............X.X
............XXX
............XX.
............XX.
............XX.
.........XXXX..
....XXXXXXXX...
.XXXXXXXXXXXX..
X.X.X......X.X.
""",
    # Ankylosaurus: club tail, low chunky body.
    "anky": """
.....X..X..X.......
....XXXXXXXXX.XXX..
.XXXXXXXXXXXXXXXXX.
XXXXXXXXXXXXXXXXXXX
.XXXXXXXXXXXXXXXXX.
.X..X.....X..X.X.X.
.X..X.....X..X.X.X.
""",
    # Spinosaurus: tall sail along the back.
    "spino": """
....X.....X......
...XXX...XXX.....
...XXX...XXX.....
..XXXXX.XXXXX.XX.
.XXXXXXXXXXXXXXXX
.XXXXXXXXXXXXXXX.
XXXXXXXXXXXXXXX..
.XXXXXXXXXXXX....
.X..X.X..X.......
.X..X.X..X.......
""",
    # Parasaurolophus: distinctive head crest curving back.
    "para": """
............XX..
...........XX...
..........XX....
.........XXXXXX.
....XXXXXXXXXXX.
.XXXXXXXXXXXXX..
XXXXXXXXXXXXX...
.X..X..X..X.....
.X..X..X..X.....
""",
    # Diplodocus: very long horizontal silhouette.
    "diplo": """
..............XXX
..............X.X
..............XXX
.XXXXXXXXXXXXXXX.
XXXXXXXXXXXXXXX..
XXXXXXXXXXXXXXX..
.X.X......X.X....
.X.X......X.X....
""",
    # --- Prehistoric scenery / relic icons used as subtle section accents in
    # the PDF report. Same ASCII-grid convention as the dinos above; kept
    # small (≤14 wide) so they embed at ~14–18 px without dominating the
    # section title.
    # Volcano with a lava crown.
    "volcano": """
......X......
.....XXX.....
....XX.XX....
...X..X..X...
..XX.....XX..
.XX.XX.XX.XX.
XX.XX..XX..XX
XXXXXXXXXXXXX
""",
    # Curled ammonite fossil.
    "fossil": """
..XXXXX..
.X.....X.
X..XXX..X
X.X...X.X
X.X.X.X.X
X.X...X.X
X..XXX..X
.X.....X.
..XXXXX..
""",
    # Palm tree with fronds + trunk (coconut era).
    "palm": """
...X.X.X...
..X.XXX.X..
.X.XXXXX.X.
XX.XXXXX.XX
.XXXXXXXXX.
....XXX....
....XXX....
....XXX....
....XXX....
....XXX....
""",
    # Dino egg with speckle.
    "egg": """
...XXX...
..XXXXX..
.XXXXXXX.
.XX.XXXX.
XXXXXXXXX
XXX.XXXXX
XXXXXXXXX
.XXXXXXX.
..XXXXX..
""",
    # 3-toed dino footprint.
    "footprint": """
.X..X..X.
XX..X..XX
XX..X..XX
.XXXXXXX.
.XXXXXXX.
..XXXXX..
...XXX...
""",
    # Fern / prehistoric leaf curl.
    "fern": """
.....X....
....XX....
...XX.....
..XX.X....
.XX..XX...
XX....XX..
XX...X.X..
XXX.XX.X..
XXXX..XX..
XXXXXXXX..
.XXXXXX...
..XX......
""",
}


def _grid_to_svg(grid: str) -> str:
    """Convert an ASCII pixel grid to a compact, crisp SVG string."""
    rows = [r for r in grid.strip("\n").split("\n") if r]
    if not rows:
        return '<svg xmlns="http://www.w3.org/2000/svg"/>'
    h = len(rows)
    w = max(len(r) for r in rows)
    rects = []
    for y, row in enumerate(rows):
        for x, ch in enumerate(row):
            if ch == "X":
                # 1.05 width avoids hairline gaps when the SVG is scaled up.
                rects.append(
                    f'<rect x="{x}" y="{y}" width="1.05" height="1.05"/>'
                )
    return (
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {w} {h}" '
        f'shape-rendering="crispEdges" fill="currentColor">'
        f'{"".join(rects)}</svg>'
    )


def get_dino_svg(name: str) -> str:
    """Return the SVG string for the named dinosaur. Falls back to the
    non-TREX default so the favicon/header stay the only places that show
    the T-Rex silhouette."""
    if name == "trex":
        return _grid_to_svg(DINO_GRIDS["trex"])
    grid = DINO_GRIDS.get(name) or DINO_GRIDS[_DEFAULT_SECTION_DINO]
    return _grid_to_svg(grid)


def _pixel_icon_png(name: str, scale: int = 6,
                     color: tuple[int, int, int] = (40, 40, 40)) -> bytes:
    """Rasterize a DINO_GRIDS icon to PNG bytes for PDF embedding.

    ReportLab accepts in-memory PNG via `Image(io.BytesIO(...))`; rendering
    our ASCII grids straight to pixels (no SVG→PDF conversion) keeps the
    PDF dependency surface small (Pillow only).
    """
    from PIL import Image as _PILImage  # local import keeps CLI startup lean
    import numpy as _np
    if name == "trex":
        grid = DINO_GRIDS["trex"]
    else:
        grid = DINO_GRIDS.get(name) or DINO_GRIDS[_DEFAULT_SECTION_DINO]
    rows = [r for r in grid.strip("\n").split("\n") if r]
    h = len(rows)
    w = max(len(r) for r in rows)
    arr = _np.zeros((h, w, 4), dtype=_np.uint8)
    for y, row in enumerate(rows):
        for x, ch in enumerate(row):
            if ch == "X":
                arr[y, x] = (*color, 255)
    img = _PILImage.fromarray(arr, mode="RGBA").resize(
        (w * scale, h * scale), _PILImage.NEAREST
    )
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def _pdf_apply_chrome(story, styles, jp_font):
    """Shared chrome for every PDF report:

      1. Appends a PageBreak + centred TREX signature page at the tail
         of `story` — the report's visual bookend. TREX is reserved
         for the app chrome everywhere else (v1.0.37); this final page
         mirrors the header so the report literally opens and closes
         with the app's signature sprite.
      2. Returns a canvas callback usable on `onFirstPage` /
         `onLaterPages` of `doc.multiBuild`, which draws a thin grey
         "dashboard4dx" at the bottom-left corner of every rendered
         page (using the caller's leftMargin so the anchor tracks
         each report's layout).
    """
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import Image, PageBreak, Paragraph, Spacer

    signature_title_style = ParagraphStyle(
        "PdfSignatureTitle", parent=styles["Title"], fontName=jp_font,
        fontSize=26, alignment=1, textColor=colors.HexColor("#3aa872"),
    )
    signature_icon = Image(
        io.BytesIO(_pixel_icon_png("trex", scale=10)),
        width=120, height=100,
    )
    signature_icon.hAlign = "CENTER"
    story.append(PageBreak())
    # Push the bookend roughly to the vertical centre of the page.
    story.append(Spacer(1, 9 * cm))
    story.append(signature_icon)
    story.append(Spacer(1, 22))
    story.append(Paragraph("dashboard4dx", signature_title_style))

    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setFont(jp_font, 8)
        canvas.setFillColor(colors.grey)
        canvas.drawString(doc.leftMargin, 0.8 * cm, "dashboard4dx")
        canvas.restoreState()
    return _footer


def dino_data_uri(name: str, color: str = "currentColor") -> str:
    """SVG data URI for inline use in <img src="..."> tags."""
    svg = get_dino_svg(name)
    if color != "currentColor":
        svg = svg.replace('fill="currentColor"', f'fill="{color}"')
    b64 = base64.b64encode(svg.encode("utf-8")).decode("ascii")
    return f"data:image/svg+xml;base64,{b64}"


# Path used for the browser favicon. Streamlit's set_page_config accepts a
# file path; we write the T-Rex SVG out at startup so it's always available.
ICON_DIR = SCRIPT_DIR / "resources" / "icons"


def ensure_favicon() -> Path:
    """(Re)write the T-Rex favicon SVG to disk, return its path."""
    ICON_DIR.mkdir(parents=True, exist_ok=True)
    target = ICON_DIR / "trex.svg"
    # Always rewrite — the grid may evolve; favicon is small.
    target.write_text(get_dino_svg("trex"), encoding="utf-8")
    return target


# Map each chart / view key to a distinct dinosaur (or era-scenery) icon.
# Keep T-Rex reserved for the page chrome (favicon + page title) so users
# instantly recognize "this is dashboard4dx" when it appears — every other
# surface picks from the remaining dino grids or the prehistoric-era
# accents (volcano / fossil / palm / egg / footprint / fern).
CHART_DINOS: dict[str, str] = {
    "chart_progress_gap":     "raptor",
    "chart_test_coverage":    "stego",
    "chart_test_density":     "palm",
    "chart_incident_rate":    "volcano",
    "chart_loc_vs_ng":        "trike",
    "chart_design_impl_gap":  "para",
    "chart_risk_heatmap":     "spino",
    "chart_loc_trend":        "diplo",
    "chart_test_trend":       "anky",
    "chart_bug_trend":        "ptero",
    "chart_defect_class":     "fossil",
    "chart_overview_compare": "fern",
    "role_analytics_title":   "footprint",
    "gantt_title":            "bronto",
    "calendar_title":         "plesio",
}
# Fallback used by section_header / icon helpers when the key isn't in
# CHART_DINOS. Must NOT be "trex" — that's reserved for the page chrome.
_DEFAULT_SECTION_DINO = "bronto"


# Module-level collector populated by section_header while an enclosing
# renderer (currently only render_charts_tab) holds an active collector.
# Each entry is (anchor_id, localized_title). None disables collection.
_CHARTS_TOC_ACTIVE: Optional[list[tuple[str, str]]] = None


def _register_toc_entry(anchor: str, label: str) -> None:
    """Append an (anchor, label) pair to the active TOC collector, if any.
    A no-op when no tab is collecting — safe to call from any header."""
    if _CHARTS_TOC_ACTIVE is not None:
        _CHARTS_TOC_ACTIVE.append((anchor, label))


def section_header(title_key: str, help_key: Optional[str] = None,
                   dino: Optional[str] = None,
                   anchor: Optional[str] = None) -> None:
    """Render a chart/section header: dino icon + localized title + help (?).

    `dino` defaults to whatever is mapped for `title_key`. Help is shown via
    Streamlit's standard tooltip on the subheader so the rich markdown header
    (with 🦕) still appears on hover.

    `anchor` is passed to st.subheader so the header gets a stable HTML id
    — lets the Charts-tab TOC link to it. Defaults to `sec-<title_key>` so
    chart-key renaming drives anchor renaming automatically.
    """
    dino_name = dino or CHART_DINOS.get(title_key, _DEFAULT_SECTION_DINO)
    icon_uri = dino_data_uri(dino_name)
    icon_col, txt_col = st.columns(
        [1, 24], gap="small", vertical_alignment="center"
    )
    effective_anchor = anchor if anchor is not None else f"sec-{title_key}"
    with icon_col:
        st.markdown(
            f'<img src="{icon_uri}" alt="{dino_name}" '
            'style="width:36px;height:36px;display:block;margin:0 auto;" />',
            unsafe_allow_html=True,
        )
    label = t(title_key)
    with txt_col:
        if help_key:
            st.subheader(label, help=t(help_key), anchor=effective_anchor)
        else:
            st.subheader(label, anchor=effective_anchor)
    _register_toc_entry(effective_anchor, label)


def _detect_csv_encoding(data: bytes) -> Optional[str]:
    """Return the first encoding (utf-8-sig / utf-8 / cp932) that decodes
    `data` strictly, or None if all fail."""
    if not data:
        return "utf-8"
    for enc in ("utf-8-sig", "utf-8", "cp932"):
        try:
            data.decode(enc)
            return enc
        except UnicodeDecodeError:
            continue
    return None


def _decode_csv_bytes(data: bytes) -> str:
    """Decode CSV bytes, trying UTF-8 (with BOM) then CP932 strictly.

    Real-world exports come in either encoding; trying UTF-8 first and falling
    back to CP932 picks the variant that decodes cleanly, so Japanese text
    isn't returned as mojibake. Raises ValueError if neither works.
    """
    if not data:
        return ""
    for enc in ("utf-8-sig", "utf-8", "cp932"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    raise ValueError(
        "could not decode CSV — expected UTF-8 (with or without BOM) or CP932"
    )


# =============================================================================
# ETL: Function master (the authoritative ID list)
# =============================================================================
def load_function_master(file_bytes: bytes) -> pd.DataFrame:
    """Parse the Function master xlsx.

    Sheet `機能一覧`, col F = Function ID, col G = Function name.

    Scan rules:
      - Data range = row 2 .. last row where col **B** holds a value.
        (Real masters have section-header / subtotal rows near the bottom that
        bound the data; openpyxl's `max_row` can also include trailing
        formatting-only rows. Bounding by col B avoids both pitfalls.)
      - Within that range, rows whose col F is empty (e.g. section breaks) are
        **skipped**, not treated as terminators.
      - Strike-through cells are NOT excluded — the spec is explicit on this.
      - A Function ID may legitimately appear with multiple distinct names;
        every unique (Function ID, Function name) pair is kept.
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    if MASTER_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{MASTER_SHEET}' not found in master file.")
    ws = wb[MASTER_SHEET]

    fid_idx = _col_to_idx(MASTER_FID_COL)
    name_idx = _col_to_idx(MASTER_NAME_COL)
    b_idx = _col_to_idx("B")

    # Buffer rows so we can identify the last B-filled row in a single pass.
    buffered: list[tuple] = list(ws.iter_rows(min_row=2, values_only=True))

    last_b_offset = -1  # offset within buffered (0 = sheet row 2)
    for i, row in enumerate(buffered):
        if row is None:
            continue
        b_val = row[b_idx - 1] if len(row) >= b_idx else None
        if b_val not in (None, ""):
            last_b_offset = i

    if last_b_offset < 0:
        return pd.DataFrame(columns=["機能ID", "機能名称"])

    rows = []
    for row in buffered[: last_b_offset + 1]:
        if row is None:
            continue
        raw_fid = row[fid_idx - 1] if len(row) >= fid_idx else None
        raw_name = row[name_idx - 1] if len(row) >= name_idx else None
        fid = _normalize_fid(raw_fid)
        if fid is None:
            # F empty (or non-ID-shaped) — skip row, keep going.
            continue
        name = "" if raw_name is None else str(raw_name).strip()
        rows.append({"機能ID": fid, "機能名称": name})

    df = pd.DataFrame(rows, columns=["機能ID", "機能名称"])
    # Drop exact duplicates only — duplicate 機能ID with different names stays.
    df = df.drop_duplicates(subset=["機能ID", "機能名称"]).reset_index(drop=True)
    return df


# =============================================================================
# ETL: WBS
# =============================================================================
@dataclass(frozen=True)
class WbsCols:
    planned_effort: str = "P"
    planned_start: str = "Q"
    planned_end: str = "R"
    actual_start: str = "S"
    actual_end: str = "T"
    actual_effort: str = "U"
    actual_progress: str = "V"
    # N column carries the assignee (担当者) on sub-task rows. Parent rows
    # often leave it blank; that's OK — role analytics downstream operate
    # solely on sub-task rows where is_subtask=True.
    assignee: str = "N"
    planned_progress: str = "AA"


WBS_COLS = WbsCols()


def _parse_phase_date(v) -> Optional[date]:
    """Parse a WBS phase anchor cell (expected 年/月/日). Returns None on
    missing/unparseable input — the caller decides whether that is fatal."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if not isinstance(v, str):
        return None
    s = v.strip()
    if not s:
        return None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d", "%Y年%m月%d日",
                "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _resolve_wbs_date(v, phase_start: date,
                      phase_end: date) -> Optional[date]:
    """Resolve a WBS per-task date cell. Real files often write these as
    ``MM/DD`` (no year); the year is inferred by picking whichever candidate
    year (from the phase window) makes the date fall inside the phase range.
    Native date/datetime and full ``YYYY/MM/DD`` strings are returned as-is."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if not isinstance(v, str):
        return None
    s = v.strip()
    if not s:
        return None
    full = _parse_phase_date(s)
    if full is not None:
        return full
    parts = re.split(r"[/\-.]", s)
    if len(parts) != 2:
        return None
    try:
        m, dd = int(parts[0]), int(parts[1])
    except ValueError:
        return None
    candidates: list[date] = []
    for y in sorted({phase_start.year, phase_end.year}):
        try:
            candidates.append(date(y, m, dd))
        except ValueError:
            continue
    if not candidates:
        return None
    for c in candidates:
        if phase_start <= c <= phase_end:
            return c
    # Outside phase range — pick the candidate closest to the window so we
    # still return something, rather than silently dropping the date.
    return min(candidates,
               key=lambda c: min(abs((c - phase_start).days),
                                 abs((c - phase_end).days)))


def _to_percent_scale(v) -> Optional[float]:
    """Normalize a WBS progress cell to percent scale (0..100).

    Real WBS files write these as Excel percent-formatted numbers (0..1),
    literal strings like ``"91%"``, or bare percent integers like ``91``.
    All three round-trip to the same scale. Fractions <= 1.5 are assumed to
    be 0..1 percent format; everything above is already percent.
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, str):
        s = v.strip().rstrip("%").replace(",", "").strip()
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    if isinstance(v, (int, float)):
        f = float(v)
        if 0 < f <= 1.5:
            return f * 100.0
        return f
    return None


def load_wbs(file_bytes: bytes) -> pd.DataFrame:
    """Parse WBS xlsm.

    Function ID is extracted by scanning columns E..I left-to-right per row,
    starting from row 16. Key columns are at fixed positions (P/Q/R/S/T/U/V/AA).

    After a parent row (one with a 機能ID), any following row *without* a
    機能ID that carries a label in the column immediately right of the parent's
    機能ID column AND at least one schedule date is emitted as a sub-task row.
    Sub-tasks inherit their parent's 機能ID and are flagged with
    `is_subtask=True` + `task_label=<label>`. The sub-task attribution ends
    the moment the next parent row appears.
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True,
                       keep_vba=False)
    if WBS_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{WBS_SHEET}' not found in WBS file.")
    ws = wb[WBS_SHEET]

    scan_idx = [_col_to_idx(c) for c in WBS_FUNC_ID_COLS]
    col_idx = {name: _col_to_idx(letter)
               for name, letter in WBS_COLS.__dict__.items()}
    date_keys = ("planned_start", "planned_end", "actual_start", "actual_end")

    # Read the phase anchors (J6 / N6). Real WBS files put the absolute
    # phase start/end here so row dates written as 月/日 can be resolved.
    ps_col = _col_to_idx(WBS_PHASE_START_CELL[0]) - 1
    pe_col = _col_to_idx(WBS_PHASE_END_CELL[0]) - 1
    phase_row_num = WBS_PHASE_START_CELL[1]
    phase_row = next(iter(ws.iter_rows(min_row=phase_row_num,
                                       max_row=phase_row_num,
                                       values_only=True)), None)
    ps_raw = (phase_row[ps_col]
              if phase_row is not None and ps_col < len(phase_row) else None)
    pe_raw = (phase_row[pe_col]
              if phase_row is not None and pe_col < len(phase_row) else None)
    phase_start = _parse_phase_date(ps_raw)
    phase_end = _parse_phase_date(pe_raw)
    if phase_start is None or phase_end is None:
        missing = []
        if phase_start is None:
            missing.append(f"{WBS_PHASE_START_CELL[0]}{phase_row_num} "
                           f"(フェーズ開始日)")
        if phase_end is None:
            missing.append(f"{WBS_PHASE_END_CELL[0]}{phase_row_num} "
                           f"(フェーズ終了日)")
        raise ValueError(
            "WBS のフェーズ日付セルが未入力または不正です: "
            + " / ".join(missing)
            + "。年/月/日 形式で入力してください "
              "(例: 2026/04/01)。各行の日付はフェーズ期間を元に年を判定します。"
        )
    if phase_end < phase_start:
        raise ValueError(
            f"WBS のフェーズ終了日 ({phase_end}) がフェーズ開始日 "
            f"({phase_start}) より前になっています。"
            f"{WBS_PHASE_START_CELL[0]}{phase_row_num} / "
            f"{WBS_PHASE_END_CELL[0]}{phase_row_num} を確認してください。"
        )

    def _build_rec(row_tuple, fid: str, *, label: Optional[str],
                   is_sub: bool) -> dict:
        rec = {"機能ID": fid, "task_label": label, "is_subtask": is_sub}
        for name, idx in col_idx.items():
            rec[name] = (row_tuple[idx - 1]
                         if idx - 1 < len(row_tuple) else None)
        for k in date_keys:
            rec[k] = _resolve_wbs_date(rec[k], phase_start, phase_end)
        for k in ("planned_progress", "actual_progress"):
            rec[k] = _to_percent_scale(rec[k])
        return rec

    try:
        attach_after_dup = bool(
            st.session_state.get("wbs_attach_after_dup", False)
        )
    except Exception:
        attach_after_dup = False

    mark_col_idx0 = _col_to_idx(WBS_SUBTASK_MARK_COL) - 1

    out = []
    seen_fids: set[str] = set()
    parent_fid: Optional[str] = None
    parent_fid_col: Optional[int] = None  # 1-based column index

    for row in ws.iter_rows(min_row=WBS_DATA_START_ROW, values_only=True):
        if row is None:
            continue

        fid = None
        fid_col = None
        for i in scan_idx:
            if i - 1 < len(row):
                candidate = _normalize_fid(row[i - 1])
                if candidate:
                    fid = candidate
                    fid_col = i
                    break

        if fid:
            if fid in seen_fids:
                # Duplicate Function ID — the first occurrence wins; this row
                # and its following sub-tasks are skipped. By default the
                # "active parent" is invalidated so following sub-task rows
                # are skipped too; the wbs_attach_after_dup setting flips
                # this to re-attach orphaned sub-tasks to the last valid
                # parent instead.
                if not attach_after_dup:
                    parent_fid = None
                    parent_fid_col = None
                continue
            seen_fids.add(fid)
            parent_fid = fid
            parent_fid_col = fid_col
            out.append(_build_rec(row, fid, label=None, is_sub=False))
            continue

        # No 機能ID on this row → only treated as a sub-task if the marker
        # column (L, fixed) is "●". Every other row without a 機能ID is
        # intentionally skipped (notes, spacers, separators, etc.).
        if parent_fid is None or parent_fid_col is None:
            continue
        mark_val = row[mark_col_idx0] if mark_col_idx0 < len(row) else None
        if mark_val is None:
            continue
        if str(mark_val).strip() != WBS_SUBTASK_MARK:
            continue
        sub_cell_idx = parent_fid_col  # column right of the parent's 機能ID
        if sub_cell_idx >= len(row):
            continue
        label_raw = row[sub_cell_idx]
        label = str(label_raw).strip() if label_raw is not None else ""
        if not label:
            continue
        rec = _build_rec(row, parent_fid, label=label, is_sub=True)
        if not any(rec[k] is not None for k in date_keys):
            continue
        out.append(rec)

    cols = (["機能ID", "task_label", "is_subtask"]
            + list(WBS_COLS.__dict__.keys()))
    return pd.DataFrame(out, columns=cols)


# =============================================================================
# ETL: Defect tracker (Redmine-style export)
# =============================================================================
DEFECT_COLS = {
    "tracker": "トラッカー",
    "status": "ステータス",
    "assignee": "担当者",
    "actual_start": "実開始日",
    "actual_end": "実終了日",
    "function_id": "機能ID",
    "problem_class": "問題分類",
}


def load_defects(file_bytes: bytes) -> pd.DataFrame:
    """Parse the defect tracker CSV. Accepts UTF-8 or CP932; defect dates are
    in MM/DD/YYYY (US-style) per the export format. Filters to
    tracker == '不具合管理' and derives `unresolved` (True when 実終了日 is empty)."""
    text = _decode_csv_bytes(file_bytes)
    df = pd.read_csv(io.StringIO(text), dtype=str).fillna("")

    missing = [v for v in DEFECT_COLS.values() if v not in df.columns]
    if missing:
        raise ValueError(f"Defect CSV missing columns: {missing}")

    df = df[df[DEFECT_COLS["tracker"]].str.strip() == DEFECT_TRACKER_FILTER].copy()
    df["機能ID"] = df[DEFECT_COLS["function_id"]].map(_normalize_fid)
    df = df[df["機能ID"].notna()].copy()

    df["実開始日"] = df[DEFECT_COLS["actual_start"]].map(_parse_us_date)
    df["実終了日"] = df[DEFECT_COLS["actual_end"]].map(_parse_us_date)
    df["unresolved"] = df["実終了日"].isna()

    return df[
        ["機能ID", "トラッカー", "ステータス", "担当者",
         "実開始日", "実終了日", "問題分類", "unresolved"]
    ].reset_index(drop=True)


# =============================================================================
# ETL: Backlog.com issue CSV
# =============================================================================
# Columns we explicitly surface in the UI. Any *extra* columns present in the
# source CSV are preserved as-is and echoed back unchanged on export so
# round-tripping to Backlog doesn't drop fields we didn't model.
BACKLOG_CORE_COLS = (
    "キーID", "ID", "種別", "状態", "カテゴリ", "件名", "詳細",
    "担当者", "開始日", "期限日", "更新日", "発生フェーズ", "顧客共有",
)
# Date columns — stored as python `date` objects internally; serialised back
# to `M/D/YYYY` (empty string for None) on export.
BACKLOG_DATE_COLS = ("開始日", "期限日", "更新日")


def _normalize_backlog_assignee(raw) -> str:
    """Collapse Backlog's three 担当者 registration patterns onto a single
    canonical form. Real exports mix `田中太郎` / `田中 太郎` / `田中　太郎`
    (no-space / half-width / full-width) for the same person; fold them all
    to `田中太郎` (no space) so filter buckets and overdue tallies don't
    spuriously split one person into three."""
    if raw is None:
        return ""
    try:
        if pd.isna(raw):
            return ""
    except (TypeError, ValueError):
        pass
    s = unicodedata.normalize("NFKC", str(raw))
    s = s.replace("　", "")  # full-width space
    s = re.sub(r"\s+", "", s)    # any remaining whitespace
    return s.strip()


def load_backlog(file_bytes: bytes) -> pd.DataFrame:
    """Parse a Backlog.com issue-list CSV export.

    Encoding: the real export is CP932 (SJIS); we attempt that first and fall
    back to UTF-8 so ad-hoc re-saves still load. Line endings may be LF or
    CRLF. All 13 core columns are expected; any extras are preserved on a
    per-row `_extras` column so `export_backlog` below can round-trip them
    back to Backlog without information loss.

    The 3 date columns are parsed as M/D/YYYY (matching Backlog's export
    format); blank cells come back as None. An additional
    `担当者_normalized` column is added — canonical form across the 3
    spacing variants Backlog users register — powering the filter / kanban
    grouping without mutating the original `担当者` string.
    """
    text = _decode_csv_bytes(file_bytes)
    # Keep every column as string so IDs and free-text fields round-trip
    # unchanged; date parsing happens after load.
    df = pd.read_csv(io.StringIO(text), dtype=str).fillna("")
    if df.empty:
        return pd.DataFrame(columns=list(BACKLOG_CORE_COLS))

    missing = [c for c in BACKLOG_CORE_COLS if c not in df.columns]
    if missing:
        raise ValueError(
            "Backlog CSV is missing required columns: "
            + ", ".join(missing)
        )

    # Parse the 3 date columns into real date objects (blank → None).
    for c in BACKLOG_DATE_COLS:
        df[c] = df[c].map(lambda v: _parse_us_date(v) if v else None)

    # Strip whitespace on enum-ish fields so "  未対応" == "未対応" in
    # filters. 件名 / 詳細 keep their internal whitespace; trimming only
    # at the ends keeps free-text intact.
    for c in ("種別", "状態", "カテゴリ", "発生フェーズ", "顧客共有"):
        df[c] = df[c].astype(str).map(lambda s: s.strip())

    df["担当者_normalized"] = df["担当者"].map(_normalize_backlog_assignee)

    # Stash any extra columns that the export carried but the UI doesn't
    # model, so on export we can write them back unchanged.
    extra_cols = [c for c in df.columns
                  if c not in BACKLOG_CORE_COLS
                  and c != "担当者_normalized"]
    if extra_cols:
        df["_extras"] = df[extra_cols].to_dict(orient="records")
    else:
        df["_extras"] = [{} for _ in range(len(df))]

    # Keep core cols in their canonical order at the front; append the
    # derived + extras columns last.
    out_cols = (list(BACKLOG_CORE_COLS)
                + ["担当者_normalized", "_extras"])
    return df[out_cols].reset_index(drop=True)


# =============================================================================
# ETL: Test counts
# =============================================================================
def load_test_counts(file_bytes: bytes) -> pd.DataFrame:
    """Parse test counts CSV. Accepts UTF-8 or CP932. Column layout is positional
    (A=機能ID, C=総設定テスト数, D=実施済, E=OK, F=NG; B intentionally unused).
    Derives 未実施 = 総設定テスト数 - 実施済."""
    text = _decode_csv_bytes(file_bytes)
    raw = pd.read_csv(io.StringIO(text), header=0, dtype=str).fillna("")
    if raw.shape[1] < 6:
        raise ValueError("Test counts CSV needs at least 6 columns (A..F).")

    df = pd.DataFrame({
        "機能ID": raw.iloc[:, 0].map(_normalize_fid),
        "総設定テスト数": pd.to_numeric(raw.iloc[:, 2], errors="coerce"),
        "実施済": pd.to_numeric(raw.iloc[:, 3], errors="coerce"),
        "OK": pd.to_numeric(raw.iloc[:, 4], errors="coerce"),
        "NG": pd.to_numeric(raw.iloc[:, 5], errors="coerce"),
    })
    df = df[df["機能ID"].notna()].copy()
    df["未実施"] = df["総設定テスト数"] - df["実施済"]
    return df.reset_index(drop=True)


# =============================================================================
# ETL: Code line counts
# =============================================================================
def load_code_counts(file_bytes: bytes) -> pd.DataFrame:
    """Parse the code-LoC xlsx. Sheet `機能ID別サマリ`, A=機能ID, B=LoC.
    The data is already aggregated; one row per Function ID is expected."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    if CODE_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{CODE_SHEET}' not found in code file.")
    ws = wb[CODE_SHEET]

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None or len(r) < 2:
            continue
        fid = _normalize_fid(r[0])
        if fid is None:
            continue
        try:
            loc = int(r[1]) if r[1] is not None else None
        except (TypeError, ValueError):
            loc = None
        rows.append({"機能ID": fid, "LoC": loc})
    return pd.DataFrame(rows)


# =============================================================================
# ETL: Roster (team membership + gear assignment)
# =============================================================================
ROSTER_SHEET = "担当者一覧"
ROSTER_COLS_ORDER = [
    "チーム名", "担当者名", "PC貸与数", "専用携帯貸与数", "VPNアカウント",
]


def _parse_bool_lenient(v) -> bool:
    """True/False from a spreadsheet cell. Accepts native booleans,
    numeric 1/0, or the mix of JP/EN strings users commonly type:
      - yes/true flavours: 1, '1', 'y','yes','true','○','有','あり'
      - anything else (including blank / None / NaN) is False.
    """
    if v is None:
        return False
    try:
        if pd.isna(v):
            return False
    except (TypeError, ValueError):
        pass
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v != 0
    s = unicodedata.normalize("NFKC", str(v)).strip().lower()
    return s in {"1", "y", "yes", "true", "○", "有", "あり", "o"}


def _parse_int_lenient(v) -> int:
    """0 for None/blank/garbage; int() otherwise."""
    if v is None:
        return 0
    try:
        if pd.isna(v):
            return 0
    except (TypeError, ValueError):
        pass
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def load_roster(file_bytes: bytes) -> pd.DataFrame:
    """Parse the 担当者一覧 xlsx.

    Sheet ``担当者一覧``, header row 1 with columns (in order):
        チーム名 / 担当者名 / PC貸与数 / 専用携帯貸与数 / VPNアカウント

    VPN flag accepts the usual yes/no variants (see _parse_bool_lenient).
    Assignee name is normalised so full-width / doubled / padded spaces
    collapse — the same rule used by the WBS N-column so roster names
    can be compared against 担当者×ロール analytics without drift.
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    if ROSTER_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{ROSTER_SHEET}' not found in roster file.")
    ws = wb[ROSTER_SHEET]
    out: list[dict] = []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame(columns=[
            "team", "assignee", "pc_count", "phone_count", "vpn_account",
        ])
    # Header row: loose — only checks that *all* expected labels are
    # somewhere in row 1 in the right order. Extra trailing columns are
    # ignored, which lets templates carry notes / formulas without
    # breaking the loader.
    header = [str(c or "").strip() for c in rows[0]]
    for want, got in zip(ROSTER_COLS_ORDER, header):
        if want != got:
            raise ValueError(
                f"担当者一覧のヘッダー行が想定と異なります。"
                f"期待: {ROSTER_COLS_ORDER} / 実際: {header[:5]}"
            )
    for r in rows[1:]:
        if r is None:
            continue
        if all(c is None or str(c).strip() == "" for c in r[:2]):
            continue  # skip wholly-blank / trailer rows
        team = str(r[0] or "").strip() if len(r) > 0 else ""
        assignee = _normalize_assignee(r[1]) if len(r) > 1 else ""
        if not assignee:
            continue
        pc = _parse_int_lenient(r[2] if len(r) > 2 else 0)
        phone = _parse_int_lenient(r[3] if len(r) > 3 else 0)
        vpn = _parse_bool_lenient(r[4] if len(r) > 4 else None)
        out.append({
            "team": team, "assignee": assignee,
            "pc_count": pc, "phone_count": phone,
            "vpn_account": vpn,
        })
    return pd.DataFrame(out, columns=[
        "team", "assignee", "pc_count", "phone_count", "vpn_account",
    ])


def generate_roster_template(sample: bool = True) -> bytes:
    """Build a blank-or-lightly-seeded 担当者一覧 xlsx the user can edit.

    `sample=True` pre-fills three example rows so first-time users see the
    expected shape (team / name / numeric gear counts / yes/no VPN flag)
    without guessing. Pass `sample=False` for a truly empty template."""
    wb = Workbook()
    ws = wb.active
    ws.title = ROSTER_SHEET
    for col_idx, h in enumerate(ROSTER_COLS_ORDER, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = Font(bold=True)
    if sample:
        seed = [
            ("フロントエンド", "田中 太郎", 1, 1, True),
            ("バックエンド",   "佐藤 花子", 1, 0, True),
            ("QA",             "鈴木 次郎", 1, 0, False),
        ]
        for i, row in enumerate(seed, start=2):
            for col_idx, v in enumerate(row, start=1):
                ws.cell(row=i, column=col_idx, value=v)
    # Reasonable default widths so the file opens readable.
    for col_idx, width in enumerate([16, 22, 12, 14, 14], start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx)
                              .column_letter].width = width
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# ETL: Calendar (global events + per-assignee non-working days)
# =============================================================================
CAL_EVENT_SHEET = "イベント"
CAL_EVENT_COLS_ORDER = ["日付", "タイトル", "説明"]
CAL_NONWORK_SHEET = "非稼働日"
CAL_NONWORK_COLS_ORDER = ["担当者名", "開始日", "終了日", "理由"]
# On-calendar display range. Editable here; used by render to jump the
# initial view when no events cross today.
CAL_DISPLAY_START = date(2024, 1, 1)
CAL_DISPLAY_END   = date(2027, 12, 31)


# -----------------------------------------------------------------------------
# Japanese national holidays seeded into the calendar template.
# Only "base" holidays are enumerated below — 振替休日 (the next non-holiday
# weekday after a Sunday holiday) and 国民の休日 (a weekday sandwiched
# between two other holidays) are derived at runtime by _jp_holidays().
# Equinox dates come from the National Astronomical Observatory of Japan's
# published values for 2024–2027; extend here if CAL_DISPLAY_END moves.
# -----------------------------------------------------------------------------
_JP_BASE_HOLIDAYS: dict[int, list[tuple[int, int, str]]] = {
    2024: [
        (1, 1, "元日"), (1, 8, "成人の日"),
        (2, 11, "建国記念の日"), (2, 23, "天皇誕生日"),
        (3, 20, "春分の日"),
        (4, 29, "昭和の日"),
        (5, 3, "憲法記念日"), (5, 4, "みどりの日"), (5, 5, "こどもの日"),
        (7, 15, "海の日"),
        (8, 11, "山の日"),
        (9, 16, "敬老の日"), (9, 22, "秋分の日"),
        (10, 14, "スポーツの日"),
        (11, 3, "文化の日"), (11, 23, "勤労感謝の日"),
    ],
    2025: [
        (1, 1, "元日"), (1, 13, "成人の日"),
        (2, 11, "建国記念の日"), (2, 23, "天皇誕生日"),
        (3, 20, "春分の日"),
        (4, 29, "昭和の日"),
        (5, 3, "憲法記念日"), (5, 4, "みどりの日"), (5, 5, "こどもの日"),
        (7, 21, "海の日"),
        (8, 11, "山の日"),
        (9, 15, "敬老の日"), (9, 23, "秋分の日"),
        (10, 13, "スポーツの日"),
        (11, 3, "文化の日"), (11, 23, "勤労感謝の日"),
    ],
    2026: [
        (1, 1, "元日"), (1, 12, "成人の日"),
        (2, 11, "建国記念の日"), (2, 23, "天皇誕生日"),
        (3, 20, "春分の日"),
        (4, 29, "昭和の日"),
        (5, 3, "憲法記念日"), (5, 4, "みどりの日"), (5, 5, "こどもの日"),
        (7, 20, "海の日"),
        (8, 11, "山の日"),
        (9, 21, "敬老の日"), (9, 23, "秋分の日"),
        (10, 12, "スポーツの日"),
        (11, 3, "文化の日"), (11, 23, "勤労感謝の日"),
    ],
    2027: [
        (1, 1, "元日"), (1, 11, "成人の日"),
        (2, 11, "建国記念の日"), (2, 23, "天皇誕生日"),
        (3, 21, "春分の日"),
        (4, 29, "昭和の日"),
        (5, 3, "憲法記念日"), (5, 4, "みどりの日"), (5, 5, "こどもの日"),
        (7, 19, "海の日"),
        (8, 11, "山の日"),
        (9, 20, "敬老の日"), (9, 23, "秋分の日"),
        (10, 11, "スポーツの日"),
        (11, 3, "文化の日"), (11, 23, "勤労感謝の日"),
    ],
}


def _jp_holidays_for_year(year: int) -> list[tuple[date, str]]:
    """Return (date, name) tuples for `year` including substitute holidays.

    Substitute rules applied automatically on top of _JP_BASE_HOLIDAYS:
      1. 振替休日 — a base holiday on Sunday pushes to the next non-holiday
         weekday (skipping any adjacent base holidays).
      2. 国民の休日 — a single non-holiday weekday sandwiched between two
         holidays (both Monday-Friday, exactly one day apart) becomes a
         holiday itself (applies to e.g. 2026-09-22 between 敬老の日 and
         秋分の日).
    """
    base = _JP_BASE_HOLIDAYS.get(year, [])
    holidays: list[tuple[date, str]] = [
        (date(year, m, d), name) for (m, d, name) in base
    ]
    holiday_dates = {d for d, _ in holidays}
    # Rule 1: Sunday → next non-holiday day
    for d, _name in list(holidays):
        if d.weekday() == 6:  # Sunday
            sub = d + timedelta(days=1)
            while sub in holiday_dates:
                sub += timedelta(days=1)
            holidays.append((sub, "振替休日"))
            holiday_dates.add(sub)
    # Rule 2: 国民の休日 (gap of exactly one weekday between two holidays)
    sorted_dates = sorted(holiday_dates)
    for i in range(len(sorted_dates) - 1):
        d1, d2 = sorted_dates[i], sorted_dates[i + 1]
        if (d2 - d1).days == 2:
            gap = d1 + timedelta(days=1)
            if gap.weekday() < 5 and gap not in holiday_dates:
                holidays.append((gap, "国民の休日"))
                holiday_dates.add(gap)
    return sorted(holidays)


def _jp_holidays_in_range(start: date, end: date) -> list[tuple[date, str]]:
    """Concatenate _jp_holidays_for_year over the year range, clipped
    to [start, end]. Used to seed the calendar template with公休."""
    out: list[tuple[date, str]] = []
    for year in range(start.year, end.year + 1):
        for d, name in _jp_holidays_for_year(year):
            if start <= d <= end:
                out.append((d, name))
    return out


def load_calendar(file_bytes: bytes) -> pd.DataFrame:
    """Parse the calendar xlsx (2 sheets) into one long-form dataframe.

    Returned columns:
      kind          "event" or "nonwork"
      assignee       担当者名 (nonwork only; '' for events)
      start_date    date
      end_date      date (same as start_date for single-day entries)
      title         event title OR nonwork reason
      description   free-text (events only; '' for nonwork)
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    out: list[dict] = []

    # --- Events sheet ------------------------------------------------------
    if CAL_EVENT_SHEET in wb.sheetnames:
        ws = wb[CAL_EVENT_SHEET]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [str(c or "").strip() for c in rows[0]]
            for want, got in zip(CAL_EVENT_COLS_ORDER, header):
                if want != got:
                    raise ValueError(
                        f"「{CAL_EVENT_SHEET}」シートのヘッダー行が想定と"
                        f"異なります。期待: {CAL_EVENT_COLS_ORDER} / "
                        f"実際: {header[:3]}"
                    )
            for r in rows[1:]:
                if r is None:
                    continue
                d = _to_date(r[0]) if len(r) > 0 else None
                title = (str(r[1] or "").strip()
                         if len(r) > 1 else "")
                if d is None or not title:
                    continue
                desc = (str(r[2] or "").strip()
                        if len(r) > 2 else "")
                out.append({
                    "kind": "event", "assignee": "",
                    "start_date": d, "end_date": d,
                    "title": title, "description": desc,
                })

    # --- Non-working days sheet -------------------------------------------
    if CAL_NONWORK_SHEET in wb.sheetnames:
        ws = wb[CAL_NONWORK_SHEET]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [str(c or "").strip() for c in rows[0]]
            for want, got in zip(CAL_NONWORK_COLS_ORDER, header):
                if want != got:
                    raise ValueError(
                        f"「{CAL_NONWORK_SHEET}」シートのヘッダー行が想定と"
                        f"異なります。期待: {CAL_NONWORK_COLS_ORDER} / "
                        f"実際: {header[:4]}"
                    )
            for r in rows[1:]:
                if r is None:
                    continue
                assignee_raw = r[0] if len(r) > 0 else None
                assignee = _normalize_assignee(assignee_raw)
                start = _to_date(r[1]) if len(r) > 1 else None
                end = _to_date(r[2]) if len(r) > 2 else None
                reason = (str(r[3] or "").strip()
                          if len(r) > 3 else "")
                if not assignee or start is None:
                    continue
                if end is None:
                    end = start
                if end < start:
                    # Swap obvious input mistakes rather than dropping.
                    start, end = end, start
                out.append({
                    "kind": "nonwork", "assignee": assignee,
                    "start_date": start, "end_date": end,
                    "title": reason, "description": "",
                })

    return pd.DataFrame(out, columns=[
        "kind", "assignee", "start_date", "end_date",
        "title", "description",
    ])


def generate_calendar_template(sample: bool = True) -> bytes:
    """Build the 2-sheet calendar template.

    Japanese national holidays for the full CAL_DISPLAY_START..END range
    are ALWAYS pre-seeded in the 「イベント」 sheet (not gated on `sample`)
    so every generated template ships with 公休 info out of the box —
    振替休日 / 国民の休日 are derived automatically.

    `sample=True` additionally seeds a few ordinary company events and
    non-working-day rows so first-time users see the expected shape.
    """
    wb = Workbook()
    # --- Events sheet ------------------------------------------------------
    ws_e = wb.active
    ws_e.title = CAL_EVENT_SHEET
    for col_idx, h in enumerate(CAL_EVENT_COLS_ORDER, start=1):
        c = ws_e.cell(row=1, column=col_idx, value=h)
        c.font = Font(bold=True)

    # Jp holidays first — users can delete rows they don't want, and the
    # "description" column is preset with 公休 so it's obvious on a glance.
    seed_rows: list[tuple[date, str, str]] = []
    for d, name in _jp_holidays_in_range(CAL_DISPLAY_START, CAL_DISPLAY_END):
        seed_rows.append((d, name, "公休（日本の祝日）"))

    if sample:
        seed_rows.extend([
            (date(2025, 4, 1),  "年度開始",        "全社キックオフ"),
            (date(2025, 10, 1), "下期キックオフ",  "半期レビュー+方針"),
            (date(2025, 12, 25),"全社MTG（年末）", ""),
        ])
    # Chronological order for editing comfort.
    seed_rows.sort(key=lambda r: r[0])
    for i, (d, title, desc) in enumerate(seed_rows, start=2):
        ws_e.cell(row=i, column=1, value=d)
        ws_e.cell(row=i, column=1).number_format = "yyyy-mm-dd"
        ws_e.cell(row=i, column=2, value=title)
        ws_e.cell(row=i, column=3, value=desc)
    for col_idx, width in enumerate([14, 28, 40], start=1):
        ws_e.column_dimensions[ws_e.cell(row=1, column=col_idx)
                                 .column_letter].width = width

    # --- Non-working days sheet -------------------------------------------
    ws_n = wb.create_sheet(CAL_NONWORK_SHEET)
    for col_idx, h in enumerate(CAL_NONWORK_COLS_ORDER, start=1):
        c = ws_n.cell(row=1, column=col_idx, value=h)
        c.font = Font(bold=True)
    if sample:
        seed_nonwork = [
            ("田中 太郎", date(2025, 5, 1),  date(2025, 5, 2),  "有給"),
            ("佐藤 花子", date(2025, 5, 7),  date(2025, 5, 7),  "半休"),
            ("鈴木 次郎", date(2025, 8, 11), date(2025, 8, 15), "夏季休暇"),
        ]
        for i, row in enumerate(seed_nonwork, start=2):
            ws_n.cell(row=i, column=1, value=row[0])
            ws_n.cell(row=i, column=2, value=row[1])
            ws_n.cell(row=i, column=2).number_format = "yyyy-mm-dd"
            ws_n.cell(row=i, column=3, value=row[2])
            ws_n.cell(row=i, column=3).number_format = "yyyy-mm-dd"
            ws_n.cell(row=i, column=4, value=row[3])
    for col_idx, width in enumerate([22, 14, 14, 28], start=1):
        ws_n.column_dimensions[ws_n.cell(row=1, column=col_idx)
                                 .column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# Step-wise validation pipeline (drives the dino-runner UI)
# =============================================================================
@dataclass
class StepResult:
    """One named pre-flight check on an uploaded file. `status` is one of:
    'ok', 'warn', 'error', 'pending'. `detail` shows extra context (e.g.
    detected encoding, row count); `message` is the user-facing reason on
    a non-ok result. `exc` carries the original exception (when any) so the
    log writer can render its traceback."""
    label_key: str
    status: str = "ok"
    message: str = ""
    detail: str = ""
    exc: Optional[BaseException] = field(default=None, repr=False)


def _step(steps: list[StepResult], label_key: str, status: str = "ok",
          message: str = "", detail: str = "",
          exc: Optional[BaseException] = None) -> None:
    steps.append(StepResult(label_key, status, message, detail, exc))


def _preflight_master(data: bytes) -> list[StepResult]:
    steps: list[StepResult] = []
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        _step(steps, "step_xlsx_parse", "ok")
    except Exception as exc:
        _step(steps, "step_xlsx_parse", "error", str(exc), exc=exc)
        return steps

    if MASTER_SHEET not in wb.sheetnames:
        _step(steps, "step_master_sheet", "error",
              f"Sheet '{MASTER_SHEET}' not found in {wb.sheetnames}")
        return steps
    _step(steps, "step_master_sheet", "ok")

    ws = wb[MASTER_SHEET]
    fid_idx = _col_to_idx(MASTER_FID_COL)
    b_idx = _col_to_idx("B")
    buffered = list(ws.iter_rows(min_row=2, values_only=True))

    last_b = -1
    for i, row in enumerate(buffered):
        if row is None:
            continue
        v = row[b_idx - 1] if len(row) >= b_idx else None
        if v not in (None, ""):
            last_b = i
    if last_b < 0:
        _step(steps, "step_master_b_col", "error",
              "B column has no data — cannot determine end of master")
        return steps
    _step(steps, "step_master_b_col", "ok",
          detail=f"last B-filled row = {last_b + 2}")

    fids: list[str] = []
    for row in buffered[: last_b + 1]:
        if row is None:
            continue
        raw = row[fid_idx - 1] if len(row) >= fid_idx else None
        fid = _normalize_fid(raw)
        if fid:
            fids.append(fid)
    if not fids:
        _step(steps, "step_master_fid", "error",
              f"No valid Function IDs in column {MASTER_FID_COL}")
        return steps
    _step(steps, "step_master_fid", "ok",
          detail=f"{len(fids)} ID rows · {len(set(fids))} unique IDs")

    dups = sum(1 for c in __import__("collections").Counter(fids).values() if c > 1)
    if dups:
        _step(steps, "step_master_dups", "warn",
              detail=f"{dups} Function IDs appear with multiple names")
    else:
        _step(steps, "step_master_dups", "ok")
    return steps


class WbsDiagnosticError(RuntimeError):
    """Carries a multi-line diagnostic dump explaining why 0 Function IDs
    were extracted from the WBS. Shown verbatim in the on-screen detail
    expander and written in full to the session log."""


def _diagnose_wbs_fid_absence(data: bytes) -> str:
    """Re-read the WBS in three modes and collect clues for why col E–I scan
    from row 16 yielded zero Function IDs. Intended for on-screen + log dump
    only (never raises — failures in sub-probes become in-line notes)."""
    scan_idx = set(_col_to_idx(c) for c in WBS_FUNC_ID_COLS)       # {5..9}
    sample_cols = set(range(1, 12))                                # A..K
    lines: list[str] = []

    def _val_repr(v) -> str:
        s = repr(v)
        return s if len(s) <= 60 else s[:57] + "…"

    # -- Probe 1: read_only=True, data_only=True (= what load_wbs uses) -------
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True,
                           read_only=True, keep_vba=False)
        ws = wb[WBS_SHEET]
        try:
            declared = ws.calculated_dimension
        except Exception:
            declared = "?"
        lines.append(f"[probe1: read_only=True, data_only=True]")
        lines.append(f"  declared dimension: {declared}")
        lines.append(f"  max_row={ws.max_row}, max_col={ws.max_column}")

        total = nonempty = ei_nonempty = 0
        samples: list[str] = []
        for row in ws.iter_rows(min_row=WBS_DATA_START_ROW):
            total += 1
            ne_cells = [c for c in row if c.value not in (None, "")]
            if not ne_cells:
                continue
            nonempty += 1
            if any(c.column in scan_idx for c in ne_cells):
                ei_nonempty += 1
            if len(samples) < 8:
                row_num = ne_cells[0].row
                parts = [f"{c.column_letter}={_val_repr(c.value)}"
                         for c in row
                         if c.value not in (None, "")
                         and getattr(c, "column", 0) in sample_cols]
                samples.append(f"    row {row_num}: " + " | ".join(parts))
        lines.append(f"  rows iterated from {WBS_DATA_START_ROW}+: {total}")
        lines.append(f"    - with any non-empty cell: {nonempty}")
        lines.append(f"    - with non-empty in E–I : {ei_nonempty}")
        if samples:
            lines.append("  first non-empty rows (cols A–K):")
            lines.extend(samples)
        else:
            lines.append("  (no non-empty rows visible in this mode)")
        try:
            wb.close()
        except Exception:
            pass
    except Exception as e:
        lines.append(f"[probe1] failed: {e}")

    # -- Probe 2: read_only=False — escapes declared-dimension mis-hints ------
    try:
        wb2 = load_workbook(io.BytesIO(data), data_only=True,
                            read_only=False, keep_vba=False)
        ws2 = wb2[WBS_SHEET]
        lines.append(f"[probe2: read_only=False, data_only=True]")
        lines.append(f"  max_row={ws2.max_row}, max_col={ws2.max_column}")

        total2 = nonempty2 = ei_nonempty2 = fid_count2 = 0
        first_fid: tuple[int, str, str] | None = None   # (row, col, fid)
        samples2: list[str] = []
        for row in ws2.iter_rows(min_row=WBS_DATA_START_ROW):
            total2 += 1
            ne_cells = [c for c in row if c.value not in (None, "")]
            if not ne_cells:
                continue
            nonempty2 += 1
            ei_cells = [c for c in row if c.column in scan_idx]
            any_ei = any(c.value not in (None, "") for c in ei_cells)
            if any_ei:
                ei_nonempty2 += 1
                fid = None
                for c in ei_cells:
                    fid = _normalize_fid(c.value)
                    if fid:
                        if first_fid is None:
                            first_fid = (c.row, c.column_letter, fid)
                        break
                if fid:
                    fid_count2 += 1
                elif len(samples2) < 8:
                    row_num = ne_cells[0].row
                    parts = [f"{c.column_letter}={_val_repr(c.value)}"
                             for c in row
                             if c.column in sample_cols
                             and c.value not in (None, "")]
                    samples2.append(f"    row {row_num}: "
                                    + " | ".join(parts))
        lines.append(f"  rows iterated from {WBS_DATA_START_ROW}+: {total2}")
        lines.append(f"    - with any non-empty cell: {nonempty2}")
        lines.append(f"    - with non-empty in E–I : {ei_nonempty2}")
        lines.append(f"    - parsed as Function ID : {fid_count2}")
        if first_fid:
            lines.append(f"  first FID: row {first_fid[0]} "
                         f"col {first_fid[1]} → {first_fid[2]}")
        if samples2:
            lines.append("  rows with E–I data but NO parsed FID (A–K):")
            lines.extend(samples2)
        try:
            wb2.close()
        except Exception:
            pass
    except Exception as e:
        lines.append(f"[probe2] failed: {e}")

    # -- Probe 3: data_only=False — detect formulas w/o cached values --------
    try:
        wb3 = load_workbook(io.BytesIO(data), data_only=False,
                            read_only=False, keep_vba=False)
        ws3 = wb3[WBS_SHEET]
        lines.append(f"[probe3: data_only=False] — formula detection")
        formula_hits: list[str] = []
        cached_none_hits = 0
        for row in ws3.iter_rows(min_row=WBS_DATA_START_ROW):
            for c in row:
                if c.column not in scan_idx:
                    continue
                if c.data_type == "f":
                    if len(formula_hits) < 5:
                        formula_hits.append(
                            f"    row {c.row} col {c.column_letter}: "
                            f"formula = {_val_repr(c.value)}")
                    cached_none_hits += 1
            if len(formula_hits) >= 5 and cached_none_hits >= 5:
                break
        if formula_hits:
            lines.append(f"  formulas present in E–I "
                         f"(≥{cached_none_hits} cells); first 5:")
            lines.extend(formula_hits)
            lines.append("  ⇒ if probe1 saw None but probe2 saw values, "
                         "cached-value table is incomplete — "
                         "open file in Excel and Save-As to refresh caches.")
        else:
            lines.append("  no formulas in E–I (cells are literal values)")
        try:
            wb3.close()
        except Exception:
            pass
    except Exception as e:
        lines.append(f"[probe3] failed: {e}")

    return "\n".join(lines)


def _preflight_wbs(data: bytes) -> list[StepResult]:
    steps: list[StepResult] = []
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True,
                           keep_vba=False)
        _step(steps, "step_xlsm_parse", "ok")
    except Exception as exc:
        _step(steps, "step_xlsm_parse", "error", str(exc), exc=exc)
        return steps

    if WBS_SHEET not in wb.sheetnames:
        _step(steps, "step_wbs_sheet", "error",
              f"Sheet '{WBS_SHEET}' not found in {wb.sheetnames}")
        return steps
    _step(steps, "step_wbs_sheet", "ok")

    ws = wb[WBS_SHEET]
    ps_col0 = _col_to_idx(WBS_PHASE_START_CELL[0]) - 1
    pe_col0 = _col_to_idx(WBS_PHASE_END_CELL[0]) - 1
    phase_row_num = WBS_PHASE_START_CELL[1]
    phase_row = next(iter(ws.iter_rows(min_row=phase_row_num,
                                       max_row=phase_row_num,
                                       values_only=True)), None)
    ps_raw = (phase_row[ps_col0]
              if phase_row is not None and ps_col0 < len(phase_row) else None)
    pe_raw = (phase_row[pe_col0]
              if phase_row is not None and pe_col0 < len(phase_row) else None)
    ps_d = _parse_phase_date(ps_raw)
    pe_d = _parse_phase_date(pe_raw)
    if ps_d is None or pe_d is None or pe_d < ps_d:
        bad = []
        if ps_d is None:
            bad.append(f"{WBS_PHASE_START_CELL[0]}{phase_row_num}"
                       f" (開始) = {ps_raw!r}")
        if pe_d is None:
            bad.append(f"{WBS_PHASE_END_CELL[0]}{phase_row_num}"
                       f" (終了) = {pe_raw!r}")
        if ps_d is not None and pe_d is not None and pe_d < ps_d:
            bad.append(f"終了 ({pe_d}) が開始 ({ps_d}) より前")
        _step(steps, "step_wbs_phase_dates", "error",
              "フェーズ日付セルが不正: " + "; ".join(bad)
              + "。年/月/日 形式 (例: 2026/04/01) で入力してください。")
        return steps
    _step(steps, "step_wbs_phase_dates", "ok",
          detail=f"phase: {ps_d} 〜 {pe_d}")

    scan_idx = [_col_to_idx(c) for c in WBS_FUNC_ID_COLS]
    fid_count = 0
    rows_seen = 0
    first_fid_row: Optional[int] = None
    for r_i, row in enumerate(
            ws.iter_rows(min_row=WBS_DATA_START_ROW, values_only=True)):
        rows_seen += 1
        if row is None:
            continue
        for i in scan_idx:
            if i - 1 < len(row) and _normalize_fid(row[i - 1]):
                fid_count += 1
                if first_fid_row is None:
                    first_fid_row = WBS_DATA_START_ROW + r_i
                break
    try:
        wb.close()
    except Exception:
        pass

    if fid_count == 0:
        diag = _diagnose_wbs_fid_absence(data)
        msg = (f"no IDs in E–I from row {WBS_DATA_START_ROW}+ "
               f"(scanned {rows_seen} rows in read_only mode) — "
               f"see detailed log entry")
        _step(steps, "step_wbs_fid", "error", msg,
              exc=WbsDiagnosticError("\n" + diag))
        return steps
    _step(steps, "step_wbs_fid", "ok",
          detail=f"{fid_count} rows with IDs "
                 f"(first at row {first_fid_row}, "
                 f"from row {WBS_DATA_START_ROW})")
    return steps


class DefectsDiagnosticError(RuntimeError):
    """Carries a multi-line diagnostic explaining why the defect-CSV dry-run
    produced zero usable rows. The same text is shown on-screen and logged
    in full to the session log."""


def _diagnose_defects_build_failure(
    raw_df: pd.DataFrame,
    stage: str,
) -> str:
    """Explain why the defect CSV would collapse to an empty DataFrame.

    `stage` identifies which step produced the zero-row outcome:
      'tracker' — tracker filter removed every row
      'fid'     — tracker filter kept rows but none parsed as Function ID
    """
    lines: list[str] = []
    lines.append(f"stage collapsing to 0 rows: {stage}")
    lines.append(f"raw CSV rows: {len(raw_df)}")
    lines.append(f"columns ({len(raw_df.columns)}): "
                 f"{list(raw_df.columns)}")

    tracker_col = DEFECT_COLS["tracker"]
    fid_col = DEFECT_COLS["function_id"]
    trackers = raw_df[tracker_col].astype(str).str.strip()
    distinct = trackers.value_counts()
    lines.append(f"tracker filter expects (exact, NFKC-insensitive): "
                 f"{DEFECT_TRACKER_FILTER!r}")
    lines.append(f"distinct tracker values in file: {distinct.size}")
    for v, c in distinct.head(10).items():
        marker = " ← match" if v == DEFECT_TRACKER_FILTER else ""
        lines.append(f"  • {v!r}: {c} rows{marker}")

    filtered = raw_df[trackers == DEFECT_TRACKER_FILTER]
    lines.append(f"rows surviving tracker filter: {len(filtered)}")

    if stage == "tracker":
        similar = distinct[distinct.index.to_series()
                           .str.contains("不具合", na=False)]
        if len(similar):
            lines.append("tracker values containing '不具合' "
                         "(likely rename candidates):")
            for v, c in similar.head(5).items():
                lines.append(f"  • {v!r}: {c} rows")
        else:
            lines.append("no tracker value contains '不具合' — "
                         "tracker column may be wired to a different "
                         "field, or the export was pre-filtered.")
        return "\n".join(lines)

    # stage == 'fid'
    raw_fids = filtered[fid_col].astype(str).str.strip()
    parsed = raw_fids.map(_normalize_fid)
    n_parsed = int(parsed.notna().sum())
    lines.append(f"rows with parseable Function ID: "
                 f"{n_parsed} / {len(filtered)}")
    unparsed = raw_fids[parsed.isna() & (raw_fids != "")]
    if not unparsed.empty:
        top = unparsed.value_counts().head(10)
        lines.append(f"top raw 機能ID values that failed to parse "
                     f"({unparsed.nunique()} distinct):")
        for v, c in top.items():
            lines.append(f"  • {v!r}: {c} rows")
    empty_fid = int((raw_fids == "").sum())
    if empty_fid:
        lines.append(f"rows with empty 機能ID cell: {empty_fid}")
    lines.append(
        "expected formats: '機能ID：XXXX', '機能ID:XXXX', "
        "'XXXX：機能名', 'XXXX:name', or bare 'XXXX' where "
        "XXXX = 1–10 ASCII letters + 1–10 ASCII digits "
        "(full-width letters/digits/colons are NFKC-normalised). "
        "Hyphens (e.g. 'AUTH-001') and other separators do NOT match.")
    return "\n".join(lines)


def _preflight_defects(data: bytes) -> list[StepResult]:
    steps: list[StepResult] = []
    enc = _detect_csv_encoding(data)
    if enc is None:
        _step(steps, "step_csv_encoding", "error",
              "could not decode CSV as UTF-8 or CP932")
        return steps
    _step(steps, "step_csv_encoding", "ok", detail=f"encoding: {enc}")

    try:
        text = data.decode(enc)
        df = pd.read_csv(io.StringIO(text), dtype=str).fillna("")
        _step(steps, "step_csv_parse", "ok", detail=f"{len(df)} rows parsed")
    except Exception as exc:
        _step(steps, "step_csv_parse", "error", str(exc), exc=exc)
        return steps

    missing = [v for v in DEFECT_COLS.values() if v not in df.columns]
    if missing:
        _step(steps, "step_defects_columns", "error",
              f"missing columns: {missing}")
        return steps
    _step(steps, "step_defects_columns", "ok")

    filtered = df[df[DEFECT_COLS["tracker"]].astype(str).str.strip()
                  == DEFECT_TRACKER_FILTER]
    if filtered.empty:
        diag = _diagnose_defects_build_failure(df, stage="tracker")
        _step(steps, "step_defects_filter", "error",
              f"0 rows match tracker = '{DEFECT_TRACKER_FILTER}' — "
              f"see detailed log entry",
              exc=DefectsDiagnosticError("\n" + diag))
        return steps
    _step(steps, "step_defects_filter", "ok",
          detail=f"{len(filtered)} defect rows after filter")

    bad_dates = 0
    for v in filtered[DEFECT_COLS["actual_start"]]:
        if v and not _parse_us_date(v):
            bad_dates += 1
    if bad_dates:
        _step(steps, "step_defects_dates", "warn",
              detail=f"{bad_dates} 実開始日 cells not in MM/DD/YYYY")
    else:
        _step(steps, "step_defects_dates", "ok")

    # Dry-run the FID extraction that load_defects performs last, so the
    # "empty dataframe" failure from step_load_failed is pre-empted with a
    # precise cause (top unparseable 機能ID samples + regex reminder).
    parsed_fid = filtered[DEFECT_COLS["function_id"]].map(_normalize_fid)
    n_fid = int(parsed_fid.notna().sum())
    if n_fid == 0:
        diag = _diagnose_defects_build_failure(df, stage="fid")
        _step(steps, "step_defects_build", "error",
              f"tracker filter kept {len(filtered)} rows but 0 had a "
              f"parseable 機能ID — see detailed log entry",
              exc=DefectsDiagnosticError("\n" + diag))
        return steps
    detail = f"{n_fid} rows will load"
    if n_fid < len(filtered):
        detail += f" ({len(filtered) - n_fid} dropped for unparseable 機能ID)"
    _step(steps, "step_defects_build", "ok", detail=detail)
    return steps


def _preflight_tests(data: bytes) -> list[StepResult]:
    steps: list[StepResult] = []
    enc = _detect_csv_encoding(data)
    if enc is None:
        _step(steps, "step_csv_encoding", "error",
              "could not decode CSV as UTF-8 or CP932")
        return steps
    _step(steps, "step_csv_encoding", "ok", detail=f"encoding: {enc}")

    try:
        text = data.decode(enc)
        raw = pd.read_csv(io.StringIO(text), header=0, dtype=str).fillna("")
        _step(steps, "step_csv_parse", "ok", detail=f"{len(raw)} rows parsed")
    except Exception as exc:
        _step(steps, "step_csv_parse", "error", str(exc), exc=exc)
        return steps

    if raw.shape[1] < 6:
        _step(steps, "step_tests_min_cols", "error",
              f"only {raw.shape[1]} columns; need ≥ 6 (A–F)")
        return steps
    _step(steps, "step_tests_min_cols", "ok")

    fids = raw.iloc[:, 0].map(_normalize_fid).dropna()
    if fids.empty:
        _step(steps, "step_tests_fid", "error",
              "no valid Function IDs in column A")
        return steps
    _step(steps, "step_tests_fid", "ok", detail=f"{len(fids)} rows with IDs")

    bad_numeric = []
    for idx, label in [(2, "C 総設定テスト数"), (3, "D 実施済"),
                       (4, "E OK"), (5, "F NG")]:
        s = pd.to_numeric(raw.iloc[:, idx], errors="coerce")
        nan_count = int(s.isna().sum())
        if nan_count:
            bad_numeric.append(f"{label} ({nan_count})")
    if bad_numeric:
        _step(steps, "step_tests_numeric", "warn",
              detail=f"non-numeric cells: {', '.join(bad_numeric)}")
    else:
        _step(steps, "step_tests_numeric", "ok")

    total = pd.to_numeric(raw.iloc[:, 2], errors="coerce").fillna(0)
    run = pd.to_numeric(raw.iloc[:, 3], errors="coerce").fillna(0)
    over = int((run > total).sum())
    if over:
        _step(steps, "step_tests_sanity", "warn",
              detail=f"{over} rows have 実施済 > 総設定テスト数")
    else:
        _step(steps, "step_tests_sanity", "ok")
    return steps


def _preflight_code(data: bytes) -> list[StepResult]:
    steps: list[StepResult] = []
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        _step(steps, "step_xlsx_parse", "ok")
    except Exception as exc:
        _step(steps, "step_xlsx_parse", "error", str(exc), exc=exc)
        return steps

    if CODE_SHEET not in wb.sheetnames:
        _step(steps, "step_code_sheet", "error",
              f"Sheet '{CODE_SHEET}' not found in {wb.sheetnames}")
        return steps
    _step(steps, "step_code_sheet", "ok")

    ws = wb[CODE_SHEET]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    fids: list[str] = []
    bad_loc = 0
    for r in rows:
        if r is None or len(r) < 2:
            continue
        fid = _normalize_fid(r[0])
        if not fid:
            continue
        fids.append(fid)
        try:
            if r[1] is not None:
                int(r[1])
        except (TypeError, ValueError):
            bad_loc += 1
    if not fids:
        _step(steps, "step_code_fid", "error",
              "no valid Function IDs in column A")
        return steps
    _step(steps, "step_code_fid", "ok", detail=f"{len(fids)} ID rows")
    if bad_loc:
        _step(steps, "step_code_loc", "warn",
              detail=f"{bad_loc} rows with non-numeric LoC")
    else:
        _step(steps, "step_code_loc", "ok")
    return steps


_PREFLIGHTS: dict[str, Callable[[bytes], list[StepResult]]] = {
    "master":  _preflight_master,
    "wbs":     _preflight_wbs,
    "defects": _preflight_defects,
    "tests":   _preflight_tests,
    "code":    _preflight_code,
}


def validate_with_steps(
    spec: dict, data: bytes
) -> tuple[Optional[pd.DataFrame], list[StepResult]]:
    """Run named pre-flight checks then (if all passed) the loader. Returns
    (df, steps) — df is None when any step has status='error'."""
    pre = _PREFLIGHTS.get(spec["key"])
    steps = pre(data) if pre else []
    if any(s.status == "error" for s in steps):
        return None, steps
    try:
        df = spec["loader"](data)
    except Exception as exc:
        steps.append(StepResult("step_load_failed", "error", str(exc), exc=exc))
        return None, steps
    if df is None or df.empty:
        steps.append(StepResult("step_load_failed", "error",
                                "loader produced an empty dataframe"))
        return None, steps
    steps.append(StepResult("step_load_ok", "ok",
                            detail=f"dataframe shape: {df.shape}"))
    return df, steps


# =============================================================================
# Integration: master-driven LEFT JOIN
# =============================================================================
def build_design_pages_df(
    master: pd.DataFrame,
    pages: dict[str, int] | None,
) -> pd.DataFrame:
    """Turn the manual design-pages input into a per-Function-ID dataframe."""
    if not pages:
        return pd.DataFrame(columns=["機能ID", "設計書ページ数"])
    return pd.DataFrame(
        [{"機能ID": fid, "設計書ページ数": n} for fid, n in pages.items()]
    )


def integrate(
    master: pd.DataFrame,
    wbs: pd.DataFrame | None = None,
    defects: pd.DataFrame | None = None,
    tests: pd.DataFrame | None = None,
    code: pd.DataFrame | None = None,
    design_pages: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """LEFT JOIN every supplied source onto the master on `機能ID`.

    Per agreed approach (A): the master keeps every (機能ID, 機能名称) pair, so
    when a Function ID has multiple names the joined sources are duplicated
    onto each name row. Aggregations downstream account for this when needed.
    """
    if master is None or master.empty:
        return pd.DataFrame(columns=["機能ID", "機能名称"])

    df = master.copy()

    # Defects need pre-aggregation per Function ID
    if defects is not None and not defects.empty:
        agg = defects.groupby("機能ID").agg(
            defect_total=("機能ID", "size"),
            defect_unresolved=("unresolved", "sum"),
        ).reset_index()
        df = df.merge(agg, on="機能ID", how="left")

    if tests is not None and not tests.empty:
        df = df.merge(
            tests[["機能ID", "総設定テスト数", "実施済", "OK", "NG", "未実施"]],
            on="機能ID", how="left",
        )

    if code is not None and not code.empty:
        df = df.merge(code[["機能ID", "LoC"]], on="機能ID", how="left")

    if design_pages is not None and not design_pages.empty:
        df = df.merge(design_pages, on="機能ID", how="left")

    if wbs is not None and not wbs.empty:
        # Sub-task rows are schedule *breakdowns* for a parent Function ID and
        # must not participate in per-FID KPI aggregates — they'd overwrite
        # the parent's dates/effort with a slice of themselves.
        wbs_parents = (wbs[~wbs["is_subtask"].fillna(False).astype(bool)]
                       if "is_subtask" in wbs.columns else wbs)
        wbs_sorted = wbs_parents.sort_values(
            "planned_end", ascending=False, na_position="last"
        )
        wbs_one = wbs_sorted.drop_duplicates(subset=["機能ID"], keep="first")
        merge_cols = [c for c in wbs_one.columns
                      if c not in ("task_label", "is_subtask")]
        df = df.merge(wbs_one[merge_cols], on="機能ID", how="left")

    return df


# =============================================================================
# KPIs
# =============================================================================
RISK_AT_RISK_THRESHOLD = 0.5  # risk_score >= this counts as "at risk"


def _safe_div(num: pd.Series, den: pd.Series) -> pd.Series:
    """num / den with NaN where den is 0 or NaN. Always returns float."""
    num_n = pd.to_numeric(num, errors="coerce")
    den_n = pd.to_numeric(den, errors="coerce")
    out = num_n / den_n.where(den_n != 0)
    return out.astype("float64")


def _delay_days(row, today: date) -> Optional[int]:
    """Days late vs planned_end. 0 if on time, NaN if planned_end missing.
    For ongoing items (no actual_end), delay accumulates against `today`."""
    pe = row.get("planned_end")
    ae = row.get("actual_end")
    if pe is None or pd.isna(pe):
        return None
    if ae is not None and not pd.isna(ae):
        return max(0, (ae - pe).days)
    if today > pe:
        return (today - pe).days
    return 0


def _delay_rate(row) -> Optional[float]:
    """delay_days normalized by planned duration; capped at 1.0."""
    delay = row.get("delay_days")
    ps = row.get("planned_start")
    pe = row.get("planned_end")
    if pd.isna(delay) or ps is None or pd.isna(ps) or pe is None or pd.isna(pe):
        return None
    duration = (pe - ps).days
    if duration <= 0:
        return None
    return float(min(1.0, max(0.0, delay / duration)))


def _normalize_max(s: pd.Series) -> pd.Series:
    """Min-zero / max-1 normalization. Empty/zero-max returns all zeros."""
    sn = pd.to_numeric(s, errors="coerce")
    m = sn.max(skipna=True)
    if pd.isna(m) or m == 0:
        return pd.Series([0.0] * len(s), index=s.index)
    return sn.fillna(0) / m


def compute_kpis(
    integrated: pd.DataFrame, today: Optional[date] = None
) -> pd.DataFrame:
    """Append derived KPI columns to the integrated dataframe.

    Per-Function-ID metrics added (when their inputs exist):
      - bug_density     = NG / LoC               (test-spec defect density;
                                                  ≠ Redmine fault count)
      - test_density    = 総設定テスト数 / 設計書ページ数  (tests per design page)
      - complexity      = LoC / 設計書ページ数      (lines per design page)
      - test_run_rate   = 実施済 / 総設定テスト数         (0..1)
      - test_pass_rate  = OK / 実施済              (0..1)
      - defect_rate     = NG / 総設定テスト数             (test-spec defect rate;
                                                  ≠ incident_rate)
      - incident_rate   = defect_total / 実施済    (Redmine fault rate;
                                                  numerator from Redmine,
                                                  denominator from test spec)
      - delay_days      = days late vs planned_end (0 if on time)
      - delay_rate      = delay_days / planned duration, capped at 1.0
      - health_score    = 実施率 - defect_rate - 遅延率 (range ~ -2..1)
      - risk_score      = weighted blend of normalized inputs (0..1):
                          0.4*defect_unresolved + 0.2*未実施 + 0.2*遅延 +
                          0.2*bug_density

    The risk_score components are min-max normalized within the dataset so the
    weights add up meaningfully across very different scales (a count vs a
    density). Weights are renormalized when some inputs are absent.
    """
    if today is None:
        today = date.today()
    df = integrated.copy()

    # Densities & rates
    if "NG" in df.columns and "LoC" in df.columns:
        df["bug_density"] = _safe_div(df["NG"], df["LoC"])
    if "総設定テスト数" in df.columns and "設計書ページ数" in df.columns:
        df["test_density"] = _safe_div(df["総設定テスト数"], df["設計書ページ数"])
    if "LoC" in df.columns and "設計書ページ数" in df.columns:
        df["complexity"] = _safe_div(df["LoC"], df["設計書ページ数"])
    if "実施済" in df.columns and "総設定テスト数" in df.columns:
        df["test_run_rate"] = _safe_div(df["実施済"], df["総設定テスト数"])
    if "OK" in df.columns and "実施済" in df.columns:
        df["test_pass_rate"] = _safe_div(df["OK"], df["実施済"])
    if "NG" in df.columns and "総設定テスト数" in df.columns:
        df["defect_rate"] = _safe_div(df["NG"], df["総設定テスト数"])
    # incident_rate is the Redmine fault count over executed tests, kept
    # deliberately separate from defect_rate (test-spec NG / 総設定テスト数).
    if "defect_total" in df.columns and "実施済" in df.columns:
        df["incident_rate"] = _safe_div(df["defect_total"], df["実施済"])

    # Delay
    if "planned_end" in df.columns:
        df["delay_days"] = df.apply(lambda r: _delay_days(r, today), axis=1)
        df["delay_days"] = pd.to_numeric(df["delay_days"], errors="coerce")
    if {"planned_start", "planned_end", "delay_days"}.issubset(df.columns):
        df["delay_rate"] = df.apply(_delay_rate, axis=1)
        df["delay_rate"] = pd.to_numeric(df["delay_rate"], errors="coerce")

    # Composite scores
    if "test_run_rate" in df.columns:
        df["health_score"] = (
            df["test_run_rate"].fillna(0)
            - (df["defect_rate"].fillna(0) if "defect_rate" in df.columns else 0)
            - (df["delay_rate"].fillna(0)  if "delay_rate"  in df.columns else 0)
        )

    components, weights = [], []
    if "defect_unresolved" in df.columns:
        components.append(_normalize_max(df["defect_unresolved"])); weights.append(0.4)
    if "未実施" in df.columns:
        components.append(_normalize_max(df["未実施"]));            weights.append(0.2)
    if "delay_days" in df.columns:
        components.append(_normalize_max(df["delay_days"]));        weights.append(0.2)
    if "bug_density" in df.columns:
        components.append(_normalize_max(df["bug_density"]));       weights.append(0.2)
    if components:
        wsum = sum(weights)
        weights = [w / wsum for w in weights]
        df["risk_score"] = sum(w * c for w, c in zip(weights, components))

    return df


COLUMN_HELP_KEYS: dict[str, str] = {
    "機能ID": "help_func_id",
    "機能名称": "help_func_name",
    "defect_total": "help_defect_total",
    "defect_unresolved": "help_defect_unresolved",
    "総設定テスト数": "help_test_total",
    "実施済": "help_test_run",
    "OK": "help_test_ok",
    "NG": "help_test_ng",
    "未実施": "help_test_notrun",
    "LoC": "help_loc",
    "設計書ページ数": "help_design_pages",
    "planned_effort": "help_planned_effort",
    "actual_effort":  "help_actual_effort",
    "planned_start":  "help_planned_start",
    "planned_end":    "help_planned_end",
    "actual_start":   "help_actual_start",
    "actual_end":     "help_actual_end",
    "actual_progress":  "help_actual_progress",
    "planned_progress": "help_planned_progress",
    "incident_rate": "help_incident_rate",
    "bug_density":   "help_bug_density",
    "test_density":  "help_test_density",
    "complexity":    "help_complexity",
    "test_run_rate": "help_test_run_rate",
    "test_pass_rate":"help_test_pass_rate",
    "defect_rate":   "help_defect_rate",
    "delay_days":    "help_delay_days",
    "delay_rate":    "help_delay_rate",
    "health_score":  "help_health_score",
    "risk_score":    "help_risk_score",
}

COLUMN_NUMERIC_FORMATS: dict[str, str] = {
    "bug_density":   "%.3f",
    "test_density":  "%.2f",
    "complexity":    "%.1f",
    "test_run_rate": "percent",
    "test_pass_rate":"percent",
    "defect_rate":   "percent",
    "incident_rate": "percent",
    "delay_days":    "%.0f",
    "delay_rate":    "percent",
    "health_score":  "%.2f",
    "risk_score":    "%.2f",
    "planned_effort":   "%.1f",
    "actual_effort":    "%.1f",
    "planned_progress": "%.0f%%",
    "actual_progress":  "%.0f%%",
}

COLUMN_LABEL_KEYS: dict[str, str] = {
    "bug_density":   "col_bug_density",
    "test_density":  "col_test_density",
    "complexity":    "col_complexity",
    "test_run_rate": "col_test_run_rate",
    "test_pass_rate":"col_test_pass_rate",
    "defect_rate":   "col_defect_rate",
    "incident_rate": "col_incident_rate",
    "defect_total":  "col_defect_total",
    "defect_unresolved": "col_defect_unresolved",
    "NG":            "col_test_ng",
    "delay_days":    "col_delay_days",
    "delay_rate":    "col_delay_rate",
    "health_score":  "col_health_score",
    "risk_score":    "col_risk_score",
}


def build_col_config(cols: list[str]) -> dict:
    """Build column_config for st.dataframe from a list of column names.

    Adds: localized label (where defined), printf-style format (for numeric
    KPI columns), and a hover-tooltip help string with provenance/definition.
    """
    config: dict = {}
    for col in cols:
        help_key = COLUMN_HELP_KEYS.get(col)
        help_text = t(help_key) if help_key else None
        label_key = COLUMN_LABEL_KEYS.get(col)
        label = t(label_key) if label_key else col
        if col in COLUMN_NUMERIC_FORMATS:
            config[col] = st.column_config.NumberColumn(
                label, format=COLUMN_NUMERIC_FORMATS[col], help=help_text,
            )
        elif help_text:
            config[col] = st.column_config.Column(label=label, help=help_text)
    return config


def project_kpi_summary(kpi_df: pd.DataFrame) -> dict:
    """Project-wide aggregates for the dashboard's metric strip."""
    def _sum(col: str) -> int:
        if col not in kpi_df.columns:
            return 0
        return int(pd.to_numeric(kpi_df[col], errors="coerce").fillna(0).sum())

    def _mean(col: str) -> Optional[float]:
        if col not in kpi_df.columns:
            return None
        s = pd.to_numeric(kpi_df[col], errors="coerce").dropna()
        return float(s.mean()) if len(s) else None

    total_loc      = _sum("LoC")
    total_tests    = _sum("総設定テスト数")
    total_run      = _sum("実施済")
    total_ok       = _sum("OK")
    total_ng       = _sum("NG")
    open_defects   = _sum("defect_unresolved")
    total_defects  = _sum("defect_total")

    run_rate  = (total_run / total_tests) if total_tests else None
    pass_rate = (total_ok  / total_run)   if total_run   else None

    at_risk = 0
    if "risk_score" in kpi_df.columns:
        at_risk = int(
            (kpi_df["risk_score"].fillna(0) >= RISK_AT_RISK_THRESHOLD).sum()
        )
    delayed = 0
    if "delay_days" in kpi_df.columns:
        delayed = int((kpi_df["delay_days"].fillna(0) > 0).sum())

    return {
        "total_loc":      total_loc,
        "total_tests":    total_tests,
        "total_run":      total_run,
        "total_ok":       total_ok,
        "total_ng":       total_ng,
        "open_defects":   open_defects,
        "total_defects":  total_defects,
        "run_rate":       run_rate,
        "pass_rate":      pass_rate,
        "avg_bug_density": _mean("bug_density"),
        "avg_test_density": _mean("test_density"),
        "avg_health":     _mean("health_score"),
        "avg_risk":       _mean("risk_score"),
        "at_risk_count":  at_risk,
        "delayed_count":  delayed,
    }


# =============================================================================
# DORA-style team-delivery metrics (computed from existing sources)
# =============================================================================
# Thresholds bucket each metric into a 3-tier badge keyed against the DORA
# 2024 research bands, collapsed for faster comprehension:
#   good   ≈ DORA Elite / High tier (above industry average)
#   normal ≈ DORA Medium tier       (around industry average)
#   bad    ≈ DORA Low tier          (below industry average)
# Tunable here if a client wants stricter/looser bands.
DORA_WINDOW_DAYS = 30

# For HIGHER-is-better metrics (deployment frequency): thresholds descend.
_DORA_FREQ_THRESHOLDS = [(1.0, "good"), (0.25, "normal")]
# For LOWER-is-better metrics: thresholds ascend.
_DORA_LEAD_THRESHOLDS        = [(7.0, "good"), (30.0, "normal")]
_DORA_CFR_THRESHOLDS         = [(30.0, "good"), (45.0, "normal")]
_DORA_RECOVERY_THRESHOLDS    = [(1.0, "good"), (7.0, "normal")]
_DORA_RELIABILITY_THRESHOLDS = [(15.0, "good"), (30.0, "normal")]


def _dora_rate_above(value: Optional[float],
                     thresholds: list[tuple[float, str]]) -> str:
    """Rating helper when HIGHER values rate better (Deployment Frequency).
    Walks the thresholds top-down and returns the first matching tier;
    anything below the lowest breakpoint collapses to "bad"."""
    if value is None:
        return "unknown"
    for cutoff, rating in thresholds:
        if value >= cutoff:
            return rating
    return "bad"


def _dora_rate_below(value: Optional[float],
                     thresholds: list[tuple[float, str]]) -> str:
    """Rating helper when LOWER values rate better (Lead time, CFR, Recovery,
    Reliability). Mirrored version of `_dora_rate_above`."""
    if value is None:
        return "unknown"
    for cutoff, rating in thresholds:
        if value <= cutoff:
            return rating
    return "bad"


def compute_dora_metrics(
    kpi_df: pd.DataFrame,
    defects_df: Optional[pd.DataFrame] = None,
    window_days: int = DORA_WINDOW_DAYS,
    today: Optional[date] = None,
) -> dict:
    """Derive the DORA 5Keys over the trailing `window_days`.

    Maps our existing sources as:
      - **Deployment frequency**: # of features whose `actual_end` lands in
        the window ÷ weeks.
      - **Lead time for changes**: median (actual_end − planned_start)
        for the same in-window cohort.
      - **Change failure rate**: % of the in-window cohort that has any
        Redmine defect recorded against it.
      - **Failed deployment recovery time**: median resolution duration
        (`実終了日 − 実開始日`) of Redmine defects CLOSED in the window.
      - **Reliability**: dataset-wide mean Redmine incident_rate.

    Returned dict has one entry per metric with value + unit + rating
    (elite / high / medium / low / unknown).
    """
    if today is None:
        today = date.today()
    window_start = today - timedelta(days=window_days)

    # Completed features landing in the window.
    ae_raw = kpi_df.get("actual_end")
    if ae_raw is None:
        in_window = kpi_df.iloc[0:0].copy()
    else:
        ae = pd.to_datetime(ae_raw, errors="coerce")
        mask = (ae.dt.date >= window_start) & (ae.dt.date <= today)
        in_window = kpi_df[mask].copy()
        in_window["_ae_dt"] = ae[mask]

    n_completed = int(len(in_window))
    weeks = max(window_days / 7.0, 1e-9)
    freq = n_completed / weeks

    # Lead time (median days).
    lead: Optional[float] = None
    if not in_window.empty and "planned_start" in in_window.columns:
        ps = pd.to_datetime(in_window["planned_start"], errors="coerce")
        pair = pd.DataFrame({"s": ps, "e": in_window["_ae_dt"]}).dropna()
        if not pair.empty:
            durations = (pair["e"] - pair["s"]).dt.days
            # Negative leads (actual before planned) shouldn't contribute.
            durations = durations[durations >= 0]
            if len(durations):
                lead = float(durations.median())

    # Change failure rate.
    cfr: Optional[float] = None
    if n_completed > 0 and "defect_total" in in_window.columns:
        dt = pd.to_numeric(in_window["defect_total"],
                           errors="coerce").fillna(0)
        cfr = float((dt > 0).sum() / n_completed * 100.0)

    # Recovery time (median days for in-window resolved defects).
    recovery: Optional[float] = None
    if defects_df is not None and not defects_df.empty:
        dfd = defects_df.copy()
        dfd["_start"] = pd.to_datetime(dfd.get("実開始日"), errors="coerce")
        dfd["_end"]   = pd.to_datetime(dfd.get("実終了日"), errors="coerce")
        resolved = dfd.dropna(subset=["_start", "_end"])
        resolved = resolved[
            (resolved["_end"].dt.date >= window_start)
            & (resolved["_end"].dt.date <= today)
        ]
        if not resolved.empty:
            dur_days = (resolved["_end"] - resolved["_start"]).dt.total_seconds() / 86400.0
            dur_days = dur_days[dur_days >= 0]
            if len(dur_days):
                recovery = float(dur_days.median())

    # Reliability: mean of incident_rate across the project (as percent).
    reliability: Optional[float] = None
    if "incident_rate" in kpi_df.columns:
        rel_vals = pd.to_numeric(
            kpi_df["incident_rate"], errors="coerce"
        ).dropna()
        if len(rel_vals):
            reliability = float(rel_vals.mean() * 100.0)

    return {
        "window_days": window_days,
        "completed": n_completed,
        "frequency": {
            "value": freq, "unit_key": "dora_unit_per_week",
            "rating": _dora_rate_above(freq, _DORA_FREQ_THRESHOLDS),
        },
        "lead_time": {
            "value": lead, "unit_key": "dora_unit_days",
            "rating": _dora_rate_below(lead, _DORA_LEAD_THRESHOLDS),
        },
        "cfr": {
            "value": cfr, "unit_key": "dora_unit_percent",
            "rating": _dora_rate_below(cfr, _DORA_CFR_THRESHOLDS),
        },
        "recovery": {
            "value": recovery, "unit_key": "dora_unit_days",
            "rating": _dora_rate_below(recovery, _DORA_RECOVERY_THRESHOLDS),
        },
        "reliability": {
            "value": reliability, "unit_key": "dora_unit_percent",
            "rating": _dora_rate_below(reliability,
                                        _DORA_RELIABILITY_THRESHOLDS),
        },
    }


# =============================================================================
# i18n
# =============================================================================
LANG_OPTIONS: list[tuple[str, str]] = [("en", "EN"), ("ja", "日本語")]
DEFAULT_LANG = "en"

TRANSLATIONS: dict[str, dict[str, str]] = {
    "en": {
        "intro_caption": "Integrated dashboard for the management team",
        "main_tab_dashboard": "📥 Inputs",
        "main_tab_charts": "📊 Charts",
        "main_tab_calendar": "📅 Calendar",
        "main_tab_alert": "🚨 Alerts",
        "main_tab_delivery": "🏁 Delivery",
        "main_tab_backlog": "📋 Backlog",
        "main_tab_design": "📐 Design pages",
        "main_tab_settings": "⚙️ Settings",
        "delivery_needs_data": (
            "Upload a Function master and WBS to see team delivery "
            "performance."
        ),
        # Backlog tab (📋) — issue list from Backlog.com
        "backlog_tab_title":   "📋 Backlog — issues & risks",
        "backlog_tab_caption": (
            "Issues exported from Backlog.com as a CSV (SJIS). Facets "
            "at the top narrow the board; each lane is one 種別 and "
            "within it cards are grouped by 状態."
        ),
        "backlog_needs_file":  (
            "Upload a Backlog issue-list CSV from the Inputs tab to "
            "unlock this view."
        ),
        "backlog_empty":       "No issues match the current filters.",
        "backlog_facet_type":          "Type (種別)",
        "backlog_facet_status":        "Status (状態)",
        "backlog_facet_category":      "Category (カテゴリ)",
        "backlog_facet_phase":         "Phase (発生フェーズ)",
        "backlog_facet_customer":      "Customer-shared (顧客共有)",
        "backlog_facet_due":           "Due date",
        "backlog_due_option_all":       "All",
        "backlog_due_option_overdue":   "Overdue",
        "backlog_due_option_this_week": "This week",
        "backlog_due_option_this_month":"This month",
        "backlog_due_option_future":    "Future",
        "backlog_due_option_none":      "No due date",
        "backlog_tile_due_label":       "Due",
        "backlog_tile_no_due":          "—",
        "backlog_tile_assignee_none":   "(unassigned)",
        "backlog_lane_empty":           "(none)",
        "backlog_status_total":         "{n} issues",
        "backlog_filter_reset":         "Reset filters",
        "dora_section_title": "🏁 Team delivery performance (DORA 5Keys)",
        "dora_section_caption": (
            "Trailing {days} days — {n} features completed in window."
        ),
        "dora_freq_title":        "Deployment frequency",
        "dora_lead_title":        "Lead time for changes",
        "dora_cfr_title":         "Change failure rate",
        "dora_recovery_title":    "Failed deployment recovery time",
        "dora_reliability_title": "Reliability (mean fault rate)",
        "dora_unit_per_week":     "features / wk",
        "dora_unit_days":         "days",
        "dora_unit_percent":      "%",
        "dora_rating_good":       "Good",
        "dora_rating_normal":     "Normal",
        "dora_rating_bad":        "Bad",
        "dora_rating_unknown":    "—",
        # Per-metric hover help for the DORA panel. "Good / Normal / Bad"
        # bands are collapsed from the DORA 2024 research bands (Elite +
        # High → Good; Medium → Normal; Low → Bad).
        "help_dora_frequency": (
            "**🦕 Deployment frequency**\n\n"
            "🧮 Features completed in the last {days} days ÷ weeks.\n\n"
            "📂 Source: WBS **column T** (actual end date).\n\n"
            "💡 vs industry average (DORA 2024): "
            "Good ≥ 1/wk, Normal ≥ 0.25/wk, Bad below."
        ),
        "help_dora_lead_time": (
            "**🦕 Lead time for changes**\n\n"
            "🧮 Median days per completed feature = (actual_end − planned_start).\n\n"
            "📂 Source: WBS **column Q** (planned start) and **column T** "
            "(actual end).\n\n"
            "💡 vs industry average (DORA 2024): "
            "Good ≤ 7 days, Normal ≤ 30 days, Bad beyond."
        ),
        "help_dora_cfr": (
            "**🦕 Change failure rate (CFR)**\n\n"
            "🧮 % of features completed in the window that have at least "
            "one Redmine defect registered.\n\n"
            "📂 Source: WBS column T (in-window completions) × Redmine "
            "defect counts per Function ID.\n\n"
            "💡 vs industry average (DORA 2024): "
            "Good ≤ 30%, Normal ≤ 45%, Bad beyond."
        ),
        "help_dora_recovery": (
            "**🦕 Failed deployment recovery time**\n\n"
            "🧮 Median (実終了日 − 実開始日) of Redmine defects closed "
            "in the window.\n\n"
            "📂 Source: Redmine defect list — 実開始日 / 実終了日.\n\n"
            "💡 vs industry average (DORA 2024): "
            "Good ≤ 1 day, Normal ≤ 7 days, Bad beyond."
        ),
        "help_dora_reliability": (
            "**🦕 Reliability (mean fault rate)**\n\n"
            "🧮 Mean of Redmine fault rate (defect_total ÷ 実施済) across "
            "all features.\n\n"
            "📂 Source: Redmine defect list × test counts column D.\n\n"
            "💡 Good ≤ 15%, Normal ≤ 30%, Bad beyond."
        ),
        # Alert tab (🚨)
        "alert_tab_title":   "Alerts — Function IDs that need attention",
        "alert_tab_caption": (
            "Flags features whose current metrics cross the configured "
            "thresholds (Redmine fault rate, test density) or break simple "
            "heuristics (>14-day delay, mostly-unexecuted test plan). "
            "Honours the global Function ID filter."
        ),
        "alert_needs_master": (
            "Upload a Function master to unlock alerts."
        ),
        "alert_all_clear": "All clear — no alerts raised for the current scope.",
        "alert_sev_high":          "HIGH",
        "alert_sev_medium":        "MED",
        "alert_sev_low":           "LOW",
        "alert_sev_high_label":    "High severity",
        "alert_sev_medium_label":  "Medium severity",
        "alert_sev_low_label":     "Low severity",
        "alert_current_label":     "current",
        "alert_threshold_label":   "threshold",
        "alert_risk_score_label":  "Risk score",
        "alert_score_legend_title": "ℹ️ How the risk score is computed",
        "alert_score_legend_body": (
            "**Risk score (0 – 1, higher = more attention needed)** — "
            "each feature is scored by combining the four breach "
            "metrics below:\n\n"
            "- **Fault rate** — weight **3.0**\n"
            "- **Delay days** — weight **1.5**\n"
            "- **Test density** — weight **1.0**\n"
            "- **Un-executed test ratio** — weight **1.0**\n\n"
            "A metric contributes `n × w` only when it crosses its "
            "threshold (otherwise 0). `n` is the breach level "
            "normalised to 0..1:\n\n"
            "- **Fault rate**: `min(1, ir ÷ (2 × threshold))`\n"
            "- **Delay**: `min(1, (days − 14) ÷ 46)` "
            "(caps at 60 days)\n"
            "- **Test density**: `max(0, 1 − td ÷ threshold)`\n"
            "- **Un-executed**: `max(0, (pct − 60) ÷ 40)` "
            "(needs ≥ 10 planned tests)\n\n"
            "Score = **Σ(n × w) ÷ Σ(w)**. Severity bucket: "
            "**> 0.70 → 🔴 HIGH**, "
            "**> 0.35 → 🟡 MED**, **otherwise 🔵 LOW**. "
            "Features with every metric in-spec do not alert.\n\n"
            "Per tile: `X` marks the metric contributing the most "
            "(the main culprit); `△` marks other metrics that also "
            "breached but contributed less."
        ),
        "help_test_notrun_short": "Un-executed test ratio",
        "alert_sort_label":      "Sort order",
        "alert_sort_severity":   "By severity (HIGH → LOW)",
        "alert_sort_date_desc":  "By date (newest first)",
        "alert_sort_date_asc":   "By date (oldest first)",
        # Per-alert date badge labels — tells the reader whether the
        # date on the tile is the planned ship date, the actual
        # completion date, or just unknown (neither column populated).
        "alert_date_label_actual":  "Actual end",
        "alert_date_label_planned": "Planned end",
        "alert_date_label_unknown": "Date unknown",
        "charts_needs_master": "Upload **Function master** in the Dashboard tab to unlock charts.",
        "chart_progress_gap": "Progress: planned vs actual",
        "chart_progress_planned": "planned",
        "chart_progress_actual": "actual",
        "chart_progress_over_marker": "⚠ over",
        "chart_test_coverage": "Test coverage (OK / NG / not run)",
        "chart_test_density": "Test density per Function ID (test count sufficiency)",
        "chart_test_density_threshold_label": "threshold",
        "chart_test_density_below_marker": "⚠ low",
        "chart_incident_rate": "Fault rate per Function ID (Redmine, defects/Executed)",
        "chart_incident_rate_threshold_label": "threshold",
        "chart_incident_rate_above_marker": "⚠ high",
        "chart_overview_compare": "Function ID overview: 4-metric comparison",
        "chart_overview_compare_filter": "Filter by Function ID",
        "chart_overview_compare_filter_help": "Empty = all. KPI cards and the chart recompute on the filtered set.",
        "chart_overview_compare_total_prefix": "Total",
        "chart_overview_compare_fids_suffix": "Function IDs in scope",
        "chart_overview_compare_empty": "No Function IDs match the current filter.",
        "help_chart_overview_compare": (
            "**🦕 4-metric overview comparison**\n\n"
            "Four side-by-side horizontal bar panels showing 設計書ページ数 / "
            "LoC / 総設定テスト数 / 障害件数（Redmine） per Function ID — each on "
            "its own X scale so absolute values stay readable, with a shared "
            "Y axis so the eye can track each FID across all four panels.\n\n"
            "📂 Source: design pages (manual), code counts (LoC), test counts "
            "(総設定テスト数), Redmine defect tracker (defect_total)."
        ),
        "chart_loc_vs_ng": "LoC × NG",
        "chart_loc_vs_ng_sub": "(size: design pages, color: risk score)",
        "chart_design_impl_gap": "Design pages × LoC",
        "chart_risk_heatmap": "Risk dimensions heatmap",
        "chart_loc_trend": "LoC trend (across snapshots)",
        "chart_test_trend": "Test counts trend (across snapshots)",
        "chart_bug_trend": "Defect trend (opened vs closed by week)",
        "chart_defect_class": "Fault root cause breakdown (Redmine 問題分類)",
        "chart_defect_class_filter": "Filter by Function ID",
        "chart_defect_class_filter_help": "Empty = every Redmine fault row.",
        "chart_defect_class_empty": "No Redmine faults match the current filter.",
        "chart_defect_class_no_class": "No 問題分類 values present in the filtered set.",
        "chart_defect_class_scope_prefix": "Faults in scope:",
        "chart_defect_class_scope_suffix": "rows",
        "chart_defect_class_top_title": "Top {n} root causes",
        "chart_defect_class_col_class": "問題分類",
        "chart_defect_class_col_count": "Count",
        "chart_defect_class_col_share": "Share",
        "global_fid_filter_title": "Function ID filter",
        "global_fid_filter_label": "Function IDs",
        "global_fid_filter_help": "Empty = every Function ID. Applied to Overview compare, Calendar, and Charts (Redmine 問題分類).",
        "global_fid_filter_scope_all": "Scope: all Function IDs",
        "global_fid_filter_scope_n": "Scope: {n} Function ID(s)",
        "global_fid_filter_upload_hint": "Upload the Function ID master to unlock the filter.",
        "kpi_missing_header": "**Cannot compute — missing inputs:**",
        "source_label_tests":   "Test counts CSV",
        "source_label_code":    "Code (LoC) XLSX",
        "source_label_design":  "Design pages (manual)",
        "source_label_defects": "Redmine defect CSV",
        "source_label_wbs":     "WBS XLSM",
        "chart_no_design_pages": "Enter design page counts in the Design pages tab to populate this chart.",
        "chart_no_history": "Need at least two snapshots to draw a trend.",
        "chart_no_defects": "Defect tracker not loaded.",
        "chart_label_ok": "OK",
        "chart_label_ng": "NG",
        "chart_label_notrun": "not run",
        "chart_label_low":    "low",
        "chart_label_opened": "opened",
        "chart_label_closed": "closed",
        "chart_label_open_cum": "open (cumulative)",
        "chart_label_loc_total": "Total LoC",
        "chart_label_total_tests": "Total tests",
        "chart_label_executed": "Executed",
        "chart_label_total": "Total",
        "chart_label_coverage": "Coverage",
        "calendar_needs_master": "Upload **Function master** in the Dashboard tab to unlock the calendar.",
        "calendar_title": "Project calendar",
        "calendar_caption": (
            "Combines WBS schedule (planned vs actual) and defect lifespans. "
            "Toggle layers below."
        ),
        "calendar_layer_planned": "WBS planned",
        "calendar_layer_actual": "WBS actual",
        "calendar_layer_defects": "Defects",
        "calendar_layer_subtasks": "Show sub-tasks",
        "calendar_filter_fid": "Filter by Function ID",
        "calendar_filter_fid_help": "Leave empty to show all",
        "calendar_no_events": "No events to display with the current selection.",
        "gantt_title": "Gantt — planned vs actual",
        "gantt_no_dates": "No WBS dates available to plot.",
        "gantt_today_label": "today",
        "calendar_section": "Calendar",
        "calendar_event_count": "{n} events",
        "settings_uploads_title": "Auto-load of previously imported files",
        "settings_uploads_caption": (
            "On app start each card auto-loads the newest file it finds under "
            "`input/<date>/<slot>/`. Resetting just stops that auto-load for "
            "this session — the historical files are kept for trend analysis."
        ),
        "settings_wbs_title": "WBS parsing behavior",
        "settings_wbs_caption": (
            "Controls how the loader handles duplicate Function IDs. "
            "Only the **first** row with a given 機能ID is kept; subsequent "
            "rows (and their `●` sub-task breakdowns in L column) are "
            "skipped by default so downstream KPIs use a single source "
            "of truth per ID."
        ),
        "settings_wbs_attach_after_dup": (
            "Re-attach sub-tasks after a duplicate to the last valid parent"
        ),
        "settings_wbs_attach_after_dup_caption": (
            "When **off** (default): ● rows that follow a duplicate 機能ID "
            "row are also skipped. Turn **on** to treat those ● rows as "
            "additional sub-tasks of the most recent valid parent 機能ID "
            "instead. Useful when a duplicate row is an accidental re-entry "
            "but its breakdown rows carry legitimate extra schedule detail."
        ),
        "settings_charts_title": "Chart thresholds",
        "settings_charts_caption": (
            "Tunable thresholds used as warning lines on the per-Function-ID "
            "charts. Bars below the configured value are flagged in red and "
            "carry a small marker."
        ),
        "settings_test_density_threshold": "Test density warning threshold (tests / page)",
        "settings_test_density_threshold_caption": (
            "Default 10. Function IDs whose 総設定テスト数 ÷ 設計書ページ数 falls "
            "below this value are highlighted on the test density chart."
        ),
        "settings_incident_rate_threshold": "Fault rate (Redmine) warning threshold (%)",
        "settings_incident_rate_threshold_caption": (
            "Default 5%. Function IDs whose Redmine defect_total ÷ "
            "実施済 exceeds this value are highlighted on the fault rate chart."
        ),
        "settings_pages_title": "Auto-load of design page counts",
        "settings_pages_caption": (
            "Same idea for `input/design_pages.json`: reset clears the in-"
            "memory state for this session; the file on disk is left alone."
        ),
        "settings_reset_btn": "Reset auto-load",
        "settings_undo_reset_btn": "Re-enable auto-load",
        "settings_show_files": "Show files",
        "settings_show_entries": "Show entries",
        "settings_files_count": "{n} files",
        "settings_no_files": "Nothing stored yet.",
        "settings_count_pages": "{n} entries stored",
        "settings_status_loaded": "auto-loaded",
        "settings_status_skipped": "auto-load reset (session)",
        "settings_status_uploaded": "user-uploaded",
        "settings_confirm_reset_msg": (
            "Stop auto-loading this source for the rest of this session? "
            "Files in `input/` stay; the card on the Dashboard tab will "
            "show 'waiting for file' until you upload one."
        ),
        "settings_confirm_pages_reset_msg": (
            "Clear the design-pages state for this session? The JSON file is "
            "kept; on next app start the values reload."
        ),
        "settings_confirm_btn": "Yes, reset",
        "settings_reset_done": "{label}: auto-load reset",
        "settings_undo_done": "{label}: auto-load re-enabled",
        "settings_delete_file_btn": "🗑",
        "settings_confirm_delete_file_msg": (
            "Permanently delete this snapshot? It will disappear from the "
            "trend charts on the next render. The other saved snapshots are "
            "left untouched."
        ),
        "settings_confirm_delete_check": (
            "I understand this is permanent and cannot be undone."
        ),
        "settings_confirm_delete_btn": "Yes, delete",
        "settings_file_deleted": "Snapshot deleted: {file}",
        # ----- Validation step labels (run inside the dino animation) -----
        "step_xlsx_parse":      "Parse Excel file",
        "step_xlsm_parse":      "Parse macro-enabled Excel file",
        "step_csv_encoding":    "Detect text encoding",
        "step_csv_parse":       "Parse CSV structure",
        "step_master_sheet":    "Find sheet '機能一覧'",
        "step_master_b_col":    "Find last B-column row",
        "step_master_fid":      "Extract Function IDs from F column",
        "step_master_dups":     "Inspect duplicate Function IDs",
        "step_wbs_sheet":       "Find sheet 'メイン'",
        "step_wbs_phase_dates": "Parse phase anchors J6 / N6 (年/月/日)",
        "step_wbs_fid":         "Extract Function IDs from cols E–I (row 16+)",
        "step_defects_columns": "Verify required columns",
        "step_defects_filter":  "Filter to '不具合管理'",
        "step_defects_dates":   "Parse MM/DD/YYYY dates",
        "step_defects_build":   "Extract Function IDs (build dataframe)",
        "step_tests_min_cols":  "Verify ≥6 columns (A–F)",
        "step_tests_fid":       "Extract Function IDs from col A",
        "step_tests_numeric":   "Numeric values in C/D/E/F",
        "step_tests_sanity":    "Sanity: 実施済 ≤ 総設定テスト数",
        "step_code_sheet":      "Find sheet '機能ID別サマリ'",
        "step_code_fid":        "Extract Function IDs from col A",
        "step_code_loc":        "Numeric LoC values in col B",
        "step_load_ok":         "Build dataframe",
        "step_load_failed":     "Build dataframe (failed)",
        # ----- Crash popup -----
        "popup_error_title": "🦖💥 Ouch!",
        "popup_error_hint": (
            "Fix the issue above, then re-drop the file. "
            "(The previously imported file is unchanged.)"
        ),
        "validation_passed": "All checks passed.",
        "validation_warnings": "Loaded with warnings — review the checklist.",
        "log_show_detail": "Show detailed log entry (≤3000 chars)",
        "log_file_caption": "Log file (this session): `{path}`",
        "log_section_title": "Session log",
        "log_section_caption": (
            "All errors raised in this session are also written to a log file "
            "under the project's `log/` folder."
        ),
        # Drill-down panel
        "drilldown_select_hint": (
            "Tip: click any row in the tables above to open a Function-ID "
            "drill-down with all its KPIs, schedule, and related defects."
        ),
        "drilldown_panel_title": "🦖 Function ID drill-down",
        "drilldown_close": "Close drill-down",
        "drilldown_section_wbs":     "Schedule (WBS)",
        "drilldown_to_deadline":     "To deadline",
        "drilldown_deadline_future": "{n} days remaining",
        "drilldown_deadline_today":  "Due today!",
        "drilldown_deadline_overdue": "+{n} days overdue",
        "drilldown_deadline_overdue_badge": "overdue",
        "drilldown_deadline_completed": "Completed ({date})",
        "drilldown_deadline_unknown":   "—",
        "drilldown_help_deadline": (
            "Days between **WBS R column (planned_end)** and today. "
            "Hidden once the feature has an actual_end (shows the "
            "completion date instead)."
        ),
        "drilldown_assignees_label": "Assignees on this feature",
        "drilldown_assignees_none":  "(no WBS sub-tasks)",
        "drilldown_role_progress_label": "Role progress",
        "drilldown_role_progress_caption": (
            "Sub-task completion per role. Helps spot features where "
            "dev is done but test execution hasn't started."
        ),
        "drilldown_subtask_expander": "🔍 Sub-task breakdown (from WBS)",
        "drilldown_subtask_none":  "No WBS sub-tasks for this Function ID.",
        "drilldown_status_completed":   "Completed",
        "drilldown_status_in_progress": "In progress",
        "drilldown_status_not_started": "Not started",
        "drilldown_subtask_col_label":    "Task",
        "drilldown_subtask_col_assignee": "Assignee",
        "drilldown_subtask_col_role":     "Role",
        "drilldown_subtask_col_planned":  "Planned",
        "drilldown_subtask_col_actual":   "Actual",
        "drilldown_subtask_col_progress": "Actual %",
        "drilldown_subtask_col_delay":    "Delay (d)",
        "drilldown_subtask_role_other":   "—",
        "drilldown_section_defects": "Defects",
        "drilldown_section_trend": "Trend (across snapshots)",
        "drilldown_section_trend_caption": (
            "Per-Function-ID history of test counts and LoC, reconstructed "
            "from every timestamped snapshot under `input/`."
        ),
        "drilldown_section_tests":   "Tests",
        "drilldown_section_code":    "Code & Design",
        "drilldown_section_scores":  "Composite scores",
        "drilldown_related_defects": "Related defect rows ({n})",
        "drilldown_no_defects":      "No defect rows for this Function ID.",
        "drilldown_no_wbs":          "No WBS schedule for this Function ID.",
        "drilldown_planned_period":  "Planned period",
        "drilldown_actual_period":   "Actual period",
        "drilldown_planned_effort":  "Planned effort",
        "drilldown_actual_effort":   "Actual effort",
        "drilldown_planned_progress": "Progress (planned)",
        "drilldown_actual_progress":  "Progress (actual)",
        "drilldown_progress":        "Progress (planned vs actual)",
        "drilldown_id_not_found":    "Function ID `{fid}` no longer in the master.",
        "drilldown_strip_title":     "Source coverage for this Function ID",
        "drilldown_source_unloaded": "(not loaded)",
        "drilldown_source_manual":   "(manual entry)",
        "src_design_label":          "Design pages",
        "sec1_title": "1. Drop your sources",
        "sec2_title": "Design page counts",
        "sec2_caption": (
            "Edit the page count per Function ID. Changes save automatically "
            "to `input/design_pages.json` and are restored on the next start. "
            "The list updates as the master changes; blank rows are treated as "
            "no value."
        ),
        "sec2_filled_ids": "Filled IDs",
        "sec2_total_pages": "Total pages entered",
        "sec2_summary_tip": (
            "These pages feed the integrated table and the design-density "
            "KPIs computed in the next step."
        ),
        "design_needs_master": (
            "Upload **Function master** in the Dashboard tab first — this "
            "editor mirrors the IDs found in the master."
        ),
        "design_last_saved": "Last saved: {ts}",
        "design_no_save_yet": "Not saved yet — edit any cell to create the file.",
        "sec3_title": "2. Integrated table",
        "sec3_caption": (
            "{n} rows · {u} unique Function IDs · split into focused tabs so "
            "you don't have to scroll horizontally"
        ),
        "tab_overview": "Overview",
        "tab_kpis": "KPIs",
        "tab_wbs": "WBS",
        "tab_defects": "Defects",
        "tab_tests": "Tests",
        "tab_code": "Code & Design",
        "tab_all": "All columns",
        "tab_all_caption": "All joined columns — horizontal scroll if needed.",
        "kpi_summary_title": "Project-wide KPIs",
        "metric_total_loc": "Total LoC",
        "metric_open_defects": "Open faults (Redmine)",
        "metric_test_run_rate": "Test run rate",
        "metric_test_pass_rate": "Test pass rate",
        "metric_avg_bug_density": "Avg defect density (test spec)",
        "metric_avg_test_density": "Avg test density",
        "metric_at_risk": "At-risk functions",
        "metric_delayed": "Delayed functions",
        "metric_avg_health": "Avg health",
        "metric_help_at_risk": (
            "**🦕 At-risk functions**\n\n"
            "🧮 Count of Function IDs with **risk_score ≥ 0.5**.\n"
            "risk_score (per feature, 0..1) = "
            "`0.4 × unresolved-defect + 0.2 × un-executed + 0.2 × "
            "delay + 0.2 × defect-density`, each normalised to 0..1 "
            "via dataset min-max.\n\n"
            "📂 Source: Redmine defects × test-spec counts × WBS × "
            "LoC per Function ID.\n\n"
            "💡 Project-wide summary of 'needs immediate attention'. "
            "Different from the alert-tab risk score (that one uses a "
            "4-metric breach scoring specifically for triage)."
        ),
        "metric_help_delayed": (
            "**🦕 Delayed functions**\n\n"
            "🧮 Count of Function IDs with **delay_days > 0**.\n"
            "delay_days = `max(0, actual_end − planned_end)`; for in-"
            "flight features, `today − planned_end` when positive.\n\n"
            "📂 Source: WBS **column R** (planned_end) and **column T** "
            "(actual_end).\n\n"
            "💡 How many features are running behind schedule. Zero is "
            "the healthy baseline."
        ),
        # ----- column / chart / calendar tooltips (hover) -----
        "help_func_id": (
            "**🦕 Function ID**\n\n"
            "Unique identifier for each function.\n\n"
            "📂 Source: Function ID master (機能一覧 sheet, col F).\n\n"
            "💡 Used as the join key across every data source."
        ),
        "help_func_name": (
            "**🦕 Function name**\n\n"
            "Human-readable name for the function.\n\n"
            "📂 Source: Function ID master (機能一覧 sheet, col G).\n\n"
            "💡 The same Function ID may legitimately appear with multiple "
            "names — every (ID, name) pair is kept."
        ),
        "help_defect_total": (
            "**🦕 Fault count (Redmine)**\n\n"
            "Number of faults logged in Redmine for this Function ID.\n\n"
            "📂 Source: Redmine defect list, filtered to tracker = '不具合管理'.\n\n"
            "💡 Includes both open and closed faults.\n\n"
            "⚠ This is **not** the test-spec NG count "
            "(see *Defect density – test spec* / *Defect rate – test spec*)."
        ),
        "help_defect_unresolved": (
            "**🦕 Unresolved faults (Redmine)**\n\n"
            "Faults whose 実終了日 is still empty.\n\n"
            "📂 Source: Redmine defect list (不具合管理).\n\n"
            "💡 Treat as the current open backlog.\n\n"
            "⚠ This is **not** the test-spec NG count."
        ),
        "help_incident_rate": (
            "**🦕 Fault rate (Redmine)**\n\n"
            "🧮 Redmine `defect_total` ÷ test-spec `実施済` "
            "(Redmine fault count over executed tests).\n\n"
            "📂 Source: Redmine defect list (tracker='不具合管理') × "
            "test counts column D.\n\n"
            "💡 How often a Redmine-tracked fault was raised per executed "
            "test case.\n\n"
            "⚠ This is **not** *Defect rate – test spec* (NG / Total tests). "
            "The numerator and the denominator come from different sources."
        ),
        "help_test_total": (
            "**🦕 Total tests (総設定テスト数)**\n\n"
            "Planned test cases for this function.\n\n"
            "📂 Source: Test counts per spec, column C.\n\n"
            "💡 Denominator for test run rate and density."
        ),
        "help_test_run": (
            "**🦕 Tests run (実施済)**\n\n"
            "Test cases that have already been executed.\n\n"
            "📂 Source: Test counts per spec, column D."
        ),
        "help_test_ok": (
            "**🦕 OK**\n\nTests that executed successfully.\n\n"
            "📂 Source: Test counts per spec, column E."
        ),
        "help_test_ng": (
            "**🦕 NG — Defect count (test spec)**\n\n"
            "Tests that failed.\n\n"
            "📂 Source: Test counts per spec, column F.\n\n"
            "💡 Numerator for defect rate (test spec) and defect density "
            "(test spec).\n\n"
            "⚠ This is **not** the Redmine fault count "
            "(*Fault count (Redmine)*)."
        ),
        "help_test_notrun": (
            "**🦕 Not run (未実施)**\n\n"
            "🧮 総設定テスト数 − 実施済.\n\n"
            "📂 Source: derived from test counts columns C and D.\n\n"
            "💡 Visible work remaining to complete the test plan."
        ),
        "help_loc": (
            "**🦕 LoC (Lines of Code)**\n\n"
            "Physical lines for this function's implementation.\n\n"
            "📂 Source: LoC per Function ID, sheet 機能ID別サマリ, col B.\n\n"
            "💡 Pre-aggregated per Function ID."
        ),
        "help_design_pages": (
            "**🦕 Design pages**\n\n"
            "Manually entered design document page count.\n\n"
            "📂 Source: Design pages tab (saved to input/design_pages.json).\n\n"
            "💡 Same value applies to every name sharing a Function ID."
        ),
        "help_planned_effort": (
            "**🦕 Planned effort**\n\n"
            "Scheduled person-hours/days.\n\n"
            "📂 Source: WBS sheet メイン, column **P**, row 16+."
        ),
        "help_actual_effort": (
            "**🦕 Actual effort**\n\n"
            "Logged person-hours/days.\n\n"
            "📂 Source: WBS column **U**."
        ),
        "help_planned_start": (
            "**🦕 Planned start**\n\n📂 Source: WBS column **Q**."
        ),
        "help_planned_end": (
            "**🦕 Planned end**\n\n📂 Source: WBS column **R**."
        ),
        "help_actual_start": (
            "**🦕 Actual start**\n\n📂 Source: WBS column **S**."
        ),
        "help_actual_end": (
            "**🦕 Actual end**\n\n"
            "📂 Source: WBS column **T**.\n\n"
            "💡 Empty cell means the work hasn't completed yet."
        ),
        "help_actual_progress": (
            "**🦕 Actual progress %**\n\n"
            "Reported percent complete.\n\n"
            "📂 Source: WBS column **V**."
        ),
        "help_planned_progress": (
            "**🦕 Planned progress %**\n\n"
            "Expected percent complete by today per the schedule.\n\n"
            "📂 Source: WBS column **AA**."
        ),
        "help_bug_density": (
            "**🦕 Defect density – test spec**\n\n"
            "🧮 NG ÷ LoC (test-spec NG count over lines of code).\n\n"
            "📂 Source: test counts column F ÷ code-LoC column B.\n\n"
            "💡 Defects per line of code. Higher = more buggy.\n\n"
            "⚠ This is **not** the Redmine fault count (*Fault count (Redmine)*)."
        ),
        "help_test_density": (
            "**🦕 Test density**\n\n"
            "🧮 総設定テスト数 ÷ 設計書ページ数.\n\n"
            "📂 Source: test counts column C ÷ design pages (manual input).\n\n"
            "💡 Tests per design page. Low values may indicate under-tested specs."
        ),
        "help_complexity": (
            "**🦕 Complexity**\n\n"
            "🧮 LoC ÷ 設計書ページ数.\n\n"
            "📂 Source: code-LoC column B ÷ design pages (manual input).\n\n"
            "💡 Implementation density per page of design."
        ),
        "help_test_run_rate": (
            "**🦕 Test run rate**\n\n"
            "🧮 実施済 ÷ 総設定テスト数.\n\n"
            "📂 Source: test counts column D ÷ column C.\n\n"
            "💡 Test execution progress. 100% = every planned test ran."
        ),
        "help_test_pass_rate": (
            "**🦕 Test pass rate**\n\n"
            "🧮 OK ÷ 実施済.\n\n"
            "📂 Source: test counts column E ÷ column D.\n\n"
            "💡 Quality of executed tests. Drops below 90% warrant investigation."
        ),
        "help_defect_rate": (
            "**🦕 Defect rate – test spec**\n\n"
            "🧮 NG ÷ 総設定テスト数 (test-spec NG over planned test cases).\n\n"
            "📂 Source: test counts column F ÷ column C.\n\n"
            "💡 Failure rate against the full test plan.\n\n"
            "⚠ This is **not** the Redmine fault rate (*Fault rate (Redmine)*)."
        ),
        "help_delay_days": (
            "**🦕 Delay (days)**\n\n"
            "🧮 max(0, actual_end − planned_end). For ongoing items: "
            "today − planned_end (if positive).\n\n"
            "📂 Source: WBS column R (planned end) and column T (actual end).\n\n"
            "💡 0 = on time. Positive numbers grow until the work completes."
        ),
        "help_delay_rate": (
            "**🦕 Delay rate**\n\n"
            "🧮 delay_days ÷ planned duration, capped at 1.0.\n\n"
            "📂 Source: WBS columns Q (planned start) and R (planned end) "
            "give the duration; delay_days is derived above.\n\n"
            "💡 0 = on time. 1.0 = at least double the planned timeline."
        ),
        "help_health_score": (
            "**🦕 Health score**\n\n"
            "🧮 test_run_rate − *Defect rate – test spec* − delay_rate "
            "(range ≈ −2…1).\n\n"
            "📂 Source: derived from the three metrics above "
            "(= test counts × WBS).\n\n"
            "💡 Higher is healthier. Negative values flag trouble."
        ),
        "help_risk_score": (
            "**🦕 Risk score**\n\n"
            "🧮 Weighted blend of normalized inputs:\n"
            "0.4×*Unresolved faults (Redmine)* + 0.2×not_run + 0.2×delay_days "
            "+ 0.2×*Defect density – test spec*.\n\n"
            "📂 Source: Redmine defects × test counts × WBS × code-LoC.\n\n"
            "💡 Each input is min-max normalized within the dataset, so 0…1. "
            "≥0.5 marks an at-risk function."
        ),
        # Charts / calendar
        "help_chart_progress_gap": (
            "**🦕 Progress: planned vs actual**\n\n"
            "Paired horizontal bars per Function ID.\n\n"
            "📂 Source: WBS columns **V** (actual %) and **AA** (planned %).\n\n"
            "💡 Wider gap with the planned bar above means slipping schedule.\n\n"
            "⚠ Bars where actual > planned are colored orange and tagged "
            "with ⚠ — actual exceeding the plan can flag over-reporting."
        ),
        "help_chart_test_coverage": (
            "**🦕 Test coverage**\n\n"
            "Stacked bars: OK / NG / not run per Function ID.\n\n"
            "📂 Source: Test counts per spec (E / F / C−D)."
        ),
        "help_chart_test_density": (
            "**🦕 Test density (test count sufficiency)**\n\n"
            "🧮 総設定テスト数 ÷ 設計書ページ数 — sorted ascending so the bottom "
            "of the chart is the under-tested specs.\n\n"
            "📂 Source: Test counts per spec (C), design pages."
        ),
        "help_chart_incident_rate": (
            "**🦕 Fault rate (Redmine)**\n\n"
            "🧮 Redmine `defect_total` ÷ test-spec `実施済` per Function ID, "
            "sorted descending so the worst rates surface at the top.\n\n"
            "📂 Source: Redmine defect tracker (defect_total) over Test "
            "counts (D 実施済).\n\n"
            "⚠ This is **not** *Defect rate – test spec* (NG / Total tests)."
        ),
        "help_chart_defect_class": (
            "**🦕 Fault root cause breakdown (Redmine 問題分類)**\n\n"
            "Donut share + Top-N table of fault rows grouped by the "
            "Redmine `問題分類` column.\n\n"
            "📂 Source: Redmine defect list (`不具合管理` tracker), 問題分類 "
            "column.\n\n"
            "💡 Use the filter to drill into a Function ID subset and see "
            "where a particular feature's faults are coming from."
        ),
        "help_chart_loc_vs_ng": (
            "**🦕 LoC × NG**\n\n"
            "Scatter: x=LoC, y=NG, size=design pages, color=risk_score.\n\n"
            "📂 Source: LoC per Function ID (B), Test counts per spec (F), "
            "design pages, risk_score.\n\n"
            "💡 Top-right + red = large + buggy + risky. Watch outliers."
        ),
        "help_chart_design_impl_gap": (
            "**🦕 Design pages × LoC**\n\n"
            "Implementation effort vs design size. Dashed line = average complexity.\n\n"
            "📂 Source: design pages (manual) and LoC per Function ID (B).\n\n"
            "💡 Above the line = denser implementation than average."
        ),
        "help_chart_risk_heatmap": (
            "**🦕 Risk dimensions heatmap**\n\n"
            "Function ID × five risk dimensions (Defect density – test spec, "
            "Fault rate (Redmine), delay_rate, test_run_rate inverted, "
            "Test density inverted), each min-max normalized.\n\n"
            "💡 Red rows are concerning across multiple dimensions."
        ),
        "chart_risk_dims_legend": "Legend — what each row means",
        # ----- Role analytics (WBS sub-task assignees × Redmine defects) -----
        "role_analytics_title":   "Assignee × role analysis",
        "help_role_analytics": (
            "**🦕 Assignee × role analysis**\n\n"
            "Cross-references WBS sub-task assignees (N column on rows marked "
            "with ● in L) against Redmine defect counts and test-spec NG, to "
            "surface who touched which feature in which role — and how the "
            "features they worked on are doing in terms of quality.\n\n"
            "📂 Source: WBS (sub-task 担当者 + task name) × Redmine defects × "
            "test counts.\n\n"
            "💡 Roles are derived from keywords in each sub-task's name: "
            "`開発` / `実装` → Development, `テスト仕様書作成` → Test-spec, "
            "`テスト実施` → Test-execution (but `再テスト実施` is excluded "
            "from test_exec)."
        ),
        "role_analytics_view1_title":   "View 1 — Feature × assignee-by-role + quality KPIs",
        "role_analytics_view1_caption": (
            "One row per Function ID. Each role column lists the distinct "
            "assignees drawn from the WBS sub-task N-column (with '/' as "
            "separator); the right-hand columns are the feature's quality "
            "signals. Sorted by Redmine fault rate (worst first)."
        ),
        "role_analytics_view2_title":   "View 2 — Assignee summary",
        "role_analytics_view2_caption": (
            "One row per assignee. Role columns are the count of WBS "
            "sub-tasks the person handled for that role, followed by the "
            "number of WBS-touched features, total Redmine defects on "
            "those features, average Redmine fault rate across them, and "
            "their top-3 Redmine 問題分類. Sorted by total Redmine "
            "defects, descending."
        ),
        "role_analytics_view3_title":   "View 3 — Assignee × 問題分類 heatmap",
        "role_analytics_view3_caption": (
            "Counts Redmine defects on each assignee's WBS-linked features, "
            "broken down by Redmine 問題分類. Use the role filter to see, "
            "e.g., what kinds of defects the developers' features attract "
            "vs. the test writers'."
        ),
        "role_analytics_view3_role_label": "Filter by role",
        "role_analytics_view3_role_all":   "All roles",
        "help_role_analytics_view3_role_filter": (
            "Limit the heatmap to a single role so you can compare, "
            "e.g., developers against each other without test-writer "
            "rows muddying the picture."
        ),
        # Bubble map + heat strip (advanced per-assignee viz).
        "role_analytics_bubble_title": (
            "Assignee bubble map (breadth × quality × defect exposure)"
        ),
        "role_analytics_bubble_caption": (
            "One bubble per assignee with a **measurable** quality "
            "signal (at least one executed test AND at least one "
            "defect on their features). X = features they're on in "
            "the WBS, Y = avg Redmine fault rate on those features, "
            "size = total Redmine defects, colour = their dominant "
            "role (derived from WBS sub-task names). Dashed lines "
            "mark the overall mean Y and the median X — top-right is "
            "the 'attention' quadrant, bottom-right is 'reliable "
            "coverage'. Assignees with no executed tests or zero "
            "defects don't belong on this scatter — they appear in "
            "the two watch-lists below instead."
        ),
        "role_analytics_bubble_color_legend": "Dominant role",
        # Core explanation — safe for both on-screen and PDF use.
        "role_analytics_dominant_role_note": (
            "**Dominant role** = the bubble is coloured by the role this "
            "assignee has the **most** WBS sub-tasks in. Tie-break order: "
            "**Dev › Test-spec › Test-exec**. A single colour hides "
            "multi-role involvement — refer to the role summary above for "
            "the full per-role counts."
        ),
        # Screen-only addendum pointing at the hover tooltip.
        "role_analytics_dominant_role_hover_hint": (
            "Hover over a bubble to see its per-role breakdown."
        ),
        # Watch-list sections below the bubble map. One for "tested but
        # no defects registered" (ambiguous — genuine quality or
        # under-reporting) and one for "tests planned but not executed"
        # (process gap, quality cannot be measured yet).
        "role_analytics_watch_zero_defect_title": (
            "⚠️ Watch-list A: zero defects registered"
        ),
        "role_analytics_watch_zero_defect_caption": (
            "These assignees had at least one test executed on their "
            "features (test-counts **column D = 実施済** ≥ 1) but "
            "**zero rows in defects.csv (Redmine export)** for any of "
            "their Function IDs. Three readings: (1) genuine high "
            "quality, (2) **defects went unreported** in Redmine, "
            "(3) the **test-spec coverage was shallow** and missed the "
            "bugs. 📂 Cross-check: defects.csv (registration reality), "
            "the test-spec sheet (coverage depth), test-counts column "
            "D (executed count)."
        ),
        "role_analytics_watch_no_exec_title": (
            "⚠️ Watch-list B: no tests executed yet"
        ),
        "role_analytics_watch_no_exec_caption": (
            "Every feature these assignees touched has **実施済 = 0** "
            "in the test-counts CSV — tests were **planned but never "
            "executed**, so the Y-axis (fault rate = defects ÷ 実施済) "
            "is undefined. Not 'bad quality', just 'not yet measured'. "
            "📂 Check: the test-spec sheet (are the plans in place), "
            "test-counts **column D** (which features stay at 0), "
            "**WBS column T** (end date — are they slipping?). "
            "Follow up on the actual test-execution schedule."
        ),
        "role_analytics_watch_empty": "None — every assignee is on the main scatter.",
        "role_analytics_watch_col_assignee": "Assignee",
        "role_analytics_watch_col_features": "Features touched",
        "role_analytics_watch_col_defects":  "Defects",
        "role_analytics_watch_col_roles":    "Role breakdown",
        "role_analytics_strip_title": (
            "Problem-class mix per assignee (stacked % of their defects)"
        ),
        "role_analytics_strip_caption": (
            "Each horizontal bar is one assignee's Redmine defect pool "
            "(100%) on their WBS-linked features, broken down by Redmine "
            "問題分類. Numbers inside each segment are the raw Redmine "
            "defect counts for that (person, category) pair; hover for "
            "'<count>件 (<pct>%)'. `n=` to the right is the assignee's "
            "total Redmine defect count the percentages divide."
        ),
        "role_analytics_strip_other": "Other",
        # In-tab navigation for the Charts tab.
        "toc_jump_label":    "Jump to",
        "toc_back_to_top":   "Back to top",
        # Unified PDF-button microcopy.
        "pdf_btn_generate_short":  "Generate PDF",
        "pdf_btn_download_short":  "Download PDF",
        "pdf_generating":          "Generating PDF…",
        # Role analytics PDF-export labels
        "ra_pdf_btn_generate":       "📄 PDF",
        "ra_pdf_btn_generate_help":  (
            "Export the Assignee × role analysis as a standalone PDF "
            "report (inputs → role rules → three tables → bubble map → "
            "problem-class strip)."
        ),
        "ra_pdf_btn_download":       "⬇ PDF",
        "ra_pdf_title":              "Assignee × role analysis report",
        "ra_pdf_filter_active":      (
            "⚠ Function ID filter is active — this report only covers "
            "the filtered subset."
        ),
        "ra_pdf_h_inputs":           "1. Inputs (where the numbers come from)",
        "ra_pdf_input_wbs":          "WBS sub-tasks (L=●, N=assignee)",
        "ra_pdf_input_wbs_note":     (
            "Sub-task names in the WBS are scanned for the role keywords; "
            "column N of the WBS carries the 担当者."
        ),
        "ra_pdf_input_defects":      "Redmine defects (trackered '不具合管理')",
        "ra_pdf_input_defects_note":  (
            "Per-feature defect counts and 問題分類 come from the Redmine "
            "defect list."
        ),
        "ra_pdf_input_tests":        "Test counts CSV",
        "ra_pdf_input_tests_note":   (
            "総設定テスト数 / 実施済 / NG columns come from the test counts CSV."
        ),
        "ra_pdf_h_rules":            "2. Role classification rules",
        "ra_pdf_rules_body": (
            "Each WBS sub-task name is matched (NFKC-normalised substring) "
            "against three fixed keywords. A WBS sub-task whose name "
            "contains multiple keywords attributes to all matching roles; "
            "WBS sub-tasks whose N-column (assignee) cell is empty surface "
            "as <b>(unassigned)</b>."
        ),
        "ra_pdf_rule_dev":       "<b>開発</b> / <b>実装</b> → Development",
        "ra_pdf_rule_test_spec": "<b>テスト仕様書作成</b> → Test-spec",
        "ra_pdf_rule_test_exec": (
            "<b>テスト実施</b> → Test-execution "
            "(excludes <b>再テスト実施</b>)"
        ),
        "ra_pdf_h_view1":            "3. Feature × assignee-by-role + quality KPIs",
        "ra_pdf_h_view2":            "4. Assignee summary",
        "ra_pdf_h_bubble":           "5. Assignee bubble map",
        "ra_pdf_h_strip":            "6. Problem-class mix per assignee",
        "ra_pdf_h_heatmap":          "7. Assignee × 問題分類 heatmap",
        "ra_pdf_no_data":            "No data to plot.",
        "ra_pdf_footer":             "Generated {when}",
        "role_analytics_no_subtasks": (
            "No WBS sub-tasks (rows marked ● in column L) are available — "
            "upload a WBS with sub-task entries to populate this section."
        ),
        "role_analytics_no_matches": (
            "No sub-task names contain the role keywords "
            "(開発 / 実装 / テスト仕様書作成 / テスト実施 — "
            "but 再テスト実施 is excluded from test-execution)."
        ),
        "role_dev":          "Development (開発 / 実装)",
        "role_test_spec":    "Test-spec (テスト仕様書作成)",
        "role_test_exec":    "Test-execution (テスト実施)",
        "role_unassigned":   "(unassigned)",
        "role_count_dev":       "Dev (WBS sub-tasks)",
        "role_count_test_spec": "Test-spec (WBS sub-tasks)",
        "role_count_test_exec": "Test-exec (WBS sub-tasks)",
        "col_feature":          "Function ID : Name",
        "col_assignee":         "Assignee",
        "col_feature_count":    "Features touched (WBS)",
        "col_avg_incident_rate": "Avg fault rate (Redmine)",
        "col_top3_problems":    "Top-3 問題分類 (Redmine)",
        "problem_class_uncategorized": "(uncategorized)",
        # Column tooltips (hover help) for the role-analytics tables.
        "help_col_feature": (
            "**🦕 Function ID : Name**\n\n"
            "Function ID followed by 機能名 (master-authoritative).\n\n"
            "📂 Source: Function master columns F (ID) + G (name)."
        ),
        "help_col_assignee": (
            "**🦕 Assignee**\n\n"
            "WBS sub-task assignee name.\n\n"
            "📂 Source: WBS column N (rows marked ● in column L).\n\n"
            "💡 Full-width / doubled / padded spaces are normalized so "
            "the same person doesn't split into two groups."
        ),
        "help_col_feature_count": (
            "**🦕 Features touched**\n\n"
            "Number of distinct Function IDs this person has a WBS "
            "sub-task on, across any role.\n\n"
            "📂 Source: WBS sub-task rows (L=● AND N has a name)."
        ),
        "help_col_avg_incident_rate": (
            "**🦕 Avg fault rate (Redmine)**\n\n"
            "🧮 Mean of this person's features' Redmine fault rates "
            "(defect_total ÷ 実施済).\n\n"
            "📂 Source: Redmine defects × test counts column D "
            "(features drawn from WBS).\n\n"
            "💡 Feel for the quality of the features they touched."
        ),
        "help_col_top3_problems": (
            "**🦕 Top-3 問題分類 (per assignee)**\n\n"
            "Top-3 Redmine 問題分類 on the related features, with counts.\n\n"
            "📂 Source: Redmine defect list 問題分類 column, "
            "filtered to the assignee's features."
        ),
        "help_col_feature_top3_problems": (
            "**🦕 Top-3 問題分類 (per feature)**\n\n"
            "Top-3 Redmine 問題分類 for defects on THIS feature only.\n\n"
            "📂 Source: Redmine defect list 問題分類 column, "
            "filtered to this Function ID."
        ),
        "help_role_count_dev": (
            "**🦕 Dev (# sub-tasks)**\n\n"
            "How many sub-tasks this person is on whose name includes "
            "`開発` or `実装`.\n\n"
            "📂 Source: WBS sub-task names (L=●) + N column (assignee)."
        ),
        "help_role_count_test_spec": (
            "**🦕 Test-spec (# sub-tasks)**\n\n"
            "How many sub-tasks this person is on whose name includes "
            "`テスト仕様書作成`.\n\n"
            "📂 Source: WBS sub-task names (L=●) + N column (assignee)."
        ),
        "help_role_count_test_exec": (
            "**🦕 Test-exec (# sub-tasks)**\n\n"
            "How many sub-tasks this person is on whose name includes "
            "`テスト実施`. **`再テスト実施` is excluded.**\n\n"
            "📂 Source: WBS sub-task names (L=●) + N column (assignee)."
        ),
        "help_role_assignees_dev": (
            "**🦕 Dev assignees**\n\n"
            "Distinct assignees on this feature's dev sub-tasks "
            "(`開発` / `実装`).\n\n"
            "📂 Source: WBS sub-task names (L=●) + N column (assignee)."
        ),
        "help_role_assignees_test_spec": (
            "**🦕 Test-spec assignees**\n\n"
            "Distinct assignees on this feature's `テスト仕様書作成` sub-tasks.\n\n"
            "📂 Source: WBS sub-task names (L=●) + N column (assignee)."
        ),
        "help_role_assignees_test_exec": (
            "**🦕 Test-exec assignees**\n\n"
            "Distinct assignees on this feature's `テスト実施` sub-tasks. "
            "**`再テスト実施` is excluded.**\n\n"
            "📂 Source: WBS sub-task names (L=●) + N column (assignee)."
        ),
        # ----- PDF report -----
        "pdf_btn_generate": "Generate PDF report",
        "pdf_btn_download": "Download PDF",
        "pdf_progress": "Building report (this may take ~20–30 s)…",
        "pdf_done": "Report ready — click Download.",
        "pdf_error": "Could not build the report: {err}",
        "pdf_step_cover":    "Building cover + KPI table…",
        "pdf_step_chart":    "Rendering chart ({i}/{n}): {title}",
        "pdf_step_gantt":    "Rendering schedule (Gantt)…",
        "pdf_step_assemble": "Assembling PDF…",
        "pdf_dialog_title":  "🦖 Generating PDF report",
        "pdf_dialog_subtitle": (
            "Per-feature bar charts render only the selected features."
        ),
        "pdf_dialog_close":  "Close",
        "chart_truncated_note": "Showing worst {shown} of {total} features",
        "pdf_select_title":   "🦖 Select features for the PDF report",
        "pdf_select_caption": (
            "Pick up to 30 Function IDs. Per-feature bar charts and the "
            "Gantt in the report will only include these rows."
        ),
        "pdf_select_label":   "Features (max 30)",
        "pdf_select_count":   "{n} / 30 selected",
        "pdf_select_error_empty": (
            "Please select at least one feature before generating."
        ),
        "pdf_btn_confirm":    "Start generation",
        "pdf_title": "dashboard4dx — Project Report",
        "pdf_generated_at": "Generated",
        "pdf_toc_title":    "Table of Contents",
        "pdf_section_kpi": "Project-wide KPIs",
        "pdf_section_charts": "Charts",
        "pdf_section_schedule": "Schedule",
        "pdf_no_chart": "No data to plot for this section.",
        "pdf_chart_definition": "Definition",
        # ----- PDF: test-density focused report -----
        "td_pdf_btn_generate":      "📄 PDF",
        "td_pdf_btn_generate_help": (
            "Export this chart as a standalone PDF report "
            "(inputs → density → below-threshold list → advice)."
        ),
        "td_pdf_btn_download":      "⬇ PDF",
        "td_pdf_title":             "Test density per Function ID — report",
        "td_pdf_filter_active":     "⚠ Function ID filter is active — this report only covers the filtered subset.",
        "td_pdf_h_inputs":          "1. Inputs (where the numbers come from)",
        "td_pdf_col_item":          "Item",
        "td_pdf_col_source":        "Source file / location",
        "td_pdf_col_method":        "Method",
        "td_pdf_input_tests":       "Total tests (総設定テスト数)",
        "td_pdf_input_pages":       "Design pages — visual count (設計書頁数（目視確認）)",
        "td_pdf_input_master":      "Function master (ID / name)",
        "td_pdf_method_auto":       "Auto-loaded CSV/XLSX",
        "td_pdf_method_manual":     "Manual (Dashboard settings)",
        "td_pdf_input_tests_note":  "(column C of the test counts CSV)",
        "td_pdf_input_pages_note":  "(input/design_pages.json, edited in the Design-pages tab)",
        "td_pdf_h_output":          "2. Output (the density metric)",
        "td_pdf_output_formula":    "<b>Test density</b> = Total tests ÷ Design pages (visual count)",
        "td_pdf_output_unit":       "Unit: tests per design-spec page",
        "td_pdf_output_meaning":    (
            "Meaning: how many planned tests exist per page of the design spec. "
            "A low value suggests the test plan under-covers the specification."
        ),
        "td_pdf_h_threshold":       "3. Warning threshold",
        "td_pdf_threshold_current":      "Threshold used in this report",
        "td_pdf_threshold_default":      "Default",
        "td_pdf_threshold_where":        "Where to change",
        "td_pdf_threshold_where_value": (
            "Dashboard → Settings tab → "
            "<b>\"Test density warning threshold\"</b>"
        ),
        "td_pdf_threshold_meaning": (
            "<b>Meaning:</b> any Function ID whose density falls below this "
            "value is flagged as <b>attention-needed</b> (red bar + ⚠) "
            "throughout this report — chart, below-threshold list, and the "
            "catch-up estimate all key on the same threshold."
        ),
        "td_pdf_threshold_unit":    "tests / page",
        "td_pdf_h_summary":         "4. Summary",
        "td_pdf_summary_total":     "Function IDs in scope",
        "td_pdf_summary_threshold": "Warning threshold",
        "td_pdf_summary_above":     "≥ threshold",
        "td_pdf_summary_below":     "< threshold (needs attention)",
        "td_pdf_summary_mean":      "Mean density",
        "td_pdf_summary_median":    "Median density",
        "td_pdf_summary_min_max":   "Min / Max density",
        "td_pdf_h_chart":           "5. Test density per Function ID (ascending)",
        "td_pdf_h_below":           "6. Function IDs below threshold",
        "td_pdf_col_fid":           "Function ID",
        "td_pdf_col_name":          "Name",
        "td_pdf_col_tests":         "Total tests",
        "td_pdf_col_pages":         "Design pages (visual)",
        "td_pdf_col_density":       "Density",
        "td_pdf_col_gap":           "Gap vs threshold",
        "td_pdf_below_none":        "No Function IDs are below the threshold — nothing to escalate.",
        "td_pdf_h_catchup":         "7. Estimated tests needed to reach the threshold",
        "td_pdf_col_current":       "Current density",
        "td_pdf_col_target":        "Target",
        "td_pdf_col_additional":    "Additional tests (recommended)",
        "td_pdf_catchup_note":      "Recommended additional tests = ceil((threshold − current density) × design pages).",
        "td_pdf_h_advice":          "8. Recommended action to exceed the threshold",
        "td_pdf_advice_body":       (
            "<b>Walk through the design spec one page at a time and audit "
            "test-viewpoint coverage.</b><br/>"
            "For each page, re-estimate how many tests are actually required "
            "across viewpoints such as normal / abnormal / boundary / "
            "permission / performance / logging."
        ),
        "td_pdf_h_notes":           "Notes",
        "td_pdf_notes_body":        (
            "・This metric gauges <b>test-count sufficiency</b>; it does not "
            "measure the quality or completeness of the tests themselves.<br/>"
            "・Design-page counts are manual visual counts — apply a consistent "
            "counting rule (e.g. exclude TOC / revision history).<br/>"
            "・Simple features may legitimately sit below the threshold. "
            "Combine this signal with feature risk before acting."
        ),
        "td_pdf_footer":            "Generated {when} · snapshot file: {src}",
        "td_pdf_pages":              "{n} / {total}",
        "help_chart_loc_trend": (
            "**🦕 LoC trend**\n\n"
            "Total LoC across saved code snapshots over time.\n\n"
            "📂 Source: input/<date>/code/*.xlsx (snapshot date from filename).\n\n"
            "💡 Need ≥2 snapshots."
        ),
        "help_chart_test_trend": (
            "**🦕 Test counts trend**\n\n"
            "Total tests vs executed across saved test snapshots.\n\n"
            "📂 Source: input/<date>/tests/*.csv.\n\n"
            "💡 Gap between Total and Executed = test backlog over time."
        ),
        "help_chart_bug_trend": (
            "**🦕 Defect trend**\n\n"
            "Weekly opened/closed bars + cumulative open line.\n\n"
            "📂 Source: Redmine defect list (実開始日 vs 実終了日, weekly buckets)."
        ),
        "help_gantt_title": (
            "**🦕 Gantt — planned vs actual**\n\n"
            "Horizontal time bars per Function ID. Gray = planned (Q–R), "
            "green = actual (S–T). Today marked with dashed yellow line.\n\n"
            "💡 Gray-only bar = work yet to start; green extending past gray = late."
        ),
        "help_calendar_title": (
            "**🦕 Calendar**\n\n"
            "Monthly / weekly / list views of WBS schedule and defect lifespans.\n\n"
            "📂 Source: WBS Q–T + Redmine defect list (実開始日, 実終了日).\n\n"
            "💡 Toggle layers above. Defects in red = unresolved."
        ),
        "col_bug_density":   "Defect density – test spec (NG/LoC)",
        "col_defect_total":  "Fault count (Redmine)",
        "col_defect_unresolved": "Unresolved faults (Redmine)",
        "col_incident_rate": "Fault rate (Redmine, defect_total/Executed)",
        "col_test_ng":       "Defect count – test spec (NG)",
        "col_test_density":  "Test density (test count sufficiency, tests/page)",
        "col_complexity":    "Complexity (LoC/page)",
        "col_test_run_rate": "Test run rate",
        "col_test_pass_rate":"Test pass rate",
        "col_defect_rate":   "Defect rate – test spec (NG/Total tests)",
        "col_delay_days":    "Delay (days)",
        "col_delay_rate":    "Delay rate",
        "col_health_score":  "Health score",
        "col_risk_score":    "Risk score",
        "raw_previews": "Raw source previews (first 10 rows each)",
        "wbs_label_short": "WBS",
        "defects_label_short": "Redmine defect list (filtered to 不具合管理)",
        "tests_label_short": "Test counts per spec",
        "code_label_short": "LoC per Function ID",
        "master_unlock_info": (
            "Upload the **Function master** card to unlock the dashboard. "
            "Other cards are optional — anything you skip is just left out of "
            "the joins."
        ),
        # Card chrome
        "badge_required": "required",
        "badge_optional": "optional",
        "card_drop_label": "Drop {label} here",
        "status_waiting": "waiting for file…",
        "status_ok": "OK · {n} rows · {u} unique IDs",
        "status_ok_no_fid": "OK · {n} rows",
        "status_failed": "validation failed",
        "origin_upload": "just uploaded",
        "origin_auto": "auto-loaded from input/",
        "origin_snapshot": "snapshot {date}",
        "origin_ingested_at": "ingested {ts}",
        "toast_loaded": "{label} loaded · saved to {path}",
        "toast_failed": "{label}: {msg}",
        "save_warn": "validated, but couldn't save to input/: {err}",
        "read_prev_warn": "failed to read previous file: {err}",
        "read_upload_err": "cannot read upload: {err}",
        # Source spec labels / hints
        "src_master_label": "Function ID master",
        "src_master_hint": "sheet 機能一覧 · col F=ID, G=name",
        "src_wbs_label": "WBS",
        "src_wbs_hint": "sheet メイン · row 16+ · ID in cols E–I",
        "src_defects_label": "Redmine defect list",
        "src_defects_hint": "tracker / status / function_id …",
        "src_tests_label": "Test counts per spec",
        "src_tests_hint": "A=ID · C=total · D=run · E=OK · F=NG",
        "src_code_label": "LoC per Function ID",
        "src_code_hint": "sheet 機能ID別サマリ · A=ID, B=LoC",
        "src_roster_label":   "Roster",
        "src_roster_hint":    "team / assignee / PC / phone / VPN (xlsx)",
        "src_calendar_label": "Calendar (events + non-working days)",
        "src_calendar_hint":  "2 sheets: events + non-working days (xlsx)",
        "src_backlog_label":  "Backlog issues",
        "src_backlog_hint":   "csv from Backlog.com (SJIS), 13 cols incl. キーID / ID / 種別 / 状態 / …",
        "card_template_dl":   "⬇ template ({label})",
        "card_dl_template_help": "Download a template",
        "card_dl_sample_help":   "Download sample data",
        "card_dl_latest_help":   "Download the most recently loaded file ({name})",
        "src_rail_hint":      "⇄ scroll horizontally to browse all sources",
        "calendar_layer_events":  "Show events",
        "calendar_layer_nonwork": "Show non-working days",
        # Validation messages
        "err_zero_rows": "parsed 0 rows — check sheet name / column layout",
        "warn_master_dups": (
            "{n} Function IDs appear with multiple names — kept all rows."
        ),
        "warn_tests_overrun": "{n} rows have 実施済 > 総設定テスト数",
        "warn_tests_nan_total": "{n} rows have non-numeric 総設定テスト数",
        "warn_code_zero_loc": "{n} rows have missing or zero LoC",
        "warn_defects_empty": "no '不具合管理' rows after filter",
    },
    "ja": {
        "intro_caption": "管理チーム用の統合ダッシュボードシステム",
        "main_tab_dashboard": "📥 インプット",
        "main_tab_charts": "📊 グラフ",
        "main_tab_calendar": "📅 カレンダー",
        "main_tab_alert": "🚨 アラート",
        "main_tab_delivery": "🏁 配信パフォーマンス",
        "main_tab_backlog": "📋 Backlog",
        "main_tab_design": "📐 設計書ページ数",
        "main_tab_settings": "⚙️ 設定",
        "delivery_needs_data": (
            "機能マスタと WBS を取り込むとチーム配信パフォーマンスが表示されます。"
        ),
        # Backlog タブ (📋) — Backlog.com の課題一覧
        "backlog_tab_title":   "📋 Backlog — 課題 & リスク",
        "backlog_tab_caption": (
            "Backlog.com から CSV (SJIS) でエクスポートした課題一覧。"
            "上部のフィルタで絞り込み、**種別ごとの縦レーン**に"
            "状態別グループで表示します。"
        ),
        "backlog_needs_file":  (
            "インプットタブから Backlog の課題一覧 CSV を取り込むと"
            "この画面が有効になります。"
        ),
        "backlog_empty":       "現在のフィルタ条件に一致する課題はありません。",
        "backlog_facet_type":          "種別",
        "backlog_facet_status":        "状態",
        "backlog_facet_category":      "カテゴリ",
        "backlog_facet_phase":         "発生フェーズ",
        "backlog_facet_customer":      "顧客共有",
        "backlog_facet_due":           "期限日",
        "backlog_due_option_all":       "すべて",
        "backlog_due_option_overdue":   "超過",
        "backlog_due_option_this_week": "今週",
        "backlog_due_option_this_month":"今月",
        "backlog_due_option_future":    "今月以降",
        "backlog_due_option_none":      "期限日なし",
        "backlog_tile_due_label":       "期限",
        "backlog_tile_no_due":          "—",
        "backlog_tile_assignee_none":   "(未割当)",
        "backlog_lane_empty":           "(なし)",
        "backlog_status_total":         "{n} 件",
        "backlog_filter_reset":         "フィルタをリセット",
        "dora_section_title": "🏁 チーム配信パフォーマンス（DORA 5Keys）",
        "dora_section_caption": (
            "過去 {days} 日間 — 期間内に完了した機能 {n} 件"
        ),
        "dora_freq_title":        "配信頻度",
        "dora_lead_title":        "リードタイム",
        "dora_cfr_title":         "変更失敗率",
        "dora_recovery_title":    "障害復旧時間",
        "dora_reliability_title": "信頼性（平均障害発生率）",
        "dora_unit_per_week":     "機能 / 週",
        "dora_unit_days":         "日",
        "dora_unit_percent":      "%",
        "dora_rating_good":       "Good",
        "dora_rating_normal":     "Normal",
        "dora_rating_bad":        "Bad",
        "dora_rating_unknown":    "—",
        # DORA 各指標のホバーツールチップ（Good / Normal / Bad の 3 段階は
        # DORA 2024 の Elite + High → Good / Medium → Normal / Low → Bad
        # を業界平均比較として集約した値）
        "help_dora_frequency": (
            "**🦕 配信頻度**\n\n"
            "🧮 過去{days}日間に完了した機能件数 ÷ 週数。\n\n"
            "📂 出典: WBS **T列**（終了実績日）。\n\n"
            "💡 業界平均比較 (DORA 2024): "
            "Good ≥ 1件/週 / Normal ≥ 0.25件/週 / Bad それ未満。"
        ),
        "help_dora_lead_time": (
            "**🦕 リードタイム**\n\n"
            "🧮 機能ごとの (終了実績日 − 開始予定日) の中央値（日）。\n\n"
            "📂 出典: WBS **Q列**（開始予定日）と **T列**（終了実績日）。\n\n"
            "💡 業界平均比較 (DORA 2024): "
            "Good ≤ 7日 / Normal ≤ 30日 / Bad それ以上。"
        ),
        "help_dora_cfr": (
            "**🦕 変更失敗率（CFR）**\n\n"
            "🧮 期間内に完了した機能のうち、Redmine 障害が1件以上登録された"
            "ものの割合。\n\n"
            "📂 出典: WBS T列（期間内完了）× Redmine 不具合一覧"
            "（機能ID別件数）。\n\n"
            "💡 業界平均比較 (DORA 2024): "
            "Good ≤ 30% / Normal ≤ 45% / Bad それ以上。"
        ),
        "help_dora_recovery": (
            "**🦕 障害復旧時間**\n\n"
            "🧮 期間内に解決した Redmine 障害の "
            "(実終了日 − 実開始日) の中央値。\n\n"
            "📂 出典: Redmine 不具合一覧 実開始日／実終了日。\n\n"
            "💡 業界平均比較 (DORA 2024): "
            "Good ≤ 1日 / Normal ≤ 7日 / Bad それ以上。"
        ),
        "help_dora_reliability": (
            "**🦕 信頼性（平均障害発生率）**\n\n"
            "🧮 全機能の障害発生率（Redmine 障害件数 ÷ 実施済）の平均。\n\n"
            "📂 出典: Redmine 不具合一覧 × 仕様書別テスト集計 D列。\n\n"
            "💡 Good ≤ 15% / Normal ≤ 30% / Bad それ以上。"
        ),
        # アラートタブ（🚨）
        "alert_tab_title":   "アラート — 要注意な機能ID",
        "alert_tab_caption": (
            "現在の数値が設定閾値（障害発生率 / テスト密度）やルール"
            "（遅延14日超・未実施60%超）を満たす機能IDを表示します。"
            "機能IDフィルタ適用中はその範囲に絞り込まれます。"
        ),
        "alert_needs_master": (
            "機能マスタを取り込むとアラートが利用できます。"
        ),
        "alert_all_clear": "対象範囲ではアラートなし。問題は検出されていません。",
        "alert_sev_high":          "重",
        "alert_sev_medium":        "中",
        "alert_sev_low":           "低",
        "alert_sev_high_label":    "重大アラート",
        "alert_sev_medium_label":  "中アラート",
        "alert_sev_low_label":     "低アラート",
        "alert_current_label":     "現在値",
        "alert_threshold_label":   "閾値",
        "alert_risk_score_label":  "リスクスコア",
        "alert_score_legend_title": "ℹ️ リスクスコアの求め方",
        "alert_score_legend_body": (
            "**リスクスコア（0〜1、高いほど要対応）** — "
            "機能ごとに下記 4 つの「閾値超過指標」を重み付け平均します：\n\n"
            "- **障害発生率** — 重み **3.0**\n"
            "- **遅延日数** — 重み **1.5**\n"
            "- **テスト密度** — 重み **1.0**\n"
            "- **テスト未実施率** — 重み **1.0**\n\n"
            "各指標は**閾値を超えたときのみ**寄与度 `n × w` を計上します"
            "（超過していなければ 0）。`n` は超過の度合いを 0..1 に"
            "正規化した値：\n\n"
            "- **障害発生率**: `min(1, ir ÷ (2×閾値))`\n"
            "- **遅延**: `min(1, (日数 − 14) ÷ 46)`（60 日で上限）\n"
            "- **テスト密度**: `max(0, 1 − td ÷ 閾値)`\n"
            "- **未実施率**: `max(0, (% − 60) ÷ 40)`"
            "（計画 10 件以上のときのみ）\n\n"
            "スコア = **Σ(n × w) ÷ Σ(w)**。重要度の分類: "
            "**> 0.70 → 🔴 HIGH**、**> 0.35 → 🟡 MED**、"
            "**以下 → 🔵 LOW**。全指標が閾値内の機能は"
            "アラートになりません。\n\n"
            "各タイルの指標行の **X** は寄与度が最大の指標（＝主犯）、"
            "**△** は主犯ではないが閾値を超えている指標です。"
        ),
        "help_test_notrun_short": "未実施率",
        "alert_sort_label":      "並び替え",
        "alert_sort_severity":   "重要度順（重 → 低）",
        "alert_sort_date_desc":  "日付順（新しい順）",
        "alert_sort_date_asc":   "日付順（古い順）",
        # タイル右端の日付バッジ用ラベル（出典: WBS）
        "alert_date_label_actual":  "終了実績日",
        "alert_date_label_planned": "終了予定日",
        "alert_date_label_unknown": "日付不明",
        "charts_needs_master": "Dashboardタブで **機能マスタ** を取り込むとグラフが利用できます。",
        "chart_progress_gap": "進捗: 計画 vs 実績",
        "chart_progress_planned": "計画",
        "chart_progress_actual": "実績",
        "chart_progress_over_marker": "⚠ 超過",
        "chart_test_coverage": "テストカバレッジ (OK / NG / 未実施)",
        "chart_test_density": "機能ID別テスト密度（テスト件数に関する充足率）",
        "chart_test_density_threshold_label": "閾値",
        "chart_test_density_below_marker": "⚠ 不足",
        "chart_incident_rate": "機能ID別 障害発生率（Redmine, 障害件数/実施済）",
        "chart_incident_rate_threshold_label": "閾値",
        "chart_incident_rate_above_marker": "⚠ 超過",
        "chart_overview_compare": "機能ID俯瞰比較（4指標）",
        "chart_overview_compare_filter": "機能IDで絞り込む",
        "chart_overview_compare_filter_help": "未選択で全件。フィルタ後の集計でカードとチャートが更新されます。",
        "chart_overview_compare_total_prefix": "合計",
        "chart_overview_compare_fids_suffix": "件の機能ID",
        "chart_overview_compare_empty": "現在のフィルタに合致する機能IDがありません。",
        "help_chart_overview_compare": (
            "**🦕 機能ID俯瞰比較（4指標）**\n\n"
            "設計書ページ数 / LoC / 総設定テスト数 / 障害件数（Redmine） を機能ID別に "
            "横棒グラフ4つで並列表示。各指標は独立したX軸スケールで絶対値が読める一方、"
            "Y軸の機能IDは共有しているので、横にスライドして同じ機能IDの4指標を比較できます。\n\n"
            "📂 出典: 設計書ページ数（手動入力）, コード行数（LoC）, "
            "テスト集計（総設定テスト数）, Redmine 障害一覧（defect_total）。"
        ),
        "chart_loc_vs_ng": "LoC × NG",
        "chart_loc_vs_ng_sub": "（サイズ: 設計ページ数、色: リスクスコア）",
        "chart_design_impl_gap": "設計ページ数 × LoC",
        "chart_risk_heatmap": "リスク要素ヒートマップ",
        "chart_loc_trend": "LoCの推移（スナップショット間）",
        "chart_test_trend": "テスト件数の推移（スナップショット間）",
        "chart_bug_trend": "不具合の推移（週次・発生 vs 解決）",
        "chart_defect_class": "障害の問題分類内訳（Redmine）",
        "chart_defect_class_filter": "機能IDで絞り込む",
        "chart_defect_class_filter_help": "未選択で Redmine の全障害行を対象。",
        "chart_defect_class_empty": "現在のフィルタに合致する障害行がありません。",
        "chart_defect_class_no_class": "フィルタ後のデータに 問題分類 の値がありません。",
        "chart_defect_class_scope_prefix": "対象障害件数:",
        "chart_defect_class_scope_suffix": "件",
        "chart_defect_class_top_title": "問題分類 トップ {n}",
        "chart_defect_class_col_class": "問題分類",
        "chart_defect_class_col_count": "件数",
        "chart_defect_class_col_share": "割合",
        "global_fid_filter_title": "機能IDフィルタ",
        "global_fid_filter_label": "機能ID",
        "global_fid_filter_help": "未選択で全機能ID。Overview比較 / カレンダー / Charts（Redmine 問題分類）に適用されます。",
        "global_fid_filter_scope_all": "対象: 全機能ID",
        "global_fid_filter_scope_n": "対象: {n} 件の機能ID",
        "global_fid_filter_upload_hint": "機能IDマスタを取り込むとフィルタが使えます。",
        "kpi_missing_header": "**計算不可 — 未入力:**",
        "source_label_tests":   "テスト集計CSV",
        "source_label_code":    "コード行数XLSX",
        "source_label_design":  "設計書ページ数（手動入力）",
        "source_label_defects": "Redmine 障害CSV",
        "source_label_wbs":     "WBS XLSM",
        "chart_no_design_pages": "設計書ページ数タブで値を入力するとこのグラフが表示されます。",
        "chart_no_history": "推移グラフには2つ以上のスナップショットが必要です。",
        "chart_no_defects": "不具合管理が未取込です。",
        "chart_label_ok": "OK",
        "chart_label_ng": "NG",
        "chart_label_notrun": "未実施",
        "chart_label_low":    "低",
        "chart_label_opened": "発生",
        "chart_label_closed": "解決",
        "chart_label_open_cum": "未解決（累積）",
        "chart_label_loc_total": "総LoC",
        "chart_label_total_tests": "総設定テスト数",
        "chart_label_executed": "実施済",
        "chart_label_total": "合計",
        "chart_label_coverage": "カバレッジ",
        "calendar_needs_master": "Dashboardタブで **機能マスタ** を取り込むとカレンダーが利用できます。",
        "calendar_title": "プロジェクトカレンダー",
        "calendar_caption": (
            "WBSの予定/実績と不具合の発生〜解決を統合表示します。下のスイッチで表示レイヤを切替えできます。"
        ),
        "calendar_layer_planned": "WBS 計画",
        "calendar_layer_actual": "WBS 実績",
        "calendar_layer_defects": "不具合",
        "calendar_layer_subtasks": "サブタスクを表示",
        "calendar_filter_fid": "機能IDで絞り込む",
        "calendar_filter_fid_help": "未選択で全件表示",
        "calendar_no_events": "選択中のレイヤに表示するイベントがありません。",
        "gantt_title": "ガント — 計画 vs 実績",
        "gantt_no_dates": "WBSの日付情報がありません。",
        "gantt_today_label": "今日",
        "calendar_section": "カレンダー",
        "calendar_event_count": "{n} 件",
        "settings_uploads_title": "前回取り込みファイルの自動取込",
        "settings_uploads_caption": (
            "起動時に各カードは `input/<日付>/<種別>/` 配下の最新ファイルを"
            "自動取り込みします。リセットしてもファイルは削除されず、"
            "今セッション中の自動取込だけを止めます（傾向分析のため履歴は保持）。"
        ),
        "settings_wbs_title": "WBS 解析の挙動",
        "settings_wbs_caption": (
            "重複した機能IDを持つ行の扱いを制御します。"
            "**最初に出現した行のみ**有効とし、以降の同一機能IDの行と、"
            "その直下の L列「●」のサブタスク行は既定でスキップされます "
            "（1機能IDあたり1ソースに保つため）。"
        ),
        "settings_wbs_attach_after_dup": (
            "重複後のサブタスクを直前の有効な親に付け替える"
        ),
        "settings_wbs_attach_after_dup_caption": (
            "**OFF（既定）**: 重複機能IDに続く ● 行もまとめてスキップ。\n\n"
            "**ON**: 重複機能IDの行だけ捨て、後続の ● 行は直前の"
            "有効な親機能IDの追加サブタスクとして取り込む。"
            "重複行が誤入力でも、そのサブ行には正当なスケジュール詳細が"
            "書かれている場合に使います。"
        ),
        "settings_charts_title": "チャート閾値",
        "settings_charts_caption": (
            "機能ID別チャートで警告線として使う閾値。"
            "閾値未満のバーは赤色＋マーカーで強調されます。"
        ),
        "settings_test_density_threshold": "テスト密度の警告閾値（テスト/ページ）",
        "settings_test_density_threshold_caption": (
            "既定値 10。総設定テスト数 ÷ 設計書ページ数 がこの値を下回る機能IDは、"
            "テスト密度チャートで赤＋⚠マーカーで強調表示されます。"
        ),
        "settings_incident_rate_threshold": "障害発生率（Redmine）の警告閾値（%）",
        "settings_incident_rate_threshold_caption": (
            "既定値 5%。Redmine `defect_total` ÷ `実施済` がこの値を超える"
            "機能IDは、障害発生率チャートで赤＋⚠マーカーで強調表示されます。"
        ),
        "settings_pages_title": "設計書ページ数の自動取込",
        "settings_pages_caption": (
            "`input/design_pages.json` も同様。リセットは今セッション中のメモリ"
            "状態のみクリアし、ディスクのファイルは温存します。"
        ),
        "settings_reset_btn": "自動取込をリセット",
        "settings_undo_reset_btn": "自動取込を再有効化",
        "settings_show_files": "ファイル一覧を表示",
        "settings_show_entries": "エントリ一覧を表示",
        "settings_files_count": "{n} 件",
        "settings_no_files": "保存ファイルはありません。",
        "settings_count_pages": "{n} 件保存済",
        "settings_status_loaded": "自動取込済",
        "settings_status_skipped": "自動取込リセット中（セッション）",
        "settings_status_uploaded": "アップロード済",
        "settings_confirm_reset_msg": (
            "今セッション中の自動取込を停止しますか？ `input/` のファイルは残ります。"
            "Dashboardタブのカードは「ファイル待ち…」表示になり、新しくアップロード"
            "するまで取込されません。"
        ),
        "settings_confirm_pages_reset_msg": (
            "今セッション中の設計書ページ数のメモリ状態をクリアしますか？"
            "JSONファイルは残るため、次回起動時には再読込されます。"
        ),
        "settings_confirm_btn": "リセットする",
        "settings_reset_done": "{label}: 自動取込をリセットしました",
        "settings_undo_done": "{label}: 自動取込を再有効化しました",
        "settings_delete_file_btn": "🗑",
        "settings_confirm_delete_file_msg": (
            "このスナップショットを完全に削除しますか？次回レンダー時にトレンド"
            "グラフからも消えます。他の保存済スナップショットには影響しません。"
        ),
        "settings_confirm_delete_check": (
            "完全に削除されることを理解しました（元に戻せません）。"
        ),
        "settings_confirm_delete_btn": "削除する",
        "settings_file_deleted": "スナップショットを削除: {file}",
        # ----- バリデーションステップラベル（恐竜アニメ内で実行） -----
        "step_xlsx_parse":      "Excelファイルを解析",
        "step_xlsm_parse":      "マクロ付Excelを解析",
        "step_csv_encoding":    "文字コードを判定",
        "step_csv_parse":       "CSV構造を解析",
        "step_master_sheet":    "シート '機能一覧' を確認",
        "step_master_b_col":    "B列の最終行を特定",
        "step_master_fid":      "F列から機能IDを抽出",
        "step_master_dups":     "機能IDの重複を確認",
        "step_wbs_sheet":       "シート 'メイン' を確認",
        "step_wbs_phase_dates": "フェーズ期間 J6 / N6 を解析（年/月/日）",
        "step_wbs_fid":         "E〜I列から機能IDを抽出（16行目以降）",
        "step_defects_columns": "必須列を確認",
        "step_defects_filter":  "「不具合管理」でフィルタ",
        "step_defects_dates":   "MM/DD/YYYY形式の日付を解析",
        "step_defects_build":   "機能IDを抽出（データフレーム構築）",
        "step_tests_min_cols":  "6列以上(A〜F)を確認",
        "step_tests_fid":       "A列から機能IDを抽出",
        "step_tests_numeric":   "C/D/E/F列が数値であることを確認",
        "step_tests_sanity":    "妥当性: 実施済 ≤ 総設定テスト数",
        "step_code_sheet":      "シート '機能ID別サマリ' を確認",
        "step_code_fid":        "A列から機能IDを抽出",
        "step_code_loc":        "B列のLoCが数値であることを確認",
        "step_load_ok":         "データフレームを構築",
        "step_load_failed":     "データフレーム構築に失敗",
        # ----- クラッシュポップアップ -----
        "popup_error_title": "🦖💥 Ouch!",
        "popup_error_hint": (
            "上記の問題を修正してから、もう一度ファイルをドロップしてください。"
            "（前回取り込んだファイルはそのまま残っています）"
        ),
        "validation_passed": "全チェックパス",
        "validation_warnings": "警告ありで取り込み完了 — チェックリストをご確認ください。",
        "log_show_detail": "詳細ログを表示（最大3000文字）",
        "log_file_caption": "ログファイル（今セッション）: `{path}`",
        "log_section_title": "セッションログ",
        "log_section_caption": (
            "今セッション中に発生したエラーはすべてプロジェクトの `log/` フォルダ"
            "配下のファイルにも記録されます。"
        ),
        # ドリルダウンパネル
        "drilldown_select_hint": (
            "ヒント: 上のテーブルで行をクリックすると、その機能IDの全KPI / "
            "スケジュール / 関連不具合をまとめたドリルダウンが開きます。"
        ),
        "drilldown_panel_title": "🦖 機能IDドリルダウン",
        "drilldown_close": "ドリルダウンを閉じる",
        "drilldown_section_wbs":     "スケジュール (WBS)",
        "drilldown_to_deadline":     "終了予定まで",
        "drilldown_deadline_future": "残 {n} 日",
        "drilldown_deadline_today":  "本日が予定日",
        "drilldown_deadline_overdue": "+{n}日超過",
        "drilldown_deadline_overdue_badge": "超過",
        "drilldown_deadline_completed": "完了 ({date})",
        "drilldown_deadline_unknown":   "—",
        "drilldown_help_deadline": (
            "**WBS R列（終了予定日）** と今日との差分。"
            "実績終了日がある機能は完了として扱い、超過日数ではなく"
            "完了日を表示します。"
        ),
        "drilldown_assignees_label": "この機能の担当者",
        "drilldown_assignees_none":  "(WBS サブタスクなし)",
        "drilldown_role_progress_label": "ロール別進捗",
        "drilldown_role_progress_caption": (
            "ロールごとのサブタスク進捗。開発だけ完了していて"
            "テスト実施が未着手、といった詰まりを見つけるための指標です。"
        ),
        "drilldown_subtask_expander": "🔍 サブタスク詳細 (WBS)",
        "drilldown_subtask_none":  "この機能 ID の WBS サブタスクはありません。",
        "drilldown_status_completed":   "完了",
        "drilldown_status_in_progress": "進行中",
        "drilldown_status_not_started": "未着手",
        "drilldown_subtask_col_label":    "タスク",
        "drilldown_subtask_col_assignee": "担当者",
        "drilldown_subtask_col_role":     "ロール",
        "drilldown_subtask_col_planned":  "予定",
        "drilldown_subtask_col_actual":   "実績",
        "drilldown_subtask_col_progress": "実績 %",
        "drilldown_subtask_col_delay":    "遅延 (日)",
        "drilldown_subtask_role_other":   "—",
        "drilldown_section_defects": "不具合",
        "drilldown_section_trend": "トレンド（スナップショット横断）",
        "drilldown_section_trend_caption": (
            "`input/` 配下のタイムスタンプ付きスナップショットを機能ID別に"
            "再構成した、テスト件数と LoC の推移。"
        ),
        "drilldown_section_tests":   "テスト",
        "drilldown_section_code":    "コード/設計",
        "drilldown_section_scores":  "合成スコア",
        "drilldown_related_defects": "関連不具合 ({n}件)",
        "drilldown_no_defects":      "この機能IDに関連する不具合はありません。",
        "drilldown_no_wbs":          "この機能IDのWBSスケジュール情報がありません。",
        "drilldown_planned_period":  "計画期間",
        "drilldown_actual_period":   "実績期間",
        "drilldown_planned_effort":  "予定工数",
        "drilldown_actual_effort":   "投入工数",
        "drilldown_planned_progress": "進捗率（計画）",
        "drilldown_actual_progress":  "進捗率（実績）",
        "drilldown_progress":        "進捗（計画 vs 実績）",
        "drilldown_id_not_found":    "機能ID `{fid}` はマスタに存在しません。",
        "drilldown_strip_title":     "この機能IDがどの入力にあるか",
        "drilldown_source_unloaded": "（未ロード）",
        "drilldown_source_manual":   "（手動入力）",
        "src_design_label":          "設計書ページ数",
        "sec1_title": "1. ソースファイル投入",
        "sec2_title": "設計書ページ数",
        "sec2_caption": (
            "機能IDごとにページ数を編集できます。変更は `input/design_pages.json` "
            "に自動保存され、次回起動時に再読込されます。マスタの増減に追従し、"
            "空欄は値なしとして扱われます。"
        ),
        "sec2_filled_ids": "入力済ID数",
        "sec2_total_pages": "合計ページ数",
        "sec2_summary_tip": (
            "ここで入力した値は統合テーブルに反映され、次工程の設計密度KPIに利用されます。"
        ),
        "design_needs_master": (
            "先にダッシュボードタブで **機能マスタ** を取り込んでください。"
            "本エディタはマスタの機能IDに連動します。"
        ),
        "design_last_saved": "最終保存: {ts}",
        "design_no_save_yet": "未保存です — 任意のセルを編集するとファイルが作成されます。",
        "sec3_title": "2. 統合テーブル",
        "sec3_caption": (
            "{n}行 · 機能ID {u}件 · 横スクロール不要のタブに分割表示"
        ),
        "tab_overview": "概要",
        "tab_kpis": "KPI",
        "tab_wbs": "WBS",
        "tab_defects": "不具合",
        "tab_tests": "テスト",
        "tab_code": "コード/設計",
        "tab_all": "全列",
        "tab_all_caption": "結合後の全列。必要に応じて横スクロールしてください。",
        "kpi_summary_title": "プロジェクト全体KPI",
        "metric_total_loc": "総LoC",
        "metric_open_defects": "未解決障害（Redmine）",
        "metric_test_run_rate": "テスト実施率",
        "metric_test_pass_rate": "テスト成功率",
        "metric_avg_bug_density": "平均不具合密度（テスト仕様書）",
        "metric_avg_test_density": "平均テスト密度",
        "metric_at_risk": "高リスク機能数",
        "metric_delayed": "遅延機能数",
        "metric_avg_health": "平均健全性",
        "metric_help_at_risk": (
            "**🦕 高リスク機能数**\n\n"
            "🧮 **リスクスコア ≥ 0.5** の機能ID件数。\n"
            "リスクスコア (機能単位、0..1) = "
            "`0.4×未解決障害 + 0.2×未実施 + 0.2×遅延 + "
            "0.2×不具合密度`（各要素はデータセット内 min-max で 0..1 に正規化）。\n\n"
            "📂 出典: Redmine不具合一覧 × 仕様書別テスト集計 × WBS × "
            "機能ID別コード行数。\n\n"
            "💡 今すぐ対処が必要な機能の件数。プロジェクト全体の品質サマリ。"
            "アラートタブのリスクスコアとは別物（あちらは 4 指標の"
            "閾値超過スコアに特化）。"
        ),
        "metric_help_delayed": (
            "**🦕 遅延機能数**\n\n"
            "🧮 **遅延日数 > 0** の機能ID件数。\n"
            "遅延日数 = `max(0, 終了実績日 − 終了予定日)`。進行中の機能の場合は "
            "`今日 − 終了予定日`（正のときのみカウント）。\n\n"
            "📂 出典: WBS **R列**（終了予定日）と **T列**（終了実績日）。\n\n"
            "💡 スケジュール超過している機能の件数。0 が健全な状態。"
        ),
        # ----- ヘルプ（ホバーツールチップ） -----
        "help_func_id": (
            "**🦕 機能ID**\n\n"
            "各機能を一意に識別するキー。\n\n"
            "📂 出典: 機能ID一覧（機能一覧シート）F列。\n\n"
            "💡 全データソースの結合キー。"
        ),
        "help_func_name": (
            "**🦕 機能名称**\n\n"
            "機能の表示名。\n\n"
            "📂 出典: 機能ID一覧（機能一覧シート）G列。\n\n"
            "💡 同一機能IDでも名称が異なる場合があり、(ID, 名称) の組合せを保持します。"
        ),
        "help_defect_total": (
            "**🦕 障害件数（Redmine）**\n\n"
            "機能ID別に Redmine で記録された障害件数。\n\n"
            "📂 出典: Redmine不具合一覧（トラッカー='不具合管理' のみ）。\n\n"
            "💡 未解決・解決済みの両方を含みます。\n\n"
            "⚠ これは**テスト仕様書の不具合件数（NG）ではありません**。"
            "（→「不具合密度（テスト仕様書）」「不具合率（テスト仕様書）」を参照）"
        ),
        "help_defect_unresolved": (
            "**🦕 未解決障害（Redmine）**\n\n"
            "実終了日が空の障害件数。\n\n"
            "📂 出典: Redmine不具合一覧（不具合管理）。\n\n"
            "💡 現時点の未解決バックログ。\n\n"
            "⚠ これは**テスト仕様書の NG 件数ではありません**。"
        ),
        "help_incident_rate": (
            "**🦕 障害発生率（Redmine）**\n\n"
            "🧮 Redmine `defect_total` ÷ 仕様書別テスト集計 `実施済` "
            "（Redmine の障害件数を実施済テスト件数で割ったもの）。\n\n"
            "📂 出典: Redmine不具合一覧（トラッカー='不具合管理'） × "
            "仕様書別テスト集計 D列。\n\n"
            "💡 実施1件あたりに Redmine 起票の障害がどれだけ出たかの目安。\n\n"
            "⚠ これは**「不具合率（テスト仕様書）」(NG / 総設定テスト数) ではありません**。"
            "分子と分母の出典が違います。"
        ),
        "help_test_total": (
            "**🦕 総設定テスト数**\n\n"
            "計画されたテストケース総数。\n\n"
            "📂 出典: 仕様書別テスト集計 C列。\n\n"
            "💡 実施率や密度の分母になります。"
        ),
        "help_test_run": (
            "**🦕 実施済**\n\n"
            "実施済みのテストケース数。\n\n"
            "📂 出典: 仕様書別テスト集計 D列。"
        ),
        "help_test_ok": (
            "**🦕 OK**\n\n成功したテスト件数。\n\n📂 出典: 仕様書別テスト集計 E列。"
        ),
        "help_test_ng": (
            "**🦕 NG — 不具合件数（テスト仕様書）**\n\n"
            "失敗したテスト件数。\n\n"
            "📂 出典: 仕様書別テスト集計 F列。\n\n"
            "💡 「不具合率（テスト仕様書）」「不具合密度（テスト仕様書）」の分子になります。\n\n"
            "⚠ これは**Redmine の障害件数（障害件数（Redmine））ではありません**。"
        ),
        "help_test_notrun": (
            "**🦕 未実施**\n\n"
            "🧮 総設定テスト数 − 実施済。\n\n"
            "📂 出典: 仕様書別テスト集計 C列 と D列 から算出。\n\n"
            "💡 残作業量の見える化。"
        ),
        "help_loc": (
            "**🦕 LoC（コード行数）**\n\n"
            "機能の実装コード総行数（物理行）。\n\n"
            "📂 出典: 機能ID別コード行数、シート 機能ID別サマリ B列。\n\n"
            "💡 機能ID単位で集約済の値。"
        ),
        "help_design_pages": (
            "**🦕 設計書ページ数**\n\n"
            "手動入力された設計書のページ数。\n\n"
            "📂 出典: 設計書ページ数タブで入力（input/design_pages.json に保存）。\n\n"
            "💡 同一機能IDを共有する複数の機能名にも同値が適用されます。"
        ),
        "help_planned_effort": (
            "**🦕 予定工数**\n\n"
            "計画された工数（人日／時間）。\n\n"
            "📂 出典: WBS シートメイン、**P列**、16行目以降。"
        ),
        "help_actual_effort": (
            "**🦕 投入工数**\n\n"
            "実投入した工数。\n\n"
            "📂 出典: WBS **U列**。"
        ),
        "help_planned_start": "**🦕 開始予定日**\n\n📂 出典: WBS **Q列**。",
        "help_planned_end":   "**🦕 終了予定日**\n\n📂 出典: WBS **R列**。",
        "help_actual_start":  "**🦕 開始実績日**\n\n📂 出典: WBS **S列**。",
        "help_actual_end": (
            "**🦕 終了実績日**\n\n"
            "📂 出典: WBS **T列**。\n\n"
            "💡 空欄の場合は進行中（未完了）扱い。"
        ),
        "help_actual_progress": (
            "**🦕 実績進捗率**\n\n"
            "報告された実績の進捗率。\n\n"
            "📂 出典: WBS **V列**。"
        ),
        "help_planned_progress": (
            "**🦕 計画進捗率**\n\n"
            "計画上の本日時点での想定進捗率。\n\n"
            "📂 出典: WBS **AA列**。"
        ),
        "help_bug_density": (
            "**🦕 不具合密度（テスト仕様書）**\n\n"
            "🧮 NG ÷ LoC（テスト仕様書の NG 件数 ÷ コード行数）。\n\n"
            "📂 出典: 仕様書別テスト集計 F列 ÷ 機能ID別コード行数 B列。\n\n"
            "💡 コード1行あたりの不具合数。高いほど不具合多。\n\n"
            "⚠ これは**Redmine の障害件数（障害件数（Redmine））ではありません**。"
        ),
        "help_test_density": (
            "**🦕 テスト密度**\n\n"
            "🧮 総設定テスト数 ÷ 設計書ページ数。\n\n"
            "📂 出典: 仕様書別テスト集計 C列 ÷ 設計書ページ数（手動入力）。\n\n"
            "💡 設計1ページあたりのテスト件数。低い場合は仕様のテスト不足の可能性。"
        ),
        "help_complexity": (
            "**🦕 複雑度**\n\n"
            "🧮 LoC ÷ 設計書ページ数。\n\n"
            "📂 出典: 機能ID別コード行数 B列 ÷ 設計書ページ数（手動入力）。\n\n"
            "💡 設計1ページあたりの実装行数。"
        ),
        "help_test_run_rate": (
            "**🦕 テスト実施率**\n\n"
            "🧮 実施済 ÷ 総設定テスト数。\n\n"
            "📂 出典: 仕様書別テスト集計 D列 ÷ C列。\n\n"
            "💡 テストの消化進捗。100% で全件実施。"
        ),
        "help_test_pass_rate": (
            "**🦕 テスト成功率**\n\n"
            "🧮 OK ÷ 実施済。\n\n"
            "📂 出典: 仕様書別テスト集計 E列 ÷ D列。\n\n"
            "💡 実施済テストの品質。90%未満は要調査。"
        ),
        "help_defect_rate": (
            "**🦕 不具合率（テスト仕様書）**\n\n"
            "🧮 NG ÷ 総設定テスト数（テスト仕様書の NG ÷ 計画テスト件数）。\n\n"
            "📂 出典: 仕様書別テスト集計 F列 ÷ C列。\n\n"
            "💡 全テスト計画に対する不合格率。\n\n"
            "⚠ これは**「障害発生率（Redmine）」ではありません**。"
        ),
        "help_delay_days": (
            "**🦕 遅延日数**\n\n"
            "🧮 max(0, 終了実績日 − 終了予定日)。進行中の場合: 今日 − 終了予定日（正のとき）。\n\n"
            "📂 出典: WBS R列（終了予定日）と T列（終了実績日）から算出。\n\n"
            "💡 0 = 予定通り。完了するまで増え続けます。"
        ),
        "help_delay_rate": (
            "**🦕 遅延率**\n\n"
            "🧮 遅延日数 ÷ 計画期間（最大1.0）。\n\n"
            "📂 出典: WBS Q列（開始予定日）〜R列（終了予定日）の計画期間 と "
            "遅延日数 から算出。\n\n"
            "💡 0 = 予定通り。1.0 = 計画期間の倍以上。"
        ),
        "help_health_score": (
            "**🦕 健全性スコア**\n\n"
            "🧮 テスト実施率 − 不具合率（テスト仕様書） − 遅延率（範囲 ≈ −2…1）。\n\n"
            "📂 出典: 上記3指標から算出（= 仕様書別テスト集計 × WBS）。\n\n"
            "💡 高いほど健全。負の値は要注意。"
        ),
        "help_risk_score": (
            "**🦕 リスクスコア**\n\n"
            "🧮 正規化された値の重み付き和:\n"
            "0.4×未解決障害（Redmine） + 0.2×未実施 + 0.2×遅延 + "
            "0.2×不具合密度（テスト仕様書）。\n\n"
            "📂 出典: Redmine不具合一覧 × 仕様書別テスト集計 × WBS × 機能ID別コード行数。\n\n"
            "💡 各要素はデータセット内 min-max 正規化で 0…1。0.5以上で高リスク機能としてカウント。"
        ),
        # チャート / カレンダー
        "help_chart_progress_gap": (
            "**🦕 進捗: 計画 vs 実績**\n\n"
            "機能ID別に計画進捗率と実績進捗率を横棒で並べて比較。\n\n"
            "📂 出典: WBS **V列**（実績）と **AA列**（計画）。\n\n"
            "💡 計画バーが実績バーより長い＝遅延傾向。\n\n"
            "⚠ 実績 > 計画 の機能IDは実績バーをオレンジ＋⚠超過マーカーで"
            "強調表示します（過剰報告の兆候）。"
        ),
        "help_chart_test_coverage": (
            "**🦕 テストカバレッジ**\n\n"
            "機能ID別に OK / NG / 未実施 件数を積み上げ表示。\n\n"
            "📂 出典: 仕様書別テスト集計（E / F / C-D）。"
        ),
        "help_chart_test_density": (
            "**🦕 テスト密度（テスト件数に関する充足率）**\n\n"
            "🧮 総設定テスト数 ÷ 設計書ページ数 — 昇順ソートで下が手薄。\n\n"
            "📂 出典: 仕様書別テスト集計（C列）, 設計書ページ数。"
        ),
        "help_chart_incident_rate": (
            "**🦕 障害発生率（Redmine）**\n\n"
            "🧮 Redmine `defect_total` ÷ 仕様書別テスト集計 `実施済` を機能ID別に表示。"
            "降順ソートで悪い方が上に並びます。\n\n"
            "📂 出典: Redmine 障害一覧（defect_total）÷ テスト集計（D列 実施済）。\n\n"
            "⚠ これは**「不具合率（テスト仕様書）」(NG / 総設定テスト数) ではありません**。"
        ),
        "help_chart_defect_class": (
            "**🦕 障害の問題分類内訳（Redmine）**\n\n"
            "Redmine の `問題分類` 列で障害行をグループ化したドーナツ＋トップN表。\n\n"
            "📂 出典: Redmine 障害一覧（`不具合管理` トラッカー）の 問題分類 列。\n\n"
            "💡 機能IDフィルタを使うと、特定機能群の障害がどんな原因に偏っているかが見えます。"
        ),
        "help_chart_loc_vs_ng": (
            "**🦕 LoC × NG**\n\n"
            "散布図: x=LoC, y=NG, 点サイズ=設計ページ数, 色=リスクスコア。\n\n"
            "📂 出典: 機能ID別コード行数(B), 仕様書別テスト集計(F), "
            "設計ページ数, リスクスコア。\n\n"
            "💡 右上+赤 = 大規模 + 不具合多 + リスク高。"
        ),
        "help_chart_design_impl_gap": (
            "**🦕 設計ページ数 × LoC**\n\n"
            "設計と実装の規模感を散布。点線は平均複雑度。\n\n"
            "📂 出典: 設計ページ数（手動入力）と 機能ID別コード行数(B)。\n\n"
            "💡 線より上 = 平均より密な実装。"
        ),
        "help_chart_risk_heatmap": (
            "**🦕 リスク要素ヒートマップ**\n\n"
            "機能ID × 5要素（不具合密度（テスト仕様書）／障害発生率（Redmine）／"
            "遅延率／実施率の反転／テスト密度の反転）、データセット内 min-max 正規化。\n\n"
            "💡 赤い行ほど複数次元で危険。"
        ),
        "chart_risk_dims_legend": "凡例 — 各行の意味",
        # ----- 担当者×ロール分析（WBSサブタスク担当者 × Redmine障害） -----
        "role_analytics_title":   "担当者 × ロール分析",
        "help_role_analytics": (
            "**🦕 担当者×ロール分析**\n\n"
            "WBSのサブタスク（L列=●）から担当者（N列）とロール（サブタスク名"
            "に含まれるキーワードで判定）を取り出して、「誰がどの機能のどの"
            "ロールを担当したか」を集計します。さらに Redmine の障害情報と"
            "突き合わせることで、機能ごとの品質KPI（障害件数・発生率・NG・"
            "実施率）や、担当者ごとに関わった機能の品質傾向を俯瞰できます。\n\n"
            "📂 出典：WBS（サブタスクの担当者・タスク名）× Redmine 障害一覧 "
            "× 仕様書別テスト集計。\n\n"
            "💡 判定キーワード：`開発` / `実装` → 開発、`テスト仕様書作成` → "
            "仕様書作成、`テスト実施` → 実施（ただし `再テスト実施` は除外）。"
            "1つのサブタスク名に複数のキーワードが含まれている場合は、"
            "該当するすべてのロールにカウントします。N列が空欄の場合は"
            "「（未割当）」として集計します。"
        ),
        "role_analytics_view1_title":   "View 1 — 機能別 担当者ロールマップ & 品質KPI",
        "role_analytics_view1_caption": (
            "1行につき1つの機能。各ロール列には WBS の該当サブタスクの"
            "N列担当者を「 / 」で区切って表示しています（重複は省いています）。"
            "右側の列は品質の指標で、Redmine の障害発生率が高い順（悪い順）に"
            "並んでいます。"
        ),
        "role_analytics_view2_title":   "View 2 — 担当者サマリ",
        "role_analytics_view2_caption": (
            "1行につき1人の担当者。ロール列は WBS サブタスクで担当した件数、"
            "そのほか WBS 上で関わった機能の数、関連する Redmine 障害の"
            "合計件数、Redmine 障害発生率の平均、Redmine 問題分類のトップ3 "
            "を表示します。Redmine 障害件数が多い順に並んでいます。"
        ),
        "role_analytics_view3_title":   "View 3 — 担当者 × 問題分類 ヒートマップ",
        "role_analytics_view3_caption": (
            "担当者が WBS で関与した機能に紐づく Redmine 障害を、"
            "Redmine の問題分類別にカウント。ロールで絞ると"
            "「開発担当者の機能はロジック系バグが多い」などの偏りが見えます。"
        ),
        "role_analytics_view3_role_label": "ロールで絞る",
        "role_analytics_view3_role_all":   "全ロール",
        "help_role_analytics_view3_role_filter": (
            "ヒートマップを特定のロールに限定します。例えば「開発」だけに"
            "絞れば、テスト担当の行が紛れ込まずに開発者同士の比較ができます。"
        ),
        # バブルマップ + 問題分類ストリップ（担当者別の先進的ビジュアル）
        "role_analytics_bubble_title": (
            "担当者バブルマップ（広さ × 品質 × 障害量）"
        ),
        "role_analytics_bubble_caption": (
            "1つのバブル＝1人の担当者（**測定可能な品質シグナルを持つ人のみ**"
            "＝テストが1件以上実施済 かつ 障害が1件以上登録されている）。"
            "横軸は WBS サブタスクで関わった機能の数、縦軸はそれらの機能の "
            "Redmine 障害発生率の平均、バブルの大きさは Redmine 障害件数の合計、"
            "色はその人のメインのロール（WBS から判定）を表します。破線は"
            "全体の平均ライン。**右上＝広く関わって障害が多い（要注意）**、"
            "**右下＝広く関わっているが品質が良い（頼れる）** というように"
            "4つのエリアで読み取れます。テスト未実施 or 障害 0 件の担当者は"
            "この散布図に載らず、**下の 2 つの要注意リスト**に分類されます。"
        ),
        "role_analytics_bubble_color_legend": "ドミナントロール",
        # 画面・PDF 共通の基本説明
        "role_analytics_dominant_role_note": (
            "**ドミナントロール** ＝ その担当者の WBS サブタスクを集計し、"
            "**最も件数が多かったロール 1 つ**でバブルを着色。同点時は "
            "**開発 ＞ テスト仕様書 ＞ テスト実施** の順で決定します。"
            "単色なので複数ロール兼務は色から読み取れません — "
            "**詳細は上の「担当者×ロール サマリ」表**でロール別件数を確認できます。"
        ),
        # 画面のみ追記する（PDFでは表示しない）ホバーの案内
        "role_analytics_dominant_role_hover_hint": (
            "バブルにカーソルを当てるとロール別件数が表示されます。"
        ),
        # バブルマップ配下の要注意リスト 2 種。
        # A = 障害 0 件（曖昧。超優秀 or レビュー漏れ or テスト浅い）
        # B = テスト未実施（実施済 = 0。プロセス上の赤旗）
        "role_analytics_watch_zero_defect_title": (
            "⚠️ 要注意リストA：障害 0 件"
        ),
        "role_analytics_watch_zero_defect_caption": (
            "関わった機能について **仕様書別テスト集計（test_counts CSV）"
            "の D 列 = 実施済** が 1 件以上 あり、かつ "
            "**Redmine 不具合一覧（defects.csv）** にその機能 ID の "
            "障害が **0 件** 登録されている担当者。3 通りの解釈が"
            "ありえます：(1) **超優秀**で本当に欠陥ゼロ、"
            "(2) Redmine への **登録漏れ**、"
            "(3) **テスト仕様書のカバレッジが浅い**ためバグを"
            "検出できていない。📂 突合先: **defects.csv**（登録実態）／"
            "**テスト仕様書**（テスト観点の十分性）／"
            "**仕様書別テスト集計 D列**（実施済件数）。"
        ),
        "role_analytics_watch_no_exec_title": (
            "⚠️ 要注意リストB：テスト未実施"
        ),
        "role_analytics_watch_no_exec_caption": (
            "関わった機能すべてが **仕様書別テスト集計（test_counts CSV）"
            "の D 列 = 実施済 = 0**（テスト計画はあるが 1 件も実施されて"
            "いない）な担当者。Y 軸（発生率 = 障害件数 ÷ 実施済）が"
            "**計算不能** なのでバブルマップに載せられません。"
            "**品質が悪いのではなく、まだ測れない**状態です — "
            "**プロセス上の要フォローアップ**に該当します。"
            "📂 確認先: **テスト仕様書**（テスト計画の整備状況）／"
            "**仕様書別テスト集計 D列**（実施済が 0 の機能特定）／"
            "**WBS T列**（終了予定日 — 遅延していないか）。"
            "テスト実施スケジュールを確認してください。"
        ),
        "role_analytics_watch_empty": "該当者なし — 全員メイン散布図に掲載されています。",
        "role_analytics_watch_col_assignee": "担当者",
        "role_analytics_watch_col_features": "関与機能数",
        "role_analytics_watch_col_defects":  "障害件数",
        "role_analytics_watch_col_roles":    "ロール内訳",
        "role_analytics_strip_title": (
            "担当者別 問題分類ミックス（その人の障害内訳を100%で積み上げ）"
        ),
        "role_analytics_strip_caption": (
            "1行につき1人の担当者。その人が WBS で関わった機能で発生した "
            "Redmine 障害を 100% として、Redmine の問題分類ごとに色分けして"
            "積み上げています。各色の中の数字はその分類の **Redmine 障害件数**、"
            "ホバーすると「N件（X.X%）」が表示されます。行の右端の「n=」は"
            "その人の Redmine 障害合計件数です。"
        ),
        "role_analytics_strip_other": "その他",
        # Charts タブのページ内ナビゲーション
        "toc_jump_label":    "ジャンプ",
        "toc_back_to_top":   "最上部へ",
        # PDF ボタンの統一マイクロコピー
        "pdf_btn_generate_short":  "PDFを生成",
        "pdf_btn_download_short":  "PDFをダウンロード",
        "pdf_generating":          "PDF を生成中…",
        # 担当者×ロール分析 PDF出力
        "ra_pdf_btn_generate":       "📄 PDF",
        "ra_pdf_btn_generate_help":  (
            "担当者×ロール分析を単体PDFレポートとして出力します"
            "（インプット→ロール判定ルール→3つの表→バブルマップ→"
            "問題分類ストリップ）。"
        ),
        "ra_pdf_btn_download":       "⬇ PDF",
        "ra_pdf_title":              "担当者 × ロール分析 レポート",
        "ra_pdf_filter_active":      (
            "⚠ 機能IDフィルタが適用されています。"
            "このレポートは絞り込み後の機能のみを対象としています。"
        ),
        "ra_pdf_h_inputs":           "1. データの出どころ",
        "ra_pdf_input_wbs":          "WBS サブタスク行（L=●, N=担当者）",
        "ra_pdf_input_wbs_note":     (
            "WBS のサブタスク名に含まれるキーワードからロールを判定し、"
            "同じく WBS の N列（担当者）を読み取ります。"
        ),
        "ra_pdf_input_defects":      "Redmine 障害一覧（トラッカー=不具合管理）",
        "ra_pdf_input_defects_note":  (
            "Redmine の不具合一覧から、機能ごとの障害件数と問題分類を取得します。"
        ),
        "ra_pdf_input_tests":        "仕様書別テスト集計CSV",
        "ra_pdf_input_tests_note":   (
            "仕様書別テスト集計CSV から、総設定テスト数数・実施済数・NG数を取得します。"
        ),
        "ra_pdf_h_rules":            "2. ロール判定ルール",
        "ra_pdf_rules_body": (
            "WBS の各サブタスク名に以下のキーワードが含まれているかを確認して"
            "ロールを判定します。1つのサブタスクに複数のキーワードが含まれていれば、"
            "該当するすべてのロールに振り分けます。WBS の N列（担当者）が空欄の"
            "場合は<b>（未割当）</b>として集計します。"
            "（全角／半角・前後の空白は自動で揃えます）"
        ),
        "ra_pdf_rule_dev":       "<b>開発</b> / <b>実装</b> → 開発",
        "ra_pdf_rule_test_spec": "<b>テスト仕様書作成</b> → 仕様書作成",
        "ra_pdf_rule_test_exec": (
            "<b>テスト実施</b> → 実施"
            "（<b>再テスト実施</b> は除外）"
        ),
        "ra_pdf_h_view1":            "3. 機能ごとの担当者と品質KPI",
        "ra_pdf_h_view2":            "4. 担当者サマリ",
        "ra_pdf_h_bubble":           "5. 担当者バブルマップ",
        "ra_pdf_h_strip":            "6. 担当者別 問題分類の内訳",
        "ra_pdf_h_heatmap":          "7. 担当者 × 問題分類 ヒートマップ",
        "ra_pdf_no_data":            "表示するデータがありません。",
        "ra_pdf_footer":             "生成: {when}",
        "role_analytics_no_subtasks": (
            "WBSのサブタスク行（L列=●）が見つかりません。"
            "サブタスク記載のあるWBSを取り込むと本セクションが表示されます。"
        ),
        "role_analytics_no_matches": (
            "サブタスク名に「開発」「実装」「テスト仕様書作成」"
            "「テスト実施」のいずれも含まれていません"
            "（※「再テスト実施」はテスト実施から除外しています）。"
        ),
        "role_dev":          "開発 / 実装",
        "role_test_spec":    "テスト仕様書作成",
        "role_test_exec":    "テスト実施",
        "role_unassigned":   "（未割当）",
        "role_count_dev":       "開発 (WBSサブタスク件)",
        "role_count_test_spec": "仕様書作成 (WBSサブタスク件)",
        "role_count_test_exec": "テスト実施 (WBSサブタスク件)",
        "col_feature":          "機能ID：機能名",
        "col_assignee":         "担当者",
        "col_feature_count":    "関与機能数 (WBS)",
        "col_avg_incident_rate": "平均障害発生率（Redmine）",
        "col_top3_problems":    "問題分類 Top3 (Redmine)",
        "problem_class_uncategorized": "（未分類）",
        # 列のマウスオーバー説明（担当者×ロール分析の各テーブル用）
        "help_col_feature": (
            "**🦕 機能ID：機能名**\n\n"
            "機能IDに機能名（機能ID一覧を正）を付けた表示ラベル。\n\n"
            "📂 出典: 機能ID一覧 F列（機能ID）＋ G列（機能名称）。"
        ),
        "help_col_assignee": (
            "**🦕 担当者**\n\n"
            "WBSサブタスクの担当者名。\n\n"
            "📂 出典: WBS N列（L列=● のサブタスク行）。\n\n"
            "💡 全角/半角・連続/前後空白は自動で揃え、同一人物として集計します。"
        ),
        "help_col_feature_count": (
            "**🦕 関与機能数**\n\n"
            "この人がWBSサブタスクで関わった一意な機能IDの数（ロール問わず）。\n\n"
            "📂 出典: WBS サブタスク行（L列=● かつ N列に担当者名）から集計。"
        ),
        "help_col_avg_incident_rate": (
            "**🦕 平均障害発生率（Redmine）**\n\n"
            "🧮 関与機能の障害発生率（障害件数 ÷ 実施済）の平均値。\n\n"
            "📂 出典: Redmine不具合一覧 × 仕様書別テスト集計 D列 "
            "（関与機能は WBS から抽出）。\n\n"
            "💡 担当した機能の平均的な障害率の肌感。"
        ),
        "help_col_top3_problems": (
            "**🦕 問題分類 Top3（担当者別）**\n\n"
            "関与機能で発生した Redmine 障害の問題分類を件数順に上位3件。\n\n"
            "📂 出典: Redmine不具合一覧「問題分類」列（関与機能でフィルタ）。"
        ),
        "help_col_feature_top3_problems": (
            "**🦕 問題分類 Top3（機能別）**\n\n"
            "この機能で発生した Redmine 障害の問題分類を件数順に上位3件。\n\n"
            "📂 出典: Redmine不具合一覧「問題分類」列（当該機能IDでフィルタ）。"
        ),
        "help_role_count_dev": (
            "**🦕 開発担当件数**\n\n"
            "この人が「開発」「実装」キーワードを含むサブタスクで"
            "担当している件数。\n\n"
            "📂 出典: WBS サブタスク名（L列=●）＋ N列（担当者）。"
        ),
        "help_role_count_test_spec": (
            "**🦕 仕様書作成担当件数**\n\n"
            "この人が「テスト仕様書作成」キーワードを含むサブタスクで"
            "担当している件数。\n\n"
            "📂 出典: WBS サブタスク名（L列=●）＋ N列（担当者）。"
        ),
        "help_role_count_test_exec": (
            "**🦕 テスト実施担当件数**\n\n"
            "この人が「テスト実施」キーワードを含むサブタスクで"
            "担当している件数。**「再テスト実施」は除外**。\n\n"
            "📂 出典: WBS サブタスク名（L列=●）＋ N列（担当者）。"
        ),
        "help_role_assignees_dev": (
            "**🦕 開発担当者**\n\n"
            "この機能の「開発」「実装」サブタスクの担当者（重複排除）。\n\n"
            "📂 出典: WBS サブタスク名（L列=●）＋ N列（担当者）。"
        ),
        "help_role_assignees_test_spec": (
            "**🦕 仕様書作成担当者**\n\n"
            "この機能の「テスト仕様書作成」サブタスクの担当者（重複排除）。\n\n"
            "📂 出典: WBS サブタスク名（L列=●）＋ N列（担当者）。"
        ),
        "help_role_assignees_test_exec": (
            "**🦕 テスト実施担当者**\n\n"
            "この機能の「テスト実施」サブタスクの担当者（重複排除）。"
            "**「再テスト実施」は除外**。\n\n"
            "📂 出典: WBS サブタスク名（L列=●）＋ N列（担当者）。"
        ),
        # ----- PDF レポート -----
        "pdf_btn_generate": "PDFレポート生成",
        "pdf_btn_download": "PDFをダウンロード",
        "pdf_progress": "レポート生成中（20〜30秒程度）…",
        "pdf_done": "生成完了 — ダウンロードボタンを押してください。",
        "pdf_error": "レポート生成に失敗しました: {err}",
        "pdf_step_cover":    "表紙 + KPI 表を生成中…",
        "pdf_step_chart":    "チャートを描画中 ({i}/{n}): {title}",
        "pdf_step_gantt":    "スケジュール (Gantt) を描画中…",
        "pdf_step_assemble": "PDF 組版中…",
        "pdf_dialog_title":  "🦖 PDFレポート生成中",
        "pdf_dialog_subtitle": "選択された機能IDのみがレポートに含まれます",
        "pdf_dialog_close":  "閉じる",
        "chart_truncated_note": "ワースト {shown} 件 / 全 {total} 件を表示",
        "pdf_select_title":   "🦖 PDFレポートに含める機能IDを選択",
        "pdf_select_caption": (
            "最大30件まで選択できます。選ばれた機能IDのみが各チャートと "
            "Gantt に含まれます。"
        ),
        "pdf_select_label":   "機能ID（最大30件）",
        "pdf_select_count":   "{n} / 30 件選択中",
        "pdf_select_error_empty": (
            "最低1件以上、機能IDを選択してください。"
        ),
        "pdf_btn_confirm":    "生成開始",
        "pdf_title": "dashboard4dx — プロジェクト報告",
        "pdf_generated_at": "生成日時",
        "pdf_toc_title":    "目次",
        "pdf_section_kpi": "プロジェクト全体KPI",
        "pdf_section_charts": "グラフ",
        "pdf_section_schedule": "スケジュール",
        "pdf_no_chart": "このセクションに表示するデータがありません。",
        "pdf_chart_definition": "定義",
        # ----- PDF: テスト密度単体レポート -----
        "td_pdf_btn_generate":      "📄 PDF",
        "td_pdf_btn_generate_help": (
            "このグラフを単体PDFレポートとして出力します"
            "（入力→密度→閾値未満一覧→アドバイス）。"
        ),
        "td_pdf_btn_download":      "⬇ PDF",
        "td_pdf_title":             "機能ID別テスト密度 レポート",
        "td_pdf_filter_active":     "⚠ 機能IDフィルタ適用中 — フィルタ後の対象のみを集計しています。",
        "td_pdf_h_inputs":          "1. インプット（数値の出どころ）",
        "td_pdf_col_item":          "項目",
        "td_pdf_col_source":        "ファイル / 保管先",
        "td_pdf_col_method":        "取得方法",
        "td_pdf_input_tests":       "総設定テスト数",
        "td_pdf_input_pages":       "設計書頁数（目視確認）",
        "td_pdf_input_master":      "機能マスタ（機能ID / 名称）",
        "td_pdf_method_auto":       "自動取込",
        "td_pdf_method_manual":     "目視確認・手動入力",
        "td_pdf_input_tests_note":  "（テスト集計CSVのC列）",
        "td_pdf_input_pages_note":  "（input/design_pages.json、設計書頁数タブで編集）",
        "td_pdf_h_output":          "2. アウトプット（算出する密度）",
        "td_pdf_output_formula":    "<b>テスト密度</b> ＝ 総設定テスト数 ÷ 設計書頁数（目視確認）",
        "td_pdf_output_unit":       "単位：テスト件数 / 設計書1ページ",
        "td_pdf_output_meaning":    (
            "意味：設計書1ページあたりのテスト計画件数。"
            "低い場合は仕様に対するテスト件数不足の疑いがあります。"
        ),
        "td_pdf_h_threshold":       "3. 警告閾値",
        "td_pdf_threshold_current":      "本レポートで使用した警告閾値",
        "td_pdf_threshold_default":      "既定値",
        "td_pdf_threshold_where":        "設定変更箇所",
        "td_pdf_threshold_where_value": (
            "Dashboard の「設定」タブ →"
            "<b>「テスト密度の警告閾値（テスト/ページ）」</b>"
        ),
        "td_pdf_threshold_meaning": (
            "<b>意味：</b>機能IDごとの密度がこの値を下回ると、本レポート全体で "
            "<b>「要対応」</b>として赤バー＋⚠ マーカーで強調表示されます"
            "（チャート／閾値未満一覧／追加テスト試算 いずれも同じ閾値を基準としています）。"
        ),
        "td_pdf_threshold_unit":    "テスト / ページ",
        "td_pdf_h_summary":         "4. サマリ",
        "td_pdf_summary_total":     "対象機能ID",
        "td_pdf_summary_threshold": "警告閾値",
        "td_pdf_summary_above":     "閾値以上",
        "td_pdf_summary_below":     "閾値未満（要対応）",
        "td_pdf_summary_mean":      "平均密度",
        "td_pdf_summary_median":    "中央値",
        "td_pdf_summary_min_max":   "最小 / 最大",
        "td_pdf_h_chart":           "5. 機能ID別テスト密度（昇順）",
        "td_pdf_h_below":           "6. 閾値未満の機能ID一覧",
        "td_pdf_col_fid":           "機能ID",
        "td_pdf_col_name":          "機能名称",
        "td_pdf_col_tests":         "総設定テスト数",
        "td_pdf_col_pages":         "設計書頁数 (目視確認)",
        "td_pdf_col_density":       "密度",
        "td_pdf_col_gap":           "閾値との差",
        "td_pdf_below_none":        "閾値を下回る機能IDはありません — 対応不要。",
        "td_pdf_h_catchup":         "7. 目標達成のために追加すべきテスト件数（試算）",
        "td_pdf_col_current":       "現密度",
        "td_pdf_col_target":        "目標",
        "td_pdf_col_additional":    "追加テスト件数（推奨）",
        "td_pdf_catchup_note":      "推奨追加テスト件数 ＝ ⌈(閾値 − 現密度) × 設計書頁数⌉。",
        "td_pdf_h_advice":          "8. 閾値を上回るためのアクション（推奨）",
        "td_pdf_advice_body":       (
            "<b>設計書を1ページずつ棚卸しし、テスト観点の抜け漏れを点検する</b><br/>"
            "各ページに対し、正常系 / 異常系 / 境界値 / 権限 / 性能 / ログ "
            "の観点別で必要テスト件数を見積り直します。"
        ),
        "td_pdf_h_notes":           "注意事項",
        "td_pdf_notes_body":        (
            "・本指標は<b>「テスト件数の充足率」</b>であり、テストの網羅性・"
            "期待値の妥当性そのものは測りません。<br/>"
            "・設計書頁数は目視確認の手動入力値です。"
            "目視確認のため、ミスがある可能性があります。<br/>"
            "・単純機能では閾値未満でも妥当な場合があります。機能特性と合わせて"
            "総合判断してください。"
        ),
        "td_pdf_footer":            "生成: {when} / スナップショット: {src}",
        "td_pdf_pages":              "{n} / {total}",
        "help_chart_loc_trend": (
            "**🦕 LoC推移**\n\n"
            "保存済 code スナップショットの総LoC推移。\n\n"
            "📂 出典: input/<日付>/code/*.xlsx（スナップショット日はファイル名から抽出）。\n\n"
            "💡 表示には2件以上のスナップショットが必要。"
        ),
        "help_chart_test_trend": (
            "**🦕 テスト件数推移**\n\n"
            "保存済 tests スナップショットの総設定テスト数数 vs 実施済推移。\n\n"
            "📂 出典: input/<日付>/tests/*.csv。\n\n"
            "💡 総数と実施済の差 = テストバックログの推移。"
        ),
        "help_chart_bug_trend": (
            "**🦕 不具合の推移**\n\n"
            "発生／解決の週次バー + 未解決累積ライン。\n\n"
            "📂 出典: Redmine不具合一覧（実開始日 vs 実終了日、週次集計）。"
        ),
        "help_gantt_title": (
            "**🦕 ガント — 計画 vs 実績**\n\n"
            "機能ID別の横バー。灰=計画(Q-R)、緑=実績(S-T)。今日は黄色破線。\n\n"
            "💡 灰のみ=未着手、緑が灰より右に伸びる=遅延。"
        ),
        "help_calendar_title": (
            "**🦕 カレンダー**\n\n"
            "WBSスケジュールと不具合期間を月／週／リスト表示。\n\n"
            "📂 出典: WBS Q-T列（計画/実績）+ Redmine不具合一覧（実開始日/実終了日）。\n\n"
            "💡 上のスイッチでレイヤ切替。赤い不具合 = 未解決。"
        ),
        "col_bug_density":   "不具合密度（テスト仕様書, NG/LoC）",
        "col_defect_total":  "障害件数（Redmine）",
        "col_defect_unresolved": "未解決障害（Redmine）",
        "col_incident_rate": "障害発生率（Redmine, 障害件数/実施済）",
        "col_test_ng":       "不具合件数（テスト仕様書, NG）",
        "col_test_density":  "テスト密度（テスト件数に関する充足率, テスト/ページ）",
        "col_complexity":    "複雑度 (LoC/ページ)",
        "col_test_run_rate": "テスト実施率",
        "col_test_pass_rate":"テスト成功率",
        "col_defect_rate":   "不具合率（テスト仕様書, NG/総設定テスト数）",
        "col_delay_days":    "遅延日数",
        "col_delay_rate":    "遅延率",
        "col_health_score":  "健全性スコア",
        "col_risk_score":    "リスクスコア",
        "raw_previews": "ソース別プレビュー（各先頭10行）",
        "wbs_label_short": "WBS",
        "defects_label_short": "Redmine不具合一覧（不具合管理のみフィルタ後）",
        "tests_label_short": "仕様書別テスト集計",
        "code_label_short": "機能ID別コード行数",
        "master_unlock_info": (
            "**機能マスタ**をアップロードするとダッシュボードが有効になります。"
            "他のカードは任意で、スキップした項目は結合から除外されます。"
        ),
        "badge_required": "必須",
        "badge_optional": "任意",
        "card_drop_label": "{label} をここにドロップ",
        "status_waiting": "ファイル待ち…",
        "status_ok": "OK · {n}行 · ID {u}件",
        "status_ok_no_fid": "OK · {n}行",
        "status_failed": "検証失敗",
        "origin_upload": "今アップロード",
        "origin_auto": "input/ から自動読込",
        "origin_snapshot": "スナップショット {date}",
        "origin_ingested_at": "投入日時 {ts}",
        "toast_loaded": "{label} 取込完了 · {path} に保存",
        "toast_failed": "{label}: {msg}",
        "save_warn": "検証はOKですが input/ への保存に失敗: {err}",
        "read_prev_warn": "前回ファイルの読込に失敗: {err}",
        "read_upload_err": "アップロード読込失敗: {err}",
        "src_master_label": "機能ID一覧",
        "src_master_hint": "シート 機能一覧 · F列=ID, G列=名称",
        "src_wbs_label": "WBS",
        "src_wbs_hint": "シート メイン · 16行目以降 · IDはE〜I列",
        "src_defects_label": "Redmine不具合一覧",
        "src_defects_hint": "トラッカー / ステータス / 機能ID 他",
        "src_tests_label": "仕様書別テスト集計",
        "src_tests_hint": "A=ID · C=総数 · D=実施 · E=OK · F=NG",
        "src_code_label": "機能ID別コード行数",
        "src_code_hint": "シート 機能ID別サマリ · A=ID, B=LoC",
        "src_roster_label":   "担当者一覧",
        "src_roster_hint":    "チーム / 担当者名 / PC / 携帯 / VPN (xlsx)",
        "src_calendar_label": "カレンダー (イベント+非稼働日)",
        "src_calendar_hint":  "2シート: イベント+非稼働日 (xlsx)",
        "src_backlog_label":  "Backlog 課題一覧",
        "src_backlog_hint":   "Backlog.com の CSV (SJIS)、13 列: キーID / ID / 種別 / 状態 / …",
        "card_template_dl":   "⬇ テンプレをDL ({label})",
        "card_dl_template_help": "テンプレートをダウンロード",
        "card_dl_sample_help":   "サンプルデータをダウンロード",
        "card_dl_latest_help":   "直近に取り込んだファイルをダウンロード（{name}）",
        "src_rail_hint":      "⇄ 横にスクロールして全ソースを閲覧",
        "calendar_layer_events":  "イベント表示",
        "calendar_layer_nonwork": "非稼働表示",
        "err_zero_rows": "0行しか読めませんでした — シート名や列構成をご確認ください",
        "warn_master_dups": "{n}件の機能IDに複数の名称がありました（全て保持しています）",
        "warn_tests_overrun": "{n}行で 実施済 > 総設定テスト数 になっています",
        "warn_tests_nan_total": "{n}行の 総設定テスト数 が数値ではありません",
        "warn_code_zero_loc": "{n}行で LoC が未入力または0です",
        "warn_defects_empty": "フィルタ後に「不具合管理」の行がありません",
    },
}


def t(key: str, **kwargs) -> str:
    """Return the localized string for `key`, formatted with kwargs.
    Falls back to English if the current language lacks the key."""
    lang = st.session_state.get("lang", DEFAULT_LANG)
    s = TRANSLATIONS.get(lang, {}).get(key) or TRANSLATIONS[DEFAULT_LANG].get(key, key)
    try:
        return s.format(**kwargs) if kwargs else s
    except (KeyError, IndexError):
        return s


def _lang_label(code: str) -> str:
    return dict(LANG_OPTIONS).get(code, code)


# =============================================================================
# UI: source spec + per-source validation
# =============================================================================
SOURCE_SPECS: list[dict] = [
    {
        "key": "master",
        "label_key": "src_master_label",
        "hint_key": "src_master_hint",
        "icon": "🗂️",
        "types": ["xlsx", "xlsm"],
        "loader": load_function_master,
        "required": True,
        "sample_filename": "function_master.xlsx",
    },
    {
        "key": "wbs",
        "label_key": "src_wbs_label",
        "hint_key": "src_wbs_hint",
        "icon": "📅",
        "types": ["xlsx", "xlsm"],
        "loader": load_wbs,
        "required": False,
        "sample_filename": "wbs.xlsm",
    },
    {
        "key": "defects",
        "label_key": "src_defects_label",
        "hint_key": "src_defects_hint",
        "icon": "🐞",
        "types": ["csv"],
        "loader": load_defects,
        "required": False,
        "sample_filename": "defects.csv",
    },
    {
        "key": "tests",
        "label_key": "src_tests_label",
        "hint_key": "src_tests_hint",
        "icon": "🧪",
        "types": ["csv"],
        "loader": load_test_counts,
        "required": False,
        "sample_filename": "test_counts_20260420090000.csv",
    },
    {
        "key": "code",
        "label_key": "src_code_label",
        "hint_key": "src_code_hint",
        "icon": "📏",
        "types": ["xlsx", "xlsm"],
        "loader": load_code_counts,
        "required": False,
        "sample_filename": "code_counts_20260420090000.xlsx",
    },
    {
        "key": "roster",
        "label_key": "src_roster_label",
        "hint_key": "src_roster_hint",
        "icon": "👥",
        "types": ["xlsx", "xlsm"],
        "loader": load_roster,
        "required": False,
        # Present? → a "テンプレDL" button appears on the upload card so
        # users can start from a filled-out shape instead of guessing.
        "template_fn": generate_roster_template,
        "template_filename": "roster_template.xlsx",
        "sample_filename": "roster.xlsx",
    },
    {
        "key": "calendar",
        "label_key": "src_calendar_label",
        "hint_key": "src_calendar_hint",
        "icon": "🗓️",
        "types": ["xlsx", "xlsm"],
        "loader": load_calendar,
        "required": False,
        "template_fn": generate_calendar_template,
        "template_filename": "calendar_template.xlsx",
        "sample_filename": "calendar.xlsx",
    },
    {
        "key": "backlog",
        "label_key": "src_backlog_label",
        "hint_key": "src_backlog_hint",
        "icon": "📋",
        "types": ["csv"],
        "loader": load_backlog,
        "required": False,
        # No template_fn on purpose — the sample CSV doubles as a
        # concrete shape reference; exposing a separate "template" on
        # top of it would be redundant.
        "sample_filename": "backlog.csv",
    },
]


# =============================================================================
# UI: styling
# =============================================================================
_CSS = """
<style>
/* Use the full viewport width — Streamlit's default `wide` layout still
   caps the block container at ~46rem; lift the cap and tighten side padding
   so cards and tables actually breathe. */
.main .block-container,
section.main > div.block-container {
  max-width: 100% !important;
  padding-left: 1.25rem !important;
  padding-right: 1.25rem !important;
  padding-top: 1.5rem !important;
}
/* Streamlit also constrains the inner column gutters; loosen them slightly */
div[data-testid="stHorizontalBlock"] {
  gap: 0.75rem !important;
}

/* Horizontal-scroll rail for the source-file upload cards on the Dashboard.
   `:has()` targets only the st.columns row whose children contain upload
   cards (marked by .d4dx-source-card-marker), so unrelated horizontal
   blocks elsewhere in the app keep their normal flex layout.
   Each card gets a fixed width so all 7 fit in a scrollable strip; modern
   browsers supporting `:has()` (Chrome/Safari/Firefox 2023+) are expected. */
div[data-testid="stHorizontalBlock"]:has(.d4dx-source-card-marker) {
  overflow-x: auto !important;
  flex-wrap: nowrap !important;
  scroll-snap-type: x mandatory;
  padding-bottom: 10px;
  scrollbar-gutter: stable;
}
div[data-testid="stHorizontalBlock"]:has(.d4dx-source-card-marker) > div {
  flex: 0 0 280px !important;
  min-width: 280px !important;
  max-width: 280px !important;
  scroll-snap-align: start;
}
/* Slim custom scrollbar so the rail doesn't visually dominate the tab. */
div[data-testid="stHorizontalBlock"]:has(.d4dx-source-card-marker)::-webkit-scrollbar {
  height: 8px;
}
div[data-testid="stHorizontalBlock"]:has(.d4dx-source-card-marker)::-webkit-scrollbar-thumb {
  background: rgba(128,128,128,0.35);
  border-radius: 4px;
}
div[data-testid="stHorizontalBlock"]:has(.d4dx-source-card-marker)::-webkit-scrollbar-thumb:hover {
  background: rgba(128,128,128,0.55);
}
/* Hint row above the rail */
.d4dx-source-rail-hint {
  font-size: 11px;
  color: #888;
  margin: -2px 0 6px;
  letter-spacing: 0.02em;
}

/* Card shell — keep just the rounded border that Streamlit draws by default,
   no gradient fill, no hover lift. The lift made the whole grid feel like it
   was floating on top of the page background, which looked cramped. */
div[data-testid="stVerticalBlockBorderWrapper"] {
  border-radius: 12px !important;
  background: transparent !important;
  box-shadow: none !important;
}

/* Drag-drop zone */
section[data-testid="stFileUploaderDropzone"] {
  border-style: dashed !important;
  border-width: 2px !important;
  border-radius: 10px !important;
  padding: 0.75rem !important;
  min-width: 0 !important;
  transition: background .15s ease, border-color .15s ease;
}
section[data-testid="stFileUploaderDropzone"]:hover {
  background: rgba(80, 200, 140, 0.08) !important;
  border-color: #4ec78a !important;
}
/* Allow the dropzone's internal flex children to shrink instead of overflow */
section[data-testid="stFileUploaderDropzone"] > div,
section[data-testid="stFileUploaderDropzone"] > div > div {
  min-width: 0 !important;
}
section[data-testid="stFileUploaderDropzone"] small,
section[data-testid="stFileUploaderDropzone"] span {
  word-break: break-word;
  overflow-wrap: anywhere;
}
/* Filename chip after upload — keep it inside the card */
div[data-testid="stFileUploaderFile"],
div[data-testid="stFileUploaderFileName"] {
  min-width: 0 !important;
  overflow: hidden;
  text-overflow: ellipsis;
  white-space: nowrap;
}

/* streamlit-calendar's iframe sets its own height from JS; if that fails
   for any reason the iframe collapses to 0 and the calendar appears blank.
   Force a sensible minimum so the user always sees the widget. The title
   attribute differs between component releases, so match all variants. */
iframe[title="streamlit_calendar.streamlit_calendar"],
iframe[title*="streamlit_calendar"],
iframe[title*="calendar"] {
  min-height: 780px !important;
  width: 100% !important;
  border: 0 !important;
}

/* Status pills */
.d4dx-pill {
  display: inline-flex; align-items: center; gap: 6px;
  padding: 4px 10px; border-radius: 999px;
  font-size: 12px; font-weight: 600; line-height: 1;
}
.d4dx-pill.idle  { background: rgba(160,160,160,0.18); color: #9aa; }
.d4dx-pill.ok    { background: rgba(78,199,138,0.20); color: #4ec78a; }
.d4dx-pill.warn  { background: rgba(255,180,40,0.20); color: #f5b400; }
.d4dx-pill.err   { background: rgba(240,80,80,0.20); color: #f05050; }

/* Happy dino — bounces on success */
@keyframes d4dx-bounce {
  0%, 100% { transform: translateY(0) scale(1); }
  20%      { transform: translateY(-10px) scale(1.05); }
  40%      { transform: translateY(0)    scale(1); }
  55%      { transform: translateY(-5px) scale(1.03); }
  70%      { transform: translateY(0)    scale(1); }
}
.d4dx-dino {
  display: inline-block;
  font-size: 34px; line-height: 1;
  animation: d4dx-bounce 1.6s ease-in-out infinite;
  filter: drop-shadow(0 4px 8px rgba(78,199,138,0.35));
}
.d4dx-dino-row {
  display: flex; align-items: center; gap: 12px;
  margin-top: 6px;
}
.d4dx-dino-meta { font-size: 13px; color: #4ec78a; font-weight: 600; }

/* Sad dino on error */
@keyframes d4dx-shake {
  0%, 100% { transform: translateX(0); }
  20%      { transform: translateX(-3px) rotate(-4deg); }
  40%      { transform: translateX(3px)  rotate(4deg); }
  60%      { transform: translateX(-2px) rotate(-2deg); }
  80%      { transform: translateX(2px)  rotate(2deg); }
}
.d4dx-dino.sad {
  animation: d4dx-shake 0.7s ease-in-out 2;
  filter: grayscale(0.6) drop-shadow(0 4px 8px rgba(240,80,80,0.35));
}

/* Card title row */
.d4dx-card-title {
  display: flex; align-items: center; justify-content: space-between;
  margin-bottom: 4px;
}
.d4dx-card-title h4 { margin: 0; font-size: 15px; }
.d4dx-card-hint { font-size: 11px; color: #888; margin: -2px 0 8px; }
</style>
"""


def _inject_styles() -> None:
    st.markdown(_CSS, unsafe_allow_html=True)


def _pill(kind: str, text: str) -> str:
    return f"<span class='d4dx-pill {kind}'>{text}</span>"


# =============================================================================
# UI: upload card
# =============================================================================
# =============================================================================
# Dino-runner animation + step checklist + crash popup
# =============================================================================
_STEP_STATUS_ICON = {"ok": "✅", "warn": "⚠️", "error": "❌", "pending": "⏳"}


def render_dino_runner(steps: list[StepResult], slot: str,
                       nonce: Optional[str] = None) -> None:
    """Embed a Chrome-dino-style canvas animation that runs over every
    validation step. The sprite (a raptor — T-Rex is reserved for the page
    chrome) jumps over OK/warning cacti and crashes into the first error
    cactus. Triggers once per (slot, nonce, step-signature).

    `nonce` lets the caller force a re-run even when the step pattern is
    identical across uploads — pass the Streamlit uploader's `file_id`
    (fresh on every new upload) or the filename so a second upload of a
    file that still validates identically still re-plays the animation
    (without it, the signature was content-derived and silently
    deduped)."""
    if not steps:
        return
    sig = (slot, nonce, tuple((s.label_key, s.status) for s in steps))
    last_sig_key = f"_dino_runner_sig_{slot}"
    if st.session_state.get(last_sig_key) == sig:
        # Already animated this exact run; skip to keep the iframe quiet.
        return
    st.session_state[last_sig_key] = sig

    steps_data = [
        {"label": t(s.label_key), "status": s.status,
         "message": s.message, "detail": s.detail}
        for s in steps
    ]
    steps_json = json.dumps(steps_data, ensure_ascii=False)
    runner_grid = [r for r in DINO_GRIDS["raptor"].strip("\n").split("\n") if r]
    runner_json = json.dumps(runner_grid)
    canvas_id = f"dinoCanvas_{slot}_{abs(hash(sig)) % 10**8}"

    html = f"""
<div style="background:transparent;padding:0;margin:0;">
<canvas id="{canvas_id}" width="800" height="120"
        style="display:block;width:100%;background:transparent;"></canvas>
</div>
<script>
(function() {{
  const STEPS = {steps_json};
  const RUNNER = {runner_json};
  const cv = document.getElementById("{canvas_id}");
  if (!cv) return;
  const ctx = cv.getContext("2d");
  const dpr = window.devicePixelRatio || 1;
  const cssW = cv.offsetWidth || 800;
  const cssH = 120;
  cv.width = cssW * dpr;
  cv.height = cssH * dpr;
  cv.style.height = cssH + "px";
  ctx.scale(dpr, dpr);
  const W = cssW, H = cssH;
  const GROUND_Y = H - 28;

  const N = STEPS.length;
  let errorIdx = -1;
  for (let i = 0; i < N; i++) {{
    if (STEPS[i].status === "error") {{ errorIdx = i; break; }}
  }}

  const startX = 38;
  const endX = W - 36;
  const totalDist = endX - startX;
  function cactusX(i) {{ return startX + (totalDist / (N + 1)) * (i + 1); }}

  let dinoX = startX - 26;
  let dinoY = 0, dinoVy = 0, jumping = false, crashed = false;
  let frameTick = 0;
  let nextCactus = 0;
  const SPEED = Math.max(1.4, totalDist / (N * 65));

  function isDark() {{
    return matchMedia("(prefers-color-scheme: dark)").matches;
  }}

  function drawSprite(grid, x, y, scale, color) {{
    ctx.fillStyle = color;
    for (let r = 0; r < grid.length; r++) {{
      for (let c = 0; c < grid[r].length; c++) {{
        if (grid[r][c] === "X") {{
          ctx.fillRect(x + c * scale, y + r * scale, scale, scale);
        }}
      }}
    }}
  }}

  function drawGround() {{
    ctx.strokeStyle = isDark() ? "#888" : "#bbb";
    ctx.lineWidth = 1;
    ctx.beginPath();
    ctx.moveTo(0, GROUND_Y + 12);
    ctx.lineTo(W, GROUND_Y + 12);
    ctx.stroke();
  }}

  function drawCacti() {{
    for (let i = 0; i < N; i++) {{
      const cx = cactusX(i);
      const status = STEPS[i].status;
      let color = "#9aa";
      if (i === errorIdx)            color = "#f05050";
      else if (status === "ok")      color = "#4ec78a";
      else if (status === "warn")    color = "#f5b400";
      ctx.fillStyle = color;
      ctx.fillRect(cx - 3, GROUND_Y - 22, 6, 22);
      ctx.fillRect(cx - 8, GROUND_Y - 14, 5, 5);
      ctx.fillRect(cx + 3, GROUND_Y - 18, 5, 5);
      ctx.fillStyle = isDark() ? "#aaa" : "#666";
      ctx.font = "9px sans-serif";
      ctx.textAlign = "center";
      ctx.fillText(String(i + 1), cx, GROUND_Y + 24);
    }}
  }}

  function drawDino() {{
    const scale = 2;
    const w = RUNNER[0].length * scale;
    const h = RUNNER.length * scale;
    const drawY = GROUND_Y - h + dinoY;
    if (crashed) {{
      ctx.save();
      ctx.translate(dinoX + w / 2, drawY + h / 2);
      ctx.rotate(0.5);
      drawSprite(RUNNER, -w / 2, -h / 2, scale, "#f05050");
      ctx.restore();
      // Stars 💫 above
      ctx.fillStyle = "#f5b400";
      ctx.font = "16px sans-serif";
      ctx.textAlign = "center";
      ctx.fillText("✦  ✷  ✦", dinoX + w / 2, drawY - 4);
    }} else {{
      const color = isDark() ? "#fafafa" : "#222";
      drawSprite(RUNNER, dinoX, drawY, scale, color);
      // Tiny running-leg flicker (alternate)
      if (!jumping && Math.floor(frameTick / 6) % 2 === 0) {{
        ctx.fillStyle = isDark() ? "#0e1117" : "#fff";
        ctx.fillRect(dinoX + 4 * scale, drawY + (RUNNER.length - 1) * scale,
                     scale, scale);
      }}
    }}
  }}

  function tick() {{
    frameTick++;
    ctx.clearRect(0, 0, W, H);
    drawGround();
    drawCacti();
    drawDino();
    if (crashed) return;

    dinoX += SPEED;
    if (jumping) {{
      dinoY += dinoVy;
      dinoVy += 0.55;
      if (dinoY >= 0) {{ dinoY = 0; dinoVy = 0; jumping = false; }}
    }}

    if (nextCactus < N) {{
      const cx = cactusX(nextCactus);
      const dist = cx - dinoX;
      if (nextCactus === errorIdx) {{
        // Don't jump — let the dino slam into it.
        if (dist < 8) {{
          crashed = true;
          drawDino();  // final crashed draw
          return;
        }}
      }} else if (!jumping && dist < 38 && dist > 18) {{
        jumping = true;
        dinoVy = -10.5;
      }}
      if (dist < -16) nextCactus++;
    }}

    if (dinoX > W + 30) return;  // left the canvas — done
    requestAnimationFrame(tick);
  }}

  tick();
}})();
</script>
"""
    st.components.v1.html(html, height=130)


def render_step_checklist(steps: list[StepResult]) -> None:
    """Render the validation steps as a vertical bullet list with status
    icons + per-step detail text. Always reflects current state (no caching)."""
    if not steps:
        return
    lines = []
    for s in steps:
        icon = _STEP_STATUS_ICON.get(s.status, "•")
        line = f"{icon} **{t(s.label_key)}**"
        if s.detail:
            line += f" — *{s.detail}*"
        if s.message:
            line += f"<br/><span style='color:#f05050;font-family:monospace;font-size:11px;'>{s.message}</span>"
        lines.append(line)
    st.markdown(
        "<div style='line-height:1.8;font-size:13px;'>"
        + "<br/>".join(lines) +
        "</div>",
        unsafe_allow_html=True,
    )


def render_crash_popup(error_step: StepResult,
                       detail_text: str = "") -> None:
    """Inline crash banner with a hurt stegosaurus, error label, and the
    underlying message. The full structured log entry is collapsed in an
    expander underneath so the page stays scannable but the detail is one
    click away. (T-Rex is reserved for the page header / favicon.)"""
    hurt_uri = dino_data_uri("stego", color="#f05050")
    st.markdown(
        f"""
<div style="border:2px solid #f05050; border-radius:10px; padding:14px;
            background:rgba(240,80,80,0.10); display:flex; align-items:center;
            gap:14px; margin-top:6px;">
  <div style="position:relative; flex:none;">
    <img src="{hurt_uri}" style="width:64px; height:64px; transform:rotate(28deg);" />
    <div style="position:absolute; top:-6px; right:-12px; font-size:18px;">💥</div>
  </div>
  <div style="min-width:0; flex:1;">
    <div style="font-weight:700; color:#f05050; font-size:14px;">
      {t("popup_error_title")}
    </div>
    <div style="font-weight:600; margin-top:4px;">
      {t(error_step.label_key)}
    </div>
    <div style="font-family:monospace; color:#fcc; margin-top:4px;
                word-break:break-word; font-size:12px;">
      {error_step.message or '(no detail)'}
    </div>
    <div style="color:#aaa; font-size:11px; margin-top:6px;">
      {t("popup_error_hint")}
    </div>
  </div>
</div>
        """,
        unsafe_allow_html=True,
    )
    if detail_text:
        with st.expander(t("log_show_detail"), expanded=False):
            st.code(detail_text, language="text")


_MIME_FOR_EXT: dict[str, str] = {
    ".csv":  "text/csv",
    ".xlsx": ("application/vnd.openxmlformats-officedocument"
              ".spreadsheetml.sheet"),
    ".xlsm": "application/vnd.ms-excel.sheet.macroEnabled.12",
    ".xlsb": "application/vnd.ms-excel.sheet.binary.macroEnabled.12",
    ".xls":  "application/vnd.ms-excel",
    ".json": "application/json",
}

_SAMPLE_DATA_DIR = SCRIPT_DIR / "sample_data"


def _mime_for_filename(name: str) -> str:
    """Pick a sensible Content-Type for an upload/sample/latest download.
    Defaults to octet-stream when the extension isn't recognised so the
    browser still offers 'Save as' instead of trying to render it."""
    ext = Path(name).suffix.lower()
    return _MIME_FOR_EXT.get(ext, "application/octet-stream")


def _render_card_download_row(spec: dict) -> None:
    """Render the per-card sample / latest download row.

    Two icon-only `st.download_button`s; either renders only when its
    source exists. Hidden entirely when neither is available.

    Template download was dropped in v1.0.53 — the sample file already
    ships with the right headers (and, for calendar, the 75 Japanese
    holidays), so users edit the sample directly instead of starting
    from a separate empty template.
    """
    # 1. Sample — only when the sample file exists on disk.
    sample_bytes: Optional[bytes] = None
    sample_filename = spec.get("sample_filename")
    sample_path: Optional[Path] = None
    if sample_filename:
        sample_path = _SAMPLE_DATA_DIR / sample_filename
        if sample_path.exists():
            try:
                sample_bytes = sample_path.read_bytes()
            except Exception as exc:
                _get_logger().warning(
                    f"[sample] {spec['key']} read failed: {exc}")
                sample_bytes = None

    # 2. Latest snapshot — only when input/<date>/<slot>/ has any file.
    latest_path = find_latest_for_slot(spec["key"])
    latest_bytes: Optional[bytes] = None
    if latest_path is not None:
        try:
            latest_bytes = latest_path.read_bytes()
        except Exception as exc:
            _get_logger().warning(
                f"[latest] {spec['key']} read failed: {exc}")
            latest_bytes = None

    # Bail cleanly when the card has nothing to offer yet.
    if sample_bytes is None and latest_bytes is None:
        return

    c_sam, c_lat = st.columns(2, gap="small")
    if sample_bytes and sample_path is not None:
        with c_sam:
            st.download_button(
                label="🧪", data=sample_bytes,
                file_name=sample_path.name,
                mime=_mime_for_filename(sample_path.name),
                key=f"sample_{spec['key']}",
                help=t("card_dl_sample_help"),
                use_container_width=True,
            )
    if latest_bytes and latest_path is not None:
        with c_lat:
            st.download_button(
                label="📥", data=latest_bytes,
                file_name=latest_path.name,
                mime=_mime_for_filename(latest_path.name),
                key=f"latest_{spec['key']}",
                help=t("card_dl_latest_help", name=latest_path.name),
                use_container_width=True,
            )


def render_upload_card(spec: dict) -> None:
    """Render a single source card with drag-drop, instant validation, and
    auto-load of the most recent saved file when the user hasn't uploaded one."""
    label = t(spec["label_key"])
    hint = t(spec["hint_key"])
    with st.container(border=True):
        # Sentinel span — lets the horizontal-scroll CSS rule
        # (.d4dx-source-card-marker → `:has()` selector) identify which
        # st.columns row is the source-cards rail, without accidentally
        # affecting unrelated horizontal blocks elsewhere.
        st.markdown(
            "<span class='d4dx-source-card-marker' style='display:none'></span>",
            unsafe_allow_html=True,
        )
        badge = (_pill("err", t("badge_required")) if spec["required"]
                 else _pill("idle", t("badge_optional")))
        st.markdown(
            f"<div class='d4dx-card-title'>"
            f"<h4>{spec['icon']} &nbsp;{label}</h4>{badge}</div>"
            f"<div class='d4dx-card-hint'>{hint}</div>",
            unsafe_allow_html=True,
        )
        file = st.file_uploader(
            label=t("card_drop_label", label=label),
            type=spec["types"],
            key=f"upload_{spec['key']}",
            label_visibility="collapsed",
            accept_multiple_files=False,
        )

        # ----- Download row: template / sample / latest ---------------------
        # Three compact icon-only buttons on a single line; only the ones
        # the slot actually supports render. Hover tooltips explain what
        # each icon does so the card stays visually quiet.
        #   📝  template (from slot's template_fn)
        #   🧪  sample   (from sample_data/<sample_filename>)
        #   📥  latest   (most recent file under input/<date>/<slot>/)
        _render_card_download_row(spec)

        # ----- Resolve the data source: explicit upload > latest from input/ -
        data: Optional[bytes] = None
        origin_kind: Optional[str] = None  # "upload" | "auto"
        origin_name: Optional[str] = None
        origin_path: Optional[Path] = None

        if file is not None:
            # An explicit upload always wins, and clears any prior session
            # "skip auto-load" flag for this slot.
            st.session_state.skip_auto_load.pop(spec["key"], None)
            try:
                data = file.getvalue()
                origin_kind = "upload"
                origin_name = file.name
            except Exception as exc:
                st.error(t("read_upload_err", err=exc), icon="🚨")
                st.session_state.dfs.pop(spec["key"], None)
                return
        elif st.session_state.skip_auto_load.get(spec["key"]):
            # User reset auto-load in Settings — wait for an explicit upload.
            data = None
        else:
            latest = find_latest_for_slot(spec["key"])
            if latest is not None:
                try:
                    data = latest.read_bytes()
                    origin_kind = "auto"
                    origin_name = latest.name
                    origin_path = latest
                except Exception as exc:
                    st.warning(t("read_prev_warn", err=exc), icon="⚠️")

        if data is None:
            st.markdown(_pill("idle", t("status_waiting")),
                        unsafe_allow_html=True)
            st.session_state.dfs.pop(spec["key"], None)
            st.session_state.errs.pop(spec["key"], None)
            return

        # ----- Run the full step-wise validation pipeline -------------------
        df, steps = validate_with_steps(spec, data)
        error_step = next((s for s in steps if s.status == "error"), None)
        warn_steps = [s for s in steps if s.status == "warn"]

        # Dino runs over each step. Plays once per (file/auto, step-
        # signature) to avoid restarting on unrelated reruns — but the
        # nonce (file_id on upload / filename on auto-load) differs
        # between a fresh upload and the previously-cached run, so a
        # new upload always re-plays even when the validation result
        # pattern happens to match the prior one.
        if origin_kind == "upload" and file is not None:
            _dino_nonce = str(getattr(file, "file_id", None) or origin_name
                              or "")
        else:
            _dino_nonce = origin_name or ""
        render_dino_runner(steps, spec["key"], nonce=_dino_nonce)

        # Always-visible checklist of what we just checked.
        render_step_checklist(steps)

        if error_step is not None:
            err_id = file.file_id if origin_kind == "upload" else (origin_name or "")
            sig = (spec["key"], origin_kind, err_id, error_step.label_key)
            # Only log/toast on a fresh error event, not every Streamlit rerun.
            if st.session_state.last_err_sig.get(spec["key"]) != sig:
                detail = log_error(
                    category=f"upload[{spec['key']}]",
                    summary=(t(error_step.label_key) + ": "
                             + (error_step.message or "validation failed")),
                    exc=error_step.exc,
                    context={
                        "slot": spec["key"],
                        "origin": origin_kind,
                        "filename": origin_name,
                        "size_bytes": len(data),
                        "step_label_key": error_step.label_key,
                    },
                )
                st.session_state[f"_err_detail_{spec['key']}"] = detail
                st.toast(
                    t("toast_failed", label=label,
                      msg=t(error_step.label_key)),
                    icon="🚨",
                )
                st.session_state.last_err_sig[spec["key"]] = sig
            detail_text = st.session_state.get(
                f"_err_detail_{spec['key']}", ""
            )
            render_crash_popup(error_step, detail_text)
            st.session_state.dfs.pop(spec["key"], None)
            st.session_state.errs[spec["key"]] = [error_step.message]
            return

        # ----- Success: persist new uploads (only after they validated) ------
        if origin_kind == "upload":
            sig = (spec["key"], file.file_id)
            if st.session_state.last_ok_sig.get(spec["key"]) != sig:
                try:
                    saved = save_uploaded_bytes(spec["key"], file.name, data)
                    origin_path = saved
                    rel_path = (
                        f"{saved.parent.parent.name}/"
                        f"{saved.parent.name}/{saved.name}"
                    )
                    st.toast(
                        t("toast_loaded", label=label, path=rel_path), icon="✅"
                    )
                except Exception as exc:
                    st.warning(t("save_warn", err=exc), icon="⚠️")
                st.session_state.last_ok_sig[spec["key"]] = sig

        # OK summary line + source provenance.
        n_str = f"{len(df):,}"
        ok_text = (t("validation_warnings") if warn_steps
                   else t("validation_passed"))
        # Not every source keys on 機能ID — roster rows are per-assignee,
        # calendar rows are per-event / per-non-working-day. Fall back to
        # a count-only line when the dataframe has no 機能ID column.
        if "機能ID" in df.columns:
            u_str = f"{df['機能ID'].nunique():,}"
            status_text = t("status_ok", n=n_str, u=u_str)
        else:
            status_text = t("status_ok_no_fid", n=n_str)
        st.markdown(
            "<div class='d4dx-dino-row'>"
            f"<span style='color:#4ec78a;font-weight:600;font-size:13px;'>"
            f"{ok_text}</span>"
            f"<span class='d4dx-dino-meta'>· "
            f"{status_text}</span>"
            "</div>",
            unsafe_allow_html=True,
        )
        src_icon = "📤" if origin_kind == "upload" else "💾"
        src_text = t("origin_upload") if origin_kind == "upload" else t("origin_auto")
        snap = _snapshot_date_from_filename(origin_name or "")
        snap_text = (" · " + t("origin_snapshot", date=snap.isoformat())
                     if snap else "")
        # Ingest timestamp = when the bytes landed on disk under input/.
        # For auto-load it's the file's mtime; for a fresh upload we
        # just saved the file above so origin_path.stat() still reflects
        # that save moment. Shows a human-readable `YYYY-MM-DD HH:MM`.
        ingest_text = ""
        if origin_path is not None:
            try:
                ts = datetime.fromtimestamp(origin_path.stat().st_mtime)
                ingest_text = (" · " + t("origin_ingested_at",
                                          ts=ts.strftime("%Y-%m-%d %H:%M")))
            except Exception:
                pass
        st.caption(
            f"{src_icon} {src_text} · `{origin_name}`{snap_text}{ingest_text}"
        )

        st.session_state.dfs[spec["key"]] = df
        st.session_state.origin_names[spec["key"]] = origin_name or ""
        st.session_state.errs.pop(spec["key"], None)


# =============================================================================
# UI: main
# =============================================================================
def _ensure_design_pages_state() -> dict[str, int]:
    """Initialize and return the session-shared design-pages state."""
    if "design_pages_state" not in st.session_state:
        st.session_state.design_pages_state = (
            {} if st.session_state.get("skip_design_pages_load")
            else load_design_pages()
        )
    return st.session_state.design_pages_state


def get_current_kpi_df() -> Optional[pd.DataFrame]:
    """Build the current integrated KPI dataframe from session_state.dfs.
    Returns None if the master is not yet loaded."""
    master = st.session_state.dfs.get("master")
    if master is None or master.empty:
        return None
    pages_state = _ensure_design_pages_state()
    master_ids = set(master["機能ID"].unique())
    pages = {fid: v for fid, v in pages_state.items() if fid in master_ids}
    design_df = build_design_pages_df(master, pages)
    integrated = integrate(
        master=master,
        wbs=st.session_state.dfs.get("wbs"),
        defects=st.session_state.dfs.get("defects"),
        tests=st.session_state.dfs.get("tests"),
        code=st.session_state.dfs.get("code"),
        design_pages=design_df if not design_df.empty else None,
    )
    return compute_kpis(integrated)


def _capture_drilldown(event, df: pd.DataFrame) -> None:
    """If the table widget reported a selected row, write its 機能ID into
    session_state. Used by every selectable table on the dashboard so the
    drill-down panel can pick up the latest pick from any tab."""
    sel = getattr(event, "selection", None) if event is not None else None
    rows = getattr(sel, "rows", None) if sel is not None else None
    if rows:
        idx = rows[0]
        if 0 <= idx < len(df):
            st.session_state.drilldown_id = str(df.iloc[idx]["機能ID"])


_DRILLDOWN_TABLE_KEYS = (
    "drill_overview", "drill_kpis", "drill_wbs",
    "drill_defects", "drill_tests", "drill_code", "drill_all",
)


def _selectable_table(df: pd.DataFrame, key: str, *,
                      column_config: dict, height: int):
    """Streamlit dataframe with single-row selection that drives the
    drill-down panel below the integrated tables."""
    return st.dataframe(
        df,
        use_container_width=True,
        height=height,
        hide_index=True,
        column_config=column_config,
        selection_mode="single-row",
        on_select="rerun",
        key=key,
    )


DRILLDOWN_SOURCE_STRIP: list[tuple[str, str, str]] = [
    # (session-dfs key, i18n label key, dino name)
    # Design pages is a special case: state lives in design_pages_state,
    # not in session_state.dfs.
    ("master",  "src_master_label",  "bronto"),
    ("wbs",     "src_wbs_label",     "raptor"),
    ("defects", "src_defects_label", "spino"),
    ("tests",   "src_tests_label",   "stego"),
    ("code",    "src_code_label",    "diplo"),
    ("design",  "src_design_label",  "ptero"),
]

_STRIP_COLORS = {"present": "#4ec78a", "absent": "#b48820",
                 "unloaded": "#7a7f88"}
_STRIP_BADGES = {"present": "✓", "absent": "⚠", "unloaded": "—"}
_STRIP_BORDERS = {"present": "#2d6a4a", "absent": "#6b5220",
                  "unloaded": "#3a3d42"}


def _fid_presence_for_slot(slot: str, fid: str) -> tuple[str, str]:
    """Return ('present'|'absent'|'unloaded', filename) for one source slot."""
    if slot == "design":
        pages = st.session_state.get("design_pages_state") or {}
        if not pages:
            return "unloaded", ""
        val = pages.get(fid)
        state = "present" if val and val > 0 else "absent"
        return state, t("drilldown_source_manual")
    df = st.session_state.dfs.get(slot)
    fname = st.session_state.get("origin_names", {}).get(slot, "") or ""
    if df is None or df.empty or "機能ID" not in df.columns:
        return "unloaded", fname
    state = "present" if fid in df["機能ID"].values else "absent"
    return state, fname


def render_drilldown_presence_strip(fid: str) -> None:
    """Horizontal strip of compact cards — one per input source — showing
    whether the selected Function ID has a row in each loaded source."""
    cards_html: list[str] = []
    for slot, label_key, dino in DRILLDOWN_SOURCE_STRIP:
        state, fname = _fid_presence_for_slot(slot, fid)
        color = _STRIP_COLORS[state]
        border = _STRIP_BORDERS[state]
        badge = _STRIP_BADGES[state]
        label = t(label_key)
        shown = fname or t("drilldown_source_unloaded")
        shown_short = (shown if len(shown) <= 24
                       else shown[:10] + "…" + shown[-12:])
        svg_uri = dino_data_uri(dino, color=color)
        file_style = (
            "color:#ccc;" if state == "present"
            else "color:#888;text-decoration:line-through;"
            if state == "absent" and fname
            else "color:#777;font-style:italic;"
        )
        tooltip = f"{label} · {shown}"
        cards_html.append(
            f'<div class="d4dx-pc-card" '
            f'style="border-color:{border};" title="{tooltip}">'
            f'<div class="d4dx-pc-head">'
            f'<img class="d4dx-pc-dino" src="{svg_uri}" alt="{dino}">'
            f'<span class="d4dx-pc-badge" '
            f'style="color:{color};border-color:{color};">{badge}</span>'
            f'</div>'
            f'<div class="d4dx-pc-src">{label}</div>'
            f'<div class="d4dx-pc-fname" style="{file_style}">{shown_short}</div>'
            f'</div>'
        )
    st.markdown(
        """
<style>
.d4dx-pc-strip { display:flex; gap:6px; flex-wrap:wrap; margin:6px 0 12px; }
.d4dx-pc-card  { flex:1 1 120px; min-width:110px; max-width:160px;
                 border:1px solid #3a3d42; border-radius:8px;
                 padding:6px 8px; background:rgba(255,255,255,0.02); }
.d4dx-pc-head  { display:flex; align-items:center; justify-content:space-between;
                 margin-bottom:4px; }
.d4dx-pc-dino  { width:22px; height:22px; flex:none; }
.d4dx-pc-badge { font-size:11px; font-weight:700; width:16px; height:16px;
                 display:inline-flex; align-items:center; justify-content:center;
                 border:1px solid; border-radius:50%; flex:none; line-height:1; }
.d4dx-pc-src   { font-weight:600; font-size:11px;
                 white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
.d4dx-pc-fname { font-family:"SF Mono",Menlo,monospace; font-size:9.5px;
                 margin-top:1px; white-space:nowrap; overflow:hidden;
                 text-overflow:ellipsis; }
</style>
""" + f'<div class="d4dx-pc-strip">{"".join(cards_html)}</div>',
        unsafe_allow_html=True,
    )


def _render_drilldown_subtask_views(subtasks: pd.DataFrame) -> None:
    """Render the three WBS-derived drill-down widgets for a feature:
    assignee roster, role-progress mini-bars, and sub-task breakdown
    table. No-op when the feature has no WBS sub-tasks."""
    if subtasks is None or subtasks.empty:
        st.caption(
            f"**{t('drilldown_assignees_label')}:** "
            f"{t('drilldown_assignees_none')}"
        )
        return

    role_labels_local = {
        "dev":       t("role_dev"),
        "test_spec": t("role_test_spec"),
        "test_exec": t("role_test_exec"),
    }
    status_colors = {
        "completed":  "#4ec78a",
        "in_progress": "#f5b400",
        "not_started": "#c0c0c0",
    }

    # ---- Assignee roster (unique (assignee, role) pairs) ---------------
    roster_pairs: list[tuple[str, str]] = []
    seen: set = set()
    for _, r in subtasks.iterrows():
        a = _normalize_assignee(r.get("assignee"))
        role = r.get("role") or ""
        if not a:
            continue
        key = (a, role)
        if key in seen:
            continue
        seen.add(key)
        roster_pairs.append(key)
    if roster_pairs:
        parts = [
            f"{a}"
            + (f" <span style='color:#888;'>({role_labels_local[r]})</span>"
               if r in role_labels_local else "")
            for a, r in roster_pairs
        ]
        st.markdown(
            f"**{t('drilldown_assignees_label')}:** 👥 "
            + " · ".join(parts),
            unsafe_allow_html=True,
        )

    # ---- Role-progress mini-bars -------------------------------------
    # For each of the 3 analytics roles, count how many of this
    # feature's sub-tasks are completed / in-progress / not-started and
    # render a short stacked bar so "dev done but test_exec hasn't
    # started" pops out visually.
    subtasks = subtasks.copy()
    subtasks["_status"] = subtasks.apply(_subtask_progress_bucket, axis=1)
    role_rows: list[str] = []
    for role in ROLE_KEYWORDS:
        sub = subtasks[subtasks["role"] == role]
        if sub.empty:
            continue
        counts = sub["_status"].value_counts()
        total = int(counts.sum())
        c_done = int(counts.get("completed", 0))
        c_prog = int(counts.get("in_progress", 0))
        c_idle = int(counts.get("not_started", 0))
        # Build a 3-segment stacked bar sized by status share.
        pct_done = (c_done / total) * 100 if total else 0
        pct_prog = (c_prog / total) * 100 if total else 0
        pct_idle = 100 - pct_done - pct_prog
        bar = (
            f"<span style='display:inline-block;width:150px;height:10px;"
            f"border-radius:3px;overflow:hidden;"
            f"background:{status_colors['not_started']};vertical-align:middle;'>"
            f"<span style='display:inline-block;width:{pct_done:.1f}%;"
            f"height:100%;background:{status_colors['completed']};'></span>"
            f"<span style='display:inline-block;width:{pct_prog:.1f}%;"
            f"height:100%;background:{status_colors['in_progress']};'></span>"
            f"</span>"
        )
        role_rows.append(
            f"<div style='display:flex;align-items:center;gap:10px;"
            f"font-size:12px;padding:1px 0;'>"
            f"<span style='min-width:100px;'>{role_labels_local[role]}</span>"
            f"{bar}"
            f"<span style='color:#666;font-size:11px;'>"
            f"{t('drilldown_status_completed')} {c_done} · "
            f"{t('drilldown_status_in_progress')} {c_prog} · "
            f"{t('drilldown_status_not_started')} {c_idle}</span>"
            f"</div>"
        )
    if role_rows:
        st.markdown(
            f"**{t('drilldown_role_progress_label')}**  \n"
            f"<span style='color:#888;font-size:11px;'>"
            f"{t('drilldown_role_progress_caption')}</span>",
            unsafe_allow_html=True,
        )
        st.markdown("\n".join(role_rows), unsafe_allow_html=True)

    # ---- Sub-task breakdown table (inside an expander so it doesn't
    # dominate the panel when the feature has many sub-tasks) ----------
    with st.expander(t("drilldown_subtask_expander"), expanded=False):
        def _fmt_date(v) -> str:
            d = _to_pydate(v)
            return d.isoformat() if d else "—"

        def _fmt_period(s, e) -> str:
            return f"{_fmt_date(s)} → {_fmt_date(e)}"

        def _subtask_delay(r) -> str:
            pe = _to_pydate(r.get("planned_end"))
            ae = _to_pydate(r.get("actual_end"))
            if pe is None:
                return "—"
            if ae is not None:
                d = (ae - pe).days
            else:
                d = (date.today() - pe).days
            return f"{d}" if d > 0 else "0"

        rows = []
        for _, r in subtasks.iterrows():
            role = r.get("role") or ""
            role_text = role_labels_local.get(
                role, t("drilldown_subtask_role_other"))
            prog = r.get("actual_progress")
            # actual_progress is already 0..100 (normalised by
            # `_to_percent_scale`) — don't multiply again.
            prog_text = (f"{float(prog):.0f}%"
                         if pd.notna(prog) else "—")
            rows.append({
                t("drilldown_subtask_col_label"): r.get("task_label") or "",
                t("drilldown_subtask_col_assignee"):
                    _normalize_assignee(r.get("assignee")) or "—",
                t("drilldown_subtask_col_role"): role_text,
                t("drilldown_subtask_col_planned"): _fmt_period(
                    r.get("planned_start"), r.get("planned_end")),
                t("drilldown_subtask_col_actual"): _fmt_period(
                    r.get("actual_start"), r.get("actual_end")),
                t("drilldown_subtask_col_progress"): prog_text,
                t("drilldown_subtask_col_delay"): _subtask_delay(r),
            })
        st.dataframe(pd.DataFrame(rows), use_container_width=True,
                     hide_index=True)


def render_drilldown_panel(kpi_df: pd.DataFrame,
                           defects_df: Optional[pd.DataFrame],
                           function_id: str) -> None:
    """All-in-one detail view for a single Function ID. Aggregates duplicate-
    name rows so the metrics are shown once even if the master holds the same
    ID with multiple names."""
    rows = kpi_df[kpi_df["機能ID"] == function_id]
    if rows.empty:
        st.info(t("drilldown_id_not_found", fid=function_id))
        return

    # ---- Aggregate per-Function-ID values (defensive: take first since
    # duplicate-name rows hold identical KPI values).
    row = rows.iloc[0]
    names = sorted(rows["機能名称"].dropna().astype(str).unique())
    name_label = " / ".join(names) if names else ""

    with st.container(border=True):
        # Header row + close button
        title_col, close_col = st.columns([10, 1], gap="small",
                                          vertical_alignment="center")
        with title_col:
            risk = row.get("risk_score")
            risk_color = ("#f05050" if pd.notna(risk) and risk >= 0.5
                          else "#f5b400" if pd.notna(risk) and risk >= 0.3
                          else "#4ec78a")
            risk_pill = (
                f"<span style='background:{risk_color};color:#fff;padding:2px 8px;"
                f"border-radius:999px;font-size:11px;font-weight:600;'>"
                f"risk {risk:.2f}</span>"
                if pd.notna(risk) else ""
            )
            st.markdown(
                f"### {t('drilldown_panel_title')} · "
                f"`{function_id}` &nbsp; {risk_pill}",
                unsafe_allow_html=True,
            )
            if name_label:
                st.caption(name_label)
        with close_col:
            if st.button("✕", key="drilldown_close_btn",
                         help=t("drilldown_close"),
                         use_container_width=True):
                st.session_state.pop("drilldown_id", None)
                # Clear all per-table selection state so a fresh click works.
                for k in _DRILLDOWN_TABLE_KEYS:
                    st.session_state.pop(k, None)
                st.rerun()

        # ---- Per-source presence strip ----------------------------------
        render_drilldown_presence_strip(function_id)

        # ---- Metric grid ------------------------------------------------
        def _f(v, fmt="{:.0f}"):
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return "—"
            try:
                return fmt.format(float(v))
            except (TypeError, ValueError):
                return str(v)

        def _pct(v):
            return f"{float(v) * 100:.1f}%" if pd.notna(v) else "—"

        def _date(v):
            d = _to_pydate(v)
            return d.isoformat() if d else "—"

        def _derived(kpi_name: str, base_help: str,
                     formatter: Callable[[object], str]) -> tuple[str, str]:
            """Return ``(value_str, help_str)`` for a derived KPI metric.

            If any required source column is missing for this row the value
            becomes ``"?"`` and the help text lists which input is blank,
            so the user knows exactly which resource to fill in.
            """
            missing = _missing_inputs_for(row, kpi_name)
            if missing:
                return "?", _compose_missing_help(base_help, missing)
            return formatter(row.get(kpi_name)), base_help

        # Schedule (WBS)
        st.markdown(f"#### {t('drilldown_section_wbs')}")
        wbs_cols = st.columns(5, gap="small")
        wbs_cols[0].metric(t("drilldown_planned_period"),
                           f"{_date(row.get('planned_start'))}",
                           f"→ {_date(row.get('planned_end'))}")
        wbs_cols[1].metric(t("drilldown_actual_period"),
                           f"{_date(row.get('actual_start'))}",
                           f"→ {_date(row.get('actual_end'))}")
        wbs_cols[2].metric(t("drilldown_planned_effort"),
                           _f(row.get("planned_effort"), "{:.1f}"))
        wbs_cols[3].metric(t("drilldown_actual_effort"),
                           _f(row.get("actual_effort"), "{:.1f}"))
        _v, _h = _derived("delay_days", t("help_delay_days"),
                          lambda v: _f(v, "{:.0f}"))
        wbs_cols[4].metric(t("col_delay_days"), _v, help=_h)

        # Progress + ⏰ To-deadline share a 3-column row so the deadline
        # readout gets ~2× the width of the 6-col version — avoids
        # truncating long values like `完了 04/12` or `本日が予定日`.
        prog_cols = st.columns(3, gap="small")
        prog_cols[0].metric(t("drilldown_planned_progress"),
                            _f(row.get("planned_progress"), "{:.0f}%"),
                            help=t("help_planned_progress"))
        prog_cols[1].metric(t("drilldown_actual_progress"),
                            _f(row.get("actual_progress"), "{:.0f}%"),
                            help=t("help_actual_progress"))
        # ⏰ To-deadline — always via `st.metric` so the "?" click
        # popover, label font, and value sizing match every sibling
        # metric in the panel. Overdue-case severity is carried on the
        # delta row ("+N日超過" in red via delta_color='inverse') rather
        # than by recolouring the main value, which keeps the DOM /
        # alignment identical to other metrics.
        _actual_end = _to_pydate(row.get("actual_end"))
        _planned_end = _to_pydate(row.get("planned_end"))
        deadline_delta = None
        deadline_delta_color = "off"
        if _actual_end is not None:
            deadline_text = t("drilldown_deadline_completed",
                              date=_actual_end.strftime("%m/%d"))
        elif _planned_end is not None:
            delta = (_planned_end - date.today()).days
            if delta > 0:
                deadline_text = t("drilldown_deadline_future", n=delta)
            elif delta == 0:
                deadline_text = t("drilldown_deadline_today")
            else:
                # Overdue: primary readout shows the signed day count;
                # the red "超過" delta line makes the severity obvious.
                deadline_text = f"+{-delta} 日"
                deadline_delta = t("drilldown_deadline_overdue_badge")
                deadline_delta_color = "inverse"
        else:
            deadline_text = t("drilldown_deadline_unknown")
        prog_cols[2].metric(
            "⏰ " + t("drilldown_to_deadline"),
            deadline_text,
            delta=deadline_delta,
            delta_color=deadline_delta_color,
            help=t("drilldown_help_deadline"),
        )

        # Per-feature WBS sub-tasks — powers the assignee list, the
        # role-progress bars, and the sub-task breakdown table below.
        # Each is a no-op when the feature has no sub-tasks.
        wbs_df = st.session_state.dfs.get("wbs")
        subtasks = _subtasks_for_function(wbs_df, function_id)
        _render_drilldown_subtask_views(subtasks)

        # Tests
        st.markdown(f"#### {t('drilldown_section_tests')}")
        t_cols = st.columns(7, gap="small")
        t_cols[0].metric("総設定テスト数", _f(row.get("総設定テスト数"), "{:.0f}"))
        t_cols[1].metric("実施済", _f(row.get("実施済"), "{:.0f}"))
        t_cols[2].metric("OK", _f(row.get("OK"), "{:.0f}"))
        t_cols[3].metric(t("col_test_ng"), _f(row.get("NG"), "{:.0f}"),
                         help=t("help_test_ng"))
        t_cols[4].metric("未実施", _f(row.get("未実施"), "{:.0f}"))
        _v, _h = _derived("test_run_rate",
                          t("help_test_run_rate"), _pct)
        t_cols[5].metric(t("col_test_run_rate"), _v, help=_h)
        _v, _h = _derived("test_pass_rate",
                          t("help_test_pass_rate"), _pct)
        t_cols[6].metric(t("col_test_pass_rate"), _v, help=_h)

        # Code & Design
        st.markdown(f"#### {t('drilldown_section_code')}")
        c_cols = st.columns(5, gap="small")
        c_cols[0].metric("LoC", _f(row.get("LoC"), "{:,.0f}"))
        c_cols[1].metric("設計書ページ数", _f(row.get("設計書ページ数"), "{:.0f}"))
        _v, _h = _derived("complexity",
                          t("help_complexity"),
                          lambda v: _f(v, "{:.1f}"))
        c_cols[2].metric(t("col_complexity"), _v, help=_h)
        _v, _h = _derived("test_density",
                          t("help_test_density"),
                          lambda v: _f(v, "{:.2f}"))
        c_cols[3].metric(t("col_test_density"), _v, help=_h)
        _v, _h = _derived("bug_density",
                          t("help_bug_density"),
                          lambda v: _f(v, "{:.3f}"))
        c_cols[4].metric(t("col_bug_density"), _v, help=_h)

        # Composite scores
        st.markdown(f"#### {t('drilldown_section_scores')}")
        s_cols = st.columns(2, gap="small")
        s_cols[0].metric(t("col_health_score"),
                         _f(row.get("health_score"), "{:.2f}"),
                         help=t("help_health_score"))
        s_cols[1].metric(t("col_risk_score"),
                         _f(row.get("risk_score"), "{:.2f}"),
                         help=t("help_risk_score"))

        # Defects
        st.markdown(f"#### {t('drilldown_section_defects')}")
        d_cols = st.columns(4, gap="small")
        d_cols[0].metric(t("col_defect_total"),
                         _f(row.get("defect_total"), "{:.0f}"),
                         help=t("help_defect_total"))
        d_cols[1].metric(t("col_defect_unresolved"),
                         _f(row.get("defect_unresolved"), "{:.0f}"),
                         help=t("help_defect_unresolved"))
        _v, _h = _derived("defect_rate", t("help_defect_rate"), _pct)
        d_cols[2].metric(t("col_defect_rate"), _v, help=_h)
        _v, _h = _derived("incident_rate",
                          t("help_incident_rate"), _pct)
        d_cols[3].metric(t("col_incident_rate"), _v, help=_h)

        if defects_df is not None and not defects_df.empty:
            related = defects_df[defects_df["機能ID"] == function_id].copy()
            if related.empty:
                st.caption(t("drilldown_no_defects"))
            else:
                st.markdown(
                    f"**{t('drilldown_related_defects', n=len(related))}**"
                )
                st.dataframe(
                    related[["トラッカー", "ステータス", "担当者",
                             "実開始日", "実終了日", "問題分類", "unresolved"]],
                    use_container_width=True,
                    hide_index=True,
                    height=min(40 + 36 * len(related), 380),
                )
        else:
            st.caption(t("drilldown_no_defects"))

        # ---- Per-FID trend (snapshots) ----------------------------------
        trend_fig = _chart_fid_trend(function_id)
        if trend_fig is not None:
            st.markdown(f"#### {t('drilldown_section_trend')}")
            st.caption(t("drilldown_section_trend_caption"))
            st.plotly_chart(trend_fig, use_container_width=True)


# =============================================================================
# Alert tab — per-Function-ID risk-score alerts
# =============================================================================
# One alert per feature. The four metrics below each contribute a normalized
# value `n ∈ [0, 1]` (0 = within spec, 1 = saturated bad); the weighted mean
# collapses into a single risk score `0..1` used to bucket severity.
#
# Risk score = Σ(n × w) ÷ Σ(w)
#
# Metric   | weight | normalization
# ---------|--------|------------------------------------------------------
# incident | 3.0    | min(1, ir ÷ (2 × threshold))  breach when ir > thr
# delay    | 1.5    | min(1, (days − 14) ÷ 46)       breach when days > 14
# density  | 1.0    | max(0, 1 − td ÷ threshold)     breach when td < thr
# notrun%  | 1.0    | max(0, (pct − 60) ÷ 40)        breach when pct > 60
#
# Severity: score > 0.70 → HIGH, > 0.35 → MED, > 0 → LOW. Features with
# all four metrics in-spec (score = 0) do NOT generate an alert.
_ALERT_SEV_ORDER = {"high": 0, "medium": 1, "low": 2}
_ALERT_SEV_COLOR = {
    "high":   "#f05050",
    "medium": "#f5b400",
    "low":    "#7ab3ff",
}
_RISK_WEIGHTS = {
    "incident_rate": 3.0,
    "delay_days":    1.5,
    "test_density":  1.0,
    "notrun_ratio":  1.0,
}
_RISK_SEV_HIGH = 0.70
_RISK_SEV_MED  = 0.35
_DELAY_CAP_DAYS   = 60     # days where n=1.0 saturates
_DELAY_FLOOR_DAYS = 14     # breach threshold
_NOTRUN_FLOOR_PCT = 0.60   # breach threshold (ratio)
_NOTRUN_MIN_PLAN  = 10     # minimum 総設定テスト数 for notrun to matter


def _risk_severity(score: float) -> str:
    if score > _RISK_SEV_HIGH:
        return "high"
    if score > _RISK_SEV_MED:
        return "medium"
    return "low"


def detect_kpi_alerts(kpi_df: pd.DataFrame) -> list[dict]:
    """One risk-scored alert per Function ID.

    Every feature is evaluated across the four breach metrics above.
    A metric contributes `n × w` only when it crosses its threshold;
    otherwise it's clamped to 0. The per-feature risk score is the
    weighted mean (denominator = constant Σw so the scale stays 0..1).

    Each alert carries: severity / risk_score / date / fid / label /
    metrics[]. `metrics` is ordered by contribution descending so the
    UI can mark the biggest one as `is_main=True` ("X" badge) and the
    rest as `is_main=False` ("△" badge).
    """
    alerts: list[dict] = []
    if kpi_df is None or kpi_df.empty:
        return alerts

    incident_thr = _incident_rate_threshold()
    density_thr  = _test_density_threshold()
    w_total = sum(_RISK_WEIGHTS.values())

    for _, r in kpi_df.iterrows():
        fid = str(r.get("機能ID") or "")
        if not fid:
            continue
        label = _label_id_name(r)
        # Relevant date for this feature — preferred order:
        #   actual_end  (terminated / already shipped — most accurate)
        #   planned_end (scheduled ship date — if not yet shipped)
        #   unknown     (neither column populated — surface honestly
        #                instead of pretending "today" is meaningful)
        _actual = _to_pydate(r.get("actual_end"))
        _planned = _to_pydate(r.get("planned_end"))
        if _actual is not None:
            ref_date, date_kind = _actual, "actual"
        elif _planned is not None:
            ref_date, date_kind = _planned, "planned"
        else:
            ref_date, date_kind = None, "unknown"

        metrics: list[dict] = []

        # --- 障害発生率 ---
        ir = r.get("incident_rate")
        if pd.notna(ir) and float(ir) > incident_thr:
            ir_f = float(ir)
            n = min(1.0, ir_f / max(1e-9, 2.0 * incident_thr))
            metrics.append({
                "key": "incident_rate",
                "label": t("col_incident_rate"),
                "display": f"{ir_f * 100:.1f} %",
                "n": n, "weight": _RISK_WEIGHTS["incident_rate"],
                "contribution": n * _RISK_WEIGHTS["incident_rate"],
            })

        # --- 遅延日数 ---
        delay = r.get("delay_days")
        if pd.notna(delay) and float(delay) > _DELAY_FLOOR_DAYS:
            d_f = float(delay)
            n = min(1.0, (d_f - _DELAY_FLOOR_DAYS) /
                    max(1.0, _DELAY_CAP_DAYS - _DELAY_FLOOR_DAYS))
            metrics.append({
                "key": "delay_days",
                "label": t("col_delay_days"),
                "display": f"{int(d_f)} 日",
                "n": n, "weight": _RISK_WEIGHTS["delay_days"],
                "contribution": n * _RISK_WEIGHTS["delay_days"],
            })

        # --- テスト密度 ---
        td = r.get("test_density")
        if pd.notna(td) and float(td) < density_thr:
            td_f = float(td)
            n = max(0.0, 1.0 - td_f / max(1e-9, density_thr))
            metrics.append({
                "key": "test_density",
                "label": t("col_test_density"),
                "display": f"{td_f:.2f}",
                "n": n, "weight": _RISK_WEIGHTS["test_density"],
                "contribution": n * _RISK_WEIGHTS["test_density"],
            })

        # --- テスト未実施率 ---
        total = r.get("総設定テスト数")
        notrun = r.get("未実施")
        if (pd.notna(total) and float(total) >= _NOTRUN_MIN_PLAN
                and pd.notna(notrun)):
            ratio = float(notrun) / float(total)
            if ratio > _NOTRUN_FLOOR_PCT:
                n = max(0.0, (ratio - _NOTRUN_FLOOR_PCT) /
                        max(1e-9, 1.0 - _NOTRUN_FLOOR_PCT))
                metrics.append({
                    "key": "notrun_ratio",
                    "label": t("help_test_notrun_short"),
                    "display": f"{ratio * 100:.0f} %",
                    "n": n, "weight": _RISK_WEIGHTS["notrun_ratio"],
                    "contribution": n * _RISK_WEIGHTS["notrun_ratio"],
                })

        if not metrics:
            continue

        # Sort by contribution desc; the first one wins the "X" badge.
        metrics.sort(key=lambda m: m["contribution"], reverse=True)
        for i, m in enumerate(metrics):
            m["is_main"] = (i == 0)

        risk_score = sum(m["contribution"] for m in metrics) / w_total
        alerts.append({
            "severity": _risk_severity(risk_score),
            "risk_score": risk_score,
            "date": ref_date, "date_kind": date_kind,
            "fid": fid, "label": label,
            "metrics": metrics,
        })

    # Default order = severity first (high → low), FID as tiebreak. The
    # render layer can re-sort based on user selection.
    alerts.sort(key=lambda a: (_ALERT_SEV_ORDER[a["severity"]], a["fid"]))
    return alerts


def render_alert_tab() -> None:
    """Dedicated 🚨 アラート tab. Shows:

    - Summary counts by severity
    - A card per alert with the feature label, metric, current value
      vs threshold, and severity badge

    Honours the global Function ID filter so narrowing on the side
    panel also narrows the alert set. When the dataset is clean the
    tab renders a calm "all clear" placeholder instead of an empty
    list — mild UX touch so the user knows the page is live.
    """
    kpi_df = get_current_kpi_df()
    if kpi_df is None or kpi_df.empty:
        st.info(t("alert_needs_master"))
        return

    selected_fids = _get_global_fids()
    if selected_fids:
        kpi_df = kpi_df[kpi_df["機能ID"].astype(str).isin(selected_fids)].copy()

    st.subheader(t("alert_tab_title"))
    st.caption(t("alert_tab_caption"))

    alerts = detect_kpi_alerts(kpi_df)

    # Severity summary bar (always visible so zero-alert state is affirming).
    sev_counts = {"high": 0, "medium": 0, "low": 0}
    for a in alerts:
        sev_counts[a["severity"]] += 1
    c_h, c_m, c_l = st.columns(3, gap="small")
    for col, sev, label_key in zip(
        (c_h, c_m, c_l),
        ("high", "medium", "low"),
        ("alert_sev_high_label",
         "alert_sev_medium_label",
         "alert_sev_low_label"),
    ):
        with col:
            n = sev_counts[sev]
            colour = _ALERT_SEV_COLOR[sev] if n else "#7aa088"
            st.markdown(
                f"""
<div style="padding:10px 12px; border-radius:8px;
            border-left:4px solid {colour};
            background:rgba(128,128,128,0.06);">
  <div style="font-size:11px; color:#888;">{t(label_key)}</div>
  <div style="font-size:22px; font-weight:700; color:inherit;">{n}</div>
</div>
                """,
                unsafe_allow_html=True,
            )

    st.markdown("")
    if not alerts:
        st.success(t("alert_all_clear"), icon="✅")
        return

    # Sort selector — lets the user flip between severity-first and
    # date-based ordering. Date = the feature's planned_end (falling
    # back to actual_end / today) so "date 新→古" surfaces the
    # latest-shipping features first.
    sort_options = [
        ("severity", t("alert_sort_severity")),
        ("date_desc", t("alert_sort_date_desc")),
        ("date_asc",  t("alert_sort_date_asc")),
    ]
    sort_choice = st.selectbox(
        t("alert_sort_label"),
        options=[k for k, _ in sort_options],
        format_func=dict(sort_options).__getitem__,
        key="alert_sort_choice",
    )
    if sort_choice == "date_desc":
        alerts = sorted(
            alerts,
            key=lambda a: (a.get("date") or date.min, a["fid"]),
            reverse=True,
        )
    elif sort_choice == "date_asc":
        alerts = sorted(
            alerts,
            key=lambda a: (a.get("date") or date.max, a["fid"]),
        )
    # else: keep default severity order from detect_kpi_alerts.

    # Scoring legend (collapsible) — explains the risk score and weights
    # once at the top so each tile below can stay compact. Default
    # collapsed; repeat-readers will likely leave it that way.
    with st.expander(t("alert_score_legend_title"), expanded=False):
        st.markdown(t("alert_score_legend_body"))

    # One tile per feature. Each tile shows the risk-score headline plus
    # the list of breaching metrics with a culprit marker (X = main
    # contributor, △ = also breaching but smaller contribution).
    for a in alerts:
        sev = a["severity"]
        colour = _ALERT_SEV_COLOR[sev]
        badge_label = t(f"alert_sev_{sev}")
        score = a.get("risk_score", 0.0)
        date_kind = a.get("date_kind", "unknown")
        date_label = t(f"alert_date_label_{date_kind}")
        if date_kind == "unknown" or a.get("date") is None:
            date_badge = date_label
        else:
            date_badge = f"{date_label}: {a['date'].isoformat()}"
        # Each metric row: "障害発生率   50.0 %   X" — value right-aligned
        # inside a fixed-width column so readings line up vertically.
        metric_rows = []
        for m in a["metrics"]:
            mark = "X" if m["is_main"] else "△"
            mark_color = "#d43a3a" if m["is_main"] else "#d98a00"
            metric_rows.append(
                f"""
  <div style="display:flex; align-items:center; gap:10px;
              font-size:13px; padding:2px 0;">
    <span style="min-width:110px; color:#333;">{m['label']}</span>
    <span style="min-width:70px; text-align:right; font-variant-numeric:tabular-nums;">{m['display']}</span>
    <span style="min-width:24px; text-align:center;
                 font-weight:700; color:{mark_color};
                 font-size:14px;">{mark}</span>
  </div>
"""
            )
        st.markdown(
            f"""
<div style="padding:12px 14px; margin-bottom:10px;
            border-radius:8px; border-left:5px solid {colour};
            background:rgba(128,128,128,0.05);">
  <div style="display:flex; align-items:center; gap:10px;
              flex-wrap:wrap;">
    <span style="display:inline-block; padding:2px 10px;
                 background:{colour}; color:#fff; border-radius:999px;
                 font-size:10px; font-weight:700; letter-spacing:0.04em;">
      {badge_label}
    </span>
    <span style="font-size:12px; color:#333;">
      <b>{t('alert_risk_score_label')} {score:.2f}</b>
    </span>
    <span style="font-size:13px; color:#222;">{a["label"]}</span>
    <span style="font-size:11px; color:#888; margin-left:auto;">
      📅 {date_badge}
    </span>
  </div>
  <div style="margin-top:8px;">
    {''.join(metric_rows)}
  </div>
</div>
            """,
            unsafe_allow_html=True,
        )


def _render_dora_panel(kpi_df: pd.DataFrame) -> None:
    """Render the DORA panel inside the Delivery tab.

    Each card shows the metric value (formatted), a tiny "completed N
    features over M days" caption, and a Good / Normal / Bad badge
    coloured by rating (vs the DORA 2024 industry bands, collapsed into
    3 tiers). Safe no-op when kpi_df is empty or missing schedule
    columns."""
    if kpi_df is None or kpi_df.empty:
        return
    defects_df = st.session_state.dfs.get("defects")
    dora = compute_dora_metrics(kpi_df, defects_df)

    st.subheader(t("dora_section_title"))
    st.caption(t("dora_section_caption",
                 days=dora["window_days"], n=dora["completed"]))

    rating_bg = {
        "good":    "#4ec78a",
        "normal":  "#f5b400",
        "bad":     "#f05050",
        "unknown": "#888888",
    }

    # (metric_key, title_key, help_key) — help_key drives the ? hover
    # tooltip on st.metric so each card carries its definition + source
    # + formula inline, matching the tooltip convention elsewhere.
    metric_specs = [
        ("lead_time",   "dora_lead_title",        "help_dora_lead_time"),
        ("cfr",         "dora_cfr_title",         "help_dora_cfr"),
        ("recovery",    "dora_recovery_title",    "help_dora_recovery"),
        ("reliability", "dora_reliability_title", "help_dora_reliability"),
    ]
    cols = st.columns(len(metric_specs), gap="small")
    for (key, title_key, help_key), col in zip(metric_specs, cols):
        entry = dora[key]
        v = entry["value"]
        rating = entry["rating"]
        unit = t(entry["unit_key"])
        # Value formatting: recovery time gets either hours (<1d) or days;
        # everything else is a compact 1-decimal number.
        if v is None:
            value_str = "—"
        elif key == "recovery" and v < 1.0:
            value_str = f"{v * 24:.1f} h"
        else:
            value_str = f"{v:.1f} {unit}"
        badge_color = rating_bg[rating]
        badge_label = t(f"dora_rating_{rating}")
        with col:
            # st.metric gives us the "?" help icon that matches the
            # rest of the dashboard's tooltip convention.
            st.metric(
                label=t(title_key),
                value=value_str,
                help=t(help_key, days=dora["window_days"]),
            )
            st.markdown(
                f"""
<div style="margin-top:-12px;">
  <span style="display:inline-block; padding:2px 10px;
               background:{badge_color}; color:#fff; border-radius:999px;
               font-size:11px; font-weight:700; letter-spacing:0.04em;">
    {badge_label}
  </span>
</div>
                """,
                unsafe_allow_html=True,
            )


def render_delivery_tab() -> None:
    """Dedicated 🏁 Delivery tab — just the DORA 5Keys panel.

    Split out from the Inputs tab so team delivery performance gets a
    top-level landing page instead of competing with the source-upload
    rail. Shows a small info banner when the upstream data isn't ready
    yet so the tab still renders in its empty state."""
    kpi_df = get_current_kpi_df()
    if kpi_df is None or kpi_df.empty:
        st.info(t("delivery_needs_data"))
        return
    _render_dora_panel(kpi_df)


def _backlog_due_bucket(d, today_d) -> str:
    """Classify a due date into one of the facet filter buckets (used by
    `render_backlog_tab`). Returns one of:
      - "overdue"     : d < today
      - "this_week"   : today ≤ d < today + 7
      - "this_month"  : today ≤ d within the current calendar month
      - "future"      : d beyond the current month
      - "none"        : no due date (None)
    A ticket can satisfy more than one bucket (e.g. a due date later
    this week also falls in this_month); the caller treats these as
    independently-queryable sets so 'This week' is a proper subset of
    'This month'."""
    if d is None:
        return "none"
    if d < today_d:
        return "overdue"
    if d < today_d + timedelta(days=7):
        return "this_week"
    if d.year == today_d.year and d.month == today_d.month:
        return "this_month"
    return "future"


def _apply_backlog_filters(
    df: pd.DataFrame,
    *,
    types: list[str],
    statuses: list[str],
    categories: list[str],
    phases: list[str],
    customers: list[str],
    due_bucket: str,
    today_d: date,
) -> pd.DataFrame:
    """Narrow the Backlog dataframe to rows matching the current facet
    selections. Each facet uses OR within itself (checked values) and AND
    across facets. Empty selections match everything (filter bypassed)."""
    out = df
    if types:
        out = out[out["種別"].isin(types)]
    if statuses:
        out = out[out["状態"].isin(statuses)]
    if categories:
        out = out[out["カテゴリ"].isin(categories)]
    if phases:
        out = out[out["発生フェーズ"].isin(phases)]
    if customers:
        out = out[out["顧客共有"].isin(customers)]
    if due_bucket and due_bucket != "all":
        if due_bucket == "none":
            out = out[out["期限日"].isna()]
        elif due_bucket == "overdue":
            out = out[out["期限日"].notna()
                      & (out["期限日"].map(
                          lambda d: d < today_d if d else False))]
        elif due_bucket == "this_week":
            out = out[out["期限日"].notna()
                      & (out["期限日"].map(
                          lambda d: (d is not None
                                     and today_d <= d <
                                     today_d + timedelta(days=7))))]
        elif due_bucket == "this_month":
            out = out[out["期限日"].notna()
                      & (out["期限日"].map(
                          lambda d: (d is not None and today_d <= d
                                     and d.year == today_d.year
                                     and d.month == today_d.month)))]
        elif due_bucket == "future":
            out = out[out["期限日"].notna()
                      & (out["期限日"].map(
                          lambda d: (d is not None
                                     and not (d.year == today_d.year
                                              and d.month == today_d.month)
                                     and d >= today_d)))]
    return out


_BACKLOG_STATUS_COLOR = {
    "未対応":   "#8a8a8a",
    "処理中":   "#f5b400",
    "処理済み": "#3c78d8",
    "完了":     "#4ec78a",
}


def _backlog_status_colour(status: str) -> str:
    """Pick a tile-border colour for a Backlog 状態. Unknown values get
    a muted grey — Backlog lets users add custom statuses so the map
    can't be exhaustive."""
    return _BACKLOG_STATUS_COLOR.get(status, "#bbbbbb")


def render_backlog_tab() -> None:
    """📋 Backlog tab — facet-filtered kanban view of Backlog.com issues.

    Phase 1: read-only display.
      - Top row: multiselect facets (種別 / 状態 / カテゴリ / 発生フェーズ /
        顧客共有) + a due-date preset selector + a Reset button.
      - Main area: one column per 種別 (after filtering). Within each
        column, cards are grouped under their 状態. Each card shows
        件名, 担当者 (falls back to "(未割当)"), and 期限日.

    Further phases layer the details dialog (Phase 2), edit/add (P3),
    CSV export (P3), Calendar layer (P4), and Gantt subsection (P5)."""
    df = st.session_state.dfs.get("backlog")
    if df is None or df.empty:
        st.info(t("backlog_needs_file"))
        return

    st.subheader(t("backlog_tab_title"))
    st.caption(t("backlog_tab_caption"))

    today_d = date.today()

    # ---- Facet bar ---------------------------------------------------------
    # Use sorted unique values from the current dataset so selection lists
    # match the uploaded CSV (Backlog users extend the option lists over
    # time; re-deriving every render keeps us current).
    def _opts(col: str) -> list[str]:
        return sorted([str(v) for v in df[col].dropna().unique() if str(v)])

    with st.container():
        f1, f2, f3, f4, f5, f6 = st.columns(6, gap="small")
        sel_type    = f1.multiselect(t("backlog_facet_type"),
                                     options=_opts("種別"),
                                     key="bl_facet_type")
        sel_status  = f2.multiselect(t("backlog_facet_status"),
                                     options=_opts("状態"),
                                     key="bl_facet_status")
        sel_cat     = f3.multiselect(t("backlog_facet_category"),
                                     options=_opts("カテゴリ"),
                                     key="bl_facet_category")
        sel_phase   = f4.multiselect(t("backlog_facet_phase"),
                                     options=_opts("発生フェーズ"),
                                     key="bl_facet_phase")
        sel_cust    = f5.multiselect(t("backlog_facet_customer"),
                                     options=_opts("顧客共有"),
                                     key="bl_facet_customer")
        due_options = [
            ("all",        t("backlog_due_option_all")),
            ("overdue",    t("backlog_due_option_overdue")),
            ("this_week",  t("backlog_due_option_this_week")),
            ("this_month", t("backlog_due_option_this_month")),
            ("future",     t("backlog_due_option_future")),
            ("none",       t("backlog_due_option_none")),
        ]
        due_keys = [k for k, _ in due_options]
        due_labels = dict(due_options)
        sel_due = f6.selectbox(
            t("backlog_facet_due"),
            options=due_keys,
            format_func=due_labels.__getitem__,
            key="bl_facet_due",
        )

    filtered = _apply_backlog_filters(
        df,
        types=sel_type, statuses=sel_status, categories=sel_cat,
        phases=sel_phase, customers=sel_cust,
        due_bucket=sel_due, today_d=today_d,
    )

    st.caption(t("backlog_status_total", n=f"{len(filtered):,}"))

    if filtered.empty:
        st.info(t("backlog_empty"))
        return

    # ---- Kanban board ------------------------------------------------------
    # Columns = 種別 (ordered by frequency desc so the fullest lane sits
    # on the left). Within each lane, cards are grouped by 状態 — using
    # the sorted unique statuses from the full dataset so lane vertical
    # positions stay stable as filters narrow results.
    type_order = (filtered["種別"]
                  .value_counts()
                  .index.tolist())
    status_order = _opts("状態")

    cols = st.columns(len(type_order), gap="small")
    for col, ttype in zip(cols, type_order):
        sub = filtered[filtered["種別"] == ttype]
        with col:
            st.markdown(
                f"<div style='font-size:13px;font-weight:700;"
                f"color:#222;padding:4px 0;'>"
                f"{ttype} "
                f"<span style='color:#888;font-weight:400;'>"
                f"({len(sub)})</span></div>",
                unsafe_allow_html=True,
            )
            for status in status_order:
                rows = sub[sub["状態"] == status]
                if rows.empty:
                    continue
                # Status header pill.
                colour = _backlog_status_colour(status)
                st.markdown(
                    f"<div style='font-size:11px;margin-top:6px;"
                    f"margin-bottom:2px;color:#555;'>"
                    f"<span style='display:inline-block;width:8px;"
                    f"height:8px;border-radius:50%;background:{colour};"
                    f"margin-right:6px;vertical-align:middle;'></span>"
                    f"{status} ({len(rows)})</div>",
                    unsafe_allow_html=True,
                )
                for _, r in rows.iterrows():
                    _render_backlog_card(r, today_d)


def _render_backlog_card(row, today_d: date) -> None:
    """Render one Backlog issue tile. Minimal Phase 1 layout — 件名,
    担当者, 期限日 — coloured by 状態."""
    title = str(row.get("件名") or "").strip() or "—"
    assignee = (str(row.get("担当者") or "").strip()
                or t("backlog_tile_assignee_none"))
    due = row.get("期限日")
    if due is None:
        due_text = t("backlog_tile_no_due")
        due_color = "#888"
    else:
        due_text = due.isoformat()
        # Overdue tiles get a red due-date; same-week gets amber; else grey.
        if due < today_d:
            due_color = "#f05050"
        elif due < today_d + timedelta(days=7):
            due_color = "#f5b400"
        else:
            due_color = "#666"
    colour = _backlog_status_colour(str(row.get("状態") or ""))
    st.markdown(
        f"""
<div style="padding:8px 10px; margin-bottom:6px;
            border-left:3px solid {colour};
            background:rgba(128,128,128,0.06);
            border-radius:4px;">
  <div style="font-size:13px; color:#222; line-height:1.35;
              word-break:break-word;">{title}</div>
  <div style="display:flex; justify-content:space-between;
              margin-top:4px; font-size:11px; color:#666;">
    <span>👤 {assignee}</span>
    <span style="color:{due_color};">
      {t('backlog_tile_due_label')} {due_text}
    </span>
  </div>
</div>
        """,
        unsafe_allow_html=True,
    )


def render_dashboard_tab() -> None:
    """Tab 1 — sources upload + the integrated tables."""
    st.subheader(t("sec1_title"))
    # Horizontal-scrolling card rail. One st.columns row with N fixed-width
    # columns (see the :has(.d4dx-source-card-marker) CSS rule above) —
    # users drag the rail left/right with a trackpad / shift-wheel to reach
    # the off-screen cards. A short hint line right above the rail tips
    # people off to the scroll affordance.
    st.markdown(
        f"<div class='d4dx-source-rail-hint'>{t('src_rail_hint')}</div>",
        unsafe_allow_html=True,
    )
    cols = st.columns(len(SOURCE_SPECS), gap="small")
    for spec, col in zip(SOURCE_SPECS, cols):
        with col:
            render_upload_card(spec)

    kpi_df = get_current_kpi_df()
    if kpi_df is None:
        st.info(t("master_unlock_info"))
        return
    summary = project_kpi_summary(kpi_df)

    st.subheader(t("sec3_title"))
    st.caption(t(
        "sec3_caption",
        n=f"{len(kpi_df):,}",
        u=f"{kpi_df['機能ID'].nunique():,}",
    ))

    # ----- Project-wide KPI strip ---------------------------------------------
    def _pct(v: Optional[float]) -> str:
        return f"{v * 100:.1f}%" if v is not None else "—"

    def _f3(v: Optional[float]) -> str:
        return f"{v:.3f}" if v is not None else "—"

    def _f2(v: Optional[float]) -> str:
        return f"{v:.2f}" if v is not None else "—"

    (m1, m2, m3, m4, m5,
     m6, m7, m8, m9) = st.columns(9, gap="small")
    m1.metric(t("metric_total_loc"),       f"{summary['total_loc']:,}",
              help=t("help_loc"))
    m2.metric(t("metric_open_defects"),    f"{summary['open_defects']:,}",
              help=t("help_defect_unresolved"))
    m3.metric(t("metric_test_run_rate"),   _pct(summary["run_rate"]),
              help=t("help_test_run_rate"))
    m4.metric(t("metric_test_pass_rate"),  _pct(summary["pass_rate"]),
              help=t("help_test_pass_rate"))
    m5.metric(t("metric_avg_bug_density"), _f3(summary["avg_bug_density"]),
              help=t("help_bug_density"))
    m6.metric(t("metric_avg_test_density"), _f2(summary["avg_test_density"]),
              help=t("help_test_density"))
    m7.metric(t("metric_avg_health"),      _f2(summary["avg_health"]),
              help=t("help_health_score"))
    m8.metric(t("metric_at_risk"),         f"{summary['at_risk_count']}",
              help=t("metric_help_at_risk"))
    m9.metric(t("metric_delayed"),         f"{summary['delayed_count']}",
              help=t("metric_help_delayed"))

    # ----- Tabbed integrated tables -------------------------------------------
    base_cols = ["機能ID", "機能名称"]
    overview_cols = base_cols + [
        c for c in [
            "defect_total", "defect_unresolved",
            "総設定テスト数", "実施済", "NG", "未実施",
            "LoC", "設計書ページ数",
            "actual_progress", "planned_progress",
            "risk_score", "health_score",
        ] if c in kpi_df.columns
    ]
    kpi_cols = base_cols + [
        c for c in [
            "bug_density", "test_density", "complexity",
            "test_run_rate", "test_pass_rate", "defect_rate",
            "delay_days", "delay_rate",
            "health_score", "risk_score",
        ] if c in kpi_df.columns
    ]
    wbs_cols = base_cols + [
        c for c in [
            "planned_effort", "actual_effort",
            "planned_start", "planned_end",
            "actual_start", "actual_end",
            "actual_progress", "planned_progress",
            "delay_days", "delay_rate",
        ] if c in kpi_df.columns
    ]
    defect_cols = base_cols + [
        c for c in ["defect_total", "defect_unresolved", "bug_density"]
        if c in kpi_df.columns
    ]
    test_cols = base_cols + [
        c for c in ["総設定テスト数", "実施済", "OK", "NG", "未実施",
                    "test_run_rate", "test_pass_rate", "defect_rate"]
        if c in kpi_df.columns
    ]
    code_cols = base_cols + [
        c for c in ["LoC", "設計書ページ数",
                    "complexity", "test_density", "bug_density"]
        if c in kpi_df.columns
    ]

    table_height = min(40 + 36 * len(kpi_df), 700)
    tabs = st.tabs([
        t("tab_overview"), t("tab_kpis"), t("tab_wbs"), t("tab_defects"),
        t("tab_tests"), t("tab_code"), t("tab_all"),
    ])
    with tabs[0]:
        df = kpi_df[overview_cols]
        ev = _selectable_table(df, "drill_overview",
                               column_config=build_col_config(overview_cols),
                               height=table_height)
        _capture_drilldown(ev, df)
    with tabs[1]:
        df = (kpi_df[kpi_cols].sort_values(
                "risk_score", ascending=False, na_position="last")
              if "risk_score" in kpi_df.columns else kpi_df[kpi_cols])
        ev = _selectable_table(df, "drill_kpis",
                               column_config=build_col_config(kpi_cols),
                               height=table_height)
        _capture_drilldown(ev, df)
    with tabs[2]:
        df = kpi_df[wbs_cols]
        ev = _selectable_table(df, "drill_wbs",
                               column_config=build_col_config(wbs_cols),
                               height=table_height)
        _capture_drilldown(ev, df)
    with tabs[3]:
        df = kpi_df[defect_cols]
        ev = _selectable_table(df, "drill_defects",
                               column_config=build_col_config(defect_cols),
                               height=table_height)
        _capture_drilldown(ev, df)
    with tabs[4]:
        df = kpi_df[test_cols]
        ev = _selectable_table(df, "drill_tests",
                               column_config=build_col_config(test_cols),
                               height=table_height)
        _capture_drilldown(ev, df)
    with tabs[5]:
        df = kpi_df[code_cols]
        ev = _selectable_table(df, "drill_code",
                               column_config=build_col_config(code_cols),
                               height=table_height)
        _capture_drilldown(ev, df)
    with tabs[6]:
        st.caption(t("tab_all_caption"))
        all_cols = list(kpi_df.columns)
        ev = _selectable_table(kpi_df, "drill_all",
                               column_config=build_col_config(all_cols),
                               height=table_height)
        _capture_drilldown(ev, kpi_df)

    # Drill-down lives below all the integrated tables so it can react to a
    # row click on any of them. If nothing is selected, surface the hint so
    # the feature is discoverable.
    if st.session_state.get("drilldown_id"):
        render_drilldown_panel(
            kpi_df,
            st.session_state.dfs.get("defects"),
            st.session_state.drilldown_id,
        )
    else:
        st.caption(t("drilldown_select_hint"))

    wbs_df = st.session_state.dfs.get("wbs")
    defects_df = st.session_state.dfs.get("defects")
    tests_df = st.session_state.dfs.get("tests")
    code_df = st.session_state.dfs.get("code")
    with st.expander(t("raw_previews")):
        if wbs_df is not None:
            st.markdown(f"**{t('wbs_label_short')}**")
            st.dataframe(wbs_df.head(10), use_container_width=True,
                         hide_index=True)
        if defects_df is not None:
            st.markdown(f"**{t('defects_label_short')}**")
            st.dataframe(defects_df.head(10), use_container_width=True,
                         hide_index=True)
        if tests_df is not None:
            st.markdown(f"**{t('tests_label_short')}**")
            st.dataframe(tests_df.head(10), use_container_width=True,
                         hide_index=True)
        if code_df is not None:
            st.markdown(f"**{t('code_label_short')}**")
            st.dataframe(code_df.head(10), use_container_width=True,
                         hide_index=True)


# =============================================================================
# Chart builders — return Plotly figures so both the UI and the PDF report
# can reuse the exact same chart definitions.
# =============================================================================
# Default outer margins for inline (dashboard) display. Generous on the left
# so horizontal bar / Gantt charts keep room for "AUTH001 · User Login"-style
# labels; `automargin=True` on the axes themselves lets Plotly grow even
# more if needed.
_INLINE_MARGIN_LONG_Y = dict(l=200, r=20, t=20, b=40)
_INLINE_MARGIN_DEFAULT = dict(l=60, r=20, t=20, b=40)
_INLINE_MARGIN_HEATMAP = dict(l=60, r=40, t=20, b=80)


# Max rows any per-Function-ID bar chart will display. Beyond this, both
# on-screen readability (431 crammed labels) and kaleido PDF rendering time
# collapse hard, so we show the worst N (sorted by the chart's native
# metric) and annotate the truncation. Management-report audiences care
# about the tail, not an unreadable all-hands scroll.
# PDF / matplotlib charts keep a hard cap on per-FID bars — reportlab's
# page area can't render 400 labels legibly and the build time blows up.
# The dashboard (Plotly) side runs uncapped: the sidebar's global 機能ID
# filter lets the user narrow scope when the full set is too large to
# read comfortably.
_BAR_CHART_MAX_ROWS = 30
# Big-enough ceiling that no realistic dashboard dataset hits it, while
# still keeping the browser from locking up on a pathological 10k+ row
# scroll. If this ever trips, the sidebar filter is the intended escape.
_INLINE_BAR_CHART_MAX_ROWS = 10_000
# Per-label max length — long 'ADM01010 · MBOM自動生成・更新（…）' strings
# force Chromium to run many glyph-metric queries when automargin retries,
# so clip them here. The drill-down panel still has the full name.
_BAR_LABEL_MAX_CHARS = 36


def _clip_label(s: str) -> str:
    return s if len(s) <= _BAR_LABEL_MAX_CHARS else s[: _BAR_LABEL_MAX_CHARS - 1] + "…"


def _truncate_note_annotation(shown: int, total: int) -> dict:
    """Small top-right annotation used when a per-feature bar chart has
    been trimmed to the worst N entries for legibility."""
    return dict(
        text=t("chart_truncated_note", shown=shown, total=total),
        xref="paper", yref="paper", x=1.0, y=1.02, xanchor="right",
        yanchor="bottom", showarrow=False,
        font=dict(size=10, color="#b48820"),
    )


def _chart_progress_gap(kpi_df: pd.DataFrame) -> Optional[go.Figure]:
    if not {"actual_progress", "planned_progress"}.issubset(kpi_df.columns):
        return None
    df = kpi_df.dropna(subset=["actual_progress", "planned_progress"]).copy()
    if df.empty:
        return None
    df["display"] = _feature_display_series(df).map(_clip_label)
    # Sort: most-behind first so head(N) keeps the worst offenders.
    df["_gap"] = df["planned_progress"] - df["actual_progress"]
    df = df.sort_values("_gap", ascending=False)
    total = len(df)
    if total > _INLINE_BAR_CHART_MAX_ROWS:
        df = df.head(_INLINE_BAR_CHART_MAX_ROWS)
    df = df.iloc[::-1]  # reverse so worst shows at the top of the bar chart
    over = df["actual_progress"] > df["planned_progress"]
    actual_colors = np.where(over, "#f5b400", "#4ec78a")
    actual_lines = np.where(over, "#a06a00", "#4ec78a")
    over_marker = t("chart_progress_over_marker")
    planned_vals = df["planned_progress"].astype(float)
    actual_vals = df["actual_progress"].astype(float)
    planned_text = [f"{v:.0f}%" for v in planned_vals]
    actual_text = [
        f"{v:.0f}% {over_marker}" if o else f"{v:.0f}%"
        for v, o in zip(actual_vals, over)
    ]
    fig = go.Figure()
    fig.add_bar(name=t("chart_progress_planned"),
                y=df["display"], x=planned_vals,
                orientation="h", marker_color="#9aa",
                text=planned_text, textposition="outside",
                textfont=dict(color="#6c6c6c", size=10),
                cliponaxis=False)
    fig.add_bar(name=t("chart_progress_actual"),
                y=df["display"], x=actual_vals,
                orientation="h",
                marker_color=actual_colors.tolist(),
                marker_line=dict(color=actual_lines.tolist(), width=1),
                text=actual_text, textposition="outside",
                textfont=dict(color="#a06a00", size=10),
                cliponaxis=False)
    fig.update_layout(barmode="group",
                      height=max(280, 28 * len(df)),
                      xaxis_title="%", yaxis_title=None,
                      margin=_INLINE_MARGIN_LONG_Y)
    fig.update_yaxes(automargin=True)
    if total > _INLINE_BAR_CHART_MAX_ROWS:
        fig.add_annotation(**_truncate_note_annotation(len(df), total))
    return fig


_OVERVIEW_COMPARE_METRICS: list[tuple[str, str, str]] = [
    # (df column, panel title, bar color)
    ("設計書ページ数", "設計書ページ数",     "#9aa0a6"),
    ("LoC",            "LoC",                "#7aaef0"),
    ("総設定テスト数",        "総設定テスト数",            "#4ec78a"),
    ("defect_total",   "障害件数（Redmine）", "#f05050"),
]


def _chart_overview_compare(kpi_df: pd.DataFrame) -> Optional[go.Figure]:
    """Small-multiples horizontal bars: 機能ID × {設計書ページ数, LoC,
    総設定テスト数, 障害件数（Redmine）}.

    Four side-by-side panels share a single Y axis (Function IDs) so each
    column's absolute magnitude stays readable on its own scale, while the
    eye still tracks each Function ID across all four metrics.
    """
    available = [(c, lbl, color)
                 for c, lbl, color in _OVERVIEW_COMPARE_METRICS
                 if c in kpi_df.columns]
    if not available:
        return None
    grp_cols = [c for c, _, _ in available]
    has_name = "機能名称" in kpi_df.columns
    agg_kw: dict[str, tuple] = {c: (c, "mean") for c in grp_cols}
    if has_name:
        agg_kw["機能名称"] = ("機能名称", "first")
    df = kpi_df.groupby("機能ID", as_index=False).agg(**agg_kw)
    df = df.dropna(subset=grp_cols, how="all")
    if df.empty:
        return None
    df = df.sort_values("機能ID", ascending=True)
    # y-axis labels: 機能ID：機能名 (from master-joined kpi_df), clipped
    # to keep long names from pushing the bars off the left margin.
    fids = _feature_display_series(df).map(_clip_label).tolist()
    n_panels = len(available)
    titles = [lbl for _, lbl, _ in available]
    fig = make_subplots(rows=1, cols=n_panels,
                        shared_yaxes=True,
                        horizontal_spacing=0.04,
                        subplot_titles=titles)
    for i, (col, lbl, color) in enumerate(available, start=1):
        raw = pd.to_numeric(df[col], errors="coerce").astype(float)
        vals = raw.tolist()
        bar_text = [
            "" if pd.isna(v) else (f"{int(v):,}" if col == "LoC"
                                   else f"{int(v)}")
            for v in vals
        ]
        fig.add_trace(
            go.Bar(
                y=fids, x=vals, orientation="h",
                marker_color=color, showlegend=False,
                text=bar_text, textposition="outside",
                textfont=dict(color="#555555", size=10),
                cliponaxis=False,
                hovertemplate=(
                    f"<b>%{{y}}</b><br>{lbl}: "
                    + ("%{x:,.0f}" if col != "LoC" else "%{x:,}")
                    + "<extra></extra>"
                ),
            ),
            row=1, col=i,
        )
    fig.update_layout(
        height=max(320, 24 * len(fids) + 100),
        margin=dict(l=140, r=20, t=60, b=40),
        bargap=0.2,
    )
    # Reverse on the (shared) y-axis so the alphabetically-first Function ID
    # sits at the top — matches every other per-FID chart in this file.
    fig.update_yaxes(autorange="reversed", automargin=True)
    fig.update_xaxes(automargin=True, rangemode="tozero")
    return fig


TEST_DENSITY_THRESHOLD_DEFAULT = 10.0
INCIDENT_RATE_THRESHOLD_DEFAULT = 0.05  # = 5%


def _read_threshold(state_key: str, default: float) -> float:
    """Common helper for reading user-configured chart thresholds out of
    Streamlit session state. Falls back to the default when called outside
    of a Streamlit session (tests, batch usage)."""
    try:
        v = st.session_state.get(state_key, default)
    except Exception:
        v = default
    try:
        return float(v) if v is not None else default
    except (TypeError, ValueError):
        return default


def _test_density_threshold() -> float:
    return _read_threshold("test_density_threshold",
                           TEST_DENSITY_THRESHOLD_DEFAULT)


def _incident_rate_threshold() -> float:
    return _read_threshold("incident_rate_threshold",
                           INCIDENT_RATE_THRESHOLD_DEFAULT)


def _chart_test_density(kpi_df: pd.DataFrame) -> Optional[go.Figure]:
    if not {"test_density", "総設定テスト数", "設計書ページ数"}.issubset(kpi_df.columns):
        return None
    df = kpi_df.dropna(subset=["test_density"]).copy()
    if df.empty:
        return None
    df["display"] = _feature_display_series(df).map(_clip_label)
    # Sort ascending so the lowest (under-tested) sit at the top of the bar
    # chart after the iloc reverse below — matches the convention used by
    # the other "attention list" charts in this file.
    df = df.sort_values("test_density", ascending=True)
    total = len(df)
    if total > _INLINE_BAR_CHART_MAX_ROWS:
        df = df.head(_INLINE_BAR_CHART_MAX_ROWS)
    df = df.iloc[::-1]
    threshold = _test_density_threshold()
    densities = df["test_density"].astype(float)
    below = densities < threshold
    bar_colors = np.where(below, "#f05050", "#7aaef0")
    bar_lines = np.where(below, "#a02020", "#7aaef0")
    customdata = np.column_stack([
        df["総設定テスト数"].fillna(0).astype(int),
        df["設計書ページ数"].fillna(0).astype(float),
        densities,
    ])
    below_marker = t("chart_test_density_below_marker")
    hover_tmpl = (
        "<b>%{y}</b><br>"
        "総設定テスト数: %{customdata[0]}  "
        "設計書ページ数: %{customdata[1]:.0f}<br>"
        f"{t('col_test_density')}: %{{customdata[2]:.2f}}"
        f" (閾値: {threshold:g})"
        "<extra></extra>"
    )
    bar_texts = [
        f"{v:.2f} {below_marker}" if b else f"{v:.2f}"
        for v, b in zip(densities, below)
    ]
    text_colors = ["#a02020" if b else "#555555" for b in below]
    fig = go.Figure()
    fig.add_bar(
        y=df["display"], x=densities,
        orientation="h",
        marker_color=bar_colors.tolist(),
        marker_line=dict(color=bar_lines.tolist(), width=1),
        customdata=customdata, hovertemplate=hover_tmpl,
        text=bar_texts,
        textposition="outside",
        textfont=dict(color=text_colors, size=10),
        cliponaxis=False,
    )
    fig.add_vline(
        x=threshold, line_width=1, line_dash="dash", line_color="#a02020",
        annotation_text=f"{t('chart_test_density_threshold_label')} {threshold:g}",
        annotation_position="top right",
        annotation_font=dict(color="#a02020", size=11),
    )
    fig.update_layout(height=max(280, 28 * len(df)),
                      xaxis_title="tests / page", yaxis_title=None,
                      margin=_INLINE_MARGIN_LONG_Y)
    fig.update_yaxes(automargin=True)
    if total > _INLINE_BAR_CHART_MAX_ROWS:
        fig.add_annotation(**_truncate_note_annotation(len(df), total))
    return fig


def _chart_incident_rate(kpi_df: pd.DataFrame) -> Optional[go.Figure]:
    """Per-機能ID horizontal bar of Redmine fault rate (defect_total ÷
    実施済). Bars *above* the configured threshold are flagged in red — the
    opposite direction from the test_density chart since here higher = bad.
    """
    needed = {"incident_rate", "defect_total", "実施済"}
    if not needed.issubset(kpi_df.columns):
        return None
    df = kpi_df.dropna(subset=["incident_rate"]).copy()
    if df.empty:
        return None
    df["display"] = _feature_display_series(df).map(_clip_label)
    df = df.sort_values("incident_rate", ascending=False)
    total = len(df)
    if total > _INLINE_BAR_CHART_MAX_ROWS:
        df = df.head(_INLINE_BAR_CHART_MAX_ROWS)
    df = df.iloc[::-1]
    threshold = _incident_rate_threshold()
    rates = df["incident_rate"].astype(float)
    above = rates > threshold
    bar_colors = np.where(above, "#f05050", "#7aaef0")
    bar_lines = np.where(above, "#a02020", "#7aaef0")
    customdata = np.column_stack([
        df["defect_total"].fillna(0).astype(int),
        df["実施済"].fillna(0).astype(int),
        rates * 100.0,
    ])
    above_marker = t("chart_incident_rate_above_marker")
    hover_tmpl = (
        "<b>%{y}</b><br>"
        f"{t('col_defect_total')}: %{{customdata[0]}}  "
        f"実施済: %{{customdata[1]}}<br>"
        f"{t('col_incident_rate')}: %{{customdata[2]:.2f}}%"
        f" (閾値: {threshold * 100:g}%)"
        "<extra></extra>"
    )
    pct_vals = rates * 100.0
    bar_texts = [
        f"{v:.1f}% {above_marker}" if a else f"{v:.1f}%"
        for v, a in zip(pct_vals, above)
    ]
    text_colors = ["#a02020" if a else "#555555" for a in above]
    fig = go.Figure()
    fig.add_bar(
        y=df["display"], x=pct_vals,
        orientation="h",
        marker_color=bar_colors.tolist(),
        marker_line=dict(color=bar_lines.tolist(), width=1),
        customdata=customdata, hovertemplate=hover_tmpl,
        text=bar_texts,
        textposition="outside",
        textfont=dict(color=text_colors, size=10),
        cliponaxis=False,
    )
    fig.add_vline(
        x=threshold * 100.0,
        line_width=1, line_dash="dash", line_color="#a02020",
        annotation_text=(
            f"{t('chart_incident_rate_threshold_label')} "
            f"{threshold * 100:g}%"
        ),
        annotation_position="top right",
        annotation_font=dict(color="#a02020", size=11),
    )
    fig.update_layout(height=max(280, 28 * len(df)),
                      xaxis_title="%", yaxis_title=None,
                      margin=_INLINE_MARGIN_LONG_Y)
    fig.update_yaxes(automargin=True)
    if total > _INLINE_BAR_CHART_MAX_ROWS:
        fig.add_annotation(**_truncate_note_annotation(len(df), total))
    return fig


def _chart_test_coverage(kpi_df: pd.DataFrame) -> Optional[go.Figure]:
    if not {"OK", "NG", "未実施"}.issubset(kpi_df.columns):
        return None
    df = kpi_df.dropna(subset=["OK", "NG", "未実施"], how="all").copy()
    if df.empty:
        return None
    df["display"] = _feature_display_series(df).map(_clip_label)
    # Worst-first by NG then 未実施 so head(N) is the attention list.
    df["_bad"] = df["NG"].fillna(0) + df["未実施"].fillna(0) * 0.5
    df = df.sort_values("_bad", ascending=False)
    total = len(df)
    if total > _INLINE_BAR_CHART_MAX_ROWS:
        df = df.head(_INLINE_BAR_CHART_MAX_ROWS)
    df = df.iloc[::-1]
    ok_vals = df["OK"].fillna(0).astype(int)
    ng_vals = df["NG"].fillna(0).astype(int)
    nr_vals = df["未実施"].fillna(0).astype(int)
    total_vals = ok_vals + ng_vals + nr_vals
    with np.errstate(divide="ignore", invalid="ignore"):
        cov_pct = np.where(total_vals > 0,
                           ok_vals / total_vals * 100, 0.0)
    customdata = np.column_stack([ok_vals, ng_vals, nr_vals,
                                  total_vals, cov_pct])
    hover_tmpl = (
        "<b>%{y}</b><br>"
        f"{t('chart_label_ok')}: %{{customdata[0]}}  "
        f"{t('chart_label_ng')}: %{{customdata[1]}}  "
        f"{t('chart_label_notrun')}: %{{customdata[2]}}<br>"
        f"{t('chart_label_total')}: %{{customdata[3]}}  "
        f"{t('chart_label_coverage')}: %{{customdata[4]:.1f}}%"
        "<extra></extra>"
    )
    def _seg_text(values):
        # Blank the label when the segment is 0 so a bunch of "0" labels
        # doesn't litter the stack.
        return [(str(int(v)) if v else "") for v in values]

    fig = go.Figure()
    fig.add_bar(name=t("chart_label_ok"),
                y=df["display"], x=ok_vals,
                orientation="h", marker_color="#4ec78a",
                text=_seg_text(ok_vals), textposition="inside",
                insidetextanchor="middle",
                textfont=dict(color="#0f3d22", size=10),
                customdata=customdata, hovertemplate=hover_tmpl)
    fig.add_bar(name=t("chart_label_ng"),
                y=df["display"], x=ng_vals,
                orientation="h", marker_color="#f05050",
                text=_seg_text(ng_vals), textposition="inside",
                insidetextanchor="middle",
                textfont=dict(color="#3d0b0b", size=10),
                customdata=customdata, hovertemplate=hover_tmpl)
    fig.add_bar(name=t("chart_label_notrun"),
                y=df["display"], x=nr_vals,
                orientation="h", marker_color="#bbbbbb",
                text=_seg_text(nr_vals), textposition="inside",
                insidetextanchor="middle",
                textfont=dict(color="#333333", size=10),
                customdata=customdata, hovertemplate=hover_tmpl)
    fig.update_layout(barmode="stack",
                      height=max(280, 28 * len(df)),
                      margin=_INLINE_MARGIN_LONG_Y)
    fig.update_yaxes(automargin=True)
    if total > _INLINE_BAR_CHART_MAX_ROWS:
        fig.add_annotation(**_truncate_note_annotation(len(df), total))
    return fig


def _chart_loc_vs_ng(kpi_df: pd.DataFrame) -> Optional[go.Figure]:
    if not {"LoC", "NG"}.issubset(kpi_df.columns):
        return None
    df = kpi_df.dropna(subset=["LoC", "NG"]).copy()
    if df.empty:
        return None
    size_col = ("設計書ページ数"
                if "設計書ページ数" in df.columns
                and df["設計書ページ数"].notna().any()
                else None)
    if size_col is not None:
        df[size_col] = pd.to_numeric(df[size_col], errors="coerce").fillna(5)
    color_col = "risk_score" if "risk_score" in df.columns else None
    df["_label"] = _feature_display_series(df)
    fig = px.scatter(
        df, x="LoC", y="NG",
        size=size_col, color=color_col,
        hover_name="_label",
        color_continuous_scale="RdYlGn_r",
    )
    fig.update_layout(height=420, margin=_INLINE_MARGIN_DEFAULT)
    fig.update_xaxes(automargin=True)
    fig.update_yaxes(automargin=True)
    return fig


def _chart_design_impl_gap(kpi_df: pd.DataFrame) -> Optional[go.Figure]:
    if not {"設計書ページ数", "LoC"}.issubset(kpi_df.columns):
        return None
    df = kpi_df.dropna(subset=["設計書ページ数", "LoC"]).copy()
    if df.empty:
        return None
    df["_label"] = _feature_display_series(df)
    fig = px.scatter(df, x="設計書ページ数", y="LoC", hover_name="_label")
    comp = pd.to_numeric(df.get("complexity"), errors="coerce").dropna()
    if len(comp):
        avg = float(comp.mean())
        xs = np.linspace(df["設計書ページ数"].min(),
                         df["設計書ページ数"].max(), 50)
        fig.add_scatter(x=xs, y=avg * xs, mode="lines",
                        line=dict(dash="dash", color="#888"),
                        name=f"avg complexity = {avg:.1f}")
    fig.update_layout(height=420, margin=_INLINE_MARGIN_DEFAULT)
    fig.update_xaxes(automargin=True)
    fig.update_yaxes(automargin=True)
    return fig


_RISK_HEATMAP_INVERTED_DIMS = ("test_run_rate", "test_density")


def _chart_risk_heatmap(kpi_df: pd.DataFrame) -> Optional[go.Figure]:
    risk_dims = [c for c in
                 ["bug_density", "incident_rate", "delay_rate",
                  "test_run_rate", "test_density"]
                 if c in kpi_df.columns]
    if not risk_dims:
        return None
    agg = kpi_df.groupby("機能ID")[risk_dims].mean(numeric_only=True)
    z_df = agg.copy()
    for c in risk_dims:
        s = z_df[c]
        m = s.max(skipna=True)
        if pd.notna(m) and m > 0:
            z_df[c] = s / m
        if c in _RISK_HEATMAP_INVERTED_DIMS:
            mask = z_df[c].notna()
            z_df.loc[mask, c] = 1 - z_df.loc[mask, c]
    z_df = z_df.sort_values(by=risk_dims[0], ascending=False,
                            na_position="last")
    dim_label = {c: t(COLUMN_LABEL_KEYS.get(c, c)) for c in risk_dims}
    if "test_run_rate" in dim_label:
        dim_label["test_run_rate"] = (
            f"{dim_label['test_run_rate']} ({t('chart_label_notrun')})"
        )
    if "test_density" in dim_label:
        dim_label["test_density"] = (
            f"{dim_label['test_density']} ({t('chart_label_low')})"
        )
    y_labels = [dim_label[c] for c in risk_dims]
    # Look up authoritative 機能名 from the master so the heatmap x-axis
    # shows 機能ID：機能名 rather than bare IDs. Fall back to the ID when
    # no master entry exists.
    name_map = _master_fid_name_map()
    if not name_map and "機能名称" in kpi_df.columns:
        dedup = kpi_df.drop_duplicates(subset=["機能ID"])
        name_map = {str(f): ("" if pd.isna(n) else str(n))
                    for f, n in zip(dedup["機能ID"], dedup["機能名称"])}
    x_labels = [_label_fid_name(f, name_map.get(str(f), "")) for f in z_df.index]
    fig = px.imshow(
        z_df.T.values, x=x_labels, y=y_labels,
        color_continuous_scale="RdYlGn_r", aspect="auto",
        labels=dict(x="機能ID：機能名", y="", color="risk"),
        zmin=0, zmax=1,
    )
    fig.update_traces(hoverongaps=False)
    fig.update_layout(height=320, margin=_INLINE_MARGIN_HEATMAP,
                      plot_bgcolor="#d0d0d0")
    fig.update_xaxes(automargin=True, tickangle=-30)
    fig.update_yaxes(automargin=True)
    return fig


# =============================================================================
# Role analytics — cross-reference WBS sub-task assignees × Redmine defects
# =============================================================================
# Map each WBS sub-task `task_label` to a role. Matching is NFKC-normalised
# substring logic — variants are tolerated — but structured so each role
# carries a list of *include* keywords (any hit → that role) and a list of
# *exclude* keywords (any hit → SKIP that role even when an include term
# also matches). The exclude list solves the "再テスト実施" problem: it
# contains "テスト実施" as a substring, which would otherwise pollute the
# first-pass test_exec tally with retest work.
#
# Multiple include keywords in one label attribute to all matching roles
# (e.g. "開発兼テスト実施" → dev + test_exec).
ROLE_KEYWORDS: dict[str, dict[str, list[str]]] = {
    "dev":       {"include": ["開発", "実装"],     "exclude": []},
    "test_spec": {"include": ["テスト仕様書作成"],  "exclude": []},
    # test_exec counts FIRST-PASS test execution only. 再テスト実施 (retest
    # after a fix) is a distinct activity and is removed from this bucket
    # even though it lexically contains "テスト実施".
    "test_exec": {"include": ["テスト実施"],        "exclude": ["再テスト実施"]},
}


def _normalize_assignee(raw) -> str:
    """Canonicalize an assignee name so cosmetic input drift doesn't split
    one person into two groups.

    NFKC folds full-width letters/digits/punctuation to ASCII; the ideographic
    space U+3000 is NOT touched by NFKC so it's replaced explicitly. Any run
    of whitespace (ASCII spaces, tabs, ideographic spaces left over from
    mixed-encoding pastes, etc.) collapses to a single ASCII space, and the
    result is trimmed. Returns '' for None / NaN / empty — the caller decides
    how to surface the unassigned case."""
    if raw is None:
        return ""
    try:
        if pd.isna(raw):
            return ""
    except (TypeError, ValueError):
        pass
    s = unicodedata.normalize("NFKC", str(raw))
    s = s.replace("　", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _subtask_role_for_label(label: str) -> str:
    """Return the analytics role ('dev' / 'test_spec' / 'test_exec') that
    matches `label`, or '' when the label doesn't match any role keyword.

    Same matching rule as `_extract_role_assignments` (NFKC normalise →
    exclude-keywords block the role → include-keywords claim it). Used by
    the per-feature drill-down to label each WBS sub-task row with its
    role without re-flattening the whole WBS."""
    label_n = unicodedata.normalize("NFKC", str(label or ""))
    for role, cfg in ROLE_KEYWORDS.items():
        if any(unicodedata.normalize("NFKC", e) in label_n
               for e in cfg.get("exclude", [])):
            continue
        if any(unicodedata.normalize("NFKC", i) in label_n
               for i in cfg.get("include", [])):
            return role
    return ""


def _subtask_progress_bucket(row) -> str:
    """Classify a WBS sub-task as one of 'completed' / 'in_progress' /
    'not_started', which drives the role-progress mini-bars in the
    drill-down.

    `actual_progress` is already normalised by `_to_percent_scale` to a
    0..100 scale (fractional 0..1 values coming out of Excel's percent
    format are multiplied to match), so the comparison uses `>= 100`
    rather than `>= 1.0`."""
    ae = row.get("actual_end")
    prog = row.get("actual_progress")
    if pd.notna(ae) or (pd.notna(prog) and float(prog) >= 100.0):
        return "completed"
    if pd.notna(prog) and float(prog) > 0:
        return "in_progress"
    return "not_started"


def _subtasks_for_function(
    wbs_df: Optional[pd.DataFrame], function_id: str,
) -> pd.DataFrame:
    """Return the WBS sub-task rows for a single Function ID, with a
    derived `role` column. Empty frame when WBS isn't loaded or the ID
    has no sub-tasks."""
    cols = ["機能ID", "task_label", "assignee",
            "planned_start", "planned_end",
            "actual_start", "actual_end",
            "planned_progress", "actual_progress",
            "planned_effort", "actual_effort",
            "role"]
    if wbs_df is None or wbs_df.empty or "is_subtask" not in wbs_df.columns:
        return pd.DataFrame(columns=cols)
    df = wbs_df[
        (wbs_df["機能ID"].astype(str) == str(function_id))
        & (wbs_df["is_subtask"].fillna(False).astype(bool))
    ].copy()
    if df.empty:
        return pd.DataFrame(columns=cols)
    df["role"] = df["task_label"].map(_subtask_role_for_label)
    keep = [c for c in cols if c != "role" and c in df.columns]
    return df[keep + ["role"]].reset_index(drop=True)


def _extract_role_assignments(wbs_df: Optional[pd.DataFrame]) -> pd.DataFrame:
    """Flatten sub-task rows into long-form (機能ID, role, assignee) records.

    Matching rule per role:
      1. NFKC-normalise the label.
      2. If any *exclude* keyword appears in the label → skip this role
         for this sub-task entirely (e.g. "再テスト実施" locks test_exec
         out even though it contains "テスト実施").
      3. If any *include* keyword appears → attribute this sub-task to
         the role.

    A single sub-task can match multiple roles (e.g. "開発兼テスト実施"
    → dev + test_exec). Blank-assignee cells surface as `(未割当)`.
    Parent rows and sub-tasks whose label matches no include keyword
    are skipped entirely.
    """
    cols = ["機能ID", "role", "assignee"]
    if wbs_df is None or wbs_df.empty:
        return pd.DataFrame(columns=cols)
    if "is_subtask" not in wbs_df.columns:
        return pd.DataFrame(columns=cols)
    subs = wbs_df[wbs_df["is_subtask"].fillna(False).astype(bool)]
    if subs.empty or "task_label" not in subs.columns:
        return pd.DataFrame(columns=cols)
    unassigned = t("role_unassigned")
    # Pre-normalise include/exclude keywords once so the per-row loop is
    # plain substring tests.
    include_norm = {
        role: [unicodedata.normalize("NFKC", kw) for kw in cfg["include"]]
        for role, cfg in ROLE_KEYWORDS.items()
    }
    exclude_norm = {
        role: [unicodedata.normalize("NFKC", kw) for kw in cfg["exclude"]]
        for role, cfg in ROLE_KEYWORDS.items()
    }

    rows: list[dict] = []
    for _, r in subs.iterrows():
        fid = str(r.get("機能ID") or "")
        label = str(r.get("task_label") or "")
        if not fid or not label:
            continue
        label_n = unicodedata.normalize("NFKC", label)
        assignee = _normalize_assignee(r.get("assignee")) or unassigned
        for role in ROLE_KEYWORDS:
            if any(ex in label_n for ex in exclude_norm[role]):
                continue
            if any(inc in label_n for inc in include_norm[role]):
                rows.append({"機能ID": fid, "role": role,
                             "assignee": assignee})
    return pd.DataFrame(rows, columns=cols)


def _top3_problem_classes_for(
    defects_df: Optional[pd.DataFrame], fids: list[str],
) -> str:
    """Compact " / "-joined "category:count" string of the top-3 Redmine
    問題分類 on the given features. Returns '—' when there are no
    matching defects or the column isn't available."""
    if (defects_df is None or defects_df.empty
            or "問題分類" not in defects_df.columns
            or "機能ID" not in defects_df.columns):
        return "—"
    d = defects_df[defects_df["機能ID"].astype(str).isin([str(f) for f in fids])]
    if d.empty:
        return "—"
    uncat = t("problem_class_uncategorized")
    vc = (d["問題分類"].fillna("").astype(str).replace("", uncat)
          .value_counts())
    if vc.empty:
        return "—"
    return " / ".join(f"{cat}:{n}" for cat, n in vc.head(3).items())


def _build_feature_role_table(
    role_df: pd.DataFrame, kpi_df: pd.DataFrame,
    defects_df: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    """Per-機能ID pivot: {dev / test_spec / test_exec} assignee lists joined
    with the feature's quality KPIs + top-3 問題分類.

    Assignees within the same (FID, role) are de-duplicated and joined with
    ' / '. The `defects_df` slice lets us compute the per-feature problem
    classification top-3; pass None to skip that column."""
    if role_df.empty:
        return pd.DataFrame()
    pivot = (role_df.groupby(["機能ID", "role"])["assignee"]
             .apply(lambda s: " / ".join(sorted(set(s))))
             .unstack("role"))
    for rk in ROLE_KEYWORDS:
        if rk not in pivot.columns:
            pivot[rk] = ""
    pivot = pivot[list(ROLE_KEYWORDS.keys())].fillna("").reset_index()
    kpi_cols = ["機能ID"]
    for c in ("機能名称", "test_run_rate", "test_density",
              "defect_total", "incident_rate", "NG"):
        if c in kpi_df.columns:
            kpi_cols.append(c)
    kpi_subset = kpi_df[kpi_cols].drop_duplicates(subset=["機能ID"])
    out = pivot.merge(kpi_subset, on="機能ID", how="left")
    # Build FID：Name label column for display (authoritative name lives on
    # the master-joined kpi_df).
    out["display"] = out.apply(_label_id_name, axis=1)
    # Per-feature top-3 問題分類 (Redmine) — single-feature variant of the
    # assignee-level rollup used in View 2.
    if defects_df is not None and not defects_df.empty:
        out["top3_problems"] = out["機能ID"].map(
            lambda fid: _top3_problem_classes_for(defects_df, [str(fid)])
        )
    else:
        out["top3_problems"] = "—"
    # Worst-first so reviewers see the problem features at the top.
    if "incident_rate" in out.columns:
        out = out.sort_values("incident_rate",
                              ascending=False, na_position="last")
    return out.reset_index(drop=True)


def _build_assignee_summary(
    role_df: pd.DataFrame, kpi_df: pd.DataFrame,
    defects_df: Optional[pd.DataFrame],
) -> pd.DataFrame:
    """Per-assignee rollup: role participation counts, feature breadth,
    defects on all features they touched, average incident rate on those
    features, and top-3 Redmine 問題分類 on those features."""
    if role_df.empty:
        return pd.DataFrame()
    role_counts = (role_df.groupby(["assignee", "role"]).size()
                   .unstack("role").fillna(0).astype(int))
    for rk in ROLE_KEYWORDS:
        if rk not in role_counts.columns:
            role_counts[rk] = 0
    role_counts = role_counts[list(ROLE_KEYWORDS.keys())]

    features_per = (role_df.groupby("assignee")["機能ID"]
                    .apply(lambda s: sorted(set(str(x) for x in s))))
    feature_counts = features_per.map(len)

    if (defects_df is not None and not defects_df.empty
            and "機能ID" in defects_df.columns):
        dfd = defects_df.copy()
        dfd["機能ID"] = dfd["機能ID"].astype(str)

        def _defect_total(fids):
            return int(dfd[dfd["機能ID"].isin(fids)].shape[0])
        defect_totals = features_per.map(_defect_total)
    else:
        dfd = None
        defect_totals = pd.Series([0] * len(features_per),
                                   index=features_per.index)

    if "incident_rate" in kpi_df.columns:
        kpi_sub = (kpi_df[["機能ID", "incident_rate"]]
                   .drop_duplicates(subset=["機能ID"])
                   .assign(機能ID=lambda d: d["機能ID"].astype(str))
                   .set_index("機能ID"))

        def _avg_rate(fids):
            vals = pd.to_numeric(
                kpi_sub.reindex(fids)["incident_rate"], errors="coerce"
            ).dropna()
            return float(vals.mean()) if len(vals) else float("nan")
        avg_rates = features_per.map(_avg_rate)
    else:
        avg_rates = pd.Series([float("nan")] * len(features_per),
                              index=features_per.index)

    if dfd is not None and "問題分類" in dfd.columns:
        uncat = t("problem_class_uncategorized")

        def _top3(fids):
            d = dfd[dfd["機能ID"].isin(fids)]
            if d.empty:
                return "—"
            vc = (d["問題分類"].fillna("").astype(str).replace("", uncat)
                  .value_counts())
            return " / ".join(f"{cat}:{n}"
                              for cat, n in vc.head(3).items())
        top3 = features_per.map(_top3)
    else:
        top3 = pd.Series(["—"] * len(features_per),
                         index=features_per.index)

    out = role_counts.copy()
    out["feature_count"] = feature_counts
    out["defect_total"] = defect_totals
    out["avg_incident_rate"] = avg_rates
    out["top3_problems"] = top3
    out = out.reset_index().rename(columns={"assignee": "担当者"})
    return out.sort_values("defect_total", ascending=False,
                            kind="stable").reset_index(drop=True)


def _build_assignee_problem_crosstab(
    role_df: pd.DataFrame,
    defects_df: Optional[pd.DataFrame],
    role_filter: Optional[str] = None,
) -> Optional[pd.DataFrame]:
    """Assignee × 問題分類 counts for features the assignee touched (via any
    role, or a specific role when `role_filter` is supplied). Returns None
    when there's nothing to plot."""
    if role_df.empty:
        return None
    if defects_df is None or defects_df.empty:
        return None
    if "問題分類" not in defects_df.columns or "機能ID" not in defects_df.columns:
        return None
    rd = role_df
    if role_filter and role_filter != "all":
        rd = rd[rd["role"] == role_filter]
    if rd.empty:
        return None
    pairs = (rd[["assignee", "機能ID"]]
             .assign(機能ID=lambda d: d["機能ID"].astype(str))
             .drop_duplicates())
    dfd = (defects_df[["機能ID", "問題分類"]].copy()
           .assign(機能ID=lambda d: d["機能ID"].astype(str)))
    uncat = t("problem_class_uncategorized")
    dfd["問題分類"] = (dfd["問題分類"].fillna("").astype(str)
                       .replace("", uncat))
    merged = pairs.merge(dfd, on="機能ID")
    if merged.empty:
        return None
    ct = (merged.groupby(["assignee", "問題分類"]).size()
          .unstack("問題分類").fillna(0).astype(int))
    # Columns ordered by overall frequency (most common class first).
    col_totals = ct.sum(axis=0).sort_values(ascending=False)
    ct = ct[col_totals.index]
    # Rows ordered by total defect count so the heaviest contributors sit
    # at the top of the heatmap.
    row_totals = ct.sum(axis=1).sort_values(ascending=False)
    return ct.loc[row_totals.index]


# ----- Advanced per-assignee viz (案A bubble + 案C problem-share strip) -----
# Plotly colours for the dominant-role categorical axis, kept in sync with
# the matplotlib fallbacks below so the PDF version matches the screen.
_ROLE_COLOR_MAP: dict[str, str] = {
    "dev":       "#4a90e2",   # blue
    "test_spec": "#4ec78a",   # green
    "test_exec": "#f5a623",   # orange
}


def _build_assignee_bubble_df(
    role_df: pd.DataFrame,
    kpi_df: pd.DataFrame,
    defects_df: Optional[pd.DataFrame],
) -> pd.DataFrame:
    """Per-assignee coordinates for the bubble map.

    Each row is one assignee with:
      - feature_count        : X-axis (breadth)
      - avg_incident_rate    : Y-axis (quality signal; 0..1)
      - defect_total         : bubble size (exposure)
      - dominant_role        : bubble colour (tie → dev > test_spec > test_exec)
      - role_count_{dev,...} : raw role counts, surfaced in the hover card
    """
    cols = ["assignee", "feature_count", "avg_incident_rate",
            "defect_total", "dominant_role"]
    if role_df.empty:
        return pd.DataFrame(columns=cols)

    counts = (role_df.groupby(["assignee", "role"]).size()
              .unstack("role").fillna(0).astype(int))
    for rk in ROLE_KEYWORDS:
        if rk not in counts.columns:
            counts[rk] = 0
    counts = counts[list(ROLE_KEYWORDS.keys())]

    # Dominant role with a stable tie-break: roles listed first in
    # ROLE_KEYWORDS win ties so dev > test_spec > test_exec.
    role_order = list(ROLE_KEYWORDS.keys())

    def _dominant(row):
        best = None
        best_n = -1
        for r in role_order:
            if int(row[r]) > best_n:
                best_n = int(row[r])
                best = r
        return best if best_n > 0 else role_order[0]
    dominant = counts.apply(_dominant, axis=1)

    features_per = (role_df.groupby("assignee")["機能ID"]
                    .apply(lambda s: sorted(set(str(x) for x in s))))
    feature_counts = features_per.map(len)

    if (defects_df is not None and not defects_df.empty
            and "機能ID" in defects_df.columns):
        dfd = defects_df.copy()
        dfd["機能ID"] = dfd["機能ID"].astype(str)

        def _defect_total(fids):
            return int(dfd[dfd["機能ID"].isin(fids)].shape[0])
        defect_totals = features_per.map(_defect_total)
    else:
        defect_totals = pd.Series([0] * len(features_per),
                                  index=features_per.index)

    if "incident_rate" in kpi_df.columns:
        kpi_sub = (kpi_df[["機能ID", "incident_rate"]]
                   .drop_duplicates(subset=["機能ID"])
                   .assign(機能ID=lambda d: d["機能ID"].astype(str))
                   .set_index("機能ID"))

        def _avg_rate(fids):
            vals = pd.to_numeric(
                kpi_sub.reindex(fids)["incident_rate"], errors="coerce"
            ).dropna()
            return float(vals.mean()) if len(vals) else float("nan")
        avg_rates = features_per.map(_avg_rate)
    else:
        avg_rates = pd.Series([float("nan")] * len(features_per),
                              index=features_per.index)

    out = pd.DataFrame({
        "assignee": counts.index,
        "feature_count": feature_counts.reindex(counts.index).fillna(0)
                        .astype(int).values,
        "avg_incident_rate": avg_rates.reindex(counts.index).values,
        "defect_total": defect_totals.reindex(counts.index).fillna(0)
                        .astype(int).values,
        "dominant_role": dominant.reindex(counts.index).values,
        "role_count_dev":       counts["dev"].values,
        "role_count_test_spec": counts["test_spec"].values,
        "role_count_test_exec": counts["test_exec"].values,
    })
    # Keep everyone with at least one feature — downstream splitters
    # decide which bucket each assignee lands in (measurable / no-defects
    # / no-test-execution). Previously this function filtered aggressively,
    # which made the bubble map's empty-data cases invisible without a
    # follow-up channel.
    return out[out["feature_count"] > 0].reset_index(drop=True)


def _split_bubble_by_signal(
    bubble_df: pd.DataFrame,
) -> dict[str, pd.DataFrame]:
    """Partition the bubble frame into three exclusive buckets so each
    gets rendered in the viz/table that actually fits its meaning:

      - `main`       — `defect_total > 0` AND `avg_incident_rate` is
        a real number. These are the only rows plotted on the main
        scatter — their X/Y/size are all populated, so the
        breadth × quality × exposure reading holds.
      - `no_exec`    — `avg_incident_rate` is NaN, which means every
        feature they touched has 実施済 = 0 (tests planned, none
        executed). Quality cannot be measured yet — this is a process
        red flag, not a quality signal.
      - `zero_defect` — has measurable incident_rate (so tests were
        executed), but `defect_total = 0`. Ambiguous: could be genuine
        quality or under-reporting / shallow test coverage.

    Priority: `no_exec` wins over `zero_defect` when both are true,
    because "no tests executed" is the root cause and the defect count
    is meaningless without executions to compare against."""
    if bubble_df is None or bubble_df.empty:
        empty = pd.DataFrame(columns=bubble_df.columns if bubble_df is not None
                             else [])
        return {"main": empty, "no_exec": empty, "zero_defect": empty}
    rate_nan = bubble_df["avg_incident_rate"].isna()
    no_exec = bubble_df[rate_nan].copy()
    rest = bubble_df[~rate_nan]
    zero_defect = rest[rest["defect_total"] == 0].copy()
    main = rest[rest["defect_total"] > 0].copy()
    return {
        "main": main.reset_index(drop=True),
        "no_exec": no_exec.reset_index(drop=True),
        "zero_defect": zero_defect.reset_index(drop=True),
    }


def _chart_assignee_bubble(bubble_df: pd.DataFrame) -> Optional[go.Figure]:
    """Plotly bubble map: X=feature_count, Y=avg_incident_rate(%),
    size=defect_total, color=dominant_role, label=assignee.

    Adds dashed reference lines at the overall mean Y and the median X so
    readers can place each person in a quadrant at a glance."""
    if bubble_df is None or bubble_df.empty:
        return None
    df = bubble_df.copy()
    df["rate_pct"] = pd.to_numeric(df["avg_incident_rate"],
                                    errors="coerce") * 100.0
    role_labels_local = {
        "dev":       t("role_dev"),
        "test_spec": t("role_test_spec"),
        "test_exec": t("role_test_exec"),
    }
    df["role_label"] = df["dominant_role"].map(role_labels_local)
    # Hover card: enumerate per-role counts so the mixed-role case reads
    # clearly ("mostly dev, but also 2 test_exec sub-tasks").
    df["hover_roles"] = df.apply(
        lambda r: (
            f"{t('role_count_dev')}: {int(r['role_count_dev'])} · "
            f"{t('role_count_test_spec')}: {int(r['role_count_test_spec'])} · "
            f"{t('role_count_test_exec')}: {int(r['role_count_test_exec'])}"
        ),
        axis=1,
    )
    color_map_local = {role_labels_local[k]: v
                        for k, v in _ROLE_COLOR_MAP.items()}
    fig = px.scatter(
        df,
        x="feature_count",
        y="rate_pct",
        size="defect_total",
        color="role_label",
        text="assignee",
        size_max=60,
        color_discrete_map=color_map_local,
        custom_data=["assignee", "feature_count", "rate_pct",
                     "defect_total", "hover_roles"],
    )
    fig.update_traces(
        textposition="top center",
        textfont=dict(size=13, color="#1f2937",
                      family="Hiragino Sans, Yu Gothic, sans-serif"),
        cliponaxis=False,
        marker=dict(line=dict(color="#444", width=0.7), opacity=0.85),
        hovertemplate=(
            "<b>%{customdata[0]}</b><br>"
            f"{t('col_feature_count')}: %{{customdata[1]}}<br>"
            f"{t('col_avg_incident_rate')}: %{{customdata[2]:.1f}}%<br>"
            f"{t('col_defect_total')}: %{{customdata[3]}}<br>"
            "%{customdata[4]}"
            "<extra></extra>"
        ),
    )

    # Reference lines: overall mean Y and the median X. These anchor the
    # quadrant interpretation in the chart caption.
    rates = df["rate_pct"].dropna()
    if len(rates):
        mean_y = float(rates.mean())
        fig.add_hline(y=mean_y, line_dash="dash", line_color="#888",
                      annotation_text=f"avg {mean_y:.1f}%",
                      annotation_position="top right",
                      annotation_font=dict(color="#888", size=10))
    fc = pd.to_numeric(df["feature_count"], errors="coerce").dropna()
    if len(fc):
        med_x = float(fc.median())
        fig.add_vline(x=med_x, line_dash="dash", line_color="#888",
                      annotation_text=f"median {med_x:g}",
                      annotation_position="top left",
                      annotation_font=dict(color="#888", size=10))

    fig.update_layout(
        height=max(380, 32 * len(df) // 3 + 320),
        margin=_INLINE_MARGIN_DEFAULT,
        xaxis_title=t("col_feature_count"),
        yaxis_title=t("col_avg_incident_rate"),
        legend_title_text=t("role_analytics_bubble_color_legend"),
    )
    fig.update_xaxes(automargin=True,
                     range=[-0.5,
                            max(1, float(df["feature_count"].max()) * 1.15)])
    y_max = float(df["rate_pct"].max()) if df["rate_pct"].notna().any() else 10.0
    fig.update_yaxes(automargin=True, range=[0, max(10.0, y_max * 1.15)])
    return fig


def _render_bubble_watchlist(
    df: pd.DataFrame,
    *,
    title_key: str,
    caption_key: str,
    hide_defect_col: bool = False,
) -> None:
    """Render one of the two bubble-map watch-lists (zero-defect or
    tests-not-executed) as a Streamlit section — bold title, a
    definition caption explaining what the bucket actually means, and
    a compact dataframe with assignee / features touched / defects /
    role breakdown.

    The caller decides which bucket by passing the translation keys
    and (for the no-exec bucket) hides the Defects column since it's
    always zero and would be misleading next to the caption that
    says the quality can't be measured yet."""
    st.markdown(f"**{t(title_key)}**")
    st.caption(t(caption_key))
    if df is None or df.empty:
        st.info(t("role_analytics_watch_empty"))
        return
    role_labels_local = {
        "dev":       t("role_dev"),
        "test_spec": t("role_test_spec"),
        "test_exec": t("role_test_exec"),
    }
    rows = []
    for _, r in df.iterrows():
        role_bits = [
            f"{role_labels_local[k]}: {int(r.get(f'role_count_{k}', 0) or 0)}"
            for k in ("dev", "test_spec", "test_exec")
        ]
        rows.append({
            t("role_analytics_watch_col_assignee"): r["assignee"],
            t("role_analytics_watch_col_features"): int(r["feature_count"]),
            t("role_analytics_watch_col_defects"):  int(r["defect_total"]),
            t("role_analytics_watch_col_roles"):    " · ".join(role_bits),
        })
    wdf = pd.DataFrame(rows)
    if hide_defect_col:
        wdf = wdf.drop(columns=[t("role_analytics_watch_col_defects")])
    st.dataframe(wdf, use_container_width=True, hide_index=True)


def _build_assignee_problem_share_df(
    role_df: pd.DataFrame,
    defects_df: Optional[pd.DataFrame],
) -> Optional[tuple[pd.DataFrame, pd.DataFrame]]:
    """Per-assignee (pct_df, count_df) pair for the strip chart.

    Both frames share the same index (assignee) and columns (問題分類);
    `pct_df` rows sum to 100 (the 100%-stacked bar width), `count_df` is
    the matching raw defect counts so each segment can display "N件"
    alongside the percent. `pct_df` carries a `_total` sentinel column —
    the raw defect total per row — which the renderer pops off and uses
    as the right-hand annotation.

    Returns None when no defects tie back to any assignee's features."""
    ct = _build_assignee_problem_crosstab(role_df, defects_df)
    if ct is None or ct.empty:
        return None
    totals = ct.sum(axis=1)
    pct = ct.div(totals, axis=0).fillna(0.0) * 100.0
    pct["_total"] = totals
    # Row order by defect total desc (ct already does this) — keep that.
    return pct, ct.copy()


def _collapse_strip_to_top_n(
    pct_df: pd.DataFrame, cnt_df: pd.DataFrame, max_cats: int = 8,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Fold columns beyond `max_cats` into a collapsed 「その他」 bucket so
    the strip stays legible with many problem classes. Keeps pct and cnt
    in lock-step so segment widths and embedded counts still agree."""
    if pct_df.shape[1] <= max_cats:
        return pct_df, cnt_df
    other_label = t("role_analytics_strip_other")
    pct_head = pct_df.iloc[:, :max_cats - 1].copy()
    pct_head[other_label] = pct_df.iloc[:, max_cats - 1:].sum(axis=1)
    cnt_head = cnt_df.iloc[:, :max_cats - 1].copy()
    cnt_head[other_label] = cnt_df.iloc[:, max_cats - 1:].sum(axis=1)
    return pct_head, cnt_head


def _chart_assignee_problem_strip(
    strip: "tuple[pd.DataFrame, pd.DataFrame] | pd.DataFrame | None",
) -> Optional[go.Figure]:
    """Horizontal 100%-stacked bar per assignee, segments = 問題分類.

    Accepts either the (pct_df, count_df) pair emitted by
    _build_assignee_problem_share_df or a legacy single pct_df (older
    callers). Each segment shows `N件` as in-bar text; the hover tooltip
    combines `N件 (X.X%)`; a right-hand `n=<total>` annotation shows the
    per-assignee total so percent scale and absolute volume read together."""
    if strip is None:
        return None
    if isinstance(strip, tuple):
        pct_df, cnt_df = strip
    else:
        pct_df, cnt_df = strip, None
    if pct_df is None or pct_df.empty:
        return None
    pct_df = pct_df.copy()
    totals = pct_df.pop("_total")
    if pct_df.empty or pct_df.shape[1] == 0:
        return None
    if cnt_df is None:
        # Fallback for the legacy single-df path — reconstruct counts from
        # pct × total (rounded to int). Not perfectly faithful at odd
        # divisions but close enough for display.
        cnt_df = (pct_df.multiply(totals, axis=0) / 100.0).round().astype(int)
    pct_df, cnt_df = _collapse_strip_to_top_n(pct_df, cnt_df)

    palette = ["#3c78d8", "#e06666", "#6aa84f", "#f1c232", "#674ea7",
                "#16a2a2", "#d5a6bd", "#a64d79"]
    assignees = pct_df.index.tolist()
    fig = go.Figure()
    for i, cat in enumerate(pct_df.columns):
        pcts = pct_df[cat].astype(float).tolist()
        cnts = cnt_df[cat].astype(int).tolist() \
            if cat in cnt_df.columns else [0] * len(pcts)
        # Per-segment in-bar text: count only (Plotly also clips the text
        # when segment is too narrow, which is fine).
        seg_text = [f"{c}" if p > 0 else "" for c, p in zip(cnts, pcts)]
        # Hover customdata carries both count and percent so one template
        # reads as "42件 (33.3%)".
        customdata = list(zip(cnts, pcts))
        fig.add_bar(
            y=assignees, x=pcts, orientation="h",
            name=str(cat),
            marker_color=palette[i % len(palette)],
            text=seg_text, textposition="inside",
            insidetextanchor="middle",
            textfont=dict(color="white", size=10),
            cliponaxis=False,
            customdata=customdata,
            hovertemplate=(
                "<b>%{y}</b><br>"
                f"{cat}: %{{customdata[0]}}件 "
                "(%{customdata[1]:.1f}%)"
                "<extra></extra>"
            ),
        )
    # `n=<total>` annotations sit just past the 100% bar; extend the
    # axis range and reserve extra right margin so the text isn't clipped
    # by the plot boundary (same idea as the matplotlib version's
    # `ax.set_xlim(0, 108)`).
    _strip_margin = dict(_INLINE_MARGIN_LONG_Y)
    _strip_margin["r"] = 70
    fig.update_layout(
        barmode="stack",
        height=max(260, 30 * len(assignees) + 120),
        margin=_strip_margin,
        xaxis_title="%",
        yaxis_title=None,
        legend_title_text=t("chart_defect_class_col_class"),
        uniformtext_minsize=9, uniformtext_mode="hide",
    )
    fig.update_xaxes(range=[0, 112], automargin=True, ticksuffix="%",
                     tickvals=[0, 25, 50, 75, 100])
    fig.update_yaxes(autorange="reversed", automargin=True)
    # Right-hand "n=<total>" annotation — absolute volume, so percent bars
    # don't mislead (a 100%-of-tiny person looks equal to a 100%-of-huge
    # person without this). Bumped font weight + darker colour + a subtle
    # label prefix so the count actually reads from across the room.
    for name, tot in zip(assignees, totals.reindex(assignees).values):
        fig.add_annotation(
            x=101, y=name, xref="x", yref="y",
            text=f"<b>n={int(tot)}</b>",
            showarrow=False,
            xanchor="left", font=dict(color="#333", size=11),
        )
    return fig


def _chart_loc_trend() -> Optional[go.Figure]:
    snaps = load_all_snapshots_for_slot("code", load_code_counts)
    if len(snaps) < 2:
        return None
    rows = []
    for snap_date, _, df_snap in snaps:
        tot = pd.to_numeric(df_snap["LoC"], errors="coerce").fillna(0).sum()
        rows.append({"date": snap_date, "value": int(tot)})
    ts = pd.DataFrame(rows)
    fig = px.line(ts, x="date", y="value", markers=True,
                  labels={"value": t("chart_label_loc_total"), "date": ""})
    fig.update_layout(height=320, margin=_INLINE_MARGIN_DEFAULT)
    fig.update_xaxes(automargin=True)
    fig.update_yaxes(automargin=True)
    return fig


def _chart_test_trend() -> Optional[go.Figure]:
    snaps = load_all_snapshots_for_slot("tests", load_test_counts)
    if len(snaps) < 2:
        return None
    rows = []
    for snap_date, _, df_snap in snaps:
        tot = pd.to_numeric(df_snap["総設定テスト数"], errors="coerce").fillna(0).sum()
        run = pd.to_numeric(df_snap["実施済"], errors="coerce").fillna(0).sum()
        rows.append({"date": snap_date,
                     t("chart_label_total_tests"): int(tot),
                     t("chart_label_executed"): int(run)})
    ts = pd.DataFrame(rows)
    fig = px.line(ts, x="date",
                  y=[t("chart_label_total_tests"), t("chart_label_executed")],
                  markers=True)
    fig.update_layout(height=320, margin=_INLINE_MARGIN_DEFAULT,
                      legend_title_text="")
    fig.update_xaxes(automargin=True)
    fig.update_yaxes(automargin=True)
    return fig


_BUG_TREND_FID_LIMIT = 10


def _bug_trend_fid_breakdown(frame: pd.DataFrame, date_col: str,
                             idx: pd.DatetimeIndex) -> list[str]:
    if frame.empty or "機能ID" not in frame.columns:
        return [""] * len(idx)
    grp = (frame.set_index(date_col)
           .groupby([pd.Grouper(freq="W"), "機能ID"]).size())
    lines_by_week: dict[pd.Timestamp, str] = {}
    for week, sub in grp.groupby(level=0):
        counts = sub.droplevel(0).sort_values(ascending=False)
        total = len(counts)
        head = counts.head(_BUG_TREND_FID_LIMIT)
        lines = [f"{fid}: {int(n)}" for fid, n in head.items()]
        if total > _BUG_TREND_FID_LIMIT:
            lines.append(f"… +{total - _BUG_TREND_FID_LIMIT}")
        lines_by_week[week] = "<br>".join(lines)
    return [lines_by_week.get(week, "") for week in idx]


def _chart_bug_trend(defects_df: Optional[pd.DataFrame]) -> Optional[go.Figure]:
    if defects_df is None or defects_df.empty:
        return None
    df = defects_df.copy()
    df["実開始日"] = pd.to_datetime(df["実開始日"], errors="coerce")
    df["実終了日"] = pd.to_datetime(df["実終了日"], errors="coerce")
    opened = df.dropna(subset=["実開始日"]).copy()
    if opened.empty:
        return None
    closed = df.dropna(subset=["実終了日"]).copy()
    wk_opened = opened.set_index("実開始日").resample("W").size()
    wk_closed = (closed.set_index("実終了日").resample("W").size()
                 if len(closed) else pd.Series(dtype=int))
    idx = wk_opened.index.union(wk_closed.index)
    wk_opened = wk_opened.reindex(idx, fill_value=0)
    wk_closed = wk_closed.reindex(idx, fill_value=0)
    cumulative_open = (wk_opened - wk_closed).cumsum().clip(lower=0)
    opened_fid_text = _bug_trend_fid_breakdown(opened, "実開始日", idx)
    closed_fid_text = _bug_trend_fid_breakdown(closed, "実終了日", idx)
    hover_opened = (
        "<b>%{x|%Y-%m-%d}</b><br>"
        f"{t('chart_label_opened')}: %{{y}}<br>"
        "%{customdata[0]}<extra></extra>"
    )
    hover_closed = (
        "<b>%{x|%Y-%m-%d}</b><br>"
        f"{t('chart_label_closed')}: %{{y}}<br>"
        "%{customdata[0]}<extra></extra>"
    )
    def _week_bar_text(values):
        return [("" if (v is None or v == 0) else str(int(v))) for v in values]

    fig = go.Figure()
    fig.add_bar(name=t("chart_label_opened"), x=idx, y=wk_opened,
                marker_color="#f05050",
                text=_week_bar_text(wk_opened.values),
                textposition="outside",
                textfont=dict(color="#6c6c6c", size=10),
                cliponaxis=False,
                customdata=np.array(opened_fid_text).reshape(-1, 1),
                hovertemplate=hover_opened)
    fig.add_bar(name=t("chart_label_closed"), x=idx, y=wk_closed,
                marker_color="#4ec78a",
                text=_week_bar_text(wk_closed.values),
                textposition="outside",
                textfont=dict(color="#6c6c6c", size=10),
                cliponaxis=False,
                customdata=np.array(closed_fid_text).reshape(-1, 1),
                hovertemplate=hover_closed)
    fig.add_scatter(name=t("chart_label_open_cum"), x=idx,
                    y=cumulative_open, mode="lines+markers",
                    line=dict(color="#f5b400", width=2), yaxis="y2")
    fig.update_layout(barmode="group", height=380,
                      margin=_INLINE_MARGIN_DEFAULT,
                      yaxis=dict(title="weekly count", automargin=True),
                      yaxis2=dict(title="open", overlaying="y", side="right",
                                  automargin=True),
                      legend_title_text="")
    fig.update_xaxes(automargin=True)
    return fig


_DEFECT_CLASS_PALETTE = [
    "#7aaef0", "#f5b400", "#f05050", "#4ec78a", "#a982f0",
    "#5fc9c9", "#d56fa6", "#9aa0a6",
]


def _defect_class_counts(defects_df: pd.DataFrame) -> pd.Series:
    """Return counts per 問題分類, dropping rows with empty/missing class."""
    s = defects_df["問題分類"].fillna("").astype(str).str.strip()
    s = s[s != ""]
    return s.value_counts()


def _chart_defect_class(defects_df: Optional[pd.DataFrame]
                        ) -> Optional[go.Figure]:
    """Pie chart of Redmine defects' 問題分類 distribution. Caller is
    expected to pre-filter `defects_df` (e.g. by 機能ID) so the chart can
    reflect a slice."""
    if defects_df is None or defects_df.empty:
        return None
    if "問題分類" not in defects_df.columns:
        return None
    counts = _defect_class_counts(defects_df)
    if counts.empty:
        return None
    palette = (_DEFECT_CLASS_PALETTE
               * (1 + len(counts) // len(_DEFECT_CLASS_PALETTE)))[:len(counts)]
    fig = go.Figure(data=[go.Pie(
        labels=counts.index.tolist(),
        values=counts.values.tolist(),
        hole=0.4,
        marker=dict(colors=palette),
        textinfo="label+percent",
        hovertemplate="<b>%{label}</b><br>件数: %{value}<br>"
                      "割合: %{percent}<extra></extra>",
        sort=False,
    )])
    fig.update_layout(height=360, margin=dict(l=20, r=20, t=20, b=20),
                      legend_title_text="")
    return fig


def _collect_fid_history(function_id: str) -> pd.DataFrame:
    """Walk every saved tests/code snapshot under ``input/`` and extract the
    row for ``function_id``. Returns a long-ish dataframe indexed by
    snapshot date with columns {総設定テスト数, 実施済, OK, NG, 未実施, LoC}
    (only those that were present). Missing cells stay NaN so the chart
    silently omits them.
    """
    rows: dict[date, dict] = {}
    for snap_date, _, df_snap in load_all_snapshots_for_slot(
        "tests", load_test_counts
    ):
        sub = df_snap[df_snap["機能ID"] == function_id]
        if sub.empty:
            continue
        r = sub.iloc[0]
        bucket = rows.setdefault(snap_date, {})
        for c in ("総設定テスト数", "実施済", "OK", "NG", "未実施"):
            if c in sub.columns:
                v = r.get(c)
                bucket[c] = float(v) if pd.notna(v) else None
    for snap_date, _, df_snap in load_all_snapshots_for_slot(
        "code", load_code_counts
    ):
        sub = df_snap[df_snap["機能ID"] == function_id]
        if sub.empty:
            continue
        r = sub.iloc[0]
        bucket = rows.setdefault(snap_date, {})
        if "LoC" in sub.columns and pd.notna(r.get("LoC")):
            bucket["LoC"] = float(r["LoC"])
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(
        [{"date": d, **vals} for d, vals in rows.items()]
    ).sort_values("date").reset_index(drop=True)
    return df


def _chart_fid_trend(function_id: str) -> Optional[go.Figure]:
    """Per-Function-ID trend: NG / 実施済 / 総設定テスト数 on the left axis
    (counts) and LoC on the right axis (lines of code). Returns None when
    fewer than two snapshots have data for this FID."""
    df = _collect_fid_history(function_id)
    if len(df) < 2:
        return None
    fig = go.Figure()
    if "NG" in df.columns and df["NG"].notna().any():
        fig.add_scatter(name="NG", x=df["date"], y=df["NG"],
                        mode="lines+markers", line=dict(color="#f05050"))
    if "実施済" in df.columns and df["実施済"].notna().any():
        fig.add_scatter(name="実施済", x=df["date"], y=df["実施済"],
                        mode="lines+markers", line=dict(color="#4ec78a"))
    if "総設定テスト数" in df.columns and df["総設定テスト数"].notna().any():
        fig.add_scatter(name="総設定テスト数", x=df["date"], y=df["総設定テスト数"],
                        mode="lines+markers", line=dict(color="#7aaef0"))
    if "LoC" in df.columns and df["LoC"].notna().any():
        fig.add_scatter(name="LoC", x=df["date"], y=df["LoC"],
                        mode="lines+markers", yaxis="y2",
                        line=dict(color="#f5b400", dash="dot"))
    # Need at least one trace worth charting
    if not fig.data:
        return None
    fig.update_layout(
        height=300, margin=dict(l=60, r=60, t=30, b=40),
        yaxis=dict(title="tests (count)", automargin=True),
        yaxis2=dict(title="LoC", overlaying="y", side="right",
                    automargin=True, showgrid=False),
        legend_title_text="",
    )
    fig.update_xaxes(automargin=True)
    return fig


def _chart_gantt(kpi_df: pd.DataFrame, today_d: date) -> Optional[go.Figure]:
    label_planned = t("calendar_layer_planned")
    label_actual = t("calendar_layer_actual")
    rows: list[dict] = []
    for _, row in kpi_df.iterrows():
        label = _clip_label(_label_id_name(row))
        ps = _to_pydate(row.get("planned_start"))
        pe = _to_pydate(row.get("planned_end"))
        ase = _to_pydate(row.get("actual_start"))
        aee = _to_pydate(row.get("actual_end"))
        if ps and pe and pe >= ps:
            rows.append({"ID": label,
                         "Start": pd.Timestamp(ps),
                         "End": pd.Timestamp(pe + timedelta(days=1)),
                         "Layer": label_planned})
        if ase:
            end = aee if aee else today_d
            if end < ase:
                end = ase
            rows.append({"ID": label,
                         "Start": pd.Timestamp(ase),
                         "End": pd.Timestamp(end + timedelta(days=1)),
                         "Layer": label_actual})
    if not rows:
        return None
    df_g = pd.DataFrame(rows)
    # Cap the row count for the same reason as the other per-feature charts:
    # a 431-feature Gantt is unreadable and makes kaleido stall. Pick the
    # features that span today (most time-relevant), breaking ties by
    # earliest actual/planned start.
    total_ids = df_g["ID"].nunique()
    if total_ids > _BAR_CHART_MAX_ROWS:
        today_ts0 = pd.Timestamp(today_d)
        status = (
            df_g.groupby("ID")
                .agg(mn=("Start", "min"), mx=("End", "max"))
                .reset_index()
        )
        status["crosses_today"] = (
            (status["mn"] <= today_ts0) & (status["mx"] >= today_ts0)
        ).astype(int)
        keep_ids = (
            status.sort_values(
                ["crosses_today", "mn"], ascending=[False, True])
                  .head(_BAR_CHART_MAX_ROWS)["ID"]
                  .tolist()
        )
        df_g = df_g[df_g["ID"].isin(keep_ids)]
    fig = px.timeline(df_g, x_start="Start", x_end="End", y="ID",
                      color="Layer",
                      color_discrete_map={label_planned: "#9aa0a6",
                                          label_actual: "#4ec78a"})
    fig.update_yaxes(autorange="reversed")
    today_ts = pd.Timestamp(today_d)
    fig.add_vline(x=today_ts, line_width=1, line_dash="dash",
                  line_color="#f5b400")
    fig.add_annotation(x=today_ts, y=1, yref="paper",
                       text=t("gantt_today_label"), showarrow=False,
                       font=dict(color="#f5b400", size=11), yanchor="bottom")
    fig.update_layout(height=max(320, 26 * df_g["ID"].nunique() + 80),
                      margin=_INLINE_MARGIN_LONG_Y,
                      xaxis_title="", yaxis_title="",
                      legend_title_text="")
    fig.update_yaxes(automargin=True)
    fig.update_xaxes(automargin=True)
    if total_ids > _BAR_CHART_MAX_ROWS:
        fig.add_annotation(**_truncate_note_annotation(
            df_g["ID"].nunique(), total_ids))
    return fig


# =============================================================================
# PDF chart builders (matplotlib)
#
# The on-screen Charts tab still uses Plotly (see _chart_* above). For the
# PDF report we render the same data with matplotlib instead, because the
# kaleido 0.2.1 Chromium subprocess fails to launch on locked-down
# corporate Windows boxes — that was the entire v1.0.2…v1.0.4 investigation.
# matplotlib is pure Python / bundled binary wheels, no subprocess, and so
# always completes.
# =============================================================================
_MPL_READY = False
_MPL_DPI = 140
_MPL_WIDTH_IN = 14.0                                 # PDF A3 landscape fits ≈ 14"
_MPL_CJK_CANDIDATES = [
    # macOS
    "Hiragino Sans", "Hiragino Maru Gothic Pro",
    # Windows (Yu Gothic is default since 8.1)
    "Yu Gothic", "Yu Gothic UI", "Meiryo", "MS Gothic", "MS UI Gothic",
    # Linux / optional
    "Noto Sans CJK JP", "Noto Sans JP", "IPAexGothic", "IPAPGothic",
    # Fallback
    "DejaVu Sans",
]


def _mpl_plt():
    """Return matplotlib.pyplot after one-time headless backend + CJK-font
    initialization. Backend is forced to Agg (no GUI) so this never needs
    a display server."""
    global _MPL_READY
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt  # noqa: E402
    if not _MPL_READY:
        from matplotlib import rcParams
        import matplotlib.font_manager as fm
        installed = {f.name for f in fm.fontManager.ttflist}
        for c in _MPL_CJK_CANDIDATES:
            if c in installed:
                rcParams["font.family"] = c
                _get_logger().info(f"[pdf_export] matplotlib font: {c}")
                break
        else:
            _get_logger().warning(
                "[pdf_export] no CJK font found; Japanese text may render "
                "as tofu boxes in the PDF.")
        rcParams["axes.unicode_minus"] = False
        _MPL_READY = True
    return plt


def _mpl_save(fig) -> tuple[bytes, int, int]:
    """Write the figure to PNG bytes and close it. Returns (bytes, w_px, h_px)
    so the PDF layout can scale to its intrinsic aspect ratio."""
    import io
    plt = _mpl_plt()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=_MPL_DPI, bbox_inches="tight",
                facecolor="white")
    w_in, h_in = fig.get_size_inches()
    plt.close(fig)
    return buf.getvalue(), int(w_in * _MPL_DPI), int(h_in * _MPL_DPI)


def _mpl_bar_height_in(n_rows: int) -> float:
    """Figure height in inches that fits `n_rows` horizontal-bar rows."""
    return max(3.0, 0.32 * n_rows + 1.2)


def _mpl_truncated_title(ax, shown: int, total: int) -> None:
    ax.set_title(t("chart_truncated_note", shown=shown, total=total),
                 fontsize=10, color="#b48820", loc="right")


def _mpl_chart_progress_gap(kpi_df: pd.DataFrame):
    if not {"actual_progress", "planned_progress"}.issubset(kpi_df.columns):
        return None
    df = kpi_df.dropna(subset=["actual_progress", "planned_progress"]).copy()
    if df.empty:
        return None
    df["display"] = _feature_display_series(df).map(_clip_label)
    df["_gap"] = df["planned_progress"] - df["actual_progress"]
    df = df.sort_values("_gap", ascending=False)
    total = len(df)
    if total > _BAR_CHART_MAX_ROWS:
        df = df.head(_BAR_CHART_MAX_ROWS)
    df = df.iloc[::-1]
    n = len(df)
    plt = _mpl_plt()
    fig, ax = plt.subplots(
        figsize=(_MPL_WIDTH_IN, _mpl_bar_height_in(n)), dpi=_MPL_DPI)
    y = np.arange(n)
    h = 0.38
    over = (df["actual_progress"] > df["planned_progress"]).values
    actual_colors = np.where(over, "#f5b400", "#4ec78a")
    actual_lines = np.where(over, "#a06a00", "#4ec78a")
    ax.barh(y - h / 2, df["planned_progress"], height=h,
            color="#9aa0a6", label=t("chart_progress_planned"))
    ax.barh(y + h / 2, df["actual_progress"], height=h,
            color=actual_colors.tolist(),
            edgecolor=actual_lines.tolist(),
            linewidth=0.6,
            label=t("chart_progress_actual"))
    over_marker = t("chart_progress_over_marker")
    for yi, val, is_over in zip(y, df["actual_progress"].values, over):
        if is_over:
            ax.text(val + 1, yi + h / 2, over_marker,
                    color="#a06a00", fontsize=8, va="center", ha="left")
    ax.set_yticks(y); ax.set_yticklabels(df["display"])
    ax.set_xlabel("%")
    ax.legend(loc="lower right", framealpha=0.9)
    ax.grid(axis="x", linestyle=":", alpha=0.3)
    if total > _BAR_CHART_MAX_ROWS:
        _mpl_truncated_title(ax, n, total)
    fig.tight_layout()
    return _mpl_save(fig)


def _mpl_chart_test_density(kpi_df: pd.DataFrame):
    if not {"test_density"}.issubset(kpi_df.columns):
        return None
    df = kpi_df.dropna(subset=["test_density"]).copy()
    if df.empty:
        return None
    df["display"] = _feature_display_series(df).map(_clip_label)
    df = df.sort_values("test_density", ascending=True)
    total = len(df)
    if total > _BAR_CHART_MAX_ROWS:
        df = df.head(_BAR_CHART_MAX_ROWS)
    df = df.iloc[::-1]
    n = len(df)
    plt = _mpl_plt()
    fig, ax = plt.subplots(
        figsize=(_MPL_WIDTH_IN, _mpl_bar_height_in(n)), dpi=_MPL_DPI)
    y = np.arange(n)
    densities = df["test_density"].values.astype(float)
    threshold = _test_density_threshold()
    colors = ["#f05050" if d < threshold else "#7aaef0" for d in densities]
    ax.barh(y, densities, color=colors)
    ax.axvline(threshold, color="#a02020", linestyle="--", linewidth=1)
    ax.text(threshold, n - 0.5,
            f" {t('chart_test_density_threshold_label')} {threshold:g}",
            color="#a02020", fontsize=9, va="bottom", ha="left")
    ax.set_yticks(y); ax.set_yticklabels(df["display"])
    ax.set_xlabel("tests / page")
    ax.grid(axis="x", linestyle=":", alpha=0.3)
    if total > _BAR_CHART_MAX_ROWS:
        _mpl_truncated_title(ax, n, total)
    fig.tight_layout()
    return _mpl_save(fig)


def _mpl_chart_incident_rate(kpi_df: pd.DataFrame):
    if not {"incident_rate"}.issubset(kpi_df.columns):
        return None
    df = kpi_df.dropna(subset=["incident_rate"]).copy()
    if df.empty:
        return None
    df["display"] = _feature_display_series(df).map(_clip_label)
    df = df.sort_values("incident_rate", ascending=False)
    total = len(df)
    if total > _BAR_CHART_MAX_ROWS:
        df = df.head(_BAR_CHART_MAX_ROWS)
    df = df.iloc[::-1]
    n = len(df)
    plt = _mpl_plt()
    fig, ax = plt.subplots(
        figsize=(_MPL_WIDTH_IN, _mpl_bar_height_in(n)), dpi=_MPL_DPI)
    y = np.arange(n)
    rates = df["incident_rate"].values.astype(float) * 100.0
    threshold_pct = _incident_rate_threshold() * 100.0
    colors = ["#f05050" if r > threshold_pct else "#7aaef0" for r in rates]
    ax.barh(y, rates, color=colors)
    ax.axvline(threshold_pct, color="#a02020", linestyle="--", linewidth=1)
    ax.text(threshold_pct, n - 0.5,
            f" {t('chart_incident_rate_threshold_label')} {threshold_pct:g}%",
            color="#a02020", fontsize=9, va="bottom", ha="left")
    ax.set_yticks(y); ax.set_yticklabels(df["display"])
    ax.set_xlabel("%")
    ax.grid(axis="x", linestyle=":", alpha=0.3)
    if total > _BAR_CHART_MAX_ROWS:
        _mpl_truncated_title(ax, n, total)
    fig.tight_layout()
    return _mpl_save(fig)


def _mpl_chart_test_coverage(kpi_df: pd.DataFrame):
    if not {"OK", "NG", "未実施"}.issubset(kpi_df.columns):
        return None
    df = kpi_df.dropna(subset=["OK", "NG", "未実施"], how="all").copy()
    if df.empty:
        return None
    df["display"] = _feature_display_series(df).map(_clip_label)
    df["_bad"] = df["NG"].fillna(0) + df["未実施"].fillna(0) * 0.5
    df = df.sort_values("_bad", ascending=False)
    total = len(df)
    if total > _BAR_CHART_MAX_ROWS:
        df = df.head(_BAR_CHART_MAX_ROWS)
    df = df.iloc[::-1]
    n = len(df)
    plt = _mpl_plt()
    fig, ax = plt.subplots(
        figsize=(_MPL_WIDTH_IN, _mpl_bar_height_in(n)), dpi=_MPL_DPI)
    y = np.arange(n)
    ok = df["OK"].fillna(0).values
    ng = df["NG"].fillna(0).values
    nr = df["未実施"].fillna(0).values
    ax.barh(y, ok, color="#4ec78a", label=t("chart_label_ok"))
    ax.barh(y, ng, left=ok, color="#f05050", label=t("chart_label_ng"))
    ax.barh(y, nr, left=ok + ng, color="#bbbbbb",
            label=t("chart_label_notrun"))
    ax.set_yticks(y); ax.set_yticklabels(df["display"])
    ax.legend(loc="lower right", framealpha=0.9)
    ax.grid(axis="x", linestyle=":", alpha=0.3)
    if total > _BAR_CHART_MAX_ROWS:
        _mpl_truncated_title(ax, n, total)
    fig.tight_layout()
    return _mpl_save(fig)


def _mpl_chart_loc_vs_ng(kpi_df: pd.DataFrame):
    if not {"LoC", "NG"}.issubset(kpi_df.columns):
        return None
    df = kpi_df.dropna(subset=["LoC", "NG"]).copy()
    if df.empty:
        return None
    if ("設計書ページ数" in df.columns
            and df["設計書ページ数"].notna().any()):
        raw = pd.to_numeric(df["設計書ページ数"], errors="coerce").fillna(5)
        m = raw.max()
        sizes = (raw / m * 240.0 + 20.0).values if m > 0 else 40.0
    else:
        sizes = 40.0
    has_risk = "risk_score" in df.columns
    colors = (df["risk_score"].fillna(0).values if has_risk else "#3aa872")
    plt = _mpl_plt()
    fig, ax = plt.subplots(figsize=(_MPL_WIDTH_IN, 5), dpi=_MPL_DPI)
    sc = ax.scatter(df["LoC"], df["NG"], s=sizes, c=colors,
                    cmap="RdYlGn_r" if has_risk else None,
                    vmin=0 if has_risk else None,
                    vmax=1 if has_risk else None,
                    alpha=0.85, edgecolors="#444", linewidth=0.5)
    if has_risk:
        fig.colorbar(sc, ax=ax, label="risk")
    ax.set_xlabel("LoC"); ax.set_ylabel("NG")
    ax.grid(True, linestyle=":", alpha=0.3)
    fig.tight_layout()
    return _mpl_save(fig)


def _mpl_chart_design_impl_gap(kpi_df: pd.DataFrame):
    if not {"設計書ページ数", "LoC"}.issubset(kpi_df.columns):
        return None
    df = kpi_df.dropna(subset=["設計書ページ数", "LoC"]).copy()
    if df.empty:
        return None
    plt = _mpl_plt()
    fig, ax = plt.subplots(figsize=(_MPL_WIDTH_IN, 5), dpi=_MPL_DPI)
    ax.scatter(df["設計書ページ数"], df["LoC"], s=40, c="#3aa872",
               edgecolors="#444", linewidth=0.5, alpha=0.85)
    comp = pd.to_numeric(df.get("complexity"), errors="coerce").dropna()
    if len(comp):
        avg = float(comp.mean())
        xs = np.linspace(float(df["設計書ページ数"].min()),
                         float(df["設計書ページ数"].max()), 50)
        ax.plot(xs, avg * xs, ls="--", color="#888",
                label=f"avg complexity = {avg:.1f}")
        ax.legend(loc="best")
    ax.set_xlabel("設計書ページ数"); ax.set_ylabel("LoC")
    ax.grid(True, linestyle=":", alpha=0.3)
    fig.tight_layout()
    return _mpl_save(fig)


def _mpl_chart_risk_heatmap(kpi_df: pd.DataFrame):
    risk_dims = [c for c in
                 ["bug_density", "incident_rate", "delay_rate",
                  "test_run_rate", "test_density"]
                 if c in kpi_df.columns]
    if not risk_dims:
        return None
    agg = kpi_df.groupby("機能ID")[risk_dims].mean(numeric_only=True)
    z_df = agg.copy()
    for c in risk_dims:
        s = z_df[c]
        m = s.max(skipna=True)
        if pd.notna(m) and m > 0:
            z_df[c] = s / m
        if c in _RISK_HEATMAP_INVERTED_DIMS:
            mask = z_df[c].notna()
            z_df.loc[mask, c] = 1 - z_df.loc[mask, c]
    z_df = z_df.sort_values(by=risk_dims[0], ascending=False,
                            na_position="last")
    dim_label = {c: t(COLUMN_LABEL_KEYS.get(c, c)) for c in risk_dims}
    if "test_run_rate" in dim_label:
        dim_label["test_run_rate"] = (
            f"{dim_label['test_run_rate']} ({t('chart_label_notrun')})"
        )
    if "test_density" in dim_label:
        dim_label["test_density"] = (
            f"{dim_label['test_density']} ({t('chart_label_low')})"
        )
    y_labels = [dim_label[c] for c in risk_dims]
    # Resolve 機能ID → 機能名 from the master (falling back to the
    # master-joined kpi_df if the master isn't in session_state) so the
    # x-axis reads as 機能ID：機能名 — same convention as the on-screen
    # Plotly heatmap at _chart_risk_heatmap.
    name_map = _master_fid_name_map()
    if not name_map and "機能名称" in kpi_df.columns:
        dedup = kpi_df.drop_duplicates(subset=["機能ID"])
        name_map = {str(f): ("" if pd.isna(n) else str(n))
                    for f, n in zip(dedup["機能ID"], dedup["機能名称"])}
    x_labels = [_label_fid_name(f, name_map.get(str(f), ""))
                for f in z_df.index]
    data = np.ma.masked_invalid(z_df.T.values.astype(float))
    plt = _mpl_plt()
    import matplotlib as _mpl
    cmap = _mpl.cm.get_cmap("RdYlGn_r").copy()
    cmap.set_bad("#d0d0d0")
    fig_h = max(3.0, 0.6 + 0.4 * len(y_labels))
    fig, ax = plt.subplots(figsize=(_MPL_WIDTH_IN, fig_h), dpi=_MPL_DPI)
    im = ax.imshow(data, aspect="auto", cmap=cmap, vmin=0, vmax=1)
    ax.set_yticks(np.arange(len(y_labels)))
    ax.set_yticklabels(y_labels)
    ax.set_xticks(np.arange(len(x_labels)))
    ax.set_xticklabels(x_labels, rotation=-30, ha="left", fontsize=9)
    ax.set_xlabel("機能ID：機能名")
    fig.colorbar(im, ax=ax, label="risk")
    fig.tight_layout()
    return _mpl_save(fig)


def _mpl_chart_loc_trend():
    snaps = load_all_snapshots_for_slot("code", load_code_counts)
    if len(snaps) < 2:
        return None
    rows = []
    for snap_date, _, df_snap in snaps:
        tot = pd.to_numeric(df_snap["LoC"], errors="coerce").fillna(0).sum()
        rows.append({"date": pd.Timestamp(snap_date), "value": int(tot)})
    ts = pd.DataFrame(rows).sort_values("date")
    plt = _mpl_plt()
    fig, ax = plt.subplots(figsize=(_MPL_WIDTH_IN, 4), dpi=_MPL_DPI)
    ax.plot(ts["date"], ts["value"], marker="o", color="#4ec78a", linewidth=2)
    ax.set_ylabel(t("chart_label_loc_total"))
    ax.grid(True, linestyle=":", alpha=0.3)
    fig.autofmt_xdate()
    fig.tight_layout()
    return _mpl_save(fig)


def _mpl_chart_test_trend():
    snaps = load_all_snapshots_for_slot("tests", load_test_counts)
    if len(snaps) < 2:
        return None
    rows = []
    for snap_date, _, df_snap in snaps:
        tot = pd.to_numeric(df_snap["総設定テスト数"], errors="coerce").fillna(0).sum()
        run = pd.to_numeric(df_snap["実施済"], errors="coerce").fillna(0).sum()
        rows.append({"date": pd.Timestamp(snap_date),
                     "total": int(tot), "executed": int(run)})
    ts = pd.DataFrame(rows).sort_values("date")
    plt = _mpl_plt()
    fig, ax = plt.subplots(figsize=(_MPL_WIDTH_IN, 4), dpi=_MPL_DPI)
    ax.plot(ts["date"], ts["total"], marker="o", color="#3aa872",
            label=t("chart_label_total_tests"), linewidth=2)
    ax.plot(ts["date"], ts["executed"], marker="s", color="#f5b400",
            label=t("chart_label_executed"), linewidth=2)
    ax.legend(loc="best")
    ax.grid(True, linestyle=":", alpha=0.3)
    fig.autofmt_xdate()
    fig.tight_layout()
    return _mpl_save(fig)


def _mpl_chart_defect_class(defects_df: Optional[pd.DataFrame]):
    if defects_df is None or defects_df.empty:
        return None
    if "問題分類" not in defects_df.columns:
        return None
    counts = _defect_class_counts(defects_df)
    if counts.empty:
        return None
    plt = _mpl_plt()
    fig, ax = plt.subplots(figsize=(_MPL_WIDTH_IN, 4.2), dpi=_MPL_DPI)
    palette = (_DEFECT_CLASS_PALETTE
               * (1 + len(counts) // len(_DEFECT_CLASS_PALETTE)))[:len(counts)]
    ax.pie(counts.values, labels=counts.index.tolist(), colors=palette,
           autopct="%1.1f%%", startangle=90, counterclock=False,
           wedgeprops=dict(width=0.45))
    ax.set_aspect("equal")
    fig.tight_layout()
    return _mpl_save(fig)


def _mpl_chart_bug_trend(defects_df: Optional[pd.DataFrame]):
    if defects_df is None or defects_df.empty:
        return None
    df = defects_df.copy()
    df["実開始日"] = pd.to_datetime(df["実開始日"], errors="coerce")
    df["実終了日"] = pd.to_datetime(df["実終了日"], errors="coerce")
    opened = df.dropna(subset=["実開始日"]).copy()
    if opened.empty:
        return None
    closed = df.dropna(subset=["実終了日"]).copy()
    wk_opened = opened.set_index("実開始日").resample("W").size()
    # Ensure wk_closed carries a DatetimeIndex even when empty, so the union
    # below stays a DatetimeIndex (otherwise .to_pydatetime() / date2num fail).
    wk_closed = (closed.set_index("実終了日").resample("W").size()
                 if len(closed)
                 else pd.Series(dtype=int,
                                index=pd.DatetimeIndex([], name="実終了日")))
    idx = pd.DatetimeIndex(wk_opened.index.union(wk_closed.index))
    wk_opened = wk_opened.reindex(idx, fill_value=0)
    wk_closed = wk_closed.reindex(idx, fill_value=0)
    cumulative_open = (wk_opened - wk_closed).cumsum().clip(lower=0)
    plt = _mpl_plt()
    fig, ax1 = plt.subplots(figsize=(_MPL_WIDTH_IN, 4.5), dpi=_MPL_DPI)
    # Bars as paired (opened / closed) per week — date2num accepts a
    # DatetimeIndex directly (no .to_pydatetime() round-trip needed).
    import matplotlib.dates as mdates
    x_num = mdates.date2num(idx)
    w = 2.8  # days
    ax1.bar(x_num - w / 2, wk_opened.values, width=w, color="#f05050",
            label=t("chart_label_opened"))
    ax1.bar(x_num + w / 2, wk_closed.values, width=w, color="#4ec78a",
            label=t("chart_label_closed"))
    ax1.set_ylabel("weekly count")
    ax2 = ax1.twinx()
    ax2.plot(x_num, cumulative_open.values, marker="o", color="#f5b400",
             linewidth=2, label=t("chart_label_open_cum"))
    ax2.set_ylabel("open")
    ax1.xaxis_date()
    l1, lbl1 = ax1.get_legend_handles_labels()
    l2, lbl2 = ax2.get_legend_handles_labels()
    ax1.legend(l1 + l2, lbl1 + lbl2, loc="upper left")
    ax1.grid(True, linestyle=":", alpha=0.3)
    fig.autofmt_xdate()
    fig.tight_layout()
    return _mpl_save(fig)


def _mpl_chart_assignee_bubble(bubble_df: pd.DataFrame):
    """Matplotlib bubble map for the role-analytics PDF export.

    Mirrors _chart_assignee_bubble: X=feature_count, Y=avg_incident_rate%,
    marker size ∝ defect_total, colour by dominant role (via _ROLE_COLOR_MAP).
    Adds mean-Y / median-X reference lines + labels each bubble with the
    assignee name."""
    if bubble_df is None or bubble_df.empty:
        return None
    df = bubble_df.copy()
    df["rate_pct"] = pd.to_numeric(df["avg_incident_rate"],
                                    errors="coerce") * 100.0
    df["_defect_total"] = pd.to_numeric(df["defect_total"],
                                         errors="coerce").fillna(0)
    plt = _mpl_plt()
    fig, ax = plt.subplots(figsize=(_MPL_WIDTH_IN, 6), dpi=_MPL_DPI)

    # Marker size: scale by defect total so the largest is legible and the
    # smallest still shows. Minimum of 80pt² means 0-defect bubbles still
    # read as a dot.
    sizes = df["_defect_total"].values.astype(float)
    max_d = float(sizes.max()) if len(sizes) and sizes.max() > 0 else 1.0
    marker_sizes = 80.0 + (sizes / max_d) * 1800.0
    colors = [_ROLE_COLOR_MAP.get(r, "#888888")
              for r in df["dominant_role"]]
    ax.scatter(
        df["feature_count"], df["rate_pct"],
        s=marker_sizes, c=colors,
        alpha=0.75, edgecolors="#333", linewidth=0.8,
    )
    # Label every bubble just above its centre. White path-effect halo
    # keeps the name readable even when it overlaps a large coloured
    # bubble or the grid lines.
    import matplotlib.patheffects as _pe
    for _, r in df.iterrows():
        if pd.isna(r["rate_pct"]):
            continue
        txt = ax.annotate(
            r["assignee"],
            xy=(r["feature_count"], r["rate_pct"]),
            xytext=(0, 11), textcoords="offset points",
            ha="center", fontsize=11, color="#1f2937", fontweight="bold",
        )
        txt.set_path_effects([
            _pe.withStroke(linewidth=2.2, foreground="white"),
        ])
    # Reference lines
    rates = df["rate_pct"].dropna()
    if len(rates):
        ax.axhline(float(rates.mean()), ls="--", color="#888", lw=1)
    fc = pd.to_numeric(df["feature_count"], errors="coerce").dropna()
    if len(fc):
        ax.axvline(float(fc.median()), ls="--", color="#888", lw=1)
    ax.set_xlabel(t("col_feature_count"))
    ax.set_ylabel(t("col_avg_incident_rate") + " (%)")
    # Y axis starts at 0; X axis starts slightly left of 0 for breathing.
    y_max = float(df["rate_pct"].max()) if df["rate_pct"].notna().any() else 10.0
    ax.set_ylim(0, max(10.0, y_max * 1.15))
    ax.set_xlim(-0.5, max(1, float(df["feature_count"].max()) * 1.15))
    ax.grid(True, linestyle=":", alpha=0.3)

    # Custom legend: one proxy marker per dominant role actually present.
    import matplotlib.lines as _mlines
    role_labels_local = {
        "dev":       t("role_dev"),
        "test_spec": t("role_test_spec"),
        "test_exec": t("role_test_exec"),
    }
    present = [r for r in ("dev", "test_spec", "test_exec")
               if r in df["dominant_role"].unique()]
    handles = [
        _mlines.Line2D([], [], marker="o", color="w",
                        markerfacecolor=_ROLE_COLOR_MAP[r],
                        markersize=10, markeredgecolor="#333",
                        label=role_labels_local[r])
        for r in present
    ]
    if handles:
        ax.legend(handles=handles, loc="upper left",
                   title=t("role_analytics_bubble_color_legend"),
                   fontsize=9, title_fontsize=9, framealpha=0.9)
    fig.tight_layout()
    return _mpl_save(fig)


def _mpl_chart_assignee_problem_strip(strip):
    """Matplotlib 100%-stacked horizontal bar per assignee — PDF counterpart
    of _chart_assignee_problem_strip. Accepts the (pct_df, count_df) pair
    and prints the raw count inside each segment wide enough to hold it."""
    if strip is None:
        return None
    if isinstance(strip, tuple):
        pct_df, cnt_df = strip
    else:
        pct_df, cnt_df = strip, None
    if pct_df is None or pct_df.empty:
        return None
    pct_df = pct_df.copy()
    totals = pct_df.pop("_total")
    if pct_df.empty or pct_df.shape[1] == 0:
        return None
    if cnt_df is None:
        cnt_df = (pct_df.multiply(totals, axis=0) / 100.0).round().astype(int)
    pct_df, cnt_df = _collapse_strip_to_top_n(pct_df, cnt_df)

    palette = ["#3c78d8", "#e06666", "#6aa84f", "#f1c232", "#674ea7",
                "#16a2a2", "#d5a6bd", "#a64d79"]
    plt = _mpl_plt()
    n = len(pct_df.index)
    fig_h = max(3.0, 0.45 * n + 1.4)
    fig, ax = plt.subplots(figsize=(_MPL_WIDTH_IN, fig_h), dpi=_MPL_DPI)
    y = np.arange(n)
    assignees = pct_df.index.tolist()
    left = np.zeros(n, dtype=float)
    # Minimum segment width (in percent units of the 100-wide X axis) a
    # count label needs to print inside without visually crowding. Tuned so
    # two digits stay readable; narrower segments drop the label.
    MIN_LABEL_PCT = 5.0
    for i, cat in enumerate(pct_df.columns):
        pct_vals = pct_df[cat].astype(float).values
        cnt_vals = (cnt_df[cat].astype(int).values
                    if cat in cnt_df.columns else np.zeros(n, dtype=int))
        ax.barh(y, pct_vals, left=left, color=palette[i % len(palette)],
                 edgecolor="white", label=str(cat))
        # In-bar count labels — only where the segment has enough width.
        for yi, (pv, cv, lf) in enumerate(zip(pct_vals, cnt_vals, left)):
            if pv >= MIN_LABEL_PCT and cv > 0:
                ax.text(lf + pv / 2, yi, f"{int(cv)}",
                        ha="center", va="center",
                        fontsize=8, color="white")
        left += pct_vals
    # Raw totals on the right
    for yi, name in enumerate(assignees):
        ax.text(101, yi, f"n={int(totals.reindex(assignees).iloc[yi])}",
                va="center", fontsize=9, color="#666")
    ax.set_yticks(y)
    ax.set_yticklabels(assignees)
    ax.invert_yaxis()
    ax.set_xlim(0, 108)   # leave room for n= annotation
    ax.set_xlabel("%")
    ax.grid(axis="x", linestyle=":", alpha=0.3)
    ax.legend(loc="upper center",
               bbox_to_anchor=(0.5, -0.08),
               ncol=min(len(pct_df.columns), 5),
               fontsize=9, frameon=False,
               title=t("chart_defect_class_col_class"))
    fig.tight_layout()
    return _mpl_save(fig)


def _mpl_chart_assignee_problem_heatmap(ct: pd.DataFrame):
    """Matplotlib version of View 3's Assignee × 問題分類 imshow, used by
    the role-analytics PDF. Mirrors the colour scale and axis orientation
    of the on-screen Plotly chart."""
    if ct is None or ct.empty:
        return None
    plt = _mpl_plt()
    n_rows = len(ct.index)
    n_cols = len(ct.columns)
    fig_h = max(3.0, 0.45 * n_rows + 1.4)
    fig, ax = plt.subplots(figsize=(_MPL_WIDTH_IN, fig_h), dpi=_MPL_DPI)
    im = ax.imshow(ct.values.astype(float), aspect="auto", cmap="YlOrRd")
    ax.set_xticks(np.arange(n_cols))
    ax.set_xticklabels(ct.columns, rotation=-30, ha="left", fontsize=9)
    ax.set_yticks(np.arange(n_rows))
    ax.set_yticklabels(ct.index)
    # Cell text — only when the grid isn't too dense to read.
    if n_rows * n_cols <= 120:
        for r_i in range(n_rows):
            for c_i in range(n_cols):
                v = int(ct.values[r_i, c_i])
                if v:
                    ax.text(c_i, r_i, str(v), ha="center", va="center",
                            fontsize=8,
                            color=("white" if v > ct.values.max() * 0.55
                                   else "#222"))
    ax.set_xlabel(t("chart_defect_class_col_class"))
    ax.set_ylabel(t("col_assignee"))
    fig.colorbar(im, ax=ax, label=t("chart_defect_class_col_count"))
    fig.tight_layout()
    return _mpl_save(fig)


def _mpl_chart_gantt(kpi_df: pd.DataFrame, today_d: date):
    label_planned = t("calendar_layer_planned")
    label_actual = t("calendar_layer_actual")
    rows: list[dict] = []
    for _, row in kpi_df.iterrows():
        label = _clip_label(_label_id_name(row))
        ps = _to_pydate(row.get("planned_start"))
        pe = _to_pydate(row.get("planned_end"))
        ase = _to_pydate(row.get("actual_start"))
        aee = _to_pydate(row.get("actual_end"))
        if ps and pe and pe >= ps:
            rows.append({"ID": label, "Start": ps,
                         "End": pe + timedelta(days=1),
                         "Layer": label_planned})
        if ase:
            end = aee if aee else today_d
            if end < ase:
                end = ase
            rows.append({"ID": label, "Start": ase,
                         "End": end + timedelta(days=1),
                         "Layer": label_actual})
    if not rows:
        return None
    df_g = pd.DataFrame(rows)
    total_ids = df_g["ID"].nunique()
    if total_ids > _BAR_CHART_MAX_ROWS:
        status = (
            df_g.groupby("ID")
                .agg(mn=("Start", "min"), mx=("End", "max"))
                .reset_index()
        )
        status["crosses_today"] = (
            (status["mn"] <= pd.Timestamp(today_d))
            & (status["mx"] >= pd.Timestamp(today_d))
        ).astype(int)
        keep_ids = (status.sort_values(
                        ["crosses_today", "mn"], ascending=[False, True])
                          .head(_BAR_CHART_MAX_ROWS)["ID"].tolist())
        df_g = df_g[df_g["ID"].isin(keep_ids)]

    ids_in_order = list(dict.fromkeys(df_g["ID"].tolist()))
    id_to_y = {i: idx for idx, i in enumerate(ids_in_order)}
    n = len(ids_in_order)

    import matplotlib.dates as mdates
    plt = _mpl_plt()
    fig, ax = plt.subplots(
        figsize=(_MPL_WIDTH_IN, _mpl_bar_height_in(n)), dpi=_MPL_DPI)
    bar_h = 0.38
    for _, r in df_g.iterrows():
        y = id_to_y[r["ID"]]
        start_num = mdates.date2num(pd.Timestamp(r["Start"]))
        end_num = mdates.date2num(pd.Timestamp(r["End"]))
        width_num = max(end_num - start_num, 0.5)
        if r["Layer"] == label_planned:
            color = "#9aa0a6"; y_off = -bar_h / 2 - 0.02
        else:
            color = "#4ec78a"; y_off = bar_h / 2 + 0.02
        ax.barh(y + y_off, width_num, height=bar_h, left=start_num,
                color=color, edgecolor="none")
    today_num = mdates.date2num(pd.Timestamp(today_d))
    ax.axvline(today_num, color="#f5b400", linestyle="--", linewidth=1)
    ax.text(today_num, -0.6, " " + t("gantt_today_label"),
            color="#f5b400", fontsize=10, va="top")
    ax.set_yticks(list(range(n)))
    ax.set_yticklabels(ids_in_order)
    ax.invert_yaxis()
    ax.xaxis_date()
    from matplotlib.patches import Patch
    ax.legend(
        handles=[Patch(color="#9aa0a6", label=label_planned),
                 Patch(color="#4ec78a", label=label_actual)],
        loc="lower right",
    )
    ax.grid(True, axis="x", linestyle=":", alpha=0.3)
    if total_ids > _BAR_CHART_MAX_ROWS:
        _mpl_truncated_title(ax, n, total_ids)
    fig.autofmt_xdate()
    fig.tight_layout()
    return _mpl_save(fig)


# =============================================================================
# PDF report builder
# =============================================================================
_PDF_EMOJI_MAP = {
    "🦕": "", "🦖": "", "📂": "[Source] ", "🧮": "[Formula] ", "💡": "[Tip] ",
}

# Font fallback chain for chart text inside the PDF. Order favors fonts
# typically installed on macOS / Windows / Linux that include CJK glyphs.
_PDF_CHART_FONT = (
    "Hiragino Sans, Hiragino Maru Gothic Pro, "
    "Yu Gothic, Meiryo, MS Gothic, "
    "Noto Sans CJK JP, Noto Sans JP, "
    "DejaVu Sans, Arial, sans-serif"
)


def _style_for_pdf(fig: go.Figure) -> go.Figure:
    """Override the tight on-screen layout the chart builders use, so that
    when the figure is rendered to PNG by kaleido the axis tick labels
    (especially Function IDs and CJK Function names) get the room they need.

    The on-screen `margin=dict(l=10, ...)` setting is intentionally minimal
    for the dashboard but starves the off-screen renderer of label space.
    Here we re-enable `automargin` and set sensible defaults so labels are
    never clipped, and switch to a print-friendly white canvas.
    """
    fig.update_layout(
        font=dict(family=_PDF_CHART_FONT, size=12, color="#222"),
        plot_bgcolor="white",
        paper_bgcolor="white",
        # Fixed margin large enough for clipped CJK labels (36 chars ≈ 290 px
        # at 11-pt CJK). automargin is DISABLED below because its iterative
        # refit pass fires repeatedly per label on tall charts (60 rows ×
        # CJK fallback font chain) and is the dominant cause of kaleido
        # stalling for minutes on the client's real dataset.
        margin=dict(l=300, r=40, t=50, b=60),
        legend=dict(font=dict(family=_PDF_CHART_FONT, size=11, color="#222")),
    )
    fig.update_xaxes(
        automargin=False,
        tickfont=dict(family=_PDF_CHART_FONT, size=11, color="#222"),
        title_font=dict(family=_PDF_CHART_FONT, size=12, color="#222"),
    )
    fig.update_yaxes(
        automargin=False,
        tickfont=dict(family=_PDF_CHART_FONT, size=11, color="#222"),
        title_font=dict(family=_PDF_CHART_FONT, size=12, color="#222"),
    )
    # Heatmap colorbar tick labels also need the font override.
    for tr in fig.data:
        if hasattr(tr, "colorbar") and tr.colorbar is not None:
            tr.colorbar.tickfont = dict(family=_PDF_CHART_FONT, size=11,
                                         color="#222")
    return fig


def _md_to_pdf(text: str) -> str:
    """Lightweight Markdown → reportlab Paragraph markup conversion.

    Converts **bold**, single newlines to <br/>, and strips emoji that the
    embedded CJK font can't render (replacing semantic ones with bracket
    labels)."""
    if not text:
        return ""
    for emoji, repl in _PDF_EMOJI_MAP.items():
        text = text.replace(emoji, repl)
    # Strip any remaining emoji-range code points
    text = re.sub(r"[\U0001F300-\U0001FAFF\u2600-\u27BF]", "", text)
    text = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", text)
    text = text.replace("\n\n", "<br/><br/>").replace("\n", "<br/>")
    return text


PDF_TOTAL_STEPS = 11  # cover + 8 charts + gantt + assemble


def _render_pdf_runner_html(step: int, total: int, msg: str,
                            done: bool = False) -> str:
    """Inner runner content for the st.dialog popup — a velociraptor
    sprinting START → FINISH along a track, advancing one cactus-jump per
    completed PDF-build step. The sprite position reflects step/total; a
    dust-puff keyframe plays on each mount so each update *reads* as the
    dino dashing forward. Pure HTML/CSS — no JS timer needed — which is
    important because Streamlit replaces the placeholder element on every
    update. Dialog chrome (border/shadow/title/✕) is provided by st.dialog.
    (T-Rex is reserved for the page header / favicon.)"""
    pct = 100 if done else int(round(step / max(total, 1) * 100))
    color = "#4ec78a" if done else "#eeeeee"
    runner_uri = dino_data_uri("raptor", color=color)
    cacti = "".join(
        f'<div class="d4dx-pdf-cactus" style="left:{(i/total)*100:.1f}%;"></div>'
        for i in range(1, total)
    )
    sub = ("🏁 " + msg) if done else f"{msg}  ·  {step}/{total}"
    return f'''
<style>
.d4dx-pdf-track {{ position:relative; height:60px; margin:10px 12px 0 40px; }}
.d4dx-pdf-track::before {{ content:""; position:absolute; left:0; right:0;
                           top:40px; height:3px; background:#2c3138;
                           border-radius:2px; }}
.d4dx-pdf-bar   {{ position:absolute; left:0; top:40px; height:3px;
                   background:#4ec78a; border-radius:2px; width:{pct}%;
                   box-shadow:0 0 8px rgba(78,199,138,0.6); }}
.d4dx-pdf-flag-s, .d4dx-pdf-flag-e {{ position:absolute; top:34px;
                                      font-size:10px; font-weight:700;
                                      color:#888; letter-spacing:1px; }}
.d4dx-pdf-flag-s {{ left:-38px; }}
.d4dx-pdf-flag-e {{ right:-24px; font-size:18px; top:28px; }}
.d4dx-pdf-cactus {{ position:absolute; top:32px; width:4px; height:12px;
                    background:#6a8e3a; border-radius:1px;
                    transform:translateX(-2px); }}
.d4dx-pdf-cactus::before {{ content:""; position:absolute; left:-3px; top:3px;
                            width:3px; height:5px; background:#6a8e3a;
                            border-radius:1px; }}
.d4dx-pdf-cactus::after {{ content:""; position:absolute; right:-3px; top:3px;
                           width:3px; height:5px; background:#6a8e3a;
                           border-radius:1px; }}
.d4dx-pdf-dino {{ position:absolute; top:6px; width:36px; height:36px;
                  left:calc({pct}% - 18px);
                  animation:d4dx-pdf-bounce 0.7s ease-in-out infinite;
                  filter:drop-shadow(0 2px 2px rgba(0,0,0,0.3)); }}
.d4dx-pdf-dino.done {{ animation:d4dx-pdf-cheer 0.9s ease-in-out infinite; }}
.d4dx-pdf-dust {{ position:absolute; top:26px; width:16px; height:10px;
                  left:calc({pct}% - 28px); opacity:0.9;
                  animation:d4dx-pdf-dust 0.65s ease-out forwards;
                  pointer-events:none; }}
.d4dx-pdf-dust span {{ display:inline-block; width:4px; height:4px;
                       margin-right:2px; border-radius:50%; background:#b48820; }}
@keyframes d4dx-pdf-bounce {{
  0%,100% {{ transform:translateY(0);    }}
  45%     {{ transform:translateY(-6px) rotate(-3deg); }}
  55%     {{ transform:translateY(-6px) rotate(3deg); }}
}}
@keyframes d4dx-pdf-cheer {{
  0%,100% {{ transform:translateY(0) scale(1);    }}
  50%     {{ transform:translateY(-5px) scale(1.12); }}
}}
@keyframes d4dx-pdf-dust {{
  0%   {{ transform:translateX(0)    scale(1);   opacity:0.9; }}
  100% {{ transform:translateX(-18px) scale(1.6); opacity:0;   }}
}}
.d4dx-pdf-caption {{ margin-top:18px; text-align:center; font-size:12.5px;
                     color:#dcdcdc; font-family:"SF Mono",Menlo,monospace;
                     letter-spacing:0.3px; }}
</style>
<div class="d4dx-pdf-track">
  <span class="d4dx-pdf-flag-s">START</span>
  {cacti}
  <span class="d4dx-pdf-flag-e">🏁</span>
  <div class="d4dx-pdf-bar"></div>
  <div class="d4dx-pdf-dust"><span></span><span></span><span></span></div>
  <img class="d4dx-pdf-dino {'done' if done else ''}" src="{runner_uri}"
       alt="raptor" width="36" height="36"/>
</div>
<div class="d4dx-pdf-caption">{sub}</div>
'''


def generate_report_pdf(
    kpi_df: pd.DataFrame,
    progress_cb: Optional[Callable[[str, int, int], None]] = None,
    defects_df: Optional[pd.DataFrame] = None,
) -> bytes:
    """Build a PDF report containing the project KPI summary plus every
    available chart and the Gantt schedule, with definition text. Tables of
    raw data are intentionally excluded.

    `progress_cb`, when provided, is called with (msg, step, total) at each
    major step so the caller can animate a progress UI.
    `defects_df` overrides the session-state defect dataframe; the caller
    passes a pre-filtered frame (matching the selected Function IDs) so
    the bug-trend chart stays consistent with the per-feature charts."""
    logger = _get_logger()
    logger.info(
        "[pdf_export] enter generate_report_pdf: "
        f"kpi_rows={len(kpi_df)} "
        f"defect_rows={0 if defects_df is None else len(defects_df)} "
        f"lang={st.session_state.get('lang')}"
    )
    step_counter = [0]
    def _progress(msg: str) -> None:
        step_counter[0] += 1
        logger.info(
            f"[pdf_export] step {step_counter[0]}/{PDF_TOTAL_STEPS}: {msg}")
        if progress_cb is not None:
            progress_cb(msg, step_counter[0], PDF_TOTAL_STEPS)
    t_start = time.time()
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import (
        Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table,
        TableStyle,
    )
    from reportlab.platypus.tableofcontents import TableOfContents
    logger.info(
        f"[pdf_export] reportlab imports done in "
        f"{time.time() - t_start:.2f}s")

    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
    JP_FONT = "HeiseiKakuGo-W5"
    # A3 landscape gives ~42 cm × 29.7 cm — twice the usable width of A4
    # portrait, so long Function IDs and crowded chart axes don't get squashed.
    page_size = landscape(A3)
    page_w, _ = page_size
    inner_w = page_w - 3 * cm  # 1.5 cm margins each side

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "PdfTitle", parent=styles["Title"], fontName=JP_FONT,
        fontSize=22, alignment=1, spaceAfter=14,
    )
    h2_style = ParagraphStyle(
        "PdfH2", parent=styles["Heading2"], fontName=JP_FONT,
        fontSize=14, spaceAfter=8, spaceBefore=4,
    )
    h3_style = ParagraphStyle(
        "PdfH3", parent=styles["Heading3"], fontName=JP_FONT,
        fontSize=11, spaceAfter=6, textColor=colors.HexColor("#3aa872"),
    )
    body_style = ParagraphStyle(
        "PdfBody", parent=styles["Normal"], fontName=JP_FONT,
        fontSize=10, leading=14,
    )
    caption_style = ParagraphStyle(
        "PdfCaption", parent=styles["Normal"], fontName=JP_FONT,
        fontSize=8, textColor=colors.grey,
    )
    # Cover + TOC styles (consistent with the two A4 standalone reports).
    cover_title_style = ParagraphStyle(
        "PdfCoverTitle", parent=styles["Title"], fontName=JP_FONT,
        fontSize=28, alignment=1, spaceAfter=20,
    )
    cover_meta_style = ParagraphStyle(
        "PdfCoverMeta", parent=styles["Normal"], fontName=JP_FONT,
        fontSize=13, alignment=1, textColor=colors.grey, leading=20,
    )
    toc_title_style = ParagraphStyle(
        "PdfTocTitle", parent=styles["Heading1"], fontName=JP_FONT,
        fontSize=20, alignment=1, spaceAfter=20,
        textColor=colors.HexColor("#2d6b4f"),
    )
    toc_entry_style = ParagraphStyle(
        "PdfTocEntry", fontName=JP_FONT, fontSize=12,
        leftIndent=30, firstLineIndent=-14, spaceBefore=4, leading=20,
    )

    buf = io.BytesIO()

    # See _TdDoc in generate_test_density_pdf for the TOC mechanism.
    class _PdfDoc(SimpleDocTemplate):
        def afterFlowable(self, flowable):  # type: ignore[override]
            if (isinstance(flowable, Paragraph)
                    and flowable.style.name == "PdfH2"):
                self.notify(
                    "TOCEntry",
                    (0, flowable.getPlainText(), self.page),
                )

    doc = _PdfDoc(
        buf, pagesize=page_size,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    story: list = []

    # --- Page 1: Cover ------------------------------------------------------
    _progress(t("pdf_step_cover"))
    cover_icon = Image(
        io.BytesIO(_pixel_icon_png("bronto", scale=12)),
        width=140, height=112,
    )
    cover_icon.hAlign = "CENTER"
    # A3 landscape is tall; push the block toward the vertical centre.
    story.append(Spacer(1, 120))
    story.append(cover_icon)
    story.append(Spacer(1, 28))
    story.append(Paragraph(t("pdf_title"), cover_title_style))
    story.append(Spacer(1, 12))
    story.append(Paragraph(
        f"{t('pdf_generated_at')}: "
        f"{datetime.now().strftime('%Y-%m-%d %H:%M')}",
        cover_meta_style,
    ))
    story.append(PageBreak())

    # --- Page 2: Table of contents ------------------------------------------
    story.append(Paragraph(t("pdf_toc_title"), toc_title_style))
    toc = TableOfContents()
    toc.levelStyles = [toc_entry_style]
    story.append(toc)
    story.append(PageBreak())

    # --- Page 3+: KPI summary + chart sections + schedule -------------------
    # KPI summary now lives on its own content page instead of sharing the
    # cover — gives it room to breathe alongside the nine KPIs.
    summary = project_kpi_summary(kpi_df)
    def _pct(v): return f"{v * 100:.1f}%" if v is not None else "—"
    def _f3(v):  return f"{v:.3f}" if v is not None else "—"
    def _f2(v):  return f"{v:.2f}" if v is not None else "—"
    kpi_rows = [
        [t("metric_total_loc"),       f"{summary['total_loc']:,}"],
        [t("metric_open_defects"),    f"{summary['open_defects']:,}"],
        [t("metric_test_run_rate"),   _pct(summary["run_rate"])],
        [t("metric_test_pass_rate"),  _pct(summary["pass_rate"])],
        [t("metric_avg_bug_density"), _f3(summary["avg_bug_density"])],
        [t("metric_avg_test_density"), _f2(summary["avg_test_density"])],
        [t("metric_avg_health"),      _f2(summary["avg_health"])],
        [t("metric_at_risk"),         f"{summary['at_risk_count']}"],
        [t("metric_delayed"),         f"{summary['delayed_count']}"],
    ]
    story.append(Paragraph(t("pdf_section_kpi"), h2_style))
    kpi_table = Table(kpi_rows, colWidths=[inner_w * 0.55, inner_w * 0.30])
    kpi_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), JP_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("LEFTPADDING",  (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f0f4f0")),
        ("ALIGN",  (1, 0), (1, -1), "RIGHT"),
        ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.lightgrey),
    ]))
    story.append(kpi_table)
    story.append(PageBreak())

    # --- Chart sections -----------------------------------------------------
    today_d = date.today()
    if defects_df is None:
        defects_df = st.session_state.dfs.get("defects")

    # Each builder returns (png_bytes, intrinsic_w_px, intrinsic_h_px) or None.
    # These are matplotlib-rendered; see _mpl_chart_* above.
    chart_specs: list[tuple[str, str,
                             Callable[[], Optional[tuple[bytes, int, int]]]]] = [
        ("chart_progress_gap",    "help_chart_progress_gap",
         lambda: _mpl_chart_progress_gap(kpi_df)),
        ("chart_test_coverage",   "help_chart_test_coverage",
         lambda: _mpl_chart_test_coverage(kpi_df)),
        ("chart_test_density",    "help_chart_test_density",
         lambda: _mpl_chart_test_density(kpi_df)),
        ("chart_incident_rate",   "help_chart_incident_rate",
         lambda: _mpl_chart_incident_rate(kpi_df)),
        ("chart_loc_vs_ng",       "help_chart_loc_vs_ng",
         lambda: _mpl_chart_loc_vs_ng(kpi_df)),
        ("chart_design_impl_gap", "help_chart_design_impl_gap",
         lambda: _mpl_chart_design_impl_gap(kpi_df)),
        ("chart_risk_heatmap",    "help_chart_risk_heatmap",
         lambda: _mpl_chart_risk_heatmap(kpi_df)),
        ("chart_loc_trend",       "help_chart_loc_trend",  _mpl_chart_loc_trend),
        ("chart_test_trend",      "help_chart_test_trend", _mpl_chart_test_trend),
        ("chart_bug_trend",       "help_chart_bug_trend",
         lambda: _mpl_chart_bug_trend(defects_df)),
        ("chart_defect_class",    "help_chart_defect_class",
         lambda: _mpl_chart_defect_class(defects_df)),
    ]

    story.append(Paragraph(t("pdf_section_charts"), h2_style))

    max_chart_h = 22 * cm  # leaves room for section title + definition above

    def embed_chart(png: bytes, w_px: int, h_px: int, label: str = "") -> None:
        t0 = time.time()
        aspect = h_px / w_px if w_px else 0.5
        disp_w = inner_w
        disp_h = disp_w * aspect
        if disp_h > max_chart_h:
            disp_h = max_chart_h
            disp_w = disp_h / aspect
        story.append(Image(io.BytesIO(png), width=disp_w, height=disp_h))
        logger.info(
            f"[pdf_export] chart embedded: {label or '?'} — "
            f"intrinsic={w_px}x{h_px}px size={len(png)//1024}KB "
            f"elapsed={time.time()-t0:.2f}s"
        )

    n_charts = len(chart_specs)
    for i, (title_key, help_key, builder) in enumerate(chart_specs, start=1):
        _progress(t("pdf_step_chart", i=i, n=n_charts, title=t(title_key)))
        result = builder()
        story.append(Paragraph(t(title_key), h2_style))
        story.append(Paragraph(t("pdf_chart_definition"), h3_style))
        story.append(Paragraph(_md_to_pdf(t(help_key)), body_style))
        story.append(Spacer(1, 6))
        if result is None:
            story.append(Paragraph(t("pdf_no_chart"), caption_style))
        else:
            png_bytes, w_px, h_px = result
            embed_chart(png_bytes, w_px, h_px, label=title_key)
        story.append(PageBreak())

    # --- Schedule (Gantt) ---------------------------------------------------
    _progress(t("pdf_step_gantt"))
    story.append(Paragraph(t("pdf_section_schedule"), h2_style))
    story.append(Paragraph(t("gantt_title"), h2_style))
    story.append(Paragraph(t("pdf_chart_definition"), h3_style))
    story.append(Paragraph(_md_to_pdf(t("help_gantt_title")), body_style))
    story.append(Spacer(1, 6))
    result = _mpl_chart_gantt(kpi_df, today_d)
    if result is None:
        story.append(Paragraph(t("pdf_no_chart"), caption_style))
    else:
        png_bytes, w_px, h_px = result
        embed_chart(png_bytes, w_px, h_px, label="gantt_title")
    # Calendar visual itself is FullCalendar (not exportable); explain that the
    # Gantt above + the calendar's data definition cover the same source data.
    story.append(Spacer(1, 10))
    story.append(Paragraph(t("calendar_title"), h2_style))
    story.append(Paragraph(t("pdf_chart_definition"), h3_style))
    story.append(Paragraph(_md_to_pdf(t("help_calendar_title")), body_style))

    _progress(t("pdf_step_assemble"))
    _footer_cb = _pdf_apply_chrome(story, styles, JP_FONT)
    doc.multiBuild(story, onFirstPage=_footer_cb, onLaterPages=_footer_cb)
    pdf = buf.getvalue()
    buf.close()
    return pdf


def generate_test_density_pdf(
    kpi_df: pd.DataFrame,
    fid_filter_active: bool = False,
) -> bytes:
    """Standalone PDF for the 機能ID別テスト密度 chart.

    Layout: title (+ red banner when the global FID filter is active) →
    input-source table → output formula → summary → bar chart PNG →
    below-threshold list → catch-up test estimate → fixed advisory → notes.
    A4 portrait; ReportLab auto-paginates when the tables grow long.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import (
        Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table,
        TableStyle,
    )
    from reportlab.platypus.tableofcontents import TableOfContents

    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
    JP_FONT = "HeiseiKakuGo-W5"

    page_size = A4
    page_w, _ = page_size
    inner_w = page_w - 3 * cm  # 1.5 cm margins each side

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TdTitle", parent=styles["Title"], fontName=JP_FONT,
        fontSize=18, alignment=1, spaceAfter=8,
    )
    filter_style = ParagraphStyle(
        "TdFilter", parent=styles["Normal"], fontName=JP_FONT,
        fontSize=11, alignment=1, textColor=colors.red,
        spaceBefore=2, spaceAfter=10,
    )
    meta_style = ParagraphStyle(
        "TdMeta", parent=styles["Normal"], fontName=JP_FONT,
        fontSize=9, alignment=1, textColor=colors.grey, spaceAfter=16,
    )
    h2_style = ParagraphStyle(
        "TdH2", parent=styles["Heading2"], fontName=JP_FONT,
        fontSize=12, spaceAfter=2, spaceBefore=0,
        textColor=colors.HexColor("#2d6b4f"),
    )
    body_style = ParagraphStyle(
        "TdBody", parent=styles["Normal"], fontName=JP_FONT,
        fontSize=9, leading=15,
    )
    caption_style = ParagraphStyle(
        "TdCaption", parent=styles["Normal"], fontName=JP_FONT,
        fontSize=8, textColor=colors.grey, leading=12,
    )
    # Cover-page styles — larger fonts + more breathing room than the body.
    cover_title_style = ParagraphStyle(
        "TdCoverTitle", parent=styles["Title"], fontName=JP_FONT,
        fontSize=24, alignment=1, spaceAfter=16,
    )
    cover_meta_style = ParagraphStyle(
        "TdCoverMeta", parent=styles["Normal"], fontName=JP_FONT,
        fontSize=12, alignment=1, textColor=colors.grey,
        spaceAfter=6, leading=18,
    )
    toc_title_style = ParagraphStyle(
        "TdTocTitle", parent=styles["Heading1"], fontName=JP_FONT,
        fontSize=18, alignment=1, spaceAfter=18,
        textColor=colors.HexColor("#2d6b4f"),
    )
    toc_entry_style = ParagraphStyle(
        "TdTocEntry", fontName=JP_FONT, fontSize=11,
        leftIndent=24, firstLineIndent=-12, spaceBefore=4, leading=18,
    )

    buf = io.BytesIO()

    # SimpleDocTemplate subclass that auto-registers a TOC entry every
    # time a section-heading paragraph (style name "TdH2") is rendered.
    # ReportLab's afterFlowable runs post-layout, so `self.page` holds
    # the real, final page number — which is why we need `multiBuild`
    # below: the first pass records entries at their initial pages, the
    # TOC flowable then renders with those entries, pages shift, and a
    # second pass rebuilds with stable numbers.
    class _TdDoc(SimpleDocTemplate):
        def afterFlowable(self, flowable):  # type: ignore[override]
            if (isinstance(flowable, Paragraph)
                    and flowable.style.name == "TdH2"):
                self.notify(
                    "TOCEntry",
                    (0, flowable.getPlainText(), self.page),
                )

    doc = _TdDoc(
        buf, pagesize=page_size,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.6 * cm, bottomMargin=1.6 * cm,
    )
    story: list = []

    # Short-lived image builder for in-PDF pixel icons; caches bytes so the
    # same accent used twice (e.g. "egg" for output + notes) only rasterizes
    # once per build.
    _icon_cache: dict[tuple[str, int], bytes] = {}

    def _icon_image(name: str, size_pt: float, scale: int = 4) -> Image:
        key = (name, scale)
        if key not in _icon_cache:
            _icon_cache[key] = _pixel_icon_png(name, scale=scale)
        # Square-ish render; actual aspect follows the grid. T-Rex is
        # reserved for the page chrome, so the fallback is a neutral dino.
        grid = (DINO_GRIDS.get(name)
                or DINO_GRIDS[_DEFAULT_SECTION_DINO])
        rows = [r for r in grid.strip("\n").split("\n") if r]
        gh = len(rows)
        gw = max(len(r) for r in rows)
        aspect = gh / gw
        return Image(io.BytesIO(_icon_cache[key]),
                     width=size_pt, height=size_pt * aspect)

    def _section_heading(story_list, icon_name: str, text: str) -> None:
        """Append a section heading. `icon_name` is accepted for
        backwards-compat with older call sites but intentionally ignored —
        inline accents next to every heading were too busy; the era-icon
        decoration now lives only in the report's footer strip."""
        del icon_name  # kept in signature for compatibility
        story_list.append(Spacer(1, 14))
        story_list.append(Paragraph(text, h2_style))

    # Resolve source filenames up-front so the cover page can reference them.
    tests_origin = (
        st.session_state.get("origin_names", {}).get("tests") or "—"
    )
    master_origin = (
        st.session_state.get("origin_names", {}).get("master") or "—"
    )

    # --- Page 1: Cover ------------------------------------------------------
    # A dedicated first page with the report title, filter banner, and the
    # generated-at / snapshot-source metadata. T-Rex stays reserved for the
    # app chrome; the cover hero icon is a bronto silhouette.
    cover_icon = Image(
        io.BytesIO(_pixel_icon_png("bronto", scale=10)),
        width=110, height=88,
    )
    cover_icon.hAlign = "CENTER"
    story.append(Spacer(1, 130))  # vertical centering-ish
    story.append(cover_icon)
    story.append(Spacer(1, 24))
    story.append(Paragraph(t("td_pdf_title"), cover_title_style))
    if fid_filter_active:
        story.append(Paragraph(t("td_pdf_filter_active"), filter_style))
    story.append(Spacer(1, 12))
    story.append(Paragraph(
        t("td_pdf_footer",
          when=datetime.now().strftime("%Y-%m-%d %H:%M"),
          src=tests_origin),
        cover_meta_style,
    ))
    story.append(PageBreak())

    # --- Page 2: Table of contents ------------------------------------------
    # Populated automatically via afterFlowable on h2_style ("TdH2")
    # paragraphs — every _section_heading call registers an entry.
    story.append(Paragraph(t("pdf_toc_title"), toc_title_style))
    toc = TableOfContents()
    toc.levelStyles = [toc_entry_style]
    story.append(toc)
    story.append(PageBreak())

    # --- Page 3+: Content sections ------------------------------------------
    # --- Inputs table -------------------------------------------------------
    _section_heading(story, "fossil", t("td_pdf_h_inputs"))
    input_rows = [
        [t("td_pdf_col_item"), t("td_pdf_col_source"),
         t("td_pdf_col_method")],
        [t("td_pdf_input_tests"),
         f"{tests_origin}<br/><font size=8 color='grey'>"
         f"{t('td_pdf_input_tests_note')}</font>",
         t("td_pdf_method_auto")],
        [t("td_pdf_input_pages"),
         f"design_pages.json<br/><font size=8 color='grey'>"
         f"{t('td_pdf_input_pages_note')}</font>",
         t("td_pdf_method_manual")],
        [t("td_pdf_input_master"),
         master_origin,
         t("td_pdf_method_auto")],
    ]
    input_rows = [
        [Paragraph(str(c), body_style) for c in row]
        for row in input_rows
    ]
    input_tbl = Table(
        input_rows,
        colWidths=[inner_w * 0.25, inner_w * 0.50, inner_w * 0.25],
    )
    input_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), JP_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8efe8")),
        ("VALIGN",    (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.grey),
        ("GRID",      (0, 0), (-1, -1), 0.25, colors.lightgrey),
    ]))
    story.append(input_tbl)

    # --- Output formula -----------------------------------------------------
    _section_heading(story, "egg", t("td_pdf_h_output"))
    story.append(Paragraph(t("td_pdf_output_formula"), body_style))
    story.append(Spacer(1, 3))
    story.append(Paragraph(t("td_pdf_output_unit"), body_style))
    story.append(Spacer(1, 3))
    story.append(Paragraph(t("td_pdf_output_meaning"), body_style))

    # --- Threshold settings -------------------------------------------------
    # Own section (was previously only a row buried inside the summary table
    # and a vline annotation on the chart). Readers asked for an explicit
    # "here's what threshold this report was built against" callout.
    threshold = _test_density_threshold()
    _section_heading(story, "volcano", t("td_pdf_h_threshold"))
    thr_unit = t("td_pdf_threshold_unit")
    thr_rows = [
        [t("td_pdf_threshold_current"),
         f"<b>{threshold:g}</b> {thr_unit}"],
        [t("td_pdf_threshold_default"),
         f"{TEST_DENSITY_THRESHOLD_DEFAULT:g} {thr_unit}"],
        [t("td_pdf_threshold_where"),
         t("td_pdf_threshold_where_value")],
    ]
    thr_rows = [
        [Paragraph(str(c), body_style) for c in row] for row in thr_rows
    ]
    thr_tbl = Table(thr_rows, colWidths=[inner_w * 0.32, inner_w * 0.68])
    thr_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), JP_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#fbe6e6")),
        ("TEXTCOLOR",  (0, 0), (0, -1), colors.HexColor("#a02020")),
        ("LEFTPADDING",  (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(thr_tbl)
    story.append(Spacer(1, 4))
    story.append(Paragraph(t("td_pdf_threshold_meaning"), body_style))

    # --- Summary ------------------------------------------------------------
    df = kpi_df.copy()
    if "test_density" in df.columns:
        df = df.dropna(subset=["test_density"])
    else:
        df = df.iloc[0:0]

    if df.empty:
        _section_heading(story, "volcano", t("td_pdf_h_summary"))
        story.append(Paragraph(t("pdf_no_chart"), caption_style))
        _footer_cb = _pdf_apply_chrome(story, styles, JP_FONT)
        doc.multiBuild(story, onFirstPage=_footer_cb,
                       onLaterPages=_footer_cb)
        pdf = buf.getvalue()
        buf.close()
        return pdf

    densities = df["test_density"].astype(float)
    below_mask = densities < threshold
    n_total = int(len(df))
    n_below = int(below_mask.sum())
    n_above = n_total - n_below

    def _pct(n: int) -> str:
        return f"{n / n_total * 100:.1f}%" if n_total else "—"

    _section_heading(story, "volcano", t("td_pdf_h_summary"))
    # No "threshold" row here any more — the 警告閾値 section above owns
    # that callout. Summary stays focused on counts + density stats.
    sum_rows = [
        [t("td_pdf_summary_total"),     f"{n_total}"],
        [t("td_pdf_summary_above"),     f"{n_above} ({_pct(n_above)})"],
        [t("td_pdf_summary_below"),     f"{n_below} ({_pct(n_below)})"],
        [t("td_pdf_summary_mean"),      f"{float(densities.mean()):.2f}"],
        [t("td_pdf_summary_median"),    f"{float(densities.median()):.2f}"],
        [t("td_pdf_summary_min_max"),
         f"{float(densities.min()):.2f} / {float(densities.max()):.2f}"],
    ]
    sum_tbl = Table(sum_rows, colWidths=[inner_w * 0.55, inner_w * 0.45])
    # Row index of the "below threshold" entry (after removing the threshold
    # row above it). Highlighted only when there's at least one such feature.
    below_row_idx = 2
    sum_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), JP_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f0f4f0")),
        ("ALIGN",  (1, 0), (1, -1), "RIGHT"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING",   (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        # Highlight the "below threshold" row so it stands out.
        ("TEXTCOLOR", (1, below_row_idx), (1, below_row_idx),
         colors.HexColor("#a02020") if n_below else colors.black),
        ("FONTNAME", (1, below_row_idx), (1, below_row_idx), JP_FONT),
    ]))
    story.append(sum_tbl)

    # --- Chart --------------------------------------------------------------
    _section_heading(story, "palm", t("td_pdf_h_chart"))
    mpl_result = _mpl_chart_test_density(kpi_df)
    if mpl_result is None:
        story.append(Paragraph(t("pdf_no_chart"), caption_style))
    else:
        png, w_px, h_px = mpl_result
        aspect = h_px / w_px if w_px else 0.5
        disp_w = inner_w
        disp_h = disp_w * aspect
        # Hard-cap the chart height so it doesn't shove the tables off the
        # first-page block; ReportLab will paginate afterwards.
        max_h = 18 * cm
        if disp_h > max_h:
            disp_h = max_h
            disp_w = disp_h / aspect
        story.append(Image(io.BytesIO(png), width=disp_w, height=disp_h))

    # --- Below-threshold list ----------------------------------------------
    below_df = (
        df[below_mask][["機能ID", "機能名称", "総設定テスト数",
                        "設計書ページ数", "test_density"]]
        .sort_values("test_density", ascending=True)
        .reset_index(drop=True)
    )
    _section_heading(
        story, "footprint", f"{t('td_pdf_h_below')} ({n_below})"
    )
    if below_df.empty:
        story.append(Paragraph(t("td_pdf_below_none"), body_style))
    else:
        header = [
            t("td_pdf_col_fid"), t("td_pdf_col_name"),
            t("td_pdf_col_tests"), t("td_pdf_col_pages"),
            t("td_pdf_col_density"), t("td_pdf_col_gap"),
        ]
        rows = [[Paragraph(h, body_style) for h in header]]
        for _, r in below_df.iterrows():
            tests_v = r.get("総設定テスト数")
            pages_v = r.get("設計書ページ数")
            rows.append([
                Paragraph(str(r["機能ID"]), body_style),
                Paragraph(str(r.get("機能名称") or ""), body_style),
                Paragraph(
                    f"{int(tests_v):,}" if pd.notna(tests_v) else "—",
                    body_style),
                Paragraph(
                    f"{float(pages_v):g}" if pd.notna(pages_v) else "—",
                    body_style),
                Paragraph(f"{float(r['test_density']):.2f}", body_style),
                Paragraph(
                    f"{float(r['test_density']) - threshold:+.2f}",
                    body_style),
            ])
        below_tbl = Table(
            rows,
            colWidths=[
                inner_w * 0.10, inner_w * 0.32, inner_w * 0.12,
                inner_w * 0.16, inner_w * 0.13, inner_w * 0.17,
            ],
            repeatRows=1,
        )
        below_tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), JP_FONT),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fbe6e6")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.HexColor("#a02020")),
            ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(below_tbl)

    # --- Catch-up estimate (only if there is something to fix) -------------
    if not below_df.empty:
        _section_heading(story, "fern", t("td_pdf_h_catchup"))
        header2 = [
            t("td_pdf_col_fid"), t("td_pdf_col_name"),
            t("td_pdf_col_current"), t("td_pdf_col_target"),
            t("td_pdf_col_additional"),
        ]
        rows2 = [[Paragraph(h, body_style) for h in header2]]
        for _, r in below_df.iterrows():
            pages_v = r.get("設計書ページ数")
            if pd.isna(pages_v) or float(pages_v) <= 0:
                additional = "—"
            else:
                gap = max(0.0, threshold - float(r["test_density"]))
                additional = f"+{math.ceil(gap * float(pages_v))}"
            rows2.append([
                Paragraph(str(r["機能ID"]), body_style),
                Paragraph(str(r.get("機能名称") or ""), body_style),
                Paragraph(f"{float(r['test_density']):.2f}", body_style),
                Paragraph(f"{threshold:g}", body_style),
                Paragraph(additional, body_style),
            ])
        catchup_tbl = Table(
            rows2,
            colWidths=[
                inner_w * 0.10, inner_w * 0.40, inner_w * 0.13,
                inner_w * 0.10, inner_w * 0.27,
            ],
            repeatRows=1,
        )
        catchup_tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), JP_FONT),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef4ec")),
            ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ]))
        story.append(catchup_tbl)
        story.append(Spacer(1, 6))
        story.append(Paragraph(t("td_pdf_catchup_note"), caption_style))

    # --- Advice + notes -----------------------------------------------------
    _section_heading(story, "raptor", t("td_pdf_h_advice"))
    story.append(Paragraph(t("td_pdf_advice_body"), body_style))

    _section_heading(story, "egg", t("td_pdf_h_notes"))
    story.append(Paragraph(t("td_pdf_notes_body"), body_style))
    story.append(Spacer(1, 18))

    # --- Footer accent: a small row of era icons centered --------------
    footer_icons = [
        _icon_image("volcano", 12),
        _icon_image("fossil", 12),
        _icon_image("palm", 12),
        _icon_image("egg", 12),
        _icon_image("footprint", 12),
        _icon_image("fern", 12),
    ]
    # Interleave icons with narrow spacer cells so they don't crowd.
    footer_cells = []
    for i, ic in enumerate(footer_icons):
        if i:
            footer_cells.append("")
        footer_cells.append(ic)
    footer_tbl = Table(
        [footer_cells],
        colWidths=[14 if c == "" else 16 for c in footer_cells],
    )
    footer_tbl.setStyle(TableStyle([
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("LEFTPADDING",   (0, 0), (-1, -1), 0),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
        ("TOPPADDING",    (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    footer_tbl.hAlign = "CENTER"
    story.append(footer_tbl)

    _footer_cb = _pdf_apply_chrome(story, styles, JP_FONT)
    doc.multiBuild(story, onFirstPage=_footer_cb, onLaterPages=_footer_cb)
    pdf = buf.getvalue()
    buf.close()
    return pdf


def _md_bold_html(text: str) -> str:
    """Convert markdown `**bold**` to ReportLab's HTML-subset `<b>bold</b>`
    so translation strings that use markdown for emphasis can be reused
    as Paragraph content in the PDF without the stars leaking through."""
    return re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", text)


def _append_watchlist_pdf_section(
    story: list,
    df: pd.DataFrame,
    *,
    title_key: str,
    caption_key: str,
    h2_style,
    caption_style,
    small_body,
    inner_w: float,
    hide_defect_col: bool = False,
) -> None:
    """Render one of the two bubble watch-lists into the PDF `story`.

    Mirrors `_render_bubble_watchlist` (on-screen) so the PDF carries the
    same title / definition caption / compact table, with the Defects
    column suppressed for the no-exec bucket (always zero and would
    contradict the caption's "quality cannot be measured" framing)."""
    from reportlab.lib import colors
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle

    story.append(Spacer(1, 8))
    story.append(Paragraph(t(title_key), h2_style))
    story.append(Paragraph(_md_bold_html(t(caption_key)), caption_style))
    story.append(Spacer(1, 3))
    if df is None or df.empty:
        story.append(Paragraph(t("role_analytics_watch_empty"), caption_style))
        return
    role_labels_local = {
        "dev":       t("role_dev"),
        "test_spec": t("role_test_spec"),
        "test_exec": t("role_test_exec"),
    }
    header = [
        t("role_analytics_watch_col_assignee"),
        t("role_analytics_watch_col_features"),
    ]
    if not hide_defect_col:
        header.append(t("role_analytics_watch_col_defects"))
    header.append(t("role_analytics_watch_col_roles"))

    rows = [[Paragraph(h, small_body) for h in header]]
    for _, r in df.iterrows():
        role_bits = " · ".join(
            f"{role_labels_local[k]}: "
            f"{int(r.get(f'role_count_{k}', 0) or 0)}"
            for k in ("dev", "test_spec", "test_exec")
        )
        cells = [
            Paragraph(str(r["assignee"]), small_body),
            Paragraph(str(int(r["feature_count"])), small_body),
        ]
        if not hide_defect_col:
            cells.append(Paragraph(str(int(r["defect_total"])), small_body))
        cells.append(Paragraph(role_bits, small_body))
        rows.append(cells)
    # Column widths: assignee / features / (defects) / roles — roles
    # takes the lion's share since it's the only wrappable column.
    if hide_defect_col:
        col_ws = [inner_w * 0.22, inner_w * 0.14, inner_w * 0.64]
    else:
        col_ws = [inner_w * 0.20, inner_w * 0.12, inner_w * 0.12,
                  inner_w * 0.56]
    tbl = Table(rows, colWidths=col_ws, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fff4e5")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("ALIGN", (1, 1), (-2, -1), "RIGHT"),
    ]))
    story.append(tbl)


def generate_role_analytics_pdf(
    kpi_df: pd.DataFrame,
    wbs_df: Optional[pd.DataFrame],
    defects_df: Optional[pd.DataFrame],
    fid_filter_active: bool = False,
) -> bytes:
    """Standalone PDF for the 担当者×ロール分析 section.

    Layout: title (+ filter banner) → inputs table → role-classification
    rules → View 1 table (feature × assignees × KPIs) → View 2 table
    (assignee summary) → bubble map PNG → problem-class strip PNG →
    View 3 heatmap PNG → era-icon footer. A4 portrait; ReportLab
    auto-paginates where the tables / charts overflow."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import (
        Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table,
        TableStyle,
    )
    from reportlab.platypus.tableofcontents import TableOfContents

    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
    JP_FONT = "HeiseiKakuGo-W5"

    page_size = A4
    page_w, _ = page_size
    inner_w = page_w - 3 * cm

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "RaTitle", parent=styles["Title"], fontName=JP_FONT,
        fontSize=18, alignment=1, spaceAfter=8,
    )
    filter_style = ParagraphStyle(
        "RaFilter", parent=styles["Normal"], fontName=JP_FONT,
        fontSize=11, alignment=1, textColor=colors.red,
        spaceBefore=2, spaceAfter=10,
    )
    meta_style = ParagraphStyle(
        "RaMeta", parent=styles["Normal"], fontName=JP_FONT,
        fontSize=9, alignment=1, textColor=colors.grey, spaceAfter=14,
    )
    h2_style = ParagraphStyle(
        "RaH2", parent=styles["Heading2"], fontName=JP_FONT,
        fontSize=12, spaceAfter=4, spaceBefore=14,
        textColor=colors.HexColor("#2d6b4f"),
    )
    body_style = ParagraphStyle(
        "RaBody", parent=styles["Normal"], fontName=JP_FONT,
        fontSize=9, leading=14,
    )
    caption_style = ParagraphStyle(
        "RaCaption", parent=styles["Normal"], fontName=JP_FONT,
        fontSize=8, textColor=colors.grey, leading=12,
    )
    small_body = ParagraphStyle(
        "RaSmallBody", parent=styles["Normal"], fontName=JP_FONT,
        fontSize=8, leading=11,
    )
    # Cover + TOC styles — mirror the test-density PDF for consistency.
    cover_title_style = ParagraphStyle(
        "RaCoverTitle", parent=styles["Title"], fontName=JP_FONT,
        fontSize=24, alignment=1, spaceAfter=16,
    )
    cover_meta_style = ParagraphStyle(
        "RaCoverMeta", parent=styles["Normal"], fontName=JP_FONT,
        fontSize=12, alignment=1, textColor=colors.grey,
        spaceAfter=6, leading=18,
    )
    toc_title_style = ParagraphStyle(
        "RaTocTitle", parent=styles["Heading1"], fontName=JP_FONT,
        fontSize=18, alignment=1, spaceAfter=18,
        textColor=colors.HexColor("#2d6b4f"),
    )
    toc_entry_style = ParagraphStyle(
        "RaTocEntry", fontName=JP_FONT, fontSize=11,
        leftIndent=24, firstLineIndent=-12, spaceBefore=4, leading=18,
    )

    buf = io.BytesIO()

    # See _TdDoc in generate_test_density_pdf for the TOC mechanism.
    class _RaDoc(SimpleDocTemplate):
        def afterFlowable(self, flowable):  # type: ignore[override]
            if (isinstance(flowable, Paragraph)
                    and flowable.style.name == "RaH2"):
                self.notify(
                    "TOCEntry",
                    (0, flowable.getPlainText(), self.page),
                )

    doc = _RaDoc(
        buf, pagesize=page_size,
        leftMargin=1.2 * cm, rightMargin=1.2 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    story: list = []

    # --- Page 1: Cover ------------------------------------------------------
    cover_icon = Image(
        io.BytesIO(_pixel_icon_png("bronto", scale=10)),
        width=110, height=88,
    )
    cover_icon.hAlign = "CENTER"
    story.append(Spacer(1, 130))
    story.append(cover_icon)
    story.append(Spacer(1, 24))
    story.append(Paragraph(t("ra_pdf_title"), cover_title_style))
    if fid_filter_active:
        story.append(Paragraph(t("ra_pdf_filter_active"), filter_style))
    story.append(Spacer(1, 12))
    story.append(Paragraph(
        t("ra_pdf_footer",
          when=datetime.now().strftime("%Y-%m-%d %H:%M")),
        cover_meta_style,
    ))
    story.append(PageBreak())

    # --- Page 2: Table of contents ------------------------------------------
    story.append(Paragraph(t("pdf_toc_title"), toc_title_style))
    toc = TableOfContents()
    toc.levelStyles = [toc_entry_style]
    story.append(toc)
    story.append(PageBreak())

    # --- Page 3+: Content sections ------------------------------------------
    # --- Inputs ------------------------------------------------------------
    story.append(Paragraph(t("ra_pdf_h_inputs"), h2_style))
    wbs_origin = (
        st.session_state.get("origin_names", {}).get("wbs") or "—"
    )
    defects_origin = (
        st.session_state.get("origin_names", {}).get("defects") or "—"
    )
    tests_origin = (
        st.session_state.get("origin_names", {}).get("tests") or "—"
    )
    inp_rows = [
        [
            Paragraph(t("ra_pdf_input_wbs"), body_style),
            Paragraph(
                f"{wbs_origin}<br/>"
                f"<font size=8 color='grey'>{t('ra_pdf_input_wbs_note')}</font>",
                body_style,
            ),
        ],
        [
            Paragraph(t("ra_pdf_input_defects"), body_style),
            Paragraph(
                f"{defects_origin}<br/>"
                f"<font size=8 color='grey'>"
                f"{t('ra_pdf_input_defects_note')}</font>",
                body_style,
            ),
        ],
        [
            Paragraph(t("ra_pdf_input_tests"), body_style),
            Paragraph(
                f"{tests_origin}<br/>"
                f"<font size=8 color='grey'>"
                f"{t('ra_pdf_input_tests_note')}</font>",
                body_style,
            ),
        ],
    ]
    inp_tbl = Table(inp_rows, colWidths=[inner_w * 0.30, inner_w * 0.70])
    inp_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), JP_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#e8efe8")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
    ]))
    story.append(inp_tbl)

    # --- Rules --------------------------------------------------------------
    story.append(Paragraph(t("ra_pdf_h_rules"), h2_style))
    story.append(Paragraph(t("ra_pdf_rules_body"), body_style))
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"・{t('ra_pdf_rule_dev')}", body_style))
    story.append(Paragraph(f"・{t('ra_pdf_rule_test_spec')}", body_style))
    story.append(Paragraph(f"・{t('ra_pdf_rule_test_exec')}", body_style))

    # --- Build role_df for all downstream views ----------------------------
    role_df = _extract_role_assignments(wbs_df)
    if role_df.empty:
        story.append(Spacer(1, 10))
        story.append(Paragraph(t("role_analytics_no_matches"), body_style))
        _footer_cb = _pdf_apply_chrome(story, styles, JP_FONT)
        doc.multiBuild(story, onFirstPage=_footer_cb,
                       onLaterPages=_footer_cb)
        pdf = buf.getvalue()
        buf.close()
        return pdf

    # --- View 1 table -------------------------------------------------------
    story.append(Paragraph(t("ra_pdf_h_view1"), h2_style))
    ft = _build_feature_role_table(role_df, kpi_df, defects_df)
    if ft.empty:
        story.append(Paragraph(t("ra_pdf_no_data"), caption_style))
    else:
        header1 = [
            t("col_feature"),
            t("role_dev"), t("role_test_spec"), t("role_test_exec"),
            t("col_test_run_rate"), t("col_test_density"),
            t("col_defect_total"), t("col_incident_rate"), t("col_test_ng"),
            t("col_top3_problems"),
        ]
        rows1 = [[Paragraph(h, small_body) for h in header1]]
        for _, r in ft.iterrows():
            def _pct_or_dash(v):
                vv = pd.to_numeric(v, errors="coerce")
                return "—" if pd.isna(vv) else f"{vv * 100:.1f}%"

            def _f2_or_dash(v):
                vv = pd.to_numeric(v, errors="coerce")
                return "—" if pd.isna(vv) else f"{vv:.2f}"

            def _int_or_dash(v):
                vv = pd.to_numeric(v, errors="coerce")
                return "—" if pd.isna(vv) else f"{int(vv)}"
            rows1.append([
                Paragraph(str(r.get("display") or ""), small_body),
                Paragraph(str(r.get("dev") or ""), small_body),
                Paragraph(str(r.get("test_spec") or ""), small_body),
                Paragraph(str(r.get("test_exec") or ""), small_body),
                Paragraph(_pct_or_dash(r.get("test_run_rate")), small_body),
                Paragraph(_f2_or_dash(r.get("test_density")), small_body),
                Paragraph(_int_or_dash(r.get("defect_total")), small_body),
                Paragraph(_pct_or_dash(r.get("incident_rate")), small_body),
                Paragraph(_int_or_dash(r.get("NG")), small_body),
                Paragraph(str(r.get("top3_problems") or "—"), small_body),
            ])
        col_ws = [
            inner_w * 0.14, inner_w * 0.09, inner_w * 0.09, inner_w * 0.09,
            inner_w * 0.07, inner_w * 0.07, inner_w * 0.06,
            inner_w * 0.07, inner_w * 0.06, inner_w * 0.26,
        ]
        v1_tbl = Table(rows1, colWidths=col_ws, repeatRows=1)
        v1_tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), JP_FONT),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef4ec")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ("ALIGN", (4, 1), (8, -1), "RIGHT"),
        ]))
        story.append(v1_tbl)

    # --- View 2 table -------------------------------------------------------
    story.append(Paragraph(t("ra_pdf_h_view2"), h2_style))
    asum = _build_assignee_summary(role_df, kpi_df, defects_df)
    if asum.empty:
        story.append(Paragraph(t("ra_pdf_no_data"), caption_style))
    else:
        header2 = [
            t("col_assignee"),
            t("role_count_dev"), t("role_count_test_spec"),
            t("role_count_test_exec"),
            t("col_feature_count"), t("col_defect_total"),
            t("col_avg_incident_rate"), t("col_top3_problems"),
        ]
        rows2 = [[Paragraph(h, small_body) for h in header2]]
        for _, r in asum.iterrows():
            air = pd.to_numeric(r.get("avg_incident_rate"), errors="coerce")
            rows2.append([
                Paragraph(str(r.get("担当者") or ""), small_body),
                Paragraph(f"{int(r.get('dev', 0) or 0)}", small_body),
                Paragraph(f"{int(r.get('test_spec', 0) or 0)}", small_body),
                Paragraph(f"{int(r.get('test_exec', 0) or 0)}", small_body),
                Paragraph(
                    f"{int(r.get('feature_count', 0) or 0)}", small_body),
                Paragraph(
                    f"{int(r.get('defect_total', 0) or 0)}", small_body),
                Paragraph(
                    "—" if pd.isna(air) else f"{air * 100:.1f}%", small_body),
                Paragraph(str(r.get("top3_problems") or "—"), small_body),
            ])
        v2_ws = [
            inner_w * 0.14, inner_w * 0.08, inner_w * 0.09, inner_w * 0.09,
            inner_w * 0.08, inner_w * 0.08, inner_w * 0.11, inner_w * 0.33,
        ]
        v2_tbl = Table(rows2, colWidths=v2_ws, repeatRows=1)
        v2_tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), JP_FONT),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef4ec")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ("ALIGN", (1, 1), (-2, -1), "RIGHT"),
        ]))
        story.append(v2_tbl)

    # --- Bubble map ---------------------------------------------------------
    bubble_df = _build_assignee_bubble_df(role_df, kpi_df, defects_df)
    if not bubble_df.empty:
        buckets = _split_bubble_by_signal(bubble_df)
        story.append(PageBreak())
        story.append(Paragraph(t("ra_pdf_h_bubble"), h2_style))
        story.append(Paragraph(
            _md_bold_html(t("role_analytics_bubble_caption")),
            caption_style))
        story.append(Spacer(1, 6))
        result = _mpl_chart_assignee_bubble(buckets["main"])
        if result is not None:
            png, w_px, h_px = result
            aspect = h_px / w_px if w_px else 0.5
            disp_w = inner_w
            disp_h = min(disp_w * aspect, 17 * cm)
            story.append(Image(
                io.BytesIO(png), width=disp_w, height=disp_h))
            story.append(Spacer(1, 4))
            # Brief explanation of how the "dominant role" colour is
            # picked so PDF readers aren't left guessing at the legend.
            # ReportLab's Paragraph parser speaks a tiny HTML subset
            # but not markdown, so convert **...** → <b>...</b>.
            _note = _md_bold_html(
                t("role_analytics_dominant_role_note")
            ).replace("›", "＞")
            story.append(Paragraph(_note, caption_style))
        else:
            story.append(Paragraph(t("ra_pdf_no_data"), caption_style))
        # Watch-lists: each bucket gets its own heading + definition
        # caption + table (or "該当者なし" info row) so the PDF carries
        # the same reasoning the on-screen section provides.
        _append_watchlist_pdf_section(
            story, buckets["zero_defect"],
            title_key="role_analytics_watch_zero_defect_title",
            caption_key="role_analytics_watch_zero_defect_caption",
            h2_style=h2_style, caption_style=caption_style,
            small_body=small_body, inner_w=inner_w,
        )
        _append_watchlist_pdf_section(
            story, buckets["no_exec"],
            title_key="role_analytics_watch_no_exec_title",
            caption_key="role_analytics_watch_no_exec_caption",
            h2_style=h2_style, caption_style=caption_style,
            small_body=small_body, inner_w=inner_w,
            hide_defect_col=True,
        )

    # --- Problem-class strip (includes raw counts per segment) -------------
    # Heatmap section was removed in v1.0.43 — the strip now carries the
    # same (assignee × 問題分類) counts in-bar, so the separate heatmap
    # was pure redundancy.
    strip = _build_assignee_problem_share_df(role_df, defects_df)
    if strip is not None:
        story.append(Paragraph(t("ra_pdf_h_strip"), h2_style))
        story.append(Paragraph(
            t("role_analytics_strip_caption"), caption_style))
        story.append(Spacer(1, 4))
        result = _mpl_chart_assignee_problem_strip(strip)
        if result is not None:
            png, w_px, h_px = result
            aspect = h_px / w_px if w_px else 0.5
            disp_w = inner_w
            disp_h = min(disp_w * aspect, 15 * cm)
            story.append(Image(
                io.BytesIO(png), width=disp_w, height=disp_h))

    # --- Footer accent (era icons) -----------------------------------------
    story.append(Spacer(1, 18))
    footer_icons = []
    for _name in ("volcano", "fossil", "palm", "egg", "footprint", "fern"):
        footer_icons.append(Image(
            io.BytesIO(_pixel_icon_png(_name, scale=4)),
            width=14, height=14,
        ))
    footer_cells = []
    for i, ic in enumerate(footer_icons):
        if i:
            footer_cells.append("")
        footer_cells.append(ic)
    footer_tbl = Table(
        [footer_cells],
        colWidths=[14 if c == "" else 16 for c in footer_cells],
    )
    footer_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    footer_tbl.hAlign = "CENTER"
    story.append(footer_tbl)

    _footer_cb = _pdf_apply_chrome(story, styles, JP_FONT)
    doc.multiBuild(story, onFirstPage=_footer_cb, onLaterPages=_footer_cb)
    pdf = buf.getvalue()
    buf.close()
    return pdf


def _render_role_analytics_header(
    kpi_df: pd.DataFrame,
    wbs_df: Optional[pd.DataFrame],
    defects_df: Optional[pd.DataFrame],
    role_df: pd.DataFrame,
) -> None:
    """Section header for 担当者×ロール分析 with an inline 📄 PDF button.

    Mirrors the test-density header pattern: invalidates cached bytes via a
    fingerprint over (FID filter, lang, role_df shape/content) so the
    download always matches what's on screen."""
    sig: tuple = (
        tuple(_get_global_fids()),
        st.session_state.get("lang", "ja"),
        int(len(role_df)),
        int(len(wbs_df)) if wbs_df is not None else 0,
        int(len(defects_df)) if defects_df is not None else 0,
        int(len(kpi_df)),
    )
    # `bool()` coerces the `and` short-circuit: when ra_pdf_bytes is None
    # (first run) the expression would otherwise be None, which is invalid
    # as a disabled= value on st.button (protobuf rejects non-bool).
    have_fresh = bool(
        st.session_state.get("ra_pdf_bytes")
        and st.session_state.get("ra_pdf_sig") == sig
    )

    dino_name = CHART_DINOS.get("role_analytics_title", _DEFAULT_SECTION_DINO)
    icon_uri = dino_data_uri(dino_name)
    icon_col, title_col, btn_col = st.columns(
        [1, 19, 5], gap="small", vertical_alignment="center",
    )
    with icon_col:
        st.markdown(
            f'<img src="{icon_uri}" alt="{dino_name}" '
            'style="width:36px;height:36px;display:block;margin:0 auto;" />',
            unsafe_allow_html=True,
        )
    ra_anchor = "sec-role_analytics_title"
    with title_col:
        st.subheader(t("role_analytics_title"),
                     help=t("help_role_analytics"),
                     anchor=ra_anchor)
    _register_toc_entry(ra_anchor, t("role_analytics_title"))
    # Unified PDF controls: icon-only 📄 (generate) + ⬇ (download), same
    # pattern as the test-density header above. Generation wrapped in
    # st.spinner so the user sees motion during the synchronous build.
    with btn_col:
        if role_df is None or role_df.empty:
            return
        gen_slot, dl_slot = st.columns([1, 1], gap="small")
        with gen_slot:
            if st.button(
                "", icon=":material/picture_as_pdf:",
                key="ra_pdf_generate",
                help=t("ra_pdf_btn_generate_help"),
                use_container_width=True,
                disabled=have_fresh,
            ):
                try:
                    with st.spinner(t("pdf_generating")):
                        pdf_bytes = generate_role_analytics_pdf(
                            kpi_df, wbs_df, defects_df,
                            fid_filter_active=bool(_get_global_fids()),
                        )
                    st.session_state.ra_pdf_bytes = pdf_bytes
                    st.session_state.ra_pdf_sig = sig
                    st.rerun()
                except Exception as exc:
                    _get_logger().exception(
                        f"[ra_pdf] build failed: {exc}"
                    )
                    st.error(t("pdf_error", err=str(exc)))
        with dl_slot:
            if have_fresh:
                fname = (
                    "role_analytics_report_"
                    f"{date.today().strftime('%Y%m%d')}.pdf"
                )
                st.download_button(
                    label="", icon=":material/download:",
                    data=st.session_state.ra_pdf_bytes,
                    file_name=fname,
                    mime="application/pdf",
                    key="ra_pdf_download",
                    help=t("pdf_btn_download_short"),
                    use_container_width=True,
                )


@st.dialog(" ")  # title set via inner markdown so we can include the emoji
def _open_pdf_dialog(kpi_df: pd.DataFrame) -> None:
    """Two-stage modal for PDF export, implemented without st.rerun()
    because calling rerun from inside an @st.dialog body triggers the
    'Could not find fragment with id ...' error in Streamlit 1.39.

    Flow: render stage 1 (selection) inside an st.empty() container.
    When the user clicks Start generation AND has a non-empty selection,
    clear the slot and render stage 2 (runner + download) in its place,
    all inside the same Python run. No rerun, no fragment id issues."""
    # Build the picker options once.
    opts_df = (kpi_df[["機能ID", "機能名称"]]
               .drop_duplicates(subset=["機能ID"])
               .fillna({"機能名称": ""}))
    label_to_fid: dict[str, str] = {}
    labels: list[str] = []
    for _, r in opts_df.iterrows():
        lab = _label_id_name(r)
        label_to_fid[lab] = str(r["機能ID"])
        labels.append(lab)

    body = st.empty()

    # --- Stage 1: selection -------------------------------------------------
    with body.container():
        st.markdown(
            f"<div style='font-weight:700;font-size:16px;margin:-4px 0 2px;'>"
            f"{t('pdf_select_title')}</div>"
            f"<div style='font-size:12px;color:#aaa;margin:0 0 14px;'>"
            f"{t('pdf_select_caption')}</div>",
            unsafe_allow_html=True,
        )
        chosen = st.multiselect(
            t("pdf_select_label"),
            options=labels,
            max_selections=30,
            key="pdf_fid_multiselect",
        )
        st.caption(t("pdf_select_count", n=len(chosen)))
        err_slot = st.empty()
        proceed = st.button(
            t("pdf_btn_confirm"), type="primary",
            key="pdf_confirm_generate", use_container_width=True,
        )

    if not proceed:
        return
    if not chosen:
        err_slot.error(t("pdf_select_error_empty"))
        return

    # --- Stage 2: generate (replaces stage 1 in `body`) ---------------------
    selected_fids = [label_to_fid[c] for c in chosen]
    # Filter kpi_df + defects_df so every chart sees only the chosen rows.
    # The per-chart _BAR_CHART_MAX_ROWS safety cap will not trigger here
    # because user selection is already ≤ 30.
    kdf = kpi_df[kpi_df["機能ID"].isin(selected_fids)].copy()
    defects_src = st.session_state.dfs.get("defects")
    ddf = (defects_src[defects_src["機能ID"].isin(selected_fids)].copy()
           if defects_src is not None else None)
    _get_logger().info(
        f"[pdf_export] stage-2 enter: {len(selected_fids)} fids selected, "
        f"kpi_rows={len(kdf)} defect_rows={0 if ddf is None else len(ddf)}"
    )

    body.empty()
    with body.container():
        st.markdown(
            f"<div style='font-weight:700;font-size:16px;margin:-4px 0 2px;'>"
            f"{t('pdf_dialog_title')}</div>"
            f"<div style='font-size:12px;color:#aaa;margin:0 0 10px;'>"
            f"{t('pdf_dialog_subtitle')}"
            f" · {t('pdf_select_count', n=len(selected_fids))}</div>",
            unsafe_allow_html=True,
        )
        slot = st.empty()
        result_slot = st.empty()
        try:
            def _cb(msg: str, step: int, total: int) -> None:
                slot.markdown(
                    _render_pdf_runner_html(step, total, msg),
                    unsafe_allow_html=True,
                )
            pdf_bytes = generate_report_pdf(
                kdf, progress_cb=_cb, defects_df=ddf)
            st.session_state.report_pdf = pdf_bytes
            st.session_state.report_pdf_lang = st.session_state.lang
            slot.markdown(
                _render_pdf_runner_html(
                    PDF_TOTAL_STEPS, PDF_TOTAL_STEPS,
                    t("pdf_done"), done=True),
                unsafe_allow_html=True,
            )
            with result_slot.container():
                lang_tag = st.session_state.lang
                fname = (
                    f"dashboard4dx_report_{date.today().strftime('%Y%m%d')}"
                    f"_{lang_tag}.pdf"
                )
                st.download_button(
                    label="📄 " + t("pdf_btn_download"),
                    data=pdf_bytes,
                    file_name=fname,
                    mime="application/pdf",
                    key="pdf_download_dialog",
                    use_container_width=True,
                )
            st.toast(t("pdf_done"), icon="📄")
        except Exception as exc:
            slot.empty()
            st.session_state.pop("report_pdf", None)
            detail = log_error(
                category="pdf_export",
                summary=str(exc),
                exc=exc,
                context={
                    "lang": st.session_state.get("lang"),
                    "rows": int(len(kdf)),
                    "selected_fids": ",".join(selected_fids)[:200],
                },
            )
            with result_slot.container():
                st.error(t("pdf_error", err=exc))
                with st.expander(t("log_show_detail"), expanded=False):
                    st.code(detail, language="text")


_DEFECT_CLASS_TOP_N = 3


def _render_defect_class_breakdown(defects_df: Optional[pd.DataFrame]
                                   ) -> None:
    """Section: Function-ID-filterable Redmine 問題分類 pie + Top-N table.

    Filter comes from the sidebar's global Function ID selection — pie
    reflects every Redmine fault row when it is empty, and narrows to
    the selected subset when any IDs are chosen (re-percentages within
    that subset).
    """
    if defects_df is None or defects_df.empty:
        return
    if "問題分類" not in defects_df.columns:
        return
    section_header("chart_defect_class", "help_chart_defect_class")
    selected = _get_global_fids()
    df = defects_df
    if selected:
        df = defects_df[defects_df["機能ID"].astype(str).isin(selected)]
    if df.empty:
        st.info(t("chart_defect_class_empty"))
        return
    counts = _defect_class_counts(df)
    if counts.empty:
        st.info(t("chart_defect_class_no_class"))
        return
    n_total = int(counts.sum())
    st.caption(
        f"{t('chart_defect_class_scope_prefix')} "
        f"{n_total} {t('chart_defect_class_scope_suffix')}"
    )
    fig = _chart_defect_class(df)
    if fig is not None:
        st.plotly_chart(fig, use_container_width=True)
    top_df = counts.head(_DEFECT_CLASS_TOP_N).reset_index()
    top_df.columns = [t("chart_defect_class_col_class"),
                      t("chart_defect_class_col_count")]
    top_df.insert(0, "#", range(1, len(top_df) + 1))
    pct_col = t("chart_defect_class_col_share")
    top_df[pct_col] = (counts.head(_DEFECT_CLASS_TOP_N).values / n_total
                       * 100.0)
    st.markdown(f"**{t('chart_defect_class_top_title', n=_DEFECT_CLASS_TOP_N)}**")
    st.dataframe(
        top_df, use_container_width=True, hide_index=True,
        column_config={
            pct_col: st.column_config.NumberColumn(
                pct_col, format="%.1f%%"
            ),
        },
    )


def _render_role_analytics(kpi_df: pd.DataFrame) -> None:
    """Section: cross-analysis views tying WBS sub-task assignees (N column
    on rows marked ● in L) to Redmine defects + test-spec NG.

    Rendered on the Charts tab. Everything honours the global FID filter by
    re-slicing wbs/defects to match — kpi_df is already filtered by the
    caller. A PDF button at the top of the section exports View 1/2 as
    tables plus the bubble map / heat-strip / View-3 heatmap as PNGs."""
    wbs_df = st.session_state.dfs.get("wbs")
    if wbs_df is None or wbs_df.empty:
        return
    if "is_subtask" not in wbs_df.columns:
        return
    if not wbs_df["is_subtask"].fillna(False).astype(bool).any():
        return

    selected_fids = _get_global_fids()
    if selected_fids:
        wbs_df = wbs_df[wbs_df["機能ID"].astype(str).isin(selected_fids)]

    defects_df = st.session_state.dfs.get("defects")
    if (defects_df is not None and not defects_df.empty
            and selected_fids):
        defects_df = defects_df[
            defects_df["機能ID"].astype(str).isin(selected_fids)
        ]

    role_df = _extract_role_assignments(wbs_df)
    _render_role_analytics_header(kpi_df, wbs_df, defects_df, role_df)
    if role_df.empty:
        st.caption(t("role_analytics_no_matches"))
        return

    role_labels = {
        "dev":       t("role_dev"),
        "test_spec": t("role_test_spec"),
        "test_exec": t("role_test_exec"),
    }
    role_count_labels = {
        "dev":       t("role_count_dev"),
        "test_spec": t("role_count_test_spec"),
        "test_exec": t("role_count_test_exec"),
    }
    role_count_helps = {
        "dev":       t("help_role_count_dev"),
        "test_spec": t("help_role_count_test_spec"),
        "test_exec": t("help_role_count_test_exec"),
    }
    role_assignee_helps = {
        "dev":       t("help_role_assignees_dev"),
        "test_spec": t("help_role_assignees_test_spec"),
        "test_exec": t("help_role_assignees_test_exec"),
    }

    def _bar_col(label: str, series: "pd.Series", fmt: str,
                 max_value: Optional[float] = None, help: Optional[str] = None):
        s = pd.to_numeric(series, errors="coerce")
        hi = (float(max_value) if max_value is not None
              else (float(s.max()) if s.notna().any() else 0.0))
        if hi and hi > 0:
            return st.column_config.ProgressColumn(
                label, format=fmt, min_value=0, max_value=hi, help=help,
            )
        return st.column_config.NumberColumn(label, format=fmt, help=help)

    # ----- View 1: Feature × assignee-by-role + KPIs + top-3 問題分類 -----
    ft = _build_feature_role_table(role_df, kpi_df, defects_df)
    if not ft.empty:
        st.markdown(f"**{t('role_analytics_view1_title')}**")
        st.caption(t("role_analytics_view1_caption"))
        ft_disp = ft.copy()
        if "test_run_rate" in ft_disp.columns:
            ft_disp["test_run_rate"] = (
                pd.to_numeric(ft_disp["test_run_rate"], errors="coerce")
                * 100.0
            )
        if "incident_rate" in ft_disp.columns:
            ft_disp["incident_rate"] = (
                pd.to_numeric(ft_disp["incident_rate"], errors="coerce")
                * 100.0
            )
        cols_v1 = ["display"] + list(ROLE_KEYWORDS.keys()) + [
            c for c in ("test_run_rate", "test_density", "defect_total",
                        "incident_rate", "NG", "top3_problems")
            if c in ft_disp.columns
        ]
        col_config_v1: dict = {
            "display": st.column_config.TextColumn(
                t("col_feature"), help=t("help_col_feature")),
            **{rk: st.column_config.TextColumn(
                role_labels[rk], help=role_assignee_helps[rk])
               for rk in ROLE_KEYWORDS},
        }
        if "test_run_rate" in ft_disp.columns:
            col_config_v1["test_run_rate"] = _bar_col(
                t("col_test_run_rate"), ft_disp["test_run_rate"],
                "%.1f%%", max_value=100.0, help=t("help_test_run_rate"))
        if "test_density" in ft_disp.columns:
            col_config_v1["test_density"] = _bar_col(
                t("col_test_density"), ft_disp["test_density"], "%.2f",
                help=t("help_test_density"))
        if "defect_total" in ft_disp.columns:
            col_config_v1["defect_total"] = _bar_col(
                t("col_defect_total"), ft_disp["defect_total"], "%d",
                help=t("help_defect_total"))
        if "incident_rate" in ft_disp.columns:
            col_config_v1["incident_rate"] = _bar_col(
                t("col_incident_rate"), ft_disp["incident_rate"],
                "%.1f%%", max_value=100.0, help=t("help_incident_rate"))
        if "NG" in ft_disp.columns:
            col_config_v1["NG"] = _bar_col(
                t("col_test_ng"), ft_disp["NG"], "%d",
                help=t("help_test_ng"))
        if "top3_problems" in ft_disp.columns:
            col_config_v1["top3_problems"] = st.column_config.TextColumn(
                t("col_top3_problems"),
                help=t("help_col_feature_top3_problems"))
        st.dataframe(
            ft_disp[cols_v1], use_container_width=True, hide_index=True,
            column_config=col_config_v1,
        )

    # ----- View 2: Assignee summary -----
    asum = _build_assignee_summary(role_df, kpi_df, defects_df)
    if not asum.empty:
        st.markdown(f"**{t('role_analytics_view2_title')}**")
        st.caption(t("role_analytics_view2_caption"))
        asum_disp = asum.copy()
        if "avg_incident_rate" in asum_disp.columns:
            asum_disp["avg_incident_rate"] = (
                pd.to_numeric(asum_disp["avg_incident_rate"],
                              errors="coerce") * 100.0
            )
        cols_v2 = ["担当者"] + list(ROLE_KEYWORDS.keys()) + [
            c for c in ("feature_count", "defect_total",
                        "avg_incident_rate", "top3_problems")
            if c in asum_disp.columns
        ]
        col_config_v2: dict = {
            "担当者": st.column_config.TextColumn(
                t("col_assignee"), help=t("help_col_assignee")),
            **{rk: _bar_col(role_count_labels[rk], asum_disp[rk],
                            "%d", help=role_count_helps[rk])
               for rk in ROLE_KEYWORDS if rk in asum_disp.columns},
            "top3_problems": st.column_config.TextColumn(
                t("col_top3_problems"), help=t("help_col_top3_problems")),
        }
        if "feature_count" in asum_disp.columns:
            col_config_v2["feature_count"] = _bar_col(
                t("col_feature_count"), asum_disp["feature_count"], "%d",
                help=t("help_col_feature_count"))
        if "defect_total" in asum_disp.columns:
            col_config_v2["defect_total"] = _bar_col(
                t("col_defect_total"), asum_disp["defect_total"], "%d",
                help=t("help_defect_total"))
        if "avg_incident_rate" in asum_disp.columns:
            col_config_v2["avg_incident_rate"] = _bar_col(
                t("col_avg_incident_rate"),
                asum_disp["avg_incident_rate"], "%.1f%%",
                max_value=100.0, help=t("help_col_avg_incident_rate"))
        st.dataframe(
            asum_disp[cols_v2], use_container_width=True, hide_index=True,
            column_config=col_config_v2,
        )

    # ----- Advanced viz 1/2: 案A 担当者バブルマップ -----
    bubble_df = _build_assignee_bubble_df(role_df, kpi_df, defects_df)
    if not bubble_df.empty:
        buckets = _split_bubble_by_signal(bubble_df)
        st.markdown(f"**{t('role_analytics_bubble_title')}**")
        st.caption(t("role_analytics_bubble_caption"))
        fig_b = _chart_assignee_bubble(buckets["main"])
        if fig_b is not None:
            st.plotly_chart(fig_b, use_container_width=True)
            st.caption(
                t("role_analytics_dominant_role_note")
                + "  \n" + t("role_analytics_dominant_role_hover_hint")
            )
        _render_bubble_watchlist(
            buckets["zero_defect"],
            title_key="role_analytics_watch_zero_defect_title",
            caption_key="role_analytics_watch_zero_defect_caption",
        )
        _render_bubble_watchlist(
            buckets["no_exec"],
            title_key="role_analytics_watch_no_exec_title",
            caption_key="role_analytics_watch_no_exec_caption",
            hide_defect_col=True,
        )

    # ----- Advanced viz 2/2: 案C 問題分類ストリップ (件数埋め込み) -----
    # Strip now embeds `N件` inside each segment and shows
    # "N件 (X.X%)" on hover, which makes the previous View 3 heatmap
    # redundant — it's been removed both on-screen and from the PDF.
    strip = _build_assignee_problem_share_df(role_df, defects_df)
    if strip is not None:
        st.markdown(f"**{t('role_analytics_strip_title')}**")
        st.caption(t("role_analytics_strip_caption"))
        fig_s = _chart_assignee_problem_strip(strip)
        if fig_s is not None:
            st.plotly_chart(fig_s, use_container_width=True)


def _render_charts_toc(items: list[tuple[str, str]]) -> None:
    """Render the Charts-tab TOC chip row. Each chip is an anchor link to
    the section's HTML id — Streamlit's st.subheader(anchor=...) puts the
    id on the heading element, so clicking scrolls the Streamlit app
    container directly to the section.

    The outer div carries id="charts-top" so the back-to-top button's
    `href="#charts-top"` resolves to this row's position (≈ top of tab)."""
    if not items:
        return
    jump_label = t("toc_jump_label")
    # Deduplicate while preserving first-seen order — several paths may
    # accidentally register the same anchor twice (e.g. a conditional
    # section_header() fallback).
    seen: set[str] = set()
    uniq: list[tuple[str, str]] = []
    for anchor, label in items:
        if anchor in seen:
            continue
        seen.add(anchor)
        uniq.append((anchor, label))
    chips_html = "".join(
        f'<a class="d4dx-toc-chip" href="#{anchor}">{label}</a>'
        for anchor, label in uniq
    )
    st.markdown(f"""
<style>
.d4dx-toc {{
  display: flex;
  flex-wrap: wrap;
  gap: 6px 8px;
  align-items: center;
  padding: 10px 12px;
  margin: 0 0 16px;
  background: rgba(58,168,114,0.06);
  border: 1px solid rgba(58,168,114,0.25);
  border-radius: 8px;
  font-size: 12px;
}}
.d4dx-toc-label {{
  font-weight: 600;
  color: #3aa872;
  margin-right: 4px;
  white-space: nowrap;
}}
.d4dx-toc-chip {{
  display: inline-block;
  padding: 3px 10px;
  background: rgba(255,255,255,0.04);
  border: 1px solid rgba(128,128,128,0.35);
  border-radius: 999px;
  color: inherit !important;
  text-decoration: none !important;
  font-size: 12px;
  line-height: 1.4;
  white-space: nowrap;
  transition: background .12s, border-color .12s, color .12s;
}}
.d4dx-toc-chip:hover {{
  background: rgba(58,168,114,0.18);
  border-color: #3aa872;
  color: #3aa872 !important;
}}
</style>
<div id="charts-top" class="d4dx-toc">
  <span class="d4dx-toc-label">🧭 {jump_label}:</span>
  {chips_html}
</div>
""", unsafe_allow_html=True)


def _render_back_to_top_button() -> None:
    """Fixed-position circular button anchored to the bottom-right of the
    viewport. href points at the Charts-tab TOC wrapper so the browser's
    native fragment-scroll behaviour does the work — no JS needed."""
    label = t("toc_back_to_top")
    st.markdown(f"""
<style>
.d4dx-backtotop {{
  position: fixed;
  right: 22px;
  bottom: 22px;
  width: 44px;
  height: 44px;
  border-radius: 50%;
  background: #3aa872;
  color: #ffffff !important;
  display: flex;
  align-items: center;
  justify-content: center;
  font-size: 20px;
  font-weight: 700;
  text-decoration: none !important;
  box-shadow: 0 2px 10px rgba(0,0,0,0.35);
  z-index: 9999;
  transition: transform .15s, background .15s;
}}
.d4dx-backtotop:hover {{
  transform: translateY(-2px);
  background: #45c28b;
}}
</style>
<a class="d4dx-backtotop" href="#charts-top" title="{label}"
   aria-label="{label}">⬆</a>
""", unsafe_allow_html=True)


def _render_overview_compare(kpi_df: pd.DataFrame) -> None:
    """Section: 機能ID-filterable KPI cards + 4-metric comparison chart.

    Filter is empty by default — KPI cards then show the totals across
    every Function ID in the master. Selecting one or more FIDs narrows
    both the cards and the small-multiples bar chart below.
    """
    available = [(c, lbl, color)
                 for c, lbl, color in _OVERVIEW_COMPARE_METRICS
                 if c in kpi_df.columns]
    if not available:
        return
    section_header("chart_overview_compare", "help_chart_overview_compare")
    selected = _get_global_fids()
    df_scope = kpi_df.copy()
    if selected:
        df_scope = df_scope[df_scope["機能ID"].astype(str).isin(selected)]
    if df_scope.empty:
        st.info(t("chart_overview_compare_empty"))
        return
    # Aggregate to one row per 機能ID *just for the card metrics*, so
    # duplicate (機能ID, 機能名称) pairs in the master don't double-count
    # their joined LoC / tests / defect counts. The chart builder below
    # receives the pre-groupby df so it can preserve 機能名称 on its own
    # groupby and render 機能ID：機能名 on the Y axis.
    grp_cols = [c for c, _, _ in available]
    df_cards = (df_scope.groupby("機能ID", as_index=False)
                .agg(**{c: (c, "mean") for c in grp_cols}))
    cards = st.columns(len(available), gap="small")
    n_fids = len(df_cards)
    fids_help = f"{n_fids} {t('chart_overview_compare_fids_suffix')}"
    total_prefix = t("chart_overview_compare_total_prefix")
    for (col, lbl, _color), card in zip(available, cards):
        s = pd.to_numeric(df_cards[col], errors="coerce").dropna()
        v = float(s.sum()) if len(s) else None
        card.metric(f"{total_prefix} {lbl}",
                    f"{int(v):,}" if v is not None else "—",
                    help=fids_help)
    fig = _chart_overview_compare(df_scope)
    if fig is not None:
        st.plotly_chart(fig, use_container_width=True)


def _render_test_density_section_header(kpi_df: pd.DataFrame) -> None:
    """Header row for the 機能ID別テスト密度 chart with an inline PDF button.

    The button generates a standalone single-topic PDF report on click; the
    resulting bytes are cached in session_state (keyed by the FID filter +
    threshold + a fingerprint of `test_density`), and replaced with a
    Download button on the next rerun. Invalidates automatically when the
    global FID filter, threshold, or underlying data changes.
    """
    sig: tuple = (
        tuple(_get_global_fids()),
        _test_density_threshold(),
        int(len(kpi_df)),
        (float(kpi_df["test_density"].fillna(0).sum())
         if "test_density" in kpi_df.columns else 0.0),
        st.session_state.get("lang", "ja"),
    )
    # See note in _render_role_analytics_header — bool() coerces the
    # `and` short-circuit so disabled= stays a real bool on first render.
    have_fresh = bool(
        st.session_state.get("td_pdf_bytes")
        and st.session_state.get("td_pdf_sig") == sig
    )

    dino_name = CHART_DINOS.get("chart_test_density", _DEFAULT_SECTION_DINO)
    icon_uri = dino_data_uri(dino_name)
    icon_col, title_col, btn_col = st.columns(
        [1, 19, 5], gap="small", vertical_alignment="center",
    )
    with icon_col:
        st.markdown(
            f'<img src="{icon_uri}" alt="{dino_name}" '
            'style="width:36px;height:36px;display:block;margin:0 auto;" />',
            unsafe_allow_html=True,
        )
    td_anchor = "sec-chart_test_density"
    with title_col:
        st.subheader(t("chart_test_density"),
                     help=t("help_chart_test_density"),
                     anchor=td_anchor)
    _register_toc_entry(td_anchor, t("chart_test_density"))
    # Unified PDF controls: one icon-only 📄 button (generation) that
    # swaps for an icon-only ⬇ button (download) after the bytes are in
    # session state. Generation itself is wrapped in st.spinner so the
    # user sees a "generating…" motion in place of the button during the
    # synchronous build.
    with btn_col:
        gen_slot, dl_slot = st.columns([1, 1], gap="small")
        with gen_slot:
            if st.button(
                "", icon=":material/picture_as_pdf:",
                key="td_pdf_generate",
                help=t("td_pdf_btn_generate_help"),
                use_container_width=True,
                disabled=have_fresh,
            ):
                try:
                    with st.spinner(t("pdf_generating")):
                        pdf_bytes = generate_test_density_pdf(
                            kpi_df,
                            fid_filter_active=bool(_get_global_fids()),
                        )
                    st.session_state.td_pdf_bytes = pdf_bytes
                    st.session_state.td_pdf_sig = sig
                    st.rerun()
                except Exception as exc:
                    _get_logger().exception(
                        f"[td_pdf] build failed: {exc}"
                    )
                    st.error(t("pdf_error", err=str(exc)))
        with dl_slot:
            if have_fresh:
                fname = (
                    "test_density_report_"
                    f"{date.today().strftime('%Y%m%d')}.pdf"
                )
                st.download_button(
                    label="", icon=":material/download:",
                    data=st.session_state.td_pdf_bytes,
                    file_name=fname,
                    mime="application/pdf",
                    key="td_pdf_download",
                    help=t("pdf_btn_download_short"),
                    use_container_width=True,
                )


def render_charts_tab() -> None:
    """Tab — visualizations of the current KPI dataframe and time-series
    derived from saved snapshots in input/."""
    kpi_df = get_current_kpi_df()
    if kpi_df is None:
        st.info(t("charts_needs_master"))
        return

    # ----- PDF export controls (top of tab, right-aligned) ----------------
    # Unified pattern across every PDF surface: icon-only 📄 button opens
    # the feature picker + progress dialog; once bytes are in session
    # state, a second ⬇ icon appears next to it for download. Both icons
    # live in a small column at the right edge of the toolbar.
    _, pdf_gen_col, pdf_dl_col = st.columns([11, 1, 1], gap="small")
    with pdf_gen_col:
        if st.button(
            "", icon=":material/picture_as_pdf:",
            key="pdf_generate",
            help=t("pdf_btn_generate_short"),
            use_container_width=True,
        ):
            _open_pdf_dialog(kpi_df)
    with pdf_dl_col:
        if st.session_state.get("report_pdf"):
            lang_tag = st.session_state.get("report_pdf_lang",
                                            st.session_state.lang)
            fname = (
                f"dashboard4dx_report_{date.today().strftime('%Y%m%d')}"
                f"_{lang_tag}.pdf"
            )
            st.download_button(
                label="", icon=":material/download:",
                data=st.session_state.report_pdf,
                file_name=fname,
                mime="application/pdf",
                key="pdf_download",
                help=t("pdf_btn_download_short"),
                use_container_width=True,
            )

    # ----- In-tab navigation (TOC chip row + floating back-to-top) -------
    # Placeholder is declared at the top of the DOM; we populate it at the
    # end of this function after every section_header() has contributed
    # its (anchor, label) to _CHARTS_TOC_ACTIVE. That way sections that
    # don't render (missing data, snapshot-only charts, etc.) are
    # automatically absent from the chip row — no manual pre-check.
    toc_slot = st.empty()
    global _CHARTS_TOC_ACTIVE
    _CHARTS_TOC_ACTIVE = []

    # All charts are now produced by the shared `_chart_*` builders above so
    # the on-screen Charts tab and the PDF report stay in lock-step. Each
    # builder already sets a sensible margin + `automargin=True` so long
    # Function-ID labels (e.g. "USER010 · Profile Edit (Admin)") don't get
    # clipped at typical viewport widths.
    _render_overview_compare(kpi_df)

    fig = _chart_progress_gap(kpi_df)
    if fig is not None:
        section_header("chart_progress_gap", "help_chart_progress_gap")
        st.plotly_chart(fig, use_container_width=True)

    fig = _chart_test_coverage(kpi_df)
    if fig is not None:
        section_header("chart_test_coverage", "help_chart_test_coverage")
        st.plotly_chart(fig, use_container_width=True)

    fig = _chart_test_density(kpi_df)
    if fig is not None:
        _render_test_density_section_header(kpi_df)
        st.plotly_chart(fig, use_container_width=True)

    fig = _chart_incident_rate(kpi_df)
    if fig is not None:
        section_header("chart_incident_rate", "help_chart_incident_rate")
        st.plotly_chart(fig, use_container_width=True)

    col1, col2 = st.columns(2, gap="medium")
    fig = _chart_loc_vs_ng(kpi_df)
    if fig is not None:
        with col1:
            section_header("chart_loc_vs_ng", "help_chart_loc_vs_ng")
            st.caption(t("chart_loc_vs_ng_sub"))
            st.plotly_chart(fig, use_container_width=True)

    with col2:
        section_header("chart_design_impl_gap", "help_chart_design_impl_gap")
        fig = _chart_design_impl_gap(kpi_df)
        if fig is None:
            st.caption(t("chart_no_design_pages"))
        else:
            st.plotly_chart(fig, use_container_width=True)

    fig = _chart_risk_heatmap(kpi_df)
    if fig is not None:
        section_header("chart_risk_heatmap", "help_chart_risk_heatmap")
        st.plotly_chart(fig, use_container_width=True)
        # Per-row legend (Plotly axis labels can't carry tooltips themselves).
        with st.expander(t("chart_risk_dims_legend")):
            risk_dims = [c for c in
                         ["bug_density", "incident_rate",
                          "delay_rate", "test_run_rate", "test_density"]
                         if c in kpi_df.columns]
            for c in risk_dims:
                st.markdown(f"- {t(COLUMN_HELP_KEYS[c])}")

    code_snaps = load_all_snapshots_for_slot("code", load_code_counts)
    if len(code_snaps) >= 2:
        section_header("chart_loc_trend", "help_chart_loc_trend")
        st.plotly_chart(_chart_loc_trend(), use_container_width=True)
    elif len(code_snaps) == 1:
        section_header("chart_loc_trend", "help_chart_loc_trend")
        st.caption(t("chart_no_history"))

    test_snaps = load_all_snapshots_for_slot("tests", load_test_counts)
    if len(test_snaps) >= 2:
        section_header("chart_test_trend", "help_chart_test_trend")
        st.plotly_chart(_chart_test_trend(), use_container_width=True)
    elif len(test_snaps) == 1:
        section_header("chart_test_trend", "help_chart_test_trend")
        st.caption(t("chart_no_history"))

    defects_df = st.session_state.dfs.get("defects")
    fig = _chart_bug_trend(defects_df)
    if fig is not None:
        section_header("chart_bug_trend", "help_chart_bug_trend")
        st.plotly_chart(fig, use_container_width=True)
    elif defects_df is None or defects_df.empty:
        section_header("chart_bug_trend", "help_chart_bug_trend")
        st.caption(t("chart_no_defects"))

    _render_defect_class_breakdown(defects_df)

    _render_role_analytics(kpi_df)

    # ----- Finalize in-tab navigation -------------------------------------
    # Populate the placeholder declared near the top with whatever
    # sections actually rendered, then emit a fixed-position back-to-top
    # button. Reset the collector so other render paths don't accidentally
    # inherit a stale list.
    toc_items = list(_CHARTS_TOC_ACTIVE or [])
    _CHARTS_TOC_ACTIVE = None
    with toc_slot.container():
        _render_charts_toc(toc_items)
    _render_back_to_top_button()


_CALENDAR_CSS = """
/* FullCalendar's default styling assumes a light page; force inheritance so
   it shows up correctly on Streamlit's dark theme too. */
.fc {
    color: inherit !important;
    font-family: inherit;
}
.fc-toolbar-title {
    color: inherit !important;
    font-size: 1.15em !important;
}
.fc-col-header-cell-cushion,
.fc-daygrid-day-number,
.fc-list-day-cushion a,
.fc-list-event-title a {
    color: inherit !important;
    text-decoration: none !important;
}
.fc-button {
    background-color: rgba(128,128,128,0.18) !important;
    border-color: rgba(128,128,128,0.35) !important;
    color: inherit !important;
    text-transform: none !important;
}
.fc-button:hover { background-color: rgba(128,128,128,0.32) !important; }
.fc-button-active,
.fc-button-primary:not(:disabled).fc-button-active {
    background-color: #4ec78a !important;
    border-color: #4ec78a !important;
    color: #0b0b0b !important;
}
.fc-day-today { background-color: rgba(78,199,138,0.10) !important; }
.fc th, .fc td { border-color: rgba(128,128,128,0.25) !important; }
.fc-list-day-cushion { background-color: rgba(128,128,128,0.15) !important; }
.fc-list-event:hover td { background-color: rgba(128,128,128,0.10) !important; }
.fc-event { cursor: default; }
"""


def _label_fid_name(fid, name) -> str:
    """Format a Function-ID / name pair as '機能ID：機能名' for display.

    The 機能名 source of truth is the Function master (機能ID一覧); callers
    should resolve it from the master-derived kpi_df or the session-state
    master df before calling. Falls back to just the ID when no name is
    available so chart axes never render a dangling '：'.
    """
    fid_s = "" if fid is None else str(fid)
    name_s = "" if name is None else str(name)
    try:
        if pd.isna(name):
            name_s = ""
    except (TypeError, ValueError):
        pass
    return f"{fid_s}：{name_s}" if name_s else fid_s


def _label_id_name(row) -> str:
    return _label_fid_name(row.get("機能ID"), row.get("機能名称"))


def _feature_display_series(df: pd.DataFrame) -> pd.Series:
    """Vectorized 'FID：Name' label series for per-Function-ID charts.
    Rows with a missing 機能名称 fall back to the ID alone. Callers
    typically chain `.map(_clip_label)` for horizontal-bar axis labels."""
    fids = df["機能ID"].astype(str)
    names = df["機能名称"].fillna("").astype(str) if "機能名称" in df.columns \
        else pd.Series([""] * len(df), index=df.index)
    return fids + names.map(lambda n: f"：{n}" if n else "")


def _master_fid_name_map() -> dict[str, str]:
    """FID → 機能名 lookup built from the in-session Function master. Empty
    dict if no master is loaded. Used by charts/events whose source table
    doesn't already carry 機能名称 (e.g. Redmine defect rows)."""
    master = st.session_state.get("dfs", {}).get("master")
    if master is None or master.empty or "機能ID" not in master.columns:
        return {}
    m = master.drop_duplicates(subset=["機能ID"])
    names = m["機能名称"] if "機能名称" in m.columns \
        else pd.Series([""] * len(m), index=m.index)
    return {str(fid): ("" if pd.isna(nm) else str(nm))
            for fid, nm in zip(m["機能ID"], names)}


def render_calendar_tab() -> None:
    """Tab — Gantt chart + FullCalendar view of WBS schedule and defects."""
    kpi_df = get_current_kpi_df()
    if kpi_df is None:
        st.info(t("calendar_needs_master"))
        return

    section_header("calendar_title", "help_calendar_title")
    st.caption(t("calendar_caption"))

    selected_fids = _get_global_fids()

    layer_cols = st.columns(6)
    with layer_cols[0]:
        show_planned = st.checkbox(t("calendar_layer_planned"), value=True,
                                   key="cal_layer_planned")
    with layer_cols[1]:
        show_actual = st.checkbox(t("calendar_layer_actual"), value=True,
                                  key="cal_layer_actual")
    with layer_cols[2]:
        show_defects = st.checkbox(t("calendar_layer_defects"), value=True,
                                   key="cal_layer_defects")
    with layer_cols[3]:
        show_subtasks = st.checkbox(t("calendar_layer_subtasks"), value=False,
                                    key="cal_layer_subtasks")
    with layer_cols[4]:
        show_events = st.checkbox(t("calendar_layer_events"), value=True,
                                  key="cal_layer_events")
    with layer_cols[5]:
        show_nonwork = st.checkbox(t("calendar_layer_nonwork"), value=True,
                                   key="cal_layer_nonwork")

    if selected_fids:
        kpi_df = kpi_df[kpi_df["機能ID"].astype(str).isin(selected_fids)].copy()

    sub_by_fid: dict[str, pd.DataFrame] = {}
    if show_subtasks:
        wbs_df = st.session_state.dfs.get("wbs")
        if (wbs_df is not None and not wbs_df.empty
                and "is_subtask" in wbs_df.columns):
            sdf = wbs_df[wbs_df["is_subtask"].fillna(False).astype(bool)].copy()
            if selected_fids:
                sdf = sdf[sdf["機能ID"].astype(str).isin(selected_fids)]
            for fid, grp in sdf.groupby("機能ID"):
                sub_by_fid[str(fid)] = grp

    today_d = date.today()

    # ----- Gantt chart (Plotly timeline) -------------------------------------
    gantt_rows: list[dict] = []
    label_planned = t("calendar_layer_planned")
    label_actual = t("calendar_layer_actual")

    def _append_schedule_bars(label: str, ps, pe, ase, aee) -> None:
        if show_planned and ps and pe and pe >= ps:
            gantt_rows.append({
                "ID": label,
                "Start": pd.Timestamp(ps),
                "End": pd.Timestamp(pe + timedelta(days=1)),
                "Layer": label_planned,
            })
        if show_actual and ase:
            end = aee if aee else today_d
            if end < ase:
                end = ase
            gantt_rows.append({
                "ID": label,
                "Start": pd.Timestamp(ase),
                "End": pd.Timestamp(end + timedelta(days=1)),
                "Layer": label_actual,
            })

    for _, row in kpi_df.iterrows():
        fid = str(row.get("機能ID", ""))
        parent_label = _label_id_name(row)
        _append_schedule_bars(
            parent_label,
            _to_pydate(row.get("planned_start")),
            _to_pydate(row.get("planned_end")),
            _to_pydate(row.get("actual_start")),
            _to_pydate(row.get("actual_end")),
        )
        for _, srow in sub_by_fid.get(fid, pd.DataFrame()).iterrows():
            sub_label = f"　└ {fid} · {srow.get('task_label', '')}"
            _append_schedule_bars(
                sub_label,
                _to_pydate(srow.get("planned_start")),
                _to_pydate(srow.get("planned_end")),
                _to_pydate(srow.get("actual_start")),
                _to_pydate(srow.get("actual_end")),
            )

    if gantt_rows:
        section_header("gantt_title", "help_gantt_title")
        df_g = pd.DataFrame(gantt_rows)
        # Preserve insertion order so each parent's sub-tasks sit directly
        # below the parent row instead of being alphabetized by Plotly.
        y_order = list(dict.fromkeys(df_g["ID"]))
        fig = px.timeline(
            df_g, x_start="Start", x_end="End", y="ID", color="Layer",
            color_discrete_map={
                label_planned: "#9aa0a6",
                label_actual:  "#4ec78a",
            },
            category_orders={"ID": y_order},
        )
        fig.update_yaxes(autorange="reversed")
        # Today marker — `add_vline(annotation_text=...)` with a Timestamp x
        # crashes Plotly (it does `sum()` on the X list internally), so draw
        # the line and the label as separate primitives.
        today_ts = pd.Timestamp(today_d)
        fig.add_vline(
            x=today_ts,
            line_width=1, line_dash="dash", line_color="#f5b400",
        )
        fig.add_annotation(
            x=today_ts, y=1, yref="paper",
            text=t("gantt_today_label"),
            showarrow=False,
            font=dict(color="#f5b400", size=11),
            yanchor="bottom",
        )
        fig.update_layout(
            height=max(320, 26 * df_g["ID"].nunique() + 80),
            margin=dict(l=10, r=10, t=50, b=10),
            xaxis_title="", yaxis_title="",
            legend_title_text="",
        )
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.caption(t("gantt_no_dates"))

    # ----- FullCalendar ------------------------------------------------------
    try:
        from streamlit_calendar import calendar
    except ImportError:
        st.error("streamlit-calendar is not installed. "
                 "Run `pip install streamlit-calendar`.")
        return

    events: list[dict] = []
    if show_planned and {"planned_start", "planned_end"}.issubset(kpi_df.columns):
        for _, row in kpi_df.iterrows():
            ps = _to_pydate(row.get("planned_start"))
            pe = _to_pydate(row.get("planned_end"))
            if ps is None or pe is None:
                continue
            events.append({
                "title": f"📅 {_label_id_name(row)}",
                "start": ps.isoformat(),
                "end": (pe + timedelta(days=1)).isoformat(),
                "backgroundColor": "rgba(150,150,150,0.35)",
                "borderColor": "#888",
                "textColor": "#ddd",
            })
    if show_actual and {"actual_start", "actual_end"}.issubset(kpi_df.columns):
        for _, row in kpi_df.iterrows():
            ase = _to_pydate(row.get("actual_start"))
            aee = _to_pydate(row.get("actual_end"))
            if ase is None:
                continue
            end = (aee or ase) + timedelta(days=1)
            events.append({
                "title": (f"✅ {_label_id_name(row)}" if aee
                          else f"▶ {_label_id_name(row)}"),
                "start": ase.isoformat(),
                "end": end.isoformat(),
                "backgroundColor": "#4ec78a",
                "borderColor": "#3aa872",
            })

    # Master-backed FID → 機能名 lookup. Used by sub-task and defect events
    # below because their source rows don't carry 機能名称 directly — the
    # authoritative name lives on the function master (機能ID一覧).
    fid_name_map = _master_fid_name_map()

    for fid, subs in sub_by_fid.items():
        feature_label = _label_fid_name(fid, fid_name_map.get(str(fid), ""))
        for _, srow in subs.iterrows():
            task = srow.get("task_label", "") or ""
            sps = _to_pydate(srow.get("planned_start"))
            spe = _to_pydate(srow.get("planned_end"))
            sase = _to_pydate(srow.get("actual_start"))
            saee = _to_pydate(srow.get("actual_end"))
            if show_planned and sps and spe:
                events.append({
                    "title": f"📅 └ {feature_label} · {task}",
                    "start": sps.isoformat(),
                    "end": (spe + timedelta(days=1)).isoformat(),
                    "backgroundColor": "rgba(150,150,150,0.25)",
                    "borderColor": "#888",
                    "textColor": "#bbb",
                })
            if show_actual and sase:
                s_end = (saee or sase) + timedelta(days=1)
                events.append({
                    "title": (f"✅ └ {feature_label} · {task}" if saee
                              else f"▶ └ {feature_label} · {task}"),
                    "start": sase.isoformat(),
                    "end": s_end.isoformat(),
                    "backgroundColor": "rgba(78,199,138,0.55)",
                    "borderColor": "#3aa872",
                })

    defects_df = st.session_state.dfs.get("defects")
    if show_defects and defects_df is not None and not defects_df.empty:
        d_iter = defects_df
        if selected_fids:
            d_iter = defects_df[
                defects_df["機能ID"].astype(str).isin(selected_fids)
            ]
        for _, row in d_iter.iterrows():
            sd = _to_pydate(row.get("実開始日"))
            ed = _to_pydate(row.get("実終了日"))
            if sd is None:
                continue
            unresolved = bool(row.get("unresolved", False))
            color = "#f05050" if unresolved else "#9aa0a6"
            end = ((ed or sd) + timedelta(days=1)).isoformat()
            fid_s = str(row.get("機能ID", "") or "")
            feature_label = _label_fid_name(fid_s, fid_name_map.get(fid_s, ""))
            category = row.get("問題分類", "") or ""
            title = (f"🐞 {feature_label} · {category}" if category
                     else f"🐞 {feature_label}")
            events.append({
                "title": title,
                "start": sd.isoformat(),
                "end": end,
                "backgroundColor": color,
                "borderColor": color,
            })

    # ----- Imported calendar entries (global events + 非稼働日) ------------
    # `calendar` is the optional new source (Dashboard upload card). Each
    # row is already discriminated by `kind`. FID filter does NOT narrow
    # these — events are global, non-working days are per-assignee.
    cal_df = st.session_state.dfs.get("calendar")
    if cal_df is not None and not cal_df.empty:
        # Deep purple for company-wide events; muted olive grey for
        # non-working days so they don't fight the WBS/defect layers.
        EV_BG = "#7c4db0"
        EV_BORDER = "#5e3a8a"
        NW_BG = "#8f9a6c"
        NW_BORDER = "#6b7552"
        if show_events:
            ev_rows = cal_df[cal_df["kind"] == "event"]
            for _, row in ev_rows.iterrows():
                sd = row.get("start_date")
                if sd is None or pd.isna(sd):
                    continue
                ed = row.get("end_date") or sd
                # FullCalendar `end` is exclusive; bump by one day so the
                # event covers its last day.
                end_iso = (ed + timedelta(days=1)).isoformat()
                title = str(row.get("title") or "").strip()
                events.append({
                    "title": f"🎉 {title}" if title else "🎉",
                    "start": sd.isoformat(),
                    "end": end_iso,
                    "backgroundColor": EV_BG,
                    "borderColor": EV_BORDER,
                    "textColor": "#ffffff",
                })
        if show_nonwork:
            nw_rows = cal_df[cal_df["kind"] == "nonwork"]
            for _, row in nw_rows.iterrows():
                sd = row.get("start_date")
                if sd is None or pd.isna(sd):
                    continue
                ed = row.get("end_date") or sd
                end_iso = (ed + timedelta(days=1)).isoformat()
                assignee = str(row.get("assignee") or "").strip() or "—"
                reason = str(row.get("title") or "").strip()
                # Herb-leaf glyph per the agreed "olive-like" styling.
                title = (f"🌿 {assignee}  ({reason})" if reason
                         else f"🌿 {assignee}")
                events.append({
                    "title": title,
                    "start": sd.isoformat(),
                    "end": end_iso,
                    "backgroundColor": NW_BG,
                    "borderColor": NW_BORDER,
                    "textColor": "#ffffff",
                })

    st.markdown(f"### {t('calendar_section')}")
    st.caption(t("calendar_event_count", n=len(events)))

    if not events:
        st.info(t("calendar_no_events"))
        return

    # Default the initial view to today when it falls inside the allowed
    # range; otherwise clamp to the nearest bound. Since the calendar
    # template now ships with 75 Japanese holidays starting 2024-01-01,
    # falling back to `min(events)` would permanently open on 2024-01
    # regardless of the user's current date — not useful.
    today_iso = today_d.isoformat()
    if CAL_DISPLAY_START <= today_d <= CAL_DISPLAY_END:
        initial_date = today_iso
    elif today_d < CAL_DISPLAY_START:
        initial_date = CAL_DISPLAY_START.isoformat()
    else:
        initial_date = CAL_DISPLAY_END.isoformat()

    options = {
        "initialView": "dayGridMonth",
        "initialDate": initial_date,
        # Keep navigation bounded to the supported 4-year window (see
        # CAL_DISPLAY_START / CAL_DISPLAY_END). FullCalendar treats the
        # `end` as exclusive, so advance by one day to include Dec 31.
        "validRange": {
            "start": CAL_DISPLAY_START.isoformat(),
            "end":   (CAL_DISPLAY_END + timedelta(days=1)).isoformat(),
        },
        "headerToolbar": {
            "left": "prev,next today",
            "center": "title",
            "right": "dayGridMonth,timeGridWeek,listMonth",
        },
        "height": 720,
        # Cap events per day cell; the rest collapse into a "+N more"
        # link so the day number + today-highlight never get buried by
        # stacked bars when the dataset is dense. dayMaxEvents (count)
        # renders fine — dayMaxEventRows (vertical fit) was the one
        # that caused the blank-iframe bug noted below.
        "dayMaxEvents": 3,
        "moreLinkClick": "popover",
    }
    # Custom CSS injected INTO the FullCalendar iframe: strengthens the
    # day-number badge and today-highlight so they stay readable even
    # with a dense event stack above them. Keeping this list short —
    # earlier combinations of dayMaxEventRows + custom_css caused the
    # inner calendar to render blank on streamlit-calendar 1.3.1.
    calendar_css = """
/* Day number: fixed-position badge so it sits above event chips. */
.fc .fc-daygrid-day-top {
  z-index: 5; position: relative;
}
.fc .fc-daygrid-day-number {
  font-weight: 600; color: #eaeaea; padding: 2px 6px;
  border-radius: 4px;
  background: rgba(0, 0, 0, 0.45);
}
/* Today's cell: amber tint + amber ring on the day badge so it
   stands out even when stacked events overlay the cell. */
.fc .fc-day-today {
  background-color: rgba(245, 180, 0, 0.18) !important;
}
.fc .fc-day-today .fc-daygrid-day-number {
  background: #f5b400; color: #1a1a1a;
  box-shadow: 0 0 0 2px rgba(245, 180, 0, 0.5);
}
/* Event rows: slight transparency so anything beneath (today's
   cell tint, grid lines) is still perceivable. */
.fc .fc-daygrid-event {
  opacity: 0.9;
}
/* "+N more" link: readable when events overflow. */
.fc .fc-more-link {
  color: #f5b400; font-weight: 600;
}
"""
    # streamlit-calendar passes events via FullCalendar's `initialEvents`,
    # which is only honored on first mount — toggling filters would otherwise
    # leave the calendar stuck on the old event list. Hash the event payload
    # into the widget key so content changes force a fresh mount.
    cal_key = "project_calendar_" + hashlib.md5(
        repr(sorted(
            (e["title"], e["start"], e["end"]) for e in events
        )).encode()
    ).hexdigest()[:12]
    calendar(events=events, options=options, custom_css=calendar_css,
             key=cal_key)


def render_design_pages_tab() -> None:
    """Tab 2 — manual page-count editor. Auto-saves to disk on each edit;
    syncs its row list with the current Function master."""
    master = st.session_state.dfs.get("master")
    if master is None or master.empty:
        st.info(t("design_needs_master"))
        return

    st.subheader(t("sec2_title"))
    st.caption(t("sec2_caption"))

    # Single source of truth across reruns. Loaded from disk once per session,
    # then mutated in-memory and mirrored to disk on changes. Reading from
    # disk on every render caused data_editor's first-keystroke edit to be
    # dropped because the widget saw the data argument flip mid-update.
    # If the user reset auto-load in Settings, this session starts empty.
    if "design_pages_state" not in st.session_state:
        st.session_state.design_pages_state = (
            {} if st.session_state.get("skip_design_pages_load")
            else load_design_pages()
        )
    state: dict[str, int] = st.session_state.design_pages_state

    unique_fids = sorted(master["機能ID"].unique())
    # Pull 機能名称 from the master (first non-null per ID). Joined back onto
    # the editor so users can see the feature name alongside its page count
    # without bouncing to the 機能ID一覧.
    name_map: dict[str, str] = {}
    if "機能名称" in master.columns:
        name_series = (
            master.dropna(subset=["機能名称"])
                  .groupby("機能ID")["機能名称"]
                  .first()
        )
        name_map = {
            str(k): str(v) for k, v in name_series.items() if str(v).strip()
        }
    # float64 + NaN is much more reliable inside data_editor than Int64 + NA;
    # the Int64 nullable variant requires two clicks to commit the first value
    # because the frontend treats the cell as text until it has a numeric type.
    initial_values = [
        float(state[fid]) if fid in state else float("nan")
        for fid in unique_fids
    ]
    initial_df = pd.DataFrame({
        "機能ID": unique_fids,
        "機能名称": [name_map.get(fid, "") for fid in unique_fids],
        "設計書ページ数": pd.array(initial_values, dtype="float64"),
    })
    editor_height = min(40 + 36 * len(unique_fids), 700)

    # Editor key keys on the master ID set so the widget resets cleanly when
    # the master gains/loses IDs.
    editor_key = "design_pages_editor::" + ",".join(unique_fids)

    edit_col, summary_col, _ = st.columns([2, 2, 3], gap="medium")
    with edit_col:
        edited_df = st.data_editor(
            initial_df,
            num_rows="fixed",
            use_container_width=True,
            height=editor_height,
            key=editor_key,
            column_config={
                "機能ID": st.column_config.TextColumn(
                    disabled=True, width="small",
                ),
                "機能名称": st.column_config.TextColumn(
                    disabled=True, width="medium",
                ),
                "設計書ページ数": st.column_config.NumberColumn(
                    min_value=0,
                    max_value=999_999,
                    step=1,
                    format="%.0f",     # display as integer (no decimal)
                    width="medium",
                ),
            },
        )

    # Reconcile editor state with the in-memory map; persist only on change.
    new_visible: dict[str, int] = {}
    for _, row in edited_df.iterrows():
        fid = str(row["機能ID"])
        v = row["設計書ページ数"]
        if pd.notna(v):
            new_visible[fid] = int(round(float(v)))

    old_visible = {fid: state[fid] for fid in unique_fids if fid in state}
    if new_visible != old_visible:
        for fid in unique_fids:
            if fid in new_visible:
                state[fid] = new_visible[fid]
            else:
                state.pop(fid, None)
        save_design_pages(
            set(unique_fids),
            {fid: new_visible.get(fid) for fid in unique_fids},
        )

    with summary_col:
        filled = len(new_visible)
        total = sum(new_visible.values())
        st.metric(t("sec2_filled_ids"), f"{filled} / {len(unique_fids)}")
        st.metric(t("sec2_total_pages"), f"{total:,}")
        if DESIGN_PAGES_FILE.exists():
            ts = datetime.fromtimestamp(DESIGN_PAGES_FILE.stat().st_mtime)
            st.caption(t("design_last_saved",
                         ts=ts.strftime("%Y-%m-%d %H:%M:%S")))
        else:
            st.caption(t("design_no_save_yet"))
        st.caption(t("sec2_summary_tip"))


def _reset_slot_auto_load(slot: str) -> None:
    """Stop auto-loading this slot for the rest of the session. Files on disk
    are NOT touched — they remain available for trend analysis later."""
    st.session_state.skip_auto_load[slot] = True
    st.session_state.dfs.pop(slot, None)
    st.session_state.errs.pop(slot, None)
    st.session_state.last_ok_sig.pop(slot, None)
    st.session_state.last_err_sig.pop(slot, None)
    st.session_state.setdefault("origin_names", {}).pop(slot, None)
    upload_key = f"upload_{slot}"
    if upload_key in st.session_state:
        # Clearing the file_uploader's stored value forces it back to empty.
        del st.session_state[upload_key]


def _reset_design_pages_session() -> None:
    """Drop the in-memory design-pages state for this session; the JSON file
    on disk is not touched."""
    st.session_state.skip_design_pages_load = True
    st.session_state.pop("design_pages_state", None)


def render_settings_tab() -> None:
    """Tab 3 — reset session-level auto-loading. Disk files are preserved."""
    # ----- Source uploads -----
    st.subheader(t("settings_uploads_title"))
    st.caption(t("settings_uploads_caption"))

    for spec in SOURCE_SPECS:
        slot = spec["key"]
        label = t(spec["label_key"])
        info = storage_summary_for_slot(slot)
        is_skipped = bool(st.session_state.skip_auto_load.get(slot))
        is_loaded = slot in st.session_state.dfs

        line = st.columns([8, 2])
        with line[0]:
            if info["count"] == 0:
                st.markdown(
                    f"{spec['icon']} &nbsp;**{label}** "
                    f"&nbsp;·&nbsp; _{t('settings_no_files')}_"
                )
            else:
                ts = info["last"].strftime("%Y-%m-%d %H:%M") if info["last"] else "—"
                if is_skipped:
                    status_pill = _pill("warn", t("settings_status_skipped"))
                elif is_loaded:
                    status_pill = _pill("ok", t("settings_status_loaded"))
                else:
                    status_pill = ""
                st.markdown(
                    f"{spec['icon']} &nbsp;**{label}** "
                    f"&nbsp;·&nbsp; {t('settings_files_count', n=info['count'])} "
                    f"&nbsp;·&nbsp; {_human_size(info['size'])} "
                    f"&nbsp;·&nbsp; {ts} &nbsp; {status_pill}",
                    unsafe_allow_html=True,
                )
        with line[1]:
            if info["count"] == 0:
                pass  # nothing to reset
            elif is_skipped:
                if st.button(
                    t("settings_undo_reset_btn"),
                    key=f"undo_reset_{slot}",
                    use_container_width=True,
                ):
                    st.session_state.skip_auto_load.pop(slot, None)
                    st.toast(t("settings_undo_done", label=label), icon="↩️")
                    st.rerun()
            else:
                with st.popover(
                    t("settings_reset_btn"), use_container_width=True
                ):
                    st.warning(t("settings_confirm_reset_msg"))
                    if st.button(
                        t("settings_confirm_btn"),
                        key=f"confirm_reset_{slot}",
                        type="primary",
                    ):
                        _reset_slot_auto_load(slot)
                        st.toast(t("settings_reset_done", label=label), icon="🚫")
                        st.rerun()

        if info["count"] > 0:
            with st.expander(t("settings_show_files"), expanded=False):
                for f in info["files"]:
                    rel = f.relative_to(INPUT_DIR)
                    sz = _human_size(f.stat().st_size)
                    ts = datetime.fromtimestamp(f.stat().st_mtime).strftime(
                        "%Y-%m-%d %H:%M"
                    )
                    file_col, del_col = st.columns([12, 1], gap="small",
                                                   vertical_alignment="center")
                    with file_col:
                        st.caption(f"`{rel}` · {sz} · {ts}")
                    with del_col:
                        # Two-step confirmation: open the popover, tick the
                        # acknowledgement checkbox, then press the (now-
                        # enabled) primary button. The popover closes itself
                        # on rerun — and resetting the checkbox key on every
                        # close avoids a previously-checked state silently
                        # arming a different file's delete button.
                        with st.popover(t("settings_delete_file_btn"),
                                        use_container_width=True):
                            st.warning(t("settings_confirm_delete_file_msg"))
                            st.caption(f"`{rel}`")
                            chk_key = f"chk_del_snap_{f}"
                            ack = st.checkbox(
                                t("settings_confirm_delete_check"),
                                key=chk_key,
                            )
                            if st.button(
                                t("settings_confirm_delete_btn"),
                                key=f"del_snap_{f}",
                                type="primary",
                                disabled=not ack,
                            ):
                                if delete_snapshot_file(f):
                                    st.toast(
                                        t("settings_file_deleted", file=rel.name),
                                        icon="🗑️",
                                    )
                                    # If we deleted the file the slot was
                                    # currently auto-loading, drop the cached
                                    # dataframe so the upload card re-resolves
                                    # against whatever's left on disk.
                                    st.session_state.dfs.pop(slot, None)
                                    st.session_state.setdefault(
                                        "origin_names", {}).pop(slot, None)
                                    # Clear the ack checkbox so the next time
                                    # this popover opens, it starts unticked.
                                    st.session_state.pop(chk_key, None)
                                    st.rerun()

    # ----- Design page counts -----
    st.divider()
    st.subheader(t("settings_pages_title"))
    st.caption(t("settings_pages_caption"))

    pages_skipped = bool(st.session_state.get("skip_design_pages_load"))
    on_disk = load_design_pages()
    line = st.columns([8, 2])
    with line[0]:
        if not on_disk:
            st.markdown(f"_{t('settings_no_files')}_")
        else:
            ts_str = "—"
            if DESIGN_PAGES_FILE.exists():
                ts_str = datetime.fromtimestamp(
                    DESIGN_PAGES_FILE.stat().st_mtime
                ).strftime("%Y-%m-%d %H:%M")
            status_pill = (
                _pill("warn", t("settings_status_skipped"))
                if pages_skipped else _pill("ok", t("settings_status_loaded"))
            )
            st.markdown(
                f"📝 &nbsp;**{t('sec2_title')}** &nbsp;·&nbsp; "
                f"{t('settings_count_pages', n=len(on_disk))} &nbsp;·&nbsp; "
                f"{ts_str} &nbsp; {status_pill}",
                unsafe_allow_html=True,
            )
    with line[1]:
        if not on_disk:
            pass
        elif pages_skipped:
            if st.button(
                t("settings_undo_reset_btn"),
                key="undo_reset_pages",
                use_container_width=True,
            ):
                st.session_state.pop("skip_design_pages_load", None)
                st.session_state.pop("design_pages_state", None)
                st.toast(
                    t("settings_undo_done", label=t("sec2_title")),
                    icon="↩️",
                )
                st.rerun()
        else:
            with st.popover(
                t("settings_reset_btn"), use_container_width=True
            ):
                st.warning(t("settings_confirm_pages_reset_msg"))
                if st.button(
                    t("settings_confirm_btn"),
                    key="confirm_reset_pages",
                    type="primary",
                ):
                    _reset_design_pages_session()
                    st.toast(
                        t("settings_reset_done", label=t("sec2_title")),
                        icon="🚫",
                    )
                    st.rerun()

    if on_disk:
        with st.expander(t("settings_show_entries"), expanded=False):
            entries_df = pd.DataFrame(
                sorted(on_disk.items()),
                columns=["機能ID", "設計書ページ数"],
            )
            entries_df["設計書ページ数"] = entries_df["設計書ページ数"].astype(int)
            # Constrain width — 2 narrow columns shouldn't stretch full-page.
            preview_col, _ = st.columns([2, 5], gap="medium")
            with preview_col:
                preview_height = min(40 + 36 * len(entries_df), 600)
                st.dataframe(
                    entries_df,
                    use_container_width=True,
                    hide_index=True,
                    height=preview_height,
                )

    # ----- WBS parsing behavior -----
    st.divider()
    st.subheader(t("settings_wbs_title"))
    st.caption(t("settings_wbs_caption"))
    st.checkbox(
        t("settings_wbs_attach_after_dup"),
        key="wbs_attach_after_dup",
        help=t("settings_wbs_attach_after_dup_caption"),
        on_change=save_user_settings,
    )
    st.caption(t("settings_wbs_attach_after_dup_caption"))

    # ----- Chart thresholds -----
    st.divider()
    st.subheader(t("settings_charts_title"))
    st.caption(t("settings_charts_caption"))
    st.number_input(
        t("settings_test_density_threshold"),
        min_value=0.0, step=0.5, format="%.2f",
        key="test_density_threshold",
        help=t("settings_test_density_threshold_caption"),
        on_change=save_user_settings,
    )
    st.caption(t("settings_test_density_threshold_caption"))
    st.number_input(
        t("settings_incident_rate_threshold"),
        min_value=0.0, max_value=100.0, step=1.0, format="%.2f",
        key="incident_rate_threshold_pct",
        help=t("settings_incident_rate_threshold_caption"),
        on_change=save_user_settings,
    )
    # The setter uses the visible % field; the chart code reads the
    # underlying fraction via `incident_rate_threshold` (= pct/100).
    st.session_state.incident_rate_threshold = (
        float(st.session_state.get("incident_rate_threshold_pct",
                                   INCIDENT_RATE_THRESHOLD_DEFAULT * 100.0))
        / 100.0
    )
    st.caption(t("settings_incident_rate_threshold_caption"))

    # ----- Session log location -----
    st.divider()
    st.subheader(t("log_section_title"))
    st.caption(t("log_section_caption"))
    log_path = _get_log_file_path()
    try:
        rel_log = log_path.relative_to(SCRIPT_DIR)
    except ValueError:
        rel_log = log_path
    st.markdown(t("log_file_caption", path=str(rel_log)))


# -----------------------------------------------------------------------------
# Derived-KPI input map. Keys are kpi_df columns that are COMPUTED from other
# columns; each value is a list of (input_column, source_key) tuples naming
# the raw inputs the computation depends on. `source_key` maps to a
# human-facing source label (test spec / WBS / Redmine / code / design).
#
# Used by the drilldown to show "?" + a tooltip naming the missing inputs
# instead of a blank "—" whenever a derived metric can't be computed —
# so the user knows exactly which resource to fill in.
# -----------------------------------------------------------------------------
_DERIVED_KPI_INPUTS: dict[str, list[tuple[str, str]]] = {
    "bug_density":    [("NG", "tests"), ("LoC", "code")],
    "test_density":   [("総設定テスト数", "tests"), ("設計書ページ数", "design")],
    "complexity":     [("LoC", "code"), ("設計書ページ数", "design")],
    "test_run_rate":  [("実施済", "tests"), ("総設定テスト数", "tests")],
    "test_pass_rate": [("OK", "tests"), ("実施済", "tests")],
    "defect_rate":    [("NG", "tests"), ("総設定テスト数", "tests")],
    "incident_rate":  [("defect_total", "defects"), ("実施済", "tests")],
    "delay_days":     [("planned_end", "wbs")],
    "delay_rate":     [("planned_start", "wbs"), ("planned_end", "wbs")],
}

_SOURCE_LABEL_KEYS: dict[str, str] = {
    "tests":   "source_label_tests",
    "code":    "source_label_code",
    "design":  "source_label_design",
    "defects": "source_label_defects",
    "wbs":     "source_label_wbs",
}


def _missing_inputs_for(row, kpi_name: str) -> list[tuple[str, str]]:
    """Return the list of (column, source_key) entries whose value is blank
    for this row. Empty list means the KPI can be computed."""
    spec = _DERIVED_KPI_INPUTS.get(kpi_name)
    if not spec:
        return []
    missing = []
    for col, src in spec:
        v = row.get(col)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            missing.append((col, src))
    return missing


def _compose_missing_help(base_help: str,
                          missing: list[tuple[str, str]]) -> str:
    """Append a localized "inputs missing" block under the existing help
    tooltip text, naming each blank input + its source file."""
    if not missing:
        return base_help
    lines = [t("kpi_missing_header")]
    for col, src in missing:
        src_key = _SOURCE_LABEL_KEYS.get(src)
        src_label = t(src_key) if src_key else src
        lines.append(f"- `{col}` — {src_label}")
    block = "\n".join(lines)
    if base_help:
        return f"{base_help}\n\n---\n{block}"
    return block


def _get_global_fids() -> list[str]:
    """Return the session-wide Function ID filter selection, as strings.
    Empty list means "no filter" (every Function ID is in scope)."""
    try:
        return [str(x) for x in
                st.session_state.get("global_fid_filter", [])]
    except Exception:
        return []


def _apply_global_fid_filter(df: Optional[pd.DataFrame]
                             ) -> Optional[pd.DataFrame]:
    """Narrow a dataframe to the sidebar's Function ID selection when the
    user has picked any IDs, otherwise return the frame untouched."""
    if df is None or df.empty:
        return df
    if "機能ID" not in df.columns:
        return df
    selected = _get_global_fids()
    if not selected:
        return df
    return df[df["機能ID"].astype(str).isin(selected)]


def _render_global_fid_filter() -> None:
    """Sidebar widget that all tabs share for narrowing by Function ID.

    Rendered LAST in main() so that tabs have already populated
    `st.session_state.dfs` via auto-load; on the next rerun the widget
    sees the freshly loaded master and offers every ID as a choice.
    """
    with st.sidebar:
        st.subheader(t("global_fid_filter_title"))
        master = st.session_state.dfs.get("master")
        if master is None or master.empty:
            st.caption(t("global_fid_filter_upload_hint"))
            return
        options = sorted(str(x) for x in master["機能ID"].dropna().unique())
        st.multiselect(
            t("global_fid_filter_label"),
            options=options, default=[],
            key="global_fid_filter",
            help=t("global_fid_filter_help"),
        )
        n = len(_get_global_fids())
        if n == 0:
            st.caption(t("global_fid_filter_scope_all"))
        else:
            st.caption(t("global_fid_filter_scope_n", n=n))


def main() -> None:
    favicon_path = ensure_favicon()
    st.set_page_config(
        page_title="dashboard4dx",
        page_icon=str(favicon_path),
        layout="wide",
        initial_sidebar_state="expanded",
    )
    _inject_styles()

    # Session state init
    st.session_state.setdefault("dfs", {})
    st.session_state.setdefault("errs", {})
    st.session_state.setdefault("last_ok_sig", {})
    st.session_state.setdefault("last_err_sig", {})
    st.session_state.setdefault("skip_auto_load", {})
    st.session_state.setdefault("origin_names", {})

    # User-facing preferences survive app restarts by seeding session
    # state from input/user_settings.json. Changes get flushed back to
    # disk via save_user_settings() on_change callbacks on each widget.
    _saved = load_user_settings()
    st.session_state.setdefault(
        "lang", _saved.get("lang", DEFAULT_LANG))
    st.session_state.setdefault(
        "test_density_threshold",
        _saved.get("test_density_threshold",
                   TEST_DENSITY_THRESHOLD_DEFAULT))
    pct = _saved.get("incident_rate_threshold_pct",
                     INCIDENT_RATE_THRESHOLD_DEFAULT * 100.0)
    st.session_state.setdefault("incident_rate_threshold_pct", pct)
    st.session_state.setdefault(
        "incident_rate_threshold", float(pct) / 100.0)
    st.session_state.setdefault(
        "wbs_attach_after_dup",
        bool(_saved.get("wbs_attach_after_dup", False)))

    # --- Header row: title (left) + language switcher (right) ----------------
    title_col, lang_col = st.columns([10, 1], gap="small")
    with title_col:
        trex_uri = dino_data_uri("trex")
        # Pure-CSS toggle (checkbox hack): clicking the T-Rex flips the hidden
        # checkbox state, the speech bubble below is shown/hidden via the
        # `:checked ~` sibling combinator. No JS, no Python state — the
        # bubble persists until clicked again or the page reloads.
        st.markdown(f"""
<style>
.d4dx-title-wrap {{
  position: relative;
  display: flex;
  align-items: center;
  gap: 14px;
}}
.d4dx-trex-toggle {{
  position: absolute;
  opacity: 0;
  pointer-events: none;
  width: 0; height: 0;
}}
.d4dx-trex-label {{
  cursor: pointer;
  display: block;
  transition: transform .15s ease;
  user-select: none;
}}
.d4dx-trex-label:hover    {{ transform: translateY(-2px); }}
.d4dx-trex-label:active   {{ transform: translateY(1px); }}
.d4dx-title-h1 {{ margin: 0; line-height: 1; }}
.d4dx-trex-bubble {{
  position: absolute;
  left: 0;
  top: 100%;
  margin-top: 14px;
  background: rgba(20, 30, 40, 0.96);
  color: #f5f5f5;
  padding: 10px 16px;
  border-radius: 10px;
  font-size: 13px;
  line-height: 1.55;
  white-space: nowrap;
  opacity: 0;
  visibility: hidden;
  transform: translateY(-6px);
  transition: opacity .18s ease, transform .18s ease,
              visibility 0s linear .18s;
  box-shadow: 0 6px 20px rgba(0, 0, 0, 0.35);
  z-index: 1000;
  pointer-events: none;
}}
.d4dx-trex-bubble::before {{
  content: "";
  position: absolute;
  left: 18px;
  top: -7px;
  border-left: 7px solid transparent;
  border-right: 7px solid transparent;
  border-bottom: 7px solid rgba(20, 30, 40, 0.96);
}}
.d4dx-trex-toggle:checked ~ .d4dx-trex-bubble {{
  opacity: 1;
  visibility: visible;
  transform: translateY(0);
  transition: opacity .18s ease, transform .18s ease;
}}
.d4dx-trex-bubble strong {{
  display: block;
  font-weight: 600;
  color: #fafafa;
}}
.d4dx-trex-bubble .ver {{
  color: #9aa;
  font-size: 11px;
  letter-spacing: .04em;
  margin-top: 2px;
  display: block;
}}
</style>
<div class="d4dx-title-wrap">
  <input type="checkbox" id="d4dx-trex-toggle" class="d4dx-trex-toggle" />
  <label for="d4dx-trex-toggle" class="d4dx-trex-label" title="dashboard4dx">
    <img src="{trex_uri}" alt="dashboard4dx"
         style="height:52px;width:auto;display:block;" />
  </label>
  <h1 class="d4dx-title-h1">dashboard4dx</h1>
  <div class="d4dx-trex-bubble">
    <strong>開発者：Shin＆Shiobara</strong>
    <span class="ver">Ver1.0.84</span>
  </div>
</div>
""", unsafe_allow_html=True)
    with lang_col:
        st.radio(
            label="lang",
            options=[code for code, _ in LANG_OPTIONS],
            format_func=_lang_label,
            key="lang",
            label_visibility="collapsed",
            horizontal=True,
            on_change=save_user_settings,
        )
    st.caption(t("intro_caption"))

    # --- Top-level tabs ------------------------------------------------------
    (tab_dashboard, tab_charts, tab_calendar, tab_alert, tab_delivery,
     tab_backlog, tab_design, tab_settings) = st.tabs([
        t("main_tab_dashboard"),
        t("main_tab_charts"),
        t("main_tab_calendar"),
        t("main_tab_alert"),
        t("main_tab_delivery"),
        t("main_tab_backlog"),
        t("main_tab_design"),
        t("main_tab_settings"),
    ])
    with tab_dashboard:
        render_dashboard_tab()
    with tab_charts:
        render_charts_tab()
    with tab_calendar:
        render_calendar_tab()
    with tab_alert:
        render_alert_tab()
    with tab_delivery:
        render_delivery_tab()
    with tab_backlog:
        render_backlog_tab()
    with tab_design:
        render_design_pages_tab()
    with tab_settings:
        render_settings_tab()

    # Rendered last so the Function ID list reflects whatever auto-load just
    # populated in session_state during the tabs above.
    _render_global_fid_filter()


if __name__ == "__main__":
    main()
